from django.shortcuts import render
from collections import defaultdict

# Create your views here.
import boto3
import requests
from rest_framework import generics, serializers, status
from datetime import timedelta, time, datetime
from .models import (
    Course,
    TextLesson,
    Lesson,
    LiveSessionLesson,
    LaserCoachingSession,
    Question,
    QuizLesson,
    FeedbackLesson,
    Assessment,
    CourseEnrollment,
    Answer,
    Certificate,
    QuizLessonResponse,
    FeedbackLessonResponse,
    CourseTemplate,
    Resources,
    PdfLesson,
    File,
    DownloadableLesson,
    ThinkificLessonCompleted,
    Nudge,
    AssignmentLesson,
    AssignmentLessonResponse,
    FacilitatorLesson,
    Feedback,
    CoachingSessionsFeedbackResponse,
)
from rest_framework.response import Response
from django.shortcuts import get_object_or_404
from django.utils import timezone
from .serializers import (
    CourseSerializer,
    CourseTemplateSerializer,
    TextLessonCreateSerializer,
    TextLessonSerializer,
    LessonSerializer,
    LiveSessionSerializer,
    LiveSessionSerializerDepthOne,
    QuestionSerializer,
    QuizLessonDepthOneSerializer,
    LaserSessionSerializerDepthOne,
    LaserCoachingSessionSerializer,
    FeedbackLessonDepthOneSerializer,
    AssessmentSerializerDepthOne,
    AssessmentSerializer,
    CourseEnrollmentDepthOneSerializer,
    AnswerSerializer,
    CertificateSerializerDepthOne,
    VideoLessonSerializerDepthOne,
    ResourcesSerializer,
    PdfLessonSerializer,
    LessonUpdateSerializer,
    FileSerializer,
    DownloadableLessonSerializer,
    NudgeSerializer,
    AssignmentSerializerDepthOne,
    AssignmentResponseSerializerDepthSix,
    AssignmentResponseSerializer,
    FacilitatorSerializer,
    FeedbackDepthOneSerializer,
    LessonSerializerForLiveSessionDateTime,
)
from django_celery_beat.models import PeriodicTask, ClockedSchedule

from rest_framework.views import APIView
from api.models import User, Learner, Profile, Role, Coach, SessionRequestCaas
from schedularApi.models import (
    LiveSession,
    SchedularBatch,
    SchedularProject,
    LiveSession as LiveSessionSchedular,
)
from schedularApi.serializers import (
    LiveSessionSerializer as LiveSessionSchedularSerializer,
)
from assessmentApi.serializers import (
    AssessmentSerializerDepthOne as AssessmentModalSerializerDepthOne,
)
from assessmentApi.models import (
    Assessment as AssessmentModal,
    ParticipantResponse,
    ParticipantUniqueId,
)
from rest_framework.decorators import api_view, permission_classes
from django.db import transaction
import random
import string
import pdfkit
import os
from django.http import HttpResponse
from django.template.loader import render_to_string
import base64
from openpyxl import Workbook
from django.db.models import Max, Q
import environ
import uuid
import logging
from rest_framework.permissions import AllowAny, IsAuthenticated

from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
import pandas as pd
from io import BytesIO
from schedularApi.tasks import (
    get_file_content,
    get_file_extension,
    get_live_session_name,
    get_nudges_of_course,
)

from django.core.mail import EmailMessage
from django.conf import settings


env = environ.Env()

logging.basicConfig(level=logging.ERROR)
logger = logging.getLogger(__name__)

from schedularApi.models import CoachingSession, SchedularSessions

wkhtmltopdf_path = os.environ.get("WKHTMLTOPDF_PATH", r"/usr/local/bin/wkhtmltopdf")

pdfkit_config = pdfkit.configuration(wkhtmltopdf=f"{wkhtmltopdf_path}")

default_feedback_questions = [
    {
        "text": "How would you rate today’s session on overall effectiveness?",
        "options": [],
        "type": "rating_1_to_5",
    },
    {
        "text": "How would you rate today’s session on the training content?",
        "options": [],
        "type": "rating_1_to_5",
    },
    {
        "text": "How would you rate today’s session on the effectiveness of the trainer?",
        "options": [],
        "type": "rating_1_to_5",
    },
    {
        "text": "How would you rate today’s session on the knowledge of the trainer?",
        "options": [],
        "type": "rating_1_to_5",
    },
    {
        "text": "Please share any other thoughts you may have about the program.",
        "options": [],
        "type": "descriptive_answer",
    },
]


nps_default_feed_questions = [
    {
        "text": "How likely are you to recommend this program to a colleague?",
        "options": [],
        "type": "rating_0_to_10",
    },
    {
        "text": "How would you rate your Facilitator?",
        "options": [],
        "type": "rating_1_to_5",
    },
    {
        "text": "Which topics covered in the training did you find most valuable?",
        "options": [],
        "type": "descriptive_answer",
    },
    {
        "text": "Kindly share 2 key take aways / actions from the program.",
        "options": [],
        "type": "descriptive_answer",
    },
    {
        "text": "Kindly share suggestions on how we can enhance the program to add more value.",
        "options": [],
        "type": "descriptive_answer",
    },
    {
        "text": "How would you rate your overall experience about the program?",
        "options": [],
        "type": "rating_1_to_5",
    },
]


def add_question_to_feedback_lesson(feedback_lesson, questions):
    for question_data in questions:
        question_serializer = QuestionSerializer(data=question_data)
        if question_serializer.is_valid():
            question = question_serializer.save()
            feedback_lesson.questions.add(question)
    feedback_lesson.save()


def create_learner(learner_name, learner_email, learner_phone=None):
    try:
        with transaction.atomic():
            learner_email = learner_email.strip().lower()
            temp_password = "".join(
                random.choices(
                    string.ascii_uppercase + string.ascii_lowercase + string.digits,
                    k=8,
                )
            )
            user = User.objects.create_user(
                username=learner_email,
                password=temp_password,
                email=learner_email,
            )
            user.save()
            learner_role, created = Role.objects.get_or_create(name="learner")
            profile = Profile.objects.create(user=user)
            profile.roles.add(learner_role)
            profile.save()

            phone = learner_phone if learner_phone else None
            learner = None
            if phone:
                learner = Learner.objects.create(
                    user=profile,
                    name=learner_name,
                    email=learner_email,
                    phone=phone,
                )
            else:
                learner = Learner.objects.create(
                    user=profile,
                    name=learner_name,
                    email=learner_email,
                )

            return learner

    except Exception as e:
        return None


def create_or_get_learner(learner_data):
    try:
        # check if the same email user exists or not
        phone = learner_data.get("phone", None)
        user = User.objects.filter(username=learner_data["email"]).first()
        if user:
            if user.profile.roles.all().filter(name="learner").exists():
                learner = Learner.objects.get(user=user.profile)
                learner.name = learner_data["name"].strip()

                if learner_data["phone"]:
                    learner.phone = learner_data["phone"]

                learner.save()
                return learner
            else:
                learner_role, created = Role.objects.get_or_create(name="learner")
                learner_profile = user.profile
                learner_profile.roles.add(learner_role)
                learner_role.save()

                learner, created = Learner.objects.get_or_create(
                    user=learner_profile,
                    defaults={
                        "name": learner_data["name"],
                        "email": learner_data["email"],
                        "phone": phone,
                    },
                )
                return learner
        else:
            learner = create_learner(learner_data["name"], learner_data["email"], phone)
            return learner
    except Exception as e:
        # Handle specific exceptions or log the error
        print(f"Error processing participant: {str(e)}")


def get_feedback_lesson_name(lesson_name):
    # Trim leading and trailing whitespaces
    trimmed_string = lesson_name.strip()
    # Convert to lowercase
    lowercased_string = trimmed_string.lower()
    # Replace spaces between words with underscores
    underscored_string = "_".join(lowercased_string.split())
    return underscored_string


def get_file_name_from_url(url):
    # Split the URL by '/' to get an array of parts
    url_parts = url.split("/")

    # Get the last part of the array, which should be the full file name with extension
    full_file_name = url_parts[-1]

    # Extract the file name without the query parameters
    file_name = full_file_name.split("?")[0]

    return file_name


def download_file_response(file_url):
    try:
        response = requests.get(file_url)
        if response.status_code == 200:
            file_content = response.content
            extension = get_file_extension(file_url)
            content_type = response.headers.get(
                "Content-Type", f"application/{extension}"
            )
            file_name = get_file_name_from_url(file_url)
            file_response = HttpResponse(file_content, content_type=content_type)
            file_response["Content-Disposition"] = 'attachment; filename="{}"'.format(
                file_name
            )
            return file_response
        else:
            return HttpResponse(
                "Failed to download the file", status=response.status_code
            )
    except Exception as e:
        return HttpResponse(status=500, content=f"Error downloading file: {str(e)}")


class CourseListView(generics.ListCreateAPIView):
    permission_classes = [IsAuthenticated]
    queryset = Course.objects.all()
    serializer_class = CourseSerializer

    def perform_create(self, serializer):
        name = serializer.validated_data.get("name", None)
        serializer.save()


class CourseTemplateListView(generics.ListCreateAPIView):
    permission_classes = [IsAuthenticated]
    queryset = CourseTemplate.objects.all()
    serializer_class = CourseTemplateSerializer

    def get_queryset(self):
        queryset = CourseTemplate.objects.all()
        # Get the status from the query parameters, if provided
        status = self.request.query_params.get("status", None)
        if status:
            queryset = queryset.filter(status=status)
        return queryset

    def perform_create(self, serializer):
        name = serializer.validated_data.get("name", None)
        if name and CourseTemplate.objects.filter(name=name.strip()).exists():
            raise serializers.ValidationError(
                "Course template with this name already exists."
            )
        serializer.save()


class CourseDetailView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAuthenticated]
    queryset = Course.objects.all()
    serializer_class = CourseSerializer

    def perform_update(self, serializer):
        name = serializer.validated_data.get("name", None)
        instance = self.get_object()
        serializer.save()


class CourseTemplateDetailView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAuthenticated]
    queryset = CourseTemplate.objects.all()
    serializer_class = CourseTemplateSerializer

    def perform_update(self, serializer):
        name = serializer.validated_data.get("name", None)
        instance = self.get_object()
        if (
            name
            and CourseTemplate.objects.exclude(pk=instance.pk)
            .filter(name=name.strip())
            .exists()
        ):
            raise serializers.ValidationError(
                "Course template with this name already exists."
            )
        serializer.save()


class DuplicateCourseAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, course_id, *args, **kwargs):
        try:
            original_course = get_object_or_404(Course, pk=course_id)

            # Duplicate the course
            new_course = Course.objects.create(
                name=f"{original_course.name}",
                description=original_course.description,
                status="draft",
            )

            # Duplicate lessons
            original_lessons = Lesson.objects.filter(course=original_course)
            for original_lesson in original_lessons:
                new_lesson = None

                # Create a new lesson only if the type is 'text', 'quiz', or 'feedback'
                if original_lesson.lesson_type in ["text", "quiz", "feedback"]:
                    new_lesson = Lesson.objects.create(
                        course=new_course,
                        name=original_lesson.name,
                        status="draft",
                        lesson_type=original_lesson.lesson_type,
                        order=original_lesson.order,
                    )

                    # Duplicate specific lesson types
                    if original_lesson.lesson_type == "text":
                        TextLesson.objects.create(
                            lesson=new_lesson,
                            content=original_lesson.textlesson.content,
                        )

                    elif original_lesson.lesson_type == "quiz":
                        new_quiz_lesson = QuizLesson.objects.create(lesson=new_lesson)
                        for question in original_lesson.quizlesson.questions.all():
                            new_question = Question.objects.create(
                                text=question.text,
                                options=question.options,
                                type=question.type,
                            )
                            new_quiz_lesson.questions.add(new_question)

                    elif original_lesson.lesson_type == "feedback":
                        unique_id = uuid.uuid4()
                        new_feedback_lesson = FeedbackLesson.objects.create(
                            lesson=new_lesson, unique_id=unique_id
                        )
                        for question in original_lesson.feedbacklesson.questions.all():
                            new_question = Question.objects.create(
                                text=question.text,
                                options=question.options,
                                type=question.type,
                            )
                            new_feedback_lesson.questions.add(new_question)

            return Response(
                {"message": "Course duplicated successfully"},
                status=status.HTTP_201_CREATED,
            )

        except Exception as e:
            return Response(
                {"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class UpdateLessonOrder(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        payload = request.data

        for lesson_id, order in payload.items():
            try:
                lesson = Lesson.objects.get(pk=lesson_id)
                lesson.order = order
                lesson.save()
            except Lesson.DoesNotExist:
                return Response(
                    {"error": f"Lesson with id {lesson_id} not found"},
                    status=status.HTTP_404_NOT_FOUND,
                )

        return Response(
            {"message": "Lesson orders updated successfully"}, status=status.HTTP_200_OK
        )


class UpdateNudgesOrder(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        payload = request.data
        for nudge_id, order in payload.items():
            try:
                nudge = Nudge.objects.get(pk=nudge_id)
                nudge.order = order
                nudge.save()
            except Nudge.DoesNotExist:
                return Response(
                    {"error": f"Nudge with id {nudge_id} not found"},
                    status=status.HTTP_404_NOT_FOUND,
                )
        return Response(
            {"message": "Nudges orders updated successfully"}, status=status.HTTP_200_OK
        )


class TextLessonCreateView(generics.CreateAPIView):
    permission_classes = [IsAuthenticated]
    queryset = TextLesson.objects.all()
    serializer_class = TextLessonCreateSerializer


class TextLessonEditView(generics.RetrieveUpdateAPIView):
    permission_classes = [IsAuthenticated]
    queryset = TextLesson.objects.all()
    serializer_class = TextLessonCreateSerializer


class LessonListView(generics.ListAPIView):
    permission_classes = [IsAuthenticated]
    queryset = Lesson.objects.all()
    serializer_class = LessonSerializerForLiveSessionDateTime

    def get_queryset(self):
        # Retrieve lessons for a specific course based on the course ID in the URL
        course_id = self.kwargs.get("course_id")
        queryset = Lesson.objects.filter(course__id=course_id)
        status = self.request.query_params.get("status", None)
        if status:
            queryset = queryset.filter(status=status)

        return queryset


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_nudges_and_course(request, course_id):
    try:
        course = Course.objects.get(id=course_id)
        course_serializer = CourseSerializer(course)
        nudges = Nudge.objects.filter(course=course)
        serializer = NudgeSerializer(nudges, many=True)
        return Response({"nudges": serializer.data, "course": course_serializer.data})
    except Course.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def create_new_nudge(request):
    serializer = NudgeSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(["PUT"])
@permission_classes([IsAuthenticated])
def update_nudge(request, nudge_id):
    try:
        nudge = Nudge.objects.get(id=nudge_id)
    except Nudge.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)
    serializer = NudgeSerializer(nudge, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(["GET"])
@permission_classes([AllowAny])
def download_nudge_file(request, nudge_id):
    nudge_obj = get_object_or_404(Nudge, id=nudge_id)
    serializer = NudgeSerializer(nudge_obj)
    return download_file_response(serializer.data["file"])


@api_view(["PUT"])
@permission_classes([IsAuthenticated])
def add_nudges_date_frequency_to_course(request, course_id):
    try:
        course = Course.objects.get(id=course_id)
        nudge_start_date = request.data.get("nudge_start_date")
        nudge_frequency = request.data.get("nudge_frequency")
        existing_nudge_start_date = course.nudge_start_date
        course.nudge_start_date = nudge_start_date
        course.nudge_frequency = nudge_frequency
        course.save()
        if course.nudge_periodic_task:
            course.nudge_periodic_task.enabled = False
            course.nudge_periodic_task.save()
        desired_time = time(18, 31)
        datetime_comined = datetime.combine(
            datetime.strptime(course.nudge_start_date, "%Y-%m-%d"), desired_time
        )
        scheduled_for = datetime_comined - timedelta(days=1)
        clocked = ClockedSchedule.objects.create(clocked_time=scheduled_for)
        periodic_task = PeriodicTask.objects.create(
            name=uuid.uuid1(),
            task="schedularApi.tasks.schedule_nudges",
            args=[course.id],
            clocked=clocked,
            one_off=True,
        )
        course.nudge_periodic_task = periodic_task
        course.save()
        return Response({"message": "Updated successfully"}, status=201)
    except Course.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)


class CourseTemplateLessonListView(generics.ListAPIView):
    permission_classes = [IsAuthenticated]
    queryset = Lesson.objects.all()
    serializer_class = LessonSerializer

    def get_queryset(self):
        # Retrieve lessons for a specific course based on the course ID in the URL
        course_template_id = self.kwargs.get("course_template_id")
        queryset = Lesson.objects.filter(course_template__id=course_template_id)
        status = self.request.query_params.get("status", None)
        if status:
            queryset = queryset.filter(status=status)
        return queryset


class LessonDetailView(generics.RetrieveAPIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        lesson_id = self.kwargs.get("lesson_id", None)
        lesson_type = self.kwargs.get("lesson_type", None)

        if lesson_id is None or lesson_type is None:
            return Response(
                {"error": "Both lesson_id and type parameters are required."},
                status=400,
            )

        try:
            lesson = Lesson.objects.get(id=lesson_id, lesson_type=lesson_type)
        except Lesson.DoesNotExist:
            return Response(
                {"error": f"Lesson doesn't exists."},
                status=404,
            )

        if lesson_type == "text":
            text_lesson = TextLesson.objects.get(lesson=lesson)
            serializer = TextLessonSerializer(text_lesson)
        elif lesson_type == "live_session":
            live_session = LiveSessionLesson.objects.get(lesson=lesson)
            serializer = LiveSessionSerializer(live_session)
        elif lesson_type == "quiz":
            quiz_lesson = QuizLesson.objects.get(lesson=lesson)
            serializer = QuizLessonDepthOneSerializer(quiz_lesson)
        elif lesson_type == "laser_coaching":
            laser_coaching = LaserCoachingSession.objects.get(lesson=lesson)
            serializer = LaserSessionSerializerDepthOne(laser_coaching)
        elif lesson_type == "assessment":
            laser_coaching = Assessment.objects.get(lesson=lesson)
            serializer = AssessmentSerializerDepthOne(laser_coaching)
        elif lesson_type == "feedback":
            laser_coaching = FeedbackLesson.objects.get(lesson=lesson)
            serializer = FeedbackLessonDepthOneSerializer(laser_coaching)
        elif lesson_type == "video":
            video = VideoLesson.objects.get(lesson=lesson)
            serializer = VideoLessonSerializerDepthOne(video)
        elif lesson_type == "ppt":
            ppt = PdfLesson.objects.get(lesson=lesson)
            serializer = PdfLessonSerializer(ppt)
        elif lesson_type == "downloadable_file":
            downloadable_file_lesson = DownloadableLesson.objects.get(lesson=lesson)
            serializer = DownloadableLessonSerializer(downloadable_file_lesson)
        elif lesson_type == "assignment":
            assignment_lesson = AssignmentLesson.objects.get(lesson=lesson)
            serializer = AssignmentSerializerDepthOne(assignment_lesson)
        elif lesson_type == "facilitator":
            facilitator_lesson = FacilitatorLesson.objects.get(lesson=lesson)
            serializer = FacilitatorSerializer(facilitator_lesson)
        else:
            return Response({"error": f"Failed to get the lessons"}, status=400)

        return Response(serializer.data)


class DeleteLessonAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def delete(self, request, lesson_id):
        try:
            lesson = Lesson.objects.get(pk=lesson_id)
        except Lesson.DoesNotExist:
            return Response(
                {"error": "Lesson not found"}, status=status.HTTP_404_NOT_FOUND
            )
        course_enrollment = CourseEnrollment.objects.filter(
            completed_lessons__contains=lesson.id
        )
        for enrollment in course_enrollment:
            enrollment.completed_lessons = [
                lesson_id
                for lesson_id in enrollment.completed_lessons
                if lesson_id != lesson.id
            ]
            enrollment.save()
        lesson.delete()
        return Response(
            {"message": "Lesson deleted successfully"},
            status=status.HTTP_204_NO_CONTENT,
        )


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def create_quiz_lesson(request):
    # Deserialize the incoming data
    data = request.data
    lesson_data = data.get("lesson")
    questions_data = data.get("questions")

    # Create the Lesson
    lesson_serializer = LessonSerializer(data=lesson_data)
    if lesson_serializer.is_valid():
        lesson = lesson_serializer.save()

        # Create Questions and associate them with the Lesson
        questions = []
        for question_data in questions_data:
            question_serializer = QuestionSerializer(data=question_data)
            if question_serializer.is_valid():
                question = question_serializer.save()
                questions.append(question)
            else:
                # If any question is invalid, delete the created lesson and return an error
                lesson.delete()
                return Response(
                    question_serializer.errors, status=status.HTTP_400_BAD_REQUEST
                )

        # Create QuizLesson and associate it with the Lesson
        quiz_lesson = QuizLesson.objects.create(lesson=lesson)
        quiz_lesson.questions.set(questions)
        lesson_serializer = LessonSerializer(lesson)
        response_data = {
            "message": "Lesson and Quiz created successfully",
            "lesson": lesson_serializer.data,
        }

        return Response(
            response_data,
            status=status.HTTP_201_CREATED,
        )
    else:
        return Response(lesson_serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(["PUT"])
@permission_classes([IsAuthenticated])
def edit_quiz_lesson(request, quiz_lesson_id):
    try:
        quiz_lesson = QuizLesson.objects.get(id=quiz_lesson_id)
    except QuizLesson.DoesNotExist:
        return Response(
            {"message": "Quiz Lesson not found"}, status=status.HTTP_404_NOT_FOUND
        )

    # Deserialize the incoming data
    data = request.data
    lesson_data = data.get("lesson")
    questions_data = data.get("questions")

    # Update Lesson details
    lesson = quiz_lesson.lesson
    lesson_serializer = LessonSerializer(lesson, data=lesson_data)
    if lesson_serializer.is_valid():
        lesson = lesson_serializer.save()
    else:
        return Response(lesson_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    # Remove existing questions from QuizLesson
    quiz_lesson.questions.clear()

    # Update or create new questions and add them to the QuizLesson
    for question_data in questions_data:
        question_id = question_data.get("_id")
        if question_id:
            try:
                existing_question = Question.objects.get(id=question_id)
                question_serializer = QuestionSerializer(
                    existing_question, data=question_data
                )
                if question_serializer.is_valid():
                    question = question_serializer.save()
                    quiz_lesson.questions.add(question)
                else:
                    return Response(
                        question_serializer.errors, status=status.HTTP_400_BAD_REQUEST
                    )
            except Question.DoesNotExist:
                # If question does not exist, create a new one
                question_serializer = QuestionSerializer(data=question_data)
                if question_serializer.is_valid():
                    question = question_serializer.save()
                    quiz_lesson.questions.add(question)
                else:
                    return Response(
                        question_serializer.errors, status=status.HTTP_400_BAD_REQUEST
                    )
        else:
            # If no question ID is provided, create a new question
            question_serializer = QuestionSerializer(data=question_data)
            if question_serializer.is_valid():
                question = question_serializer.save()
                quiz_lesson.questions.add(question)
            else:
                return Response(
                    question_serializer.errors, status=status.HTTP_400_BAD_REQUEST
                )
    lesson_serializer = LessonSerializer(lesson)
    response_data = {
        "message": "Quiz Lesson updated successfully",
        "lesson": lesson_serializer.data,
    }

    return Response(response_data, status=status.HTTP_200_OK)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def create_feedback_lesson(request):
    # Deserialize the incoming data
    data = request.data
    lesson_data = data.get("lesson")
    questions_data = data.get("questions")

    # Create the Lesson
    lesson_serializer = LessonSerializer(data=lesson_data)
    if lesson_serializer.is_valid():
        lesson = lesson_serializer.save()

        # Create Questions and associate them with the Lesson
        questions = []
        for question_data in questions_data:
            question_serializer = QuestionSerializer(data=question_data)
            if question_serializer.is_valid():
                question = question_serializer.save()
                questions.append(question)
            else:
                # If any question is invalid, delete the created lesson and return an error
                lesson.delete()
                return Response(
                    question_serializer.errors, status=status.HTTP_400_BAD_REQUEST
                )

        # Create QuizLesson and associate it with the Lesson
        unique_id = uuid.uuid4()
        feedback_lesson = FeedbackLesson.objects.create(
            lesson=lesson, unique_id=unique_id
        )
        feedback_lesson.questions.set(questions)
        lesson_serializer = LessonSerializer(lesson)
        response_data = {
            "message": "Lesson and Feedback created successfully",
            "lesson": lesson_serializer.data,
        }

        return Response(
            response_data,
            status=status.HTTP_201_CREATED,
        )
    else:
        return Response(lesson_serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(["PUT"])
@permission_classes([IsAuthenticated])
def edit_feedback_lesson(request, feedback_lesson_id):
    try:
        feedback_lesson = FeedbackLesson.objects.get(id=feedback_lesson_id)
    except QuizLesson.DoesNotExist:
        return Response(
            {"message": "Feedback Lesson not found"}, status=status.HTTP_404_NOT_FOUND
        )
    feedback_lesson_response = FeedbackLessonResponse.objects.filter(
        feedback_lesson=feedback_lesson
    )
    if feedback_lesson_response:
        return Response(
            {
                "message": "Feedback editing is unavailable as responses have already been received for this lesson."
            },
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )
    # Deserialize the incoming data
    data = request.data
    lesson_data = data.get("lesson")
    questions_data = data.get("questions")

    # Update Lesson details
    lesson = feedback_lesson.lesson
    lesson_serializer = LessonSerializer(lesson, data=lesson_data)
    if lesson_serializer.is_valid():
        lesson = lesson_serializer.save()
    else:
        return Response(lesson_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    # Remove existing questions from QuizLesson
    feedback_lesson.questions.clear()

    # Update or create new questions and add them to the QuizLesson
    for question_data in questions_data:
        question_id = question_data.get("_id")
        if question_id:
            try:
                existing_question = Question.objects.get(id=question_id)
                question_serializer = QuestionSerializer(
                    existing_question, data=question_data
                )
                if question_serializer.is_valid():
                    question = question_serializer.save()
                    feedback_lesson.questions.add(question)
                else:
                    return Response(
                        question_serializer.errors, status=status.HTTP_400_BAD_REQUEST
                    )
            except Question.DoesNotExist:
                # If question does not exist, create a new one
                question_serializer = QuestionSerializer(data=question_data)
                if question_serializer.is_valid():
                    question = question_serializer.save()
                    feedback_lesson.questions.add(question)
                else:
                    return Response(
                        question_serializer.errors, status=status.HTTP_400_BAD_REQUEST
                    )
        else:
            # If no question ID is provided, create a new question
            question_serializer = QuestionSerializer(data=question_data)
            if question_serializer.is_valid():
                question = question_serializer.save()
                feedback_lesson.questions.add(question)
            else:
                return Response(
                    question_serializer.errors, status=status.HTTP_400_BAD_REQUEST
                )
    lesson_serializer = LessonSerializer(lesson)
    response_data = {
        "message": "Feedback Lesson updated successfully",
        "lesson": lesson_serializer.data,
    }

    return Response(response_data, status=status.HTTP_200_OK)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def create_lesson_with_live_session(request):
    lesson_data = request.data.get("lesson")
    live_session_data = request.data.get("live_session")

    lesson_serializer = LessonSerializer(data=lesson_data)
    lesson_serializer.is_valid(raise_exception=True)
    lesson_instance = lesson_serializer.save()  # Save Lesson

    # Update live_session_data to include lesson_id
    live_session_data["lesson"] = lesson_instance.id

    live_session_serializer = LiveSessionSerializer(data=live_session_data)
    live_session_serializer.is_valid(raise_exception=True)
    live_session_instance = (
        live_session_serializer.save()
    )  # Save LiveSession linked to Lesson

    return Response(
        {
            "lesson_id": lesson_instance.id,
            "live_session_id": live_session_instance.id,
            "message": "Lesson and LiveSession created successfully!",
        },
        status=status.HTTP_201_CREATED,
    )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_live_sessions_for_lesson(request, lesson_id, course_id):
    try:
        live_sessions = LiveSessionLesson.objects.filter(
            lesson__id=lesson_id, lesson__course__id=course_id
        )
        serializer = LiveSessionSerializerDepthOne(live_sessions, many=True)
        return Response(serializer.data)
    except LiveSessionLesson.DoesNotExist:
        return Response(status=404)


@api_view(["PUT"])
@permission_classes([IsAuthenticated])
def update_live_session(request, course_id, lesson_id):
    try:
        lesson = Lesson.objects.get(pk=lesson_id, course__id=course_id)
        live_session = LiveSessionLesson.objects.get(lesson=lesson)
    except (Lesson.DoesNotExist, LiveSessionLesson.DoesNotExist):
        return Response(
            {"message": "Live session does not exist for this lesson"},
            status=status.HTTP_404_NOT_FOUND,
        )

    if request.method == "PUT":
        lesson_data = request.data.get("lesson")
        # live_session_data = request.data.get("live_session")
        # Update Lesson instance fields
        lesson.name = lesson_data.get("name")
        lesson.status = lesson_data.get("status")
        lesson.lesson_type = lesson_data.get("lesson_type")
        lesson.save()
        # Update LiveSession instance fields
        # live_session.description = live_session_data.get("description")
        # live_session.meeting_link = live_session_data.get("meeting_link")
        # live_session.date = live_session_data.get("date")
        # live_session.start_time = live_session_data.get("start_time")
        # live_session.end_time = live_session_data.get("end_time")
        # live_session.save()
        # Serialize the updated LiveSession instance
        serializer = LiveSessionSerializer(live_session)
        return Response(serializer.data, status=status.HTTP_200_OK)

    return Response(
        {"message": "Invalid request method"}, status=status.HTTP_400_BAD_REQUEST
    )


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def create_laser_booking_lesson(request):
    lesson_data = request.data.get("lesson")
    coaching_session_data = request.data.get("laser_coaching_session")

    # Create a Lesson instance
    lesson = Lesson.objects.create(
        course_id=lesson_data["course"],
        name=lesson_data["name"],
        status=lesson_data["status"],
        lesson_type=lesson_data["lesson_type"],
        order=lesson_data["order"],
    )

    # Create a LaserCoachingSession instance associated with the created Lesson
    coaching_session = LaserCoachingSession.objects.create(
        lesson=lesson,
        description=coaching_session_data["description"],
        booking_link=coaching_session_data["booking_link"],
    )

    # Optionally, return a success response
    return Response(
        "Laser coaching lesson created successfully", status=status.HTTP_201_CREATED
    )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_laser_coaching_sessions(request, lesson_id, course_id):
    try:
        laser_sessions = LaserCoachingSession.objects.filter(
            lesson__id=lesson_id, lesson__course__id=course_id
        )
        serializer = LaserSessionSerializerDepthOne(laser_sessions, many=True)
        return Response(serializer.data)
    except LaserCoachingSession.DoesNotExist:
        return Response(status=404)


@api_view(["PUT"])
@permission_classes([IsAuthenticated])
def update_laser_coaching_session(request, course_id, lesson_id, session_id):
    try:
        lesson = Lesson.objects.get(course_id=course_id, id=lesson_id)
        coaching_session = LaserCoachingSession.objects.get(
            lesson_id=lesson_id, id=session_id
        )
    except (Lesson.DoesNotExist, LaserCoachingSession.DoesNotExist):
        return Response(
            "Lesson or Laser Coaching Session not found",
            status=status.HTTP_404_NOT_FOUND,
        )
    lesson_data = request.data.get("lesson")
    # session_data = request.data.get("laser_coaching_session")
    lesson_serializer = LessonSerializer(lesson, data=lesson_data, partial=True)
    # session_serializer = LaserCoachingSessionSerializer(
    #     coaching_session, data=session_data, partial=True
    # )

    if lesson_serializer.is_valid():
        # and session_serializer.is_valid():
        lesson_serializer.save()
        # session_serializer.save()
        return Response(
            {
                "lesson": lesson_serializer.data,
                # "laser_coaching_session": session_serializer.data,
            }
        )

    return Response(
        {
            "lesson_errors": lesson_serializer.errors,
            # "session_errors": session_serializer.errors,
        },
        status=status.HTTP_400_BAD_REQUEST,
    )


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def create_assessment_and_lesson(request):
    with transaction.atomic():
        lesson_data = request.data.get("lesson")
        lesson_serializer1 = LessonSerializer(data=lesson_data)
        lesson_serializer2 = LessonSerializer(data=lesson_data)
        if lesson_serializer1.is_valid() and lesson_serializer2.is_valid():
            lesson1 = lesson_serializer1.save()
            lesson2 = lesson_serializer2.save()

            lesson1.name = f"Pre {lesson1.name}"
            lesson2.name = f"Post {lesson2.name}"

            lesson1.save()
            lesson2.save()

            assessment1 = Assessment.objects.create(lesson=lesson1, type="pre")

            assessment2 = Assessment.objects.create(lesson=lesson2, type="post")
            return Response(
                "Assessment lesson created successfully", status=status.HTTP_201_CREATED
            )
        else:
            return Response(
                {lesson_serializer1.errors, lesson_serializer2.errors},
                status=status.HTTP_400_BAD_REQUEST,
            )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_assessment_lesson(request, lesson_id):
    try:
        assessment = Assessment.objects.filter(lesson__id=lesson_id)
        serializer = AssessmentSerializerDepthOne(assessment, many=True)
        return Response(serializer.data)
    except Assessment.DoesNotExist:
        return Response(status=404)


@api_view(["PUT"])
@permission_classes([IsAuthenticated])
def update_assessment_lesson(request, lesson_id, session_id):
    try:
        lesson = Lesson.objects.get(id=lesson_id)
        assessment = Assessment.objects.get(lesson_id=lesson_id, id=session_id)
    except (Lesson.DoesNotExist, Assessment.DoesNotExist):
        return Response(
            "Assessment lesson not found",
            status=status.HTTP_404_NOT_FOUND,
        )

    lesson_data = request.data.get("lesson")
    assessment.save()
    # Check if 'lesson' data is provided in the request and update the lesson name
    if lesson_data and "name" in lesson_data:
        lesson.name = lesson_data["name"]

    # Update other fields of the lesson if present in the request
    if lesson_data:
        lesson_serializer = LessonSerializer(lesson, data=lesson_data, partial=True)
        if lesson_serializer.is_valid():
            lesson_serializer.save()
        else:
            return Response(
                {"lesson_errors": lesson_serializer.errors},
                status=status.HTTP_400_BAD_REQUEST,
            )

    session_data = (
        {}
    )  # Empty data for assessment session as not specified in the request

    session_serializer = AssessmentSerializer(
        assessment, data=session_data, partial=True
    )

    if session_serializer.is_valid():
        session_serializer.save()
        return Response(
            {
                "lesson": LessonSerializer(lesson).data,
                "assessment_lesson": session_serializer.data,
            }
        )

    return Response(
        {
            "session_errors": session_serializer.errors,
        },
        status=status.HTTP_400_BAD_REQUEST,
    )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_course_enrollment(request, course_id, learner_id):
    try:
        course_enrollment = CourseEnrollment.objects.get(
            course__id=course_id, learner__id=learner_id
        )
        course_enrollment_serializer = CourseEnrollmentDepthOneSerializer(
            course_enrollment
        )
        lessons = Lesson.objects.filter(
            Q(course=course_enrollment.course),
            Q(status="public"),
            ~Q(lesson_type="feedback"),
        )
        lessons_serializer = LessonSerializerForLiveSessionDateTime(lessons, many=True)

        return Response(
            {
                "course_enrollment": course_enrollment_serializer.data,
                "lessons": lessons_serializer.data,
                "is_certificate_allowed": course_enrollment.is_certificate_allowed,
            }
        )
    except CourseEnrollment.DoesNotExist:
        return Response(status=404)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_course_enrollment_for_pmo_preview(request, course_id):
    try:
        course = Course.objects.get(id=course_id)
        course_serializer = CourseSerializer(course)
        lessons = Lesson.objects.filter(
            Q(course=course),
            Q(status="public"),
            ~Q(lesson_type="feedback"),
        )
        lessons_serializer = LessonSerializerForLiveSessionDateTime(lessons, many=True)
        completed_lessons = []
        return Response(
            {
                "course_enrollment": {
                    "completed_lessons": completed_lessons,
                    "course": course_serializer.data,
                },
                "lessons": lessons_serializer.data,
            }
        )
    except Exception as e:
        return Response(status=404)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_course_enrollment_for_pmo_preview_for_course_template(
    request, course_template_id
):
    try:
        course_template = CourseTemplate.objects.get(id=course_template_id)
        course_serializer = CourseTemplateSerializer(course_template)
        lessons = Lesson.objects.filter(
            Q(course_template=course_template),
            Q(status="public"),
            ~Q(lesson_type="feedback"),
        )
        lessons_serializer = LessonSerializer(lessons, many=True)
        completed_lessons = []
        return Response(
            {
                "course_enrollment": {
                    "completed_lessons": completed_lessons,
                    "course": course_serializer.data,
                },
                "lessons": lessons_serializer.data,
            }
        )
    except Course.DoesNotExist:
        return Response(status=404)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_course_enrollments_of_learner(request, learner_id):
    try:
        course_enrollments = CourseEnrollment.objects.filter(
            learner__id=learner_id, course__status="public"
        )
        res = []
        for course_enrollment in course_enrollments:
            course_enrollment_serializer = CourseEnrollmentDepthOneSerializer(
                course_enrollment
            )
            lessons = Lesson.objects.filter(
                Q(course=course_enrollment.course),
                Q(status="public"),
                ~Q(lesson_type="feedback"),
            )

            lessons_serializer = LessonSerializer(lessons, many=True)
            data = {
                "course_enrollment": course_enrollment_serializer.data,
                "lessons": lessons_serializer.data,
            }
            res.append(data)
        return Response(res)
    except CourseEnrollment.DoesNotExist:
        return Response(status=404)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def submit_quiz_answers(request, quiz_lesson_id, learner_id):
    try:
        quiz_lesson = get_object_or_404(QuizLesson, id=quiz_lesson_id)
        course_enrollment = get_object_or_404(
            CourseEnrollment, course=quiz_lesson.lesson.course, learner__id=learner_id
        )
        learner = get_object_or_404(Learner, id=learner_id)
    except (
        QuizLesson.DoesNotExist,
        CourseEnrollment.DoesNotExist,
        Learner.DoesNotExist,
    ) as e:
        return Response(
            {"error": "Failed to submit quiz."}, status=status.HTTP_404_NOT_FOUND
        )

    answers_data = request.data
    serializer = AnswerSerializer(data=answers_data, many=True)

    if serializer.is_valid():
        answers = serializer.save()
        quiz_lesson_response = QuizLessonResponse.objects.create(
            quiz_lesson=quiz_lesson, learner=learner
        )
        quiz_lesson_response.answers.set(answers)
        quiz_lesson_response.save()
        course_enrollment.completed_lessons.append(quiz_lesson.lesson.id)
        course_enrollment.save()
        return Response(
            {"detail": "Quiz submitted successfully"}, status=status.HTTP_200_OK
        )
    else:
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


def calculate_quiz_result(quiz_lesson, quiz_lesson_response):
    correct_answers = 0
    total_questions = quiz_lesson.questions.count()
    for question in quiz_lesson.questions.all():
        is_correct = False
        try:
            answer = quiz_lesson_response.answers.get(question=question)
            selected_options = answer.selected_options
        except Answer.DoesNotExist:
            selected_options = []

        if question.type == "single_correct_answer":
            correct_option = next(
                (opt["option"] for opt in question.options if opt["is_correct"]), None
            )
            is_correct = correct_option in selected_options
        elif question.type == "multiple_correct_answer":
            correct_options = [
                opt["option"] for opt in question.options if opt["is_correct"]
            ]
            is_correct = set(selected_options) == set(correct_options)
        if is_correct:
            correct_answers += 1
    return {
        "correct_answers": correct_answers,
        "total_questions": total_questions,
        "percentage": (
            round((correct_answers / total_questions) * 100)
            if total_questions > 0
            else 0
        ),
    }


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_quiz_result(request, quiz_lesson_id, learner_id):
    quiz_lesson = QuizLesson.objects.get(id=quiz_lesson_id)
    quiz_lesson_response = QuizLessonResponse.objects.get(
        learner__id=learner_id, quiz_lesson=quiz_lesson
    )
    quiz_result = calculate_quiz_result(quiz_lesson, quiz_lesson_response)
    return Response(
        {
            "correct_answers": quiz_result["correct_answers"],
            "total_questions": quiz_result["total_questions"],
        }
    )


@api_view(["POST"])
@permission_classes([AllowAny])
def submit_feedback_answers(request, feedback_lesson_id, learner_id):
    try:
        feedback_lesson = get_object_or_404(FeedbackLesson, id=feedback_lesson_id)
        learner = get_object_or_404(Learner, id=learner_id)
    except (
        FeedbackLesson.DoesNotExist,
        Learner.DoesNotExist,
    ) as e:
        return Response(
            {"error": "Failed to submit feedback."}, status=status.HTTP_404_NOT_FOUND
        )

    answers_data = request.data
    serializer = AnswerSerializer(data=answers_data, many=True)

    if serializer.is_valid():
        answers = serializer.save()
        feedback_lesson_response = FeedbackLessonResponse.objects.create(
            feedback_lesson=feedback_lesson, learner=learner
        )
        feedback_lesson_response.answers.set(answers)
        feedback_lesson_response.save()
        return Response(
            {"detail": "Feedback submitted successfully"}, status=status.HTTP_200_OK
        )
    else:
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class CertificateListAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, format=None):
        certificates = Certificate.objects.all()
        serializer = CertificateSerializerDepthOne(certificates, many=True)
        return Response(serializer.data)

    def post(self, request):
        try:
            name = request.data["name"]
            content = request.data["content"]

            if Certificate.objects.filter(name=name).exists():
                return Response(
                    {"error": f"Certificate with the name '{name}' already exists."},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                )
            new_certificate = Certificate.objects.create(name=name, content=content)
            new_certificate.save()
            return Response(
                {
                    "message": "Created Sucessfully.",
                },
                status=status.HTTP_200_OK,
            )
        except Exception as e:
            print(str(e))
            return Response(
                {"error": "Failed to create cretificate."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    def put(self, request):
        try:
            name = request.data["name"]
            content = request.data["content"]
            certificate_id = request.data.get("certificate_id")
            if (
                Certificate.objects.filter(name=name)
                .exclude(id=certificate_id)
                .exists()
            ):
                return Response(
                    {"error": f"Certificate with the name '{name}' already exists."},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                )
            certificate = Certificate.objects.get(id=certificate_id)
            certificate.name = name
            certificate.content = content
            certificate.save()
            return Response(
                {
                    "message": "Certificate updated Sucessfully.",
                },
                status=status.HTTP_200_OK,
            )
        except Exception as e:
            print(str(e))
            return Response(
                {"error": "Failed to update cretificate."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class GetFilteredCoursesForCertificate(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        certificates = Certificate.objects.all()
        all_courses = Course.objects.all()

        courses_in_certificates = set()
        for certificate in certificates:
            courses_in_certificates.update(
                certificate.courses.values_list("id", flat=True)
            )

        available_courses = all_courses.exclude(id__in=courses_in_certificates)

        serializer = CourseSerializer(available_courses, many=True)
        return Response(serializer.data)


class AssignCoursesToCertificate(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            courses = request.data.get("courses", [])
            certificate_id = request.data.get("certificate_id")

            certificate = Certificate.objects.get(id=certificate_id)

            courses_to_assign = Course.objects.filter(id__in=courses)

            certificate.courses.add(*courses_to_assign)

            serializers = CertificateSerializerDepthOne(certificate)
            return Response(
                {
                    "message": "Courses assigned successfully.",
                    "certificate_data": serializers.data,
                },
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            print(str(e))
            return Response(
                {"error": "Failed to assign courses."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class DeleteCourseFromCertificate(APIView):
    permission_classes = [IsAuthenticated]

    def delete(self, request):
        try:
            course_id = request.data.get("course_id")
            certificate_id = request.data.get("certificate_id")

            certificate = Certificate.objects.get(id=certificate_id)

            course = Course.objects.get(id=course_id)

            certificate.courses.remove(course)

            serializer = CertificateSerializerDepthOne(certificate)

            return Response(
                {
                    "message": "Course removed successfully.",
                    "certificate_data": serializer.data,
                },
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            print(str(e))
            return Response(
                {"error": "Failed to remove the course."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class LessonMarkAsCompleteAndNotComplete(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            lesson_id = request.data.get("lesson_id")
            learner_id = request.data.get("learner_id")

            lesson = Lesson.objects.get(id=lesson_id)
            course_enrollment = CourseEnrollment.objects.get(
                course=lesson.course, learner__id=learner_id
            )
            completed_lessons = course_enrollment.completed_lessons

            if lesson_id in completed_lessons:
                completed_lessons.remove(lesson_id)
                message = "Lesson marked as Incomplete."
            else:
                completed_lessons.append(lesson_id)
                message = "Lesson marked as Completed."

            course_enrollment.save()

            return Response({"message": message}, status=status.HTTP_200_OK)

        except Exception as e:
            print(str(e))
            return Response(
                {"error": "Failed to update lesson status."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class DownloadLessonCertificate(APIView):
    permission_classes = [AllowAny]

    def get(self, request, lesson_id, learner_id):
        try:
            lesson = Lesson.objects.get(id=lesson_id)
            content = {}
            course_enrollment = CourseEnrollment.objects.get(
                course=lesson.course, learner__id=learner_id
            )

            content["learner_name"] = course_enrollment.learner.name
            content["course_name"] = lesson.course.name
            try:
                certificate = Certificate.objects.filter(courses=lesson.course).first()
            except Certificate.DoesNotExist:
                return Response(
                    {"error": "Certificate not found for the given course"},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                )

            email_message = certificate.content
            for key, value in content.items():
                email_message = email_message.replace(f"{{{{{key}}}}}", str(value))

            pdf = pdfkit.from_string(
                email_message,
                False,
                configuration=pdfkit_config,
                options={"orientation": "Landscape"},
            )
            response = HttpResponse(pdf, content_type="application/pdf")
            response["Content-Disposition"] = 'attachment; filename="Certificate.pdf"'
            return response

        except Exception as e:
            print(str(e))
            return Response(
                {"error": "Failed to download certificate."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class GetCertificateForCourse(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, course_id):
        try:
            course = get_object_or_404(Course, id=course_id)

            certificate = Certificate.objects.filter(courses=course).first()

            if certificate:
                return Response({"certificate_present": True})
            else:
                return Response({"certificate_present": False})
        except Exception as e:
            return Response(
                {"error": "Failed to retrieve certificates for the given course."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


from rest_framework.response import Response
from rest_framework.decorators import api_view
from .models import Video
from .serializers import VideoSerializer  # Import your Video model serializer


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def create_videos(request):
    name = request.data.get("name")  # Extract 'name' from request data
    video_file = request.data.get("video")  # Extract 'video' file from request data

    # You can perform validation or error checking here as needed

    # Assuming you have a VideoSerializer defined, you can create a new Video instance
    video_instance = Video(name=name, video=video_file)

    # Save the video instance to the database
    video_instance.save()

    # Serialize the newly created video instance
    serialized_video = VideoSerializer(video_instance)

    return Response(serialized_video.data)


# views.py
from rest_framework.response import Response
from rest_framework.decorators import api_view
from .models import Video
from .serializers import VideoSerializer


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_all_videos(request):
    videos = Video.objects.all()  # Retrieve all videos from the database

    # Serialize the video queryset
    serializer = VideoSerializer(videos, many=True)

    return Response(serializer.data)


# views.py
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from .models import Lesson, Video, VideoLesson
from .serializers import LessonSerializer, VideoLessonSerializer


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def create_video_lesson(request):
    if request.method == "POST":
        lesson_data = request.data.get("lesson")
        video_id = request.data.get("video")
        content = request.data.get("content", "")

        # Create or update lesson
        lesson_serializer = LessonSerializer(data=lesson_data)
        if not lesson_serializer.is_valid():
            return Response(
                lesson_serializer.errors, status=status.HTTP_400_BAD_REQUEST
            )

        # Save the lesson instance
        lesson_instance = lesson_serializer.save()

        try:
            # Retrieve video instance
            video_instance = Video.objects.get(id=video_id)

            # Create VideoLesson object
            video_lesson_data = {
                "lesson": lesson_instance.id,
                "video": video_instance.id,
                "content": content,
            }
            video_lesson_serializer = VideoLessonSerializer(data=video_lesson_data)
            if video_lesson_serializer.is_valid():
                video_lesson_serializer.save()
                return Response(
                    video_lesson_serializer.data, status=status.HTTP_201_CREATED
                )
            return Response(
                video_lesson_serializer.errors, status=status.HTTP_400_BAD_REQUEST
            )
        except Video.DoesNotExist:
            return Response(
                {"error": "Video does not exist"}, status=status.HTTP_404_NOT_FOUND
            )


# views.py

from rest_framework.decorators import api_view
from rest_framework.response import Response
from .models import Video
from .serializers import VideoSerializer


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_all_videos(request):
    if request.method == "GET":
        videos = Video.objects.all()
        serializer = VideoSerializer(videos, many=True)
        return Response(serializer.data)


from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from .models import VideoLesson, Lesson, Video
from .serializers import VideoLessonSerializer, LessonSerializer


@api_view(["PUT"])
@permission_classes([IsAuthenticated])
def update_video_lesson(request, lesson_id):
    try:
        lesson = Lesson.objects.get(pk=lesson_id)
    except Lesson.DoesNotExist:
        return Response(
            {"message": "Lesson does not exist"}, status=status.HTTP_404_NOT_FOUND
        )

    video_id = request.data.get("video")
    content = request.data.get("content", "")

    try:
        video = Video.objects.get(pk=video_id)
    except Video.DoesNotExist:
        return Response(
            {"message": "Video does not exist"}, status=status.HTTP_404_NOT_FOUND
        )

    lesson_data = request.data.get("lesson")
    if lesson_data:
        lesson_serializer = LessonSerializer(lesson, data=lesson_data, partial=True)
        if lesson_serializer.is_valid():
            lesson_serializer.save()

    video_lesson_data = {"lesson": lesson_id, "video": video_id, "content": content}

    try:
        video_lesson = VideoLesson.objects.get(lesson_id=lesson_id)
        serializer = VideoLessonSerializer(video_lesson, data=video_lesson_data)
    except VideoLesson.DoesNotExist:
        serializer = VideoLessonSerializer(data=video_lesson_data)

    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_200_OK)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


from rest_framework import status
from rest_framework.response import Response
from rest_framework.decorators import api_view
from .models import Video
from .serializers import VideoSerializer


@api_view(["PUT"])
@permission_classes([IsAuthenticated])
def update_video(request, pk):
    try:
        video = Video.objects.get(pk=pk)
    except Video.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)

    serializer = VideoSerializer(video, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class GetLaserCoachingTime(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, laser_coaching_id, participant_email):
        try:
            laser_coaching = LaserCoachingSession.objects.get(id=laser_coaching_id)
            coaching_session = laser_coaching.coaching_session
            learner = get_object_or_404(Learner, email=participant_email)
            existing_session = SchedularSessions.objects.filter(
                learner=learner, coaching_session=coaching_session
            ).first()
            return Response(
                {
                    "start_time": existing_session.availibility.start_time,
                    "end_time": existing_session.availibility.end_time,
                    "room_id": existing_session.availibility.coach.room_id,
                }
            )
        except Exception as e:
            return Response(
                {"error": "Failed to retrieve certificates for the given course."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_all_courses_progress(request):
    res = []
    courses = Course.objects.filter(status="public")
    hr_id = request.query_params.get("hr", None)
    if hr_id:
        courses = courses.filter(batch__project__hr__id=hr_id)
    for course in courses:
        course_serializer = CourseSerializer(course)
        lessons = Lesson.objects.filter(status="public", course=course)
        course_enrollments = CourseEnrollment.objects.filter(course=course)
        total_lessons_completed_by_learner = 0
        for course_enrollment in course_enrollments:
            total_lessons_completed_by_learner += len(
                course_enrollment.completed_lessons
            )
        completion_percentage = (
            (
                total_lessons_completed_by_learner
                / (lessons.count() * course_enrollments.count())
            )
            * 100
            if course_enrollments.count() > 0 and lessons.count() > 0
            else 0
        )

        res.append(
            {
                **course_serializer.data,
                "batch_name": course.batch.name,
                "total_learners": course_enrollments.count(),
                "completion_percentage": round(completion_percentage),
            }
        )
    return Response(res)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_course_progress(request, course_id):
    course_enrollments = CourseEnrollment.objects.filter(course__id=course_id)
    course_enrollments_serializer = CourseEnrollmentDepthOneSerializer(
        course_enrollments, many=True
    )
    lessons_count = Lesson.objects.filter(course__id=course_id, status="public").count()
    res = []
    for course_enrollment in course_enrollments_serializer.data:
        res.append({**course_enrollment, "total_lessons": lessons_count})
    return Response(res)


@api_view(["GET"])
@permission_classes([AllowAny])
def course_report_download(request, course_id):
    course_enrollments = CourseEnrollment.objects.filter(course__id=course_id)
    course_enrollments_serializer = CourseEnrollmentDepthOneSerializer(
        course_enrollments, many=True
    )
    lessons_count = Lesson.objects.filter(course__id=course_id, status="public").count()
    res = []
    for course_enrollment in course_enrollments_serializer.data:
        res.append(
            {
                **course_enrollment,
                "total_lessons": lessons_count,
                "progress": round(
                    (len(course_enrollment["completed_lessons"]) / lessons_count) * 100
                ),
            }
        )

    # Create a new workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Write headers to the worksheet
    headers = [
        "Participant Name",
        "Completed Lessons",
        "Total Lessons",
        "Progress",
    ]

    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    # Write data to the worksheet
    for row_num, course_enrollment_data in enumerate(res, 2):
        ws.append(
            [
                course_enrollment_data["learner"]["name"],
                len(course_enrollment_data["completed_lessons"]),
                course_enrollment_data["total_lessons"],
                str(course_enrollment_data["progress"]) + "%",
            ]
        )

    # Create a response with the Excel file
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=course_report.xlsx"
    wb.save(response)

    return response


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_all_quizes_report(request):
    res = []
    quizes = QuizLesson.objects.filter(
        lesson__status="public", lesson__course__status="public"
    )
    hr_id = request.query_params.get("hr", None)
    if hr_id:
        quizes = quizes.filter(lesson__course__batch__project__hr__id=hr_id)
    for quiz in quizes:
        course_enrollments = CourseEnrollment.objects.filter(course=quiz.lesson.course)
        total_participants = course_enrollments.count()
        quiz_lesson_responses = QuizLessonResponse.objects.filter(quiz_lesson=quiz)
        total_responses = quiz_lesson_responses.count()
        total_percentage = 0
        for quiz_lesson_response in quiz_lesson_responses:
            quiz_result = calculate_quiz_result(quiz, quiz_lesson_response)
            total_percentage += quiz_result["percentage"]
        average_percentage = (
            (total_percentage / total_responses) if total_responses > 0 else 0
        )
        res.append(
            {
                "id": quiz.id,
                "quiz_name": quiz.lesson.name,
                "course_name": quiz.lesson.course.name,
                "batch_name": quiz.lesson.course.batch.name,
                "total_participants": total_participants,
                "total_responses": total_responses,
                "average_percentage": round(average_percentage),
            }
        )

    return Response(res)


def get_quiz_report_data(quiz_lesson):
    quiz_lesson_responses = QuizLessonResponse.objects.filter(quiz_lesson=quiz_lesson)
    res = []
    for quiz_lesson_response in quiz_lesson_responses:
        quiz_result = calculate_quiz_result(quiz_lesson, quiz_lesson_response)
        res.append({"learner": quiz_lesson_response.learner.name, **quiz_result})
    return res


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_quiz_report(request, quiz_id):
    quiz_lesson = QuizLesson.objects.get(id=quiz_id)
    data = get_quiz_report_data(quiz_lesson)
    return Response(data)


@api_view(["GET"])
@permission_classes([AllowAny])
def quiz_report_download(request, quiz_id):
    quiz_lesson = QuizLesson.objects.get(id=quiz_id)
    quiz_data = get_quiz_report_data(quiz_lesson)
    # Create a new workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Write headers to the worksheet
    headers = [
        "Participant Name",
        "Correct Answers",
        "Total Questions",
        "Percentage",
    ]

    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    # Write data to the worksheet
    for row_num, quiz_data_item in enumerate(quiz_data, 2):
        ws.append(
            [
                quiz_data_item["learner"],
                quiz_data_item["correct_answers"],
                quiz_data_item["total_questions"],
                str(quiz_data_item["percentage"]) + "%",
            ]
        )

    # Create a response with the Excel file
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=quiz_report.xlsx"
    wb.save(response)

    return response


def calculate_nps(ratings):
    promoters = sum(rating >= 9 for rating in ratings)
    detractors = sum(rating <= 6 for rating in ratings)
    nps = ((promoters - detractors) / len(ratings)) * 100
    return nps


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_all_feedbacks_report(request):
    res = []
    feedbacks = FeedbackLesson.objects.filter(
        lesson__status="public", lesson__course__status="public"
    )
    hr_id = request.query_params.get("hr", None)
    facilitator_id = request.query_params.get("facilitator_id")
    if hr_id:
        feedbacks = feedbacks.filter(lesson__course__batch__project__hr__id=hr_id)
    if facilitator_id:
        feedbacks = feedbacks.filter(live_session__facilitator__id=facilitator_id)
    for feedback in feedbacks:
        facilitator_name = ""
        if feedback.live_session and feedback.live_session.facilitator:
            facilitator_name = (
                feedback.live_session.facilitator.first_name
                + " "
                + feedback.live_session.facilitator.last_name
            )
        # course_enrollments = CourseEnrollment.objects.filter(
        #     course=feedback.lesson.course
        # )
        total_participants = feedback.lesson.course.batch.learners.count()
        feedback_lesson_responses = FeedbackLessonResponse.objects.filter(
            feedback_lesson=feedback
        )
        total_responses = feedback_lesson_responses.count()
        response_percentage = (
            ((total_responses / total_participants) * 100)
            if total_participants > 0
            else 0
        )
        nps = None
        answers = Answer.objects.filter(
            question__type="rating_0_to_10", question__feedbacklesson=feedback
        )
        answer_ratings = []
        for answer in answers:
            answer_ratings.append(answer.rating)
        if len(answer_ratings) > 0:
            nps = calculate_nps(answer_ratings)
        res.append(
            {
                "id": feedback.id,
                "feedback_name": feedback.lesson.name,
                "course_name": feedback.lesson.course.name,
                "batch_name": feedback.lesson.course.batch.name,
                "total_participants": total_participants,
                "total_responses": total_responses,
                "response_percentage": round(response_percentage),
                "nps": nps,
                "facilitator_name": facilitator_name,
            }
        )

    return Response(res)


@api_view(["GET"])
def get_consolidated_feedback_report(request):
    try:
        data = {}
        all_projects = SchedularProject.objects.all()

        for project in all_projects:
            all_batches = SchedularBatch.objects.filter(project=project)
            total_participant_count = 0
            for batch in all_batches:
                total_participant_count += batch.learners.count()
                # Get live sessions for the current batch
                live_sessions = LiveSession.objects.filter(batch=batch)
                for live_session in live_sessions:
                    # Now, you can access the associated Course through the SchedularBatch
                    course = Course.objects.filter(batch=batch).first()
                    if course:
                        feedback_lesson = FeedbackLesson.objects.filter(
                            lesson__course=course, live_session=live_session
                        ).first()
                        facilitator_id = request.query_params.get("facilitator_id")
                        if facilitator_id:
                            if (
                                feedback_lesson
                                and feedback_lesson.live_session
                                and feedback_lesson.live_session.facilitator
                                and not feedback_lesson.live_session.facilitator.id
                                == int(facilitator_id)
                            ):
                                continue
                        if feedback_lesson:
                            total_responses = FeedbackLessonResponse.objects.filter(
                                feedback_lesson=feedback_lesson
                            ).count()
                            percentage_responded = (
                                round(
                                    (total_responses / total_participant_count) * 100,
                                    2,
                                )
                                if total_participant_count
                                else 0
                            )
                            session_name = get_live_session_name(
                                live_session.session_type
                            )
                            live_session_key = f"{project.name} {session_name} {live_session.live_session_number}"
                            if live_session_key not in data:
                                data[live_session_key] = {
                                    "live_session_id": live_session.id,
                                    "project_name": project.name,
                                    "session_name": f"{session_name} {live_session.live_session_number}",
                                    "total_participant": total_participant_count,
                                    "total_responses": total_responses,
                                    "percentage_responded": percentage_responded,
                                }
                            else:
                                data[live_session_key][
                                    "total_responses"
                                ] += total_responses
                                data[live_session_key]["percentage_responded"] = round(
                                    (
                                        data[live_session_key]["total_responses"]
                                        / data[live_session_key]["total_participant"]
                                    )
                                    * 100,
                                    2,
                                )

        return Response(list(data.values()))
    except Exception as e:
        print(str(e))
        return Response({"error": str(e)}, status=500)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_feedback_report(request, feedback_id):
    try:
        feedback_lesson = FeedbackLesson.objects.get(id=feedback_id)
        questions_serializer = QuestionSerializer(feedback_lesson.questions, many=True)
        question_data = {
            question["id"]: {**question, "descriptive_answers": [], "ratings": []}
            for question in questions_serializer.data
        }
        feedback_lesson_responses = FeedbackLessonResponse.objects.filter(
            feedback_lesson=feedback_lesson
        )
        for response in feedback_lesson_responses:
            for answer in response.answers.all():
                question_id = answer.question.id
                if answer.question.type.startswith("rating"):
                    question_data[question_id]["ratings"].append(answer.rating)
                elif answer.question.type == "descriptive_answer":
                    question_data[question_id]["descriptive_answers"].append(
                        answer.text_answer
                    )
        for question_id, data in question_data.items():
            # Calculate average rating for each question
            ratings = data["ratings"]
            if ratings:
                if data["type"] == "rating_0_to_10":
                    # Calculate NPS

                    data["nps"] = calculate_nps(ratings)

                else:
                    # Calculate average rating
                    data["average_rating"] = sum(ratings) / len(ratings)
            else:
                if data["type"] == "rating_0_to_10":
                    data["nps"] = 0  # Default NPS value if no ratings
                else:
                    data["average_rating"] = 0  # default rating 0 if not ratings

        return Response(question_data.values())
    except FeedbackLesson.DoesNotExist:
        return Response(
            {"error": "Feedback lesson not found"}, status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["GET"])
def get_consolidated_feedback_report_response(request, lesson_id):
    try:
        live_session = LiveSession.objects.get(id=lesson_id)
        # Dictionary to store aggregated feedback data for each question
        question_data = {}
        facilitator_id = request.query_params.get("facilitator_id")
        feedback_lesson_responses = FeedbackLessonResponse.objects.filter(
            feedback_lesson__lesson__course__batch__project=live_session.batch.project,
            feedback_lesson__live_session__session_type=live_session.session_type,
            feedback_lesson__live_session__live_session_number=live_session.live_session_number,
        )
        if facilitator_id:
            feedback_lesson_responses = feedback_lesson_responses.filter(
                feedback_lesson__live_session__facilitator__id=facilitator_id
            )
        for response in feedback_lesson_responses:
            for answer in response.answers.all():
                question_text = answer.question.text

                if question_text not in question_data:
                    question_data[question_text] = {
                        **QuestionSerializer(answer.question).data,
                        "descriptive_answers": [],
                        "ratings": [],
                        "average_rating": 0,
                        "nps": 0,
                    }

                if answer.question.type.startswith("rating"):
                    question_data[question_text]["ratings"].append(answer.rating)
                elif answer.question.type == "descriptive_answer":
                    question_data[question_text]["descriptive_answers"].append(
                        answer.text_answer
                    )

        for question_text, data in question_data.items():
            # Calculate average rating for each question
            ratings = data["ratings"]
            if ratings:
                if data["type"] == "rating_0_to_10":
                    # Calculate NPS
                    promoters = sum(rating >= 9 for rating in ratings)
                    detractors = sum(rating <= 6 for rating in ratings)
                    nps = ((promoters - detractors) / len(ratings)) * 100
                    data["nps"] = nps
                else:
                    # Calculate average rating
                    data["average_rating"] = sum(ratings) / len(ratings)
            else:
                if data["type"] == "rating_0_to_10":
                    data["nps"] = 0  # Default NPS value if no ratings
                else:
                    data["average_rating"] = 0  # Default rating 0 if no ratings

        return Response(list(question_data.values()))
    except LiveSession.DoesNotExist:
        return Response(
            {"error": "Live session not found"}, status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class AssignCourseTemplateToBatch(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, course_template_id, batch_id):
        try:
            with transaction.atomic():
                course_template = get_object_or_404(
                    CourseTemplate, pk=course_template_id
                )
                batch = get_object_or_404(SchedularBatch, pk=batch_id)
                # Duplicate the course
                new_course = Course.objects.create(
                    name=f"{course_template.name}",
                    description=course_template.description,
                    status="draft",
                    course_template=course_template,
                    batch=batch,
                )
                # Duplicate lessons
                original_lessons = Lesson.objects.filter(
                    course_template=course_template
                )
                facilitator_lesson_creation = Lesson.objects.create(
                    course=new_course,
                    name="Facilitator Lesson",
                    status="draft",
                    lesson_type="facilitator",
                    # Duplicate specific lesson types
                    order=1,
                )
                FacilitatorLesson.objects.create(
                    lesson=facilitator_lesson_creation,
                )
                assessment_creation = False
                if not original_lessons.filter(lesson_type="assessment").exists():
                    if batch.project.pre_post_assessment:
                        assessment_creation = True
                        lesson1 = Lesson.objects.create(
                            course=new_course,
                            name="Pre Assessment",
                            status="draft",
                            lesson_type="assessment",
                            # Duplicate specific lesson types
                            order=2,
                        )
                        assessment1 = Assessment.objects.create(
                            lesson=lesson1, type="pre"
                        )
                for original_lesson in original_lessons:
                    new_lesson = None
                    # Create a new lesson only if the type is 'text', 'quiz', or 'feedback'
                    if original_lesson.lesson_type not in [
                        "live_session",
                        "laser_coaching",
                    ]:
                        updated_order = original_lesson.order + 1
                        if assessment_creation:
                            updated_order = original_lesson.order + 2
                        new_lesson = Lesson.objects.create(
                            course=new_course,
                            name=original_lesson.name,
                            status=original_lesson.status,
                            lesson_type=original_lesson.lesson_type,
                            # Duplicate specific lesson types
                            order=updated_order,
                        )
                        if original_lesson.lesson_type == "text":
                            TextLesson.objects.create(
                                lesson=new_lesson,
                                content=original_lesson.textlesson.content,
                            )
                        elif original_lesson.lesson_type == "video":
                            VideoLesson.objects.create(
                                lesson=new_lesson,
                                video=original_lesson.videolesson.video,
                                content=original_lesson.videolesson.content,
                            )
                        elif original_lesson.lesson_type == "ppt":
                            PdfLesson.objects.create(
                                lesson=new_lesson,
                                pdf=original_lesson.pdflesson.pdf,
                                content=original_lesson.pdflesson.content,
                            )
                        elif original_lesson.lesson_type == "downloadable_file":
                            DownloadableLesson.objects.create(
                                lesson=new_lesson,
                                file=original_lesson.downloadablelesson.file,
                                description=original_lesson.downloadablelesson.description,
                            )
                        elif original_lesson.lesson_type == "assignment":
                            AssignmentLesson.objects.create(
                                lesson=new_lesson,
                                name=original_lesson.assignmentlesson.name,
                                description=original_lesson.assignmentlesson.description,
                            )
                        elif original_lesson.lesson_type == "assessment":
                            assessment = Assessment.objects.filter(
                                lesson=original_lesson
                            ).first()
                            Assessment.objects.create(
                                lesson=new_lesson, type=assessment.type
                            )
                        elif original_lesson.lesson_type == "quiz":
                            new_quiz_lesson = QuizLesson.objects.create(
                                lesson=new_lesson
                            )
                            for question in original_lesson.quizlesson.questions.all():
                                new_question = Question.objects.create(
                                    text=question.text,
                                    options=question.options,
                                    type=question.type,
                                )
                                new_quiz_lesson.questions.add(new_question)
                        elif original_lesson.lesson_type == "feedback":
                            unique_id = uuid.uuid4()
                            new_feedback_lesson = FeedbackLesson.objects.create(
                                lesson=new_lesson, unique_id=unique_id
                            )
                            for (
                                question
                            ) in original_lesson.feedbacklesson.questions.all():
                                new_question = Question.objects.create(
                                    text=question.text,
                                    options=question.options,
                                    type=question.type,
                                )
                                new_feedback_lesson.questions.add(new_question)
                learners = batch.learners.all()
                for learner in learners:
                    course_enrollments = CourseEnrollment.objects.filter(
                        learner=learner, course=new_course
                    )
                    if not course_enrollments.exists():
                        datetime = timezone.now()
                        CourseEnrollment.objects.create(
                            learner=learner, course=new_course, enrollment_date=datetime
                        )
                live_sessions = LiveSessionSchedular.objects.filter(batch__id=batch_id)
                training_class_sessions = LiveSession.objects.filter(
                    session_type__in=["in_person_session", "virtual_session"]
                )
                max_order_of_training_class_sessions = (
                    training_class_sessions.aggregate(Max("order"))["order__max"]
                )
                coaching_sessions = CoachingSession.objects.filter(batch__id=batch_id)
                max_order = (
                    Lesson.objects.filter(course=new_course).aggregate(Max("order"))[
                        "order__max"
                    ]
                    or 0
                )
                for live_session in live_sessions:
                    max_order = max_order + 1
                    session_name = get_live_session_name(live_session.session_type)

                    new_lesson = Lesson.objects.create(
                        course=new_course,
                        name=f"{session_name} {live_session.live_session_number}",
                        status="draft",
                        lesson_type="live_session",
                        order=max_order,
                    )
                    LiveSessionLesson.objects.create(
                        lesson=new_lesson, live_session=live_session
                    )
                    max_order_feedback = (
                        Lesson.objects.filter(course=new_course).aggregate(
                            Max("order")
                        )["order__max"]
                        or 0
                    )
                    new_feedback_lesson = Lesson.objects.create(
                        course=new_course,
                        name=f"Feedback for {session_name} {live_session.live_session_number}",
                        status="draft",
                        lesson_type="feedback",
                        order=max_order_feedback,
                    )
                    unique_id = uuid.uuid4()
                    feedback_lesson = FeedbackLesson.objects.create(
                        lesson=new_feedback_lesson,
                        unique_id=unique_id,
                        live_session=live_session,
                    )
                    if live_session.session_type in [
                        "in_person_session",
                        "virtual_session",
                    ]:
                        if int(max_order_of_training_class_sessions) == int(
                            live_session.order
                        ):
                            add_question_to_feedback_lesson(
                                feedback_lesson, nps_default_feed_questions
                            )
                        else:
                            add_question_to_feedback_lesson(
                                feedback_lesson, default_feedback_questions
                            )

                for coaching_session in coaching_sessions:
                    max_order = max_order + 1
                    session_name = None
                    if coaching_session.session_type == "laser_coaching_session":
                        session_name = "Laser coaching"
                    elif coaching_session.session_type == "mentoring_session":
                        session_name = "Mentoring session"
                    new_lesson = Lesson.objects.create(
                        course=new_course,
                        name=f"{session_name} {coaching_session.coaching_session_number}",
                        status="draft",
                        lesson_type="laser_coaching",
                        order=max_order,
                    )
                    LaserCoachingSession.objects.create(
                        lesson=new_lesson, coaching_session=coaching_session
                    )

                if assessment_creation:
                    max_order = (
                        Lesson.objects.filter(course=new_course).aggregate(
                            Max("order")
                        )["order__max"]
                        or 0
                    )

                    lesson2 = Lesson.objects.create(
                        course=new_course,
                        name="Post Assessment",
                        status="draft",
                        lesson_type="assessment",
                        # Duplicate specific lesson types
                        order=max_order + 1,
                    )

                    assessment2 = Assessment.objects.create(lesson=lesson2, type="post")

            return Response(
                {
                    "message": "Course assigned successfully.",
                },
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            print(str(e))
            return Response(
                {"error": "Failed to assign course template."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_resources(request):
    resources = Resources.objects.all()
    serializer = ResourcesSerializer(resources, many=True)
    return Response(serializer.data)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def create_resource(request):
    pdf_name = request.data.get("pdfName")  # Extracting pdfName from request data
    pdf_file = request.data.get("pdfFile")  # Extracting pdfFile from request data

    # Create a dictionary containing the data for the Resources model instance
    resource_data = {
        "name": pdf_name,
        "pdf_file": pdf_file,
    }

    # Assuming you have a serializer for the Resources model
    serializer = ResourcesSerializer(data=resource_data)

    if serializer.is_valid():
        # Save the validated data to create a new Resources instance
        serializer.save()
        return Response(serializer.data, status=201)  # Return the serialized data
    else:
        return Response(
            serializer.errors, status=400
        )  # Return errors if validation fails


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def create_pdf_lesson(request):

    try:
        with transaction.atomic():
            lesson_data = request.data.get("lesson")
            pdf_id = request.data.get("pdf_id")
            content = request.data.get("content", "")
            course_template_id = lesson_data.get("course_template", "")
            course_id = lesson_data.get("course", "")

            resources = Resources.objects.filter(id=pdf_id).first()
            if course_id:
                course_instance = Course.objects.get(id=course_id)
                course_template_instance = course_instance.course_template
                live_session_id = lesson_data["live_session"]
                live_session = None
                print(live_session_id)
                if live_session_id:
                    live_session = LiveSessionSchedular.objects.get(id=live_session_id)

                lesson_instance = Lesson.objects.create(
                    course=course_instance,
                    name=lesson_data["name"],
                    status=lesson_data["status"],
                    lesson_type=lesson_data["lesson_type"],
                    order=lesson_data["order"],
                    drip_date=lesson_data["drip_date"],
                    live_session=live_session,
                )

                pdf_lesson_instance = PdfLesson.objects.create(
                    lesson=lesson_instance, content=content, pdf=resources
                )

                return Response({"message": "PDF lesson created successfully."})

            elif course_template_id:
                course_template_instance = CourseTemplate.objects.get(
                    id=course_template_id
                )

                lesson_instance = Lesson.objects.create(
                    course_template=course_template_instance,
                    name=lesson_data["name"],
                    status=lesson_data["status"],
                    lesson_type=lesson_data["lesson_type"],
                    order=lesson_data["order"],
                )

                pdf_lesson_instance = PdfLesson.objects.create(
                    lesson=lesson_instance, content=content, pdf=resources
                )

                return Response({"message": "PDF lesson created successfully."})

            else:
                return Response(
                    {"message": "Neither Course ID nor Course Template ID provided."}
                )

    except Course.DoesNotExist:
        return Response({"message": "Course does not exist."})
    except CourseTemplate.DoesNotExist:
        return Response({"message": "Course Template does not exist."})
    except Exception as e:
        return Response({"message": "Failed to create pdf lesson."})


@api_view(["PUT"])
@permission_classes([IsAuthenticated])
def update_pdf_lesson(request, pk):
    try:
        pdf_lesson = PdfLesson.objects.get(id=pk)
    except PdfLesson.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)

    # Extract data from request
    lesson_data = request.data.get("lesson", {})
    pdf_id = request.data.get("pdf_id")
    content = request.data.get("content", "")

    try:
        lesson = Lesson.objects.get(id=lesson_data.get("id"))
    except Lesson.DoesNotExist:
        return Response({"error": "Lesson not found"}, status=status.HTTP_404_NOT_FOUND)

    # Update Lesson instance
    lesson_serializer = LessonSerializer(lesson, data=lesson_data)
    if lesson_serializer.is_valid():
        lesson_serializer.save()

        # Update PdfLesson instance
        pdf_data = {"lesson": lesson_data, "content": content, "pdf": pdf_id}
        pdf_serializer = PdfLessonSerializer(pdf_lesson, data=pdf_data)
        if pdf_serializer.is_valid():
            pdf_serializer.save()

            # Update pdf reference
            try:
                pdf_resource = Resources.objects.get(pk=pdf_id)
                pdf_lesson.pdf = pdf_resource
                pdf_lesson.save()
            except Resources.DoesNotExist:
                return Response(
                    {"error": "PDF resource not found"},
                    status=status.HTTP_404_NOT_FOUND,
                )

            return Response(pdf_serializer.data)
        return Response(pdf_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    return Response(lesson_serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(["PUT"])
@permission_classes([IsAuthenticated])
def update_course_template_status(request):
    course_template_id = request.data.get("course_template_id")
    selected_status = request.data.get("status")

    try:
        course_template = CourseTemplate.objects.get(id=course_template_id)
    except CourseTemplate.DoesNotExist:
        return Response(
            {"error": "Course Template not found"}, status=status.HTTP_404_NOT_FOUND
        )

    # Update the fields if data exists in the request
    if selected_status:
        course_template.status = selected_status

    course_template.save()

    serializer = CourseTemplateSerializer(course_template)
    return Response(serializer.data, status=status.HTTP_200_OK)


@api_view(["PUT"])
@permission_classes([IsAuthenticated])
def update_course_status(request):
    try:
        course_id = request.data.get("course_id")
        status = request.data.get("status")

        # Fetch the course object to update
        course = Course.objects.get(pk=course_id)

        # Update the status
        course.status = status
        course.save()

        return Response({"message": "Course status updated successfully"}, status=200)
    except Course.DoesNotExist:
        return Response({"message": "Course not found"}, status=404)
    except Exception as e:
        return Response({"message": str(e)}, status=500)


@api_view(["PUT"])
@permission_classes([AllowAny])
def lesson_update_status(request):
    if request.method == "PUT":
        serializer = LessonUpdateSerializer(data=request.data)
        if serializer.is_valid():
            lesson_id = serializer.validated_data["lesson_id"]
            status_value = serializer.validated_data["status"]

            try:
                lesson = Lesson.objects.get(id=lesson_id)
                lesson.status = status_value
                lesson.save()

                if lesson.lesson_type == "assessment":
                    assessment = Assessment.objects.filter(lesson=lesson).first()

                    assessment_modal = AssessmentModal.objects.get(
                        id=assessment.assessment_modal.id
                    )

                    if lesson.status == "draft":
                        assessment_modal.status = "draft"

                    if lesson.status == "public":
                        assessment_modal.status = "ongoing"
                    assessment_modal.save()

                return Response(
                    {"message": f"Lesson status updated."}, status=status.HTTP_200_OK
                )
            except Lesson.DoesNotExist:
                return Response(
                    {"message": f"Lesson does not exist"},
                    status=status.HTTP_404_NOT_FOUND,
                )
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    else:
        return Response(
            {"message": "Invalid method"}, status=status.HTTP_405_METHOD_NOT_ALLOWED
        )


class FileListAPIView(generics.ListAPIView):
    queryset = File.objects.all()
    serializer_class = FileSerializer
    permission_classes = [IsAuthenticated]


class FileUploadView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        serializer = FileSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(["PUT"])
@permission_classes([IsAuthenticated])
def update_file(request, file_id):
    try:
        file = File.objects.get(id=file_id)
    except File.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)

    serializer = FileSerializer(file, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class FileDownloadView(APIView):
    permission_classes = [AllowAny]

    def get(self, request, file_id):
        file_obj = get_object_or_404(File, id=file_id)
        serializer = FileSerializer(file_obj)
        return download_file_response(serializer.data["file"])


class DownloadableLessonCreateView(generics.CreateAPIView):
    queryset = DownloadableLesson.objects.all()
    serializer_class = DownloadableLessonSerializer
    permission_classes = [IsAuthenticated]


class DownloadableLessonUpdateView(generics.UpdateAPIView):
    queryset = DownloadableLesson.objects.all()
    serializer_class = DownloadableLessonSerializer
    permission_classes = [IsAuthenticated]


class FeedbackEmailValidation(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        try:
            unique_id = request.data.get("unique_id")
            email = request.data.get("email").strip().lower()

            feedback_lesson = FeedbackLesson.objects.get(unique_id=unique_id)
            lesson = feedback_lesson.lesson

            learner = Learner.objects.filter(email=email).first()

            if learner:
                feedback_lesson_response = FeedbackLessonResponse.objects.filter(
                    feedback_lesson=feedback_lesson, learner__id=learner.id
                ).first()
                if not feedback_lesson_response:
                    lesson_serializer = LessonSerializer(lesson)
                    feedback_lesson_serializer = FeedbackLessonDepthOneSerializer(
                        feedback_lesson
                    )

                    return Response(
                        {
                            "message": "Validation Successful",
                            "participant_exists": True,
                            "lesson_details": lesson_serializer.data,
                            "feedback_lesson_details": feedback_lesson_serializer.data,
                            "participant_id": learner.id,
                        },
                        status=status.HTTP_200_OK,
                    )
                else:
                    return Response(
                        {"error": "Already Responded."},
                        status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                    )

            return Response(
                {"error": "User does not exist."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        except FeedbackLesson.DoesNotExist:
            return Response(
                {"error": "FeedbackLesson not found."},
                status=status.HTTP_404_NOT_FOUND,
            )
        except Exception as e:
            print(str(e))
            return Response(
                {"error": "Failed to validate email."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class GetFeedbackForm(APIView):
    permission_classes = [AllowAny]

    def get(self, request, unique_id):
        try:
            feedback_lesson = FeedbackLesson.objects.get(unique_id=unique_id)

            return Response(
                {
                    "lesson_name": feedback_lesson.lesson.name,
                    "lesson_status": feedback_lesson.lesson.status,
                },
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            print(str(e))
            return Response(
                {"error": "Failed to get feedback lesson details."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class EditAllowedFeedbackLesson(APIView):
    permission_classes = [AllowAny]

    def get(self, request, feedback_lesson_id):
        try:
            feedback_lesson = FeedbackLesson.objects.get(id=feedback_lesson_id)
            feedback_lesson_response = FeedbackLessonResponse.objects.filter(
                feedback_lesson=feedback_lesson
            )

            if feedback_lesson_response:
                return Response(
                    {
                        "edit_allowed": False,
                    },
                    status=status.HTTP_200_OK,
                )

            return Response(
                {
                    "edit_allowed": True,
                },
                status=status.HTTP_200_OK,
            )
        except Exception as e:
            print(str(e))
            return Response(
                {"error": "Failed to get details."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class DuplicateLesson(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        try:
            with transaction.atomic():
                is_course_tepmlate = request.data.get("is_course_tepmlate")
                duplicate_to_course_template_or_course_id = request.data.get(
                    "duplicate_to_course_template_or_course_id"
                )
                duplicate_from_lesson_id = request.data.get("duplicate_from_lesson_id")
                course = None

                if is_course_tepmlate:
                    course = CourseTemplate.objects.get(
                        id=duplicate_to_course_template_or_course_id
                    )
                else:
                    course = Course.objects.get(
                        id=duplicate_to_course_template_or_course_id
                    )

                lesson = Lesson.objects.get(id=duplicate_from_lesson_id)
                new_lesson = None
                # Create a new lesson only if the type is 'text', 'quiz', or 'feedback'
                if lesson.lesson_type not in [
                    "live_session",
                    "laser_coaching",
                ]:
                    if not is_course_tepmlate:
                        max_order = (
                            Lesson.objects.filter(course=course).aggregate(
                                Max("order")
                            )["order__max"]
                            or 0
                        )
                        new_lesson = Lesson.objects.create(
                            course=course,
                            name=lesson.name,
                            status="draft",
                            lesson_type=lesson.lesson_type,
                            order=max_order + 1,
                        )

                    else:
                        max_order = (
                            Lesson.objects.filter(course_template=course).aggregate(
                                Max("order")
                            )["order__max"]
                            or 0
                        )
                        new_lesson = Lesson.objects.create(
                            course_template=course,
                            name=lesson.name,
                            status="draft",
                            lesson_type=lesson.lesson_type,
                            order=max_order + 1,
                        )
                    if lesson.lesson_type == "text":
                        TextLesson.objects.create(
                            lesson=new_lesson,
                            content=lesson.textlesson.content,
                        )
                    elif lesson.lesson_type == "video":
                        VideoLesson.objects.create(
                            lesson=new_lesson,
                            video=lesson.videolesson.video,
                            content=lesson.videolesson.content,
                        )
                    elif lesson.lesson_type == "ppt":
                        PdfLesson.objects.create(
                            lesson=new_lesson,
                            pdf=lesson.pdflesson.pdf,
                            content=lesson.pdflesson.content,
                        )
                    elif lesson.lesson_type == "downloadable_file":
                        DownloadableLesson.objects.create(
                            lesson=new_lesson,
                            file=lesson.downloadablelesson.file,
                            description=lesson.downloadablelesson.description,
                        )
                    elif lesson.lesson_type == "assignment":
                        AssignmentLesson.objects.create(
                            lesson=new_lesson,
                            name=lesson.assignmentlesson.name,
                            description=lesson.assignmentlesson.description,
                        )
                    elif lesson.lesson_type == "assessment":
                        Assessment.objects.create(lesson=new_lesson)
                    elif lesson.lesson_type == "quiz":
                        new_quiz_lesson = QuizLesson.objects.create(lesson=new_lesson)
                        for question in lesson.quizlesson.questions.all():
                            new_question = Question.objects.create(
                                text=question.text,
                                options=question.options,
                                type=question.type,
                            )
                            new_quiz_lesson.questions.add(new_question)
                    elif lesson.lesson_type == "feedback":
                        unique_id = uuid.uuid4()
                        new_feedback_lesson = FeedbackLesson.objects.create(
                            lesson=new_lesson, unique_id=unique_id
                        )
                        for question in lesson.feedbacklesson.questions.all():
                            new_question = Question.objects.create(
                                text=question.text,
                                options=question.options,
                                type=question.type,
                            )
                            new_feedback_lesson.questions.add(new_question)

            return Response(
                {"message": "Lesson duplicated Sucessfully."},
                status=status.HTTP_200_OK,
            )
        except Exception as e:
            print(str(e))
            return Response(
                {"error": "Failed to duplicate lesson."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class LessonCompletedWebhook(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        try:
            payload = request.data.get("payload", {})

            lesson_name = payload.get("lesson", {}).get("name", "")
            user = payload.get("user", {})
            student_name = f"{user.get('first_name', '')} {user.get('last_name', '')}"
            course_name = payload.get("course", {}).get("name", "")
            completion_data = request.data

            ThinkificLessonCompleted.objects.create(
                lesson_name=lesson_name,
                student_name=student_name,
                course_name=course_name,
                completion_data=completion_data,
            )

            return Response(status=status.HTTP_200_OK)
        except KeyError as key_error:
            # Handle KeyError (missing key in dictionary)
            return Response(
                {"error": f"KeyError: {str(key_error)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        except Exception as e:
            # Handle other exceptions and log the error
            logger.error(f"An error occurred: {str(e)}")
            return Response(
                {"error": "Failed to duplicate lesson."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class GetUniqueIdParticipantFromCourse(APIView):
    permission_classes = [AllowAny]

    def get(self, request, user_id, assessment_id):
        try:
            learner = Learner.objects.get(id=user_id)

            assessment = AssessmentModal.objects.get(id=assessment_id)

            participant_unique_id = ParticipantUniqueId.objects.filter(
                participant=learner, assessment=assessment
            ).first()

            participant_response = ParticipantResponse.objects.filter(
                participant=learner, assessment=assessment
            ).first()

            return Response(
                {
                    "unique_id": participant_unique_id.unique_id,
                    "responded": bool(participant_response),
                },
            )

        except Exception as e:
            print(str(e))
            return Response(
                {"error": "Failed to get unique id."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class GetAssessmentsOfBatch(APIView):
    permission_classes = [AllowAny]

    def get(self, request, project_or_batch, id):
        try:
            batches = None
            if project_or_batch == "project":
                batches = SchedularBatch.objects.filter(project__id=id)
            else:
                batches = SchedularBatch.objects.filter(id=id)

            assessment_list = []

            for batch in batches:
                assessments = AssessmentModal.objects.filter(
                    assessment_modal__lesson__course__batch=batch
                )

                for assessment in assessments:
                    total_responses_count = ParticipantResponse.objects.filter(
                        assessment=assessment
                    ).count()
                    assessment_data = {
                        "id": assessment.id,
                        "name": assessment.name,
                        "participant_view_name": assessment.participant_view_name,
                        "assessment_type": assessment.assessment_type,
                        "assessment_timing": assessment.assessment_timing,
                        "assessment_start_date": assessment.assessment_start_date,
                        "assessment_end_date": assessment.assessment_end_date,
                        "status": assessment.status,
                        "total_learners_count": assessment.participants_observers.count(),
                        "total_responses_count": total_responses_count,
                        "created_at": assessment.created_at,
                        "automated_reminder": assessment.automated_reminder,
                        "batch_name": batch.name,
                        "questionnaire": assessment.questionnaire.id,
                        "organisation": assessment.organisation.id,
                        "hr": list(assessment.hr.all().values_list("id", flat=True)),
                        "pre_assessment": (
                            assessment.pre_assessment.id
                            if assessment.assessment_timing == "post"
                            else None
                        ),
                    }

                    assessment_list.append(assessment_data)

            return Response(assessment_list)

        except Exception as e:
            return Response(
                {"error": "Failed to ger data"}, status.HTTP_500_INTERNAL_SERVER_ERROR
            )


@api_view(["GET"])
@permission_classes([AllowAny])
def get_all_feedbacks_download_report(request, feedback_id):
    try:
        feedback_lesson = FeedbackLesson.objects.get(id=feedback_id)
        data = []
        # Populate the list of dictionaries with data from FeedbackLesson and FeedbackLessonResponse
        for response in FeedbackLessonResponse.objects.filter(
            feedback_lesson=feedback_lesson
        ):
            participant_name = response.learner.name
            participant_email = response.learner.email
            temp_data = {
                "Participant": participant_name,
                "Participant Email": participant_email,
            }
            if response:
                for answer in response.answers.all():
                    question_text = answer.question.text
                    answer_value = (
                        answer.text_answer if answer.text_answer else answer.rating
                    )
                    temp_data[question_text] = answer_value
            else:
                # If participant did not provide feedback, populate with empty values
                for question in feedback_lesson.questions.all():
                    temp_data[question.text] = "-"
            data.append(temp_data)

        # Create a DataFrame from the list of dictionaries
        df = pd.DataFrame(data)

        # Save the DataFrame to an Excel file in-memory
        excel_data = BytesIO()
        df.to_excel(excel_data, index=False)
        excel_data.seek(0)

        # Create the response with the Excel file
        response = HttpResponse(
            excel_data.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f"attachment; filename=Feedback_Report.xlsx"

        return response

    except FeedbackLesson.DoesNotExist:
        return Response(
            {"error": "Feedback lesson not found"}, status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["GET"])
@permission_classes([AllowAny])
def get_consolidated_feedback_download_report(request, live_session_id):
    # Create a new workbook and add a worksheet
    live_session = LiveSession.objects.get(id=live_session_id)
    wb = Workbook()
    ws = wb.active
    # Write headers to the worksheet
    headers = set()
    feedback_lesson_responses = FeedbackLessonResponse.objects.filter(
        feedback_lesson__lesson__course__batch__project__id=live_session.batch.project.id,
        feedback_lesson__live_session__live_session_number=live_session.live_session_number,
        feedback_lesson__live_session__session_type=live_session.session_type,
    )
    facilitator_id = request.query_params.get("facilitator_id")
    if facilitator_id:
        feedback_lesson_responses = feedback_lesson_responses.filter(
            feedback_lesson__live_session__facilitator__id=facilitator_id
        )
    total_participants_in_project = Learner.objects.filter(
        schedularbatch__project__id=live_session.batch.project.id
    ).distinct()
    participants_who_responsded = set()
    for feedback_lesson_response in feedback_lesson_responses:
        participants_who_responsded.add(feedback_lesson_response.learner.id)
        for answer in feedback_lesson_response.answers.all():
            headers.add(answer.question.text)
    headers_list = list(headers)
    headers_list.insert(0, "Participant Name")
    headers_list.insert(1, "Feedback Batch")
    headers_list.insert(2, "Facilitator")

    for col_num, header in enumerate(headers_list, 1):
        ws.cell(row=1, column=col_num, value=header)

    for feedback_lesson_response in feedback_lesson_responses:
        data = ["-" for _ in headers_list]
        participant_index = headers_list.index("Participant Name")
        feedback_batch_index = headers_list.index("Feedback Batch")
        facilitator_index = headers_list.index("Facilitator")
        data[participant_index] = feedback_lesson_response.learner.name
        data[feedback_batch_index] = (
            feedback_lesson_response.feedback_lesson.lesson.course.batch.name
        )
        # if (
        #     feedback_lesson_response
        #     and feedback_lesson_response.feedback_lesson
        #     and feedback_lesson_response.feedback_lesson.live_session
        #     and feedback_lesson_response.feedback_lesson.live_session.facilitator
        # ):
        #     data[facilitator_index] = (
        #         feedback_lesson_response.feedback_lesson.live_session.facilitator.first_name
        #         + " "
        #         + feedback_lesson_response.feedback_lesson.live_session.facilitator.last_name
        #     )
        # else:
        #     data[facilitator_index] = "N/A"
        for answer in feedback_lesson_response.answers.all():
            question_index_in_headers = headers_list.index(answer.question.text)
            if answer.question.type == "descriptive_answer":
                data[question_index_in_headers] = answer.text_answer
            elif (
                answer.question.type == "rating_1_to_5"
                or answer.question.type == "rating_1_to_10"
                or answer.question.type == "rating_0_to_10"
            ):
                data[question_index_in_headers] = answer.rating
        ws.append(data)
    participants_who_responsded_list = list(participants_who_responsded)
    participants_not_responded = total_participants_in_project.exclude(
        id__in=participants_who_responsded_list
    )

    for learner in participants_not_responded:
        data = ["-" for _ in headers_list]
        participant_index = headers_list.index("Participant Name")
        data[participant_index] = learner.name
        ws.append(data)

    # Create a response with the Excel file
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        "attachment; filename=Project_feedback_report.xlsx"
    )
    wb.save(response)
    return response


@api_view(["GET"])
@permission_classes([AllowAny])
def feedback_reports_project_wise_consolidated(request):
    projects = SchedularProject.objects.all()
    res = []
    for project in projects:
        feedback_lesson_responses = FeedbackLessonResponse.objects.filter(
            feedback_lesson__lesson__course__batch__project=project
        )
        facilitator_id = request.query_params.get("facilitator_id")
        if facilitator_id:
            feedback_lesson_responses = feedback_lesson_responses.filter(
                feedback_lesson__live_session__facilitator__id=facilitator_id
            )
        if feedback_lesson_responses.exists():
            total_participants_count = (
                Learner.objects.filter(schedularbatch__project=project)
                .distinct()
                .count()
            )
            percentage_responded = (
                round(
                    (feedback_lesson_responses.count() / total_participants_count)
                    * 100,
                    2,
                )
                if total_participants_count
                else 0
            )
            data = {
                "project_name": project.name,
                "total_participants": total_participants_count,
                "total_responses": feedback_lesson_responses.count(),
                "project_id": project.id,
                "percentage_responded": percentage_responded,
            }
            res.append(data)
    return Response(res)


@api_view(["GET"])
@permission_classes([AllowAny])
def download_consolidated_project_report(request, project_id):
    # Create a new workbook and add a worksheet
    wb = Workbook()
    ws = wb.active
    # Write headers to the worksheet
    headers = set()
    feedback_lesson_responses = FeedbackLessonResponse.objects.filter(
        feedback_lesson__lesson__course__batch__project__id=project_id
    )
    facilitator_id = request.query_params.get("facilitator_id")
    if facilitator_id:
        feedback_lesson_responses = feedback_lesson_responses.filter(
            feedback_lesson__live_session__facilitator__id=facilitator_id
        )
    total_participants_in_project = Learner.objects.filter(
        schedularbatch__project__id=project_id
    ).distinct()
    participants_who_responsded = set()
    for feedback_lesson_response in feedback_lesson_responses:
        participants_who_responsded.add(feedback_lesson_response.learner.id)
        for answer in feedback_lesson_response.answers.all():
            headers.add(answer.question.text)
    headers_list = list(headers)
    headers_list.insert(0, "Participant Name")

    for col_num, header in enumerate(headers_list, 1):
        ws.cell(row=1, column=col_num, value=header)

    for feedback_lesson_response in feedback_lesson_responses:
        data = ["-" for _ in headers_list]
        participant_index = headers_list.index("Participant Name")
        data[participant_index] = feedback_lesson_response.learner.name
        for answer in feedback_lesson_response.answers.all():
            question_index_in_headers = headers_list.index(answer.question.text)
            if answer.question.type == "descriptive_answer":
                data[question_index_in_headers] = answer.text_answer
            elif (
                answer.question.type == "rating_1_to_5"
                or answer.question.type == "rating_1_to_10"
                or answer.question.type == "rating_0_to_10"
            ):
                data[question_index_in_headers] = answer.rating
        ws.append(data)
    participants_who_responsded_list = list(participants_who_responsded)
    participants_not_responded = total_participants_in_project.exclude(
        id__in=participants_who_responsded_list
    )

    for learner in participants_not_responded:
        data = ["-" for _ in headers_list]
        participant_index = headers_list.index("Participant Name")
        data[participant_index] = learner.name
        ws.append(data)

    # Create a response with the Excel file
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        "attachment; filename=Project_feedback_report.xlsx"
    )
    wb.save(response)
    return response


@api_view(["GET"])
def get_nudges_by_project_id(request, project_id):
    # Retrieve nudges filtered by project_id
    nudges = Nudge.objects.filter(course__batch__project__id=project_id)
    serializer = NudgeSerializer(nudges, many=True)
    return Response(serializer.data)


@api_view(["POST"])
def send_nudge_to_email(request, nudge_id):
    email = request.data.get("email")
    try:
        nudge = Nudge.objects.get(id=nudge_id)
    except Nudge.DoesNotExist:
        return Response({"error": "Nudge not found"}, status=404)

    subject = f"New Nudge: {nudge.name}"
    message = nudge.content
    email_msg = EmailMessage(subject, message, settings.DEFAULT_FROM_EMAIL, [email])
    if nudge.file:
        attachment_path = nudge.file.url
        file_content = get_file_content(nudge.file.url)
        extension = get_file_extension(nudge.file.url)
        file_name = f"Attachment.{extension}"
        email_msg.attach(file_name, file_content, f"application/{extension}")
    email_msg.content_subtype = "html"
    email_msg.send()
    return Response({"message": "Nudge sent successfully"})


@api_view(["POST"])
def duplicate_nudge(request, nudge_id, course_id):
    order = request.data.get("order")
    try:
        original_nudge = Nudge.objects.get(id=nudge_id)
        course = Course.objects.get(id=course_id)  # Fetch the course instance
        duplicated_nudge = Nudge.objects.create(
            name=f"{original_nudge.name}",
            content=original_nudge.content,
            file=original_nudge.file,
            order=order,
            course=course,  # Use the fetched course instance
            is_sent=False,  # Assuming the duplicated nudge is not sent yet
        )
        return Response({"message": "Nudge duplicated successfully."})
    except Nudge.DoesNotExist:
        return Response({"error": "Nudge not found"}, status=404)
    except Course.DoesNotExist:
        return Response({"error": "Course not found"}, status=404)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_nps_project_wise(request):
    projects = SchedularProject.objects.all()
    res = {}
    for project in projects:
        nps = 0
        total_questions = 0  # Track the total number of questions
        feedback_lessons = FeedbackLesson.objects.filter(
            lesson__course__batch__project=project
        )
        for feedback_lesson in feedback_lessons:
            feedback_lesson_reponses = FeedbackLessonResponse.objects.filter(
                feedback_lesson=feedback_lesson
            )

            questions_serializer = QuestionSerializer(
                feedback_lesson.questions, many=True
            )
            question_data = {
                question["id"]: {**question, "descriptive_answers": [], "ratings": []}
                for question in questions_serializer.data
            }
            for response in feedback_lesson_reponses:
                for answer in response.answers.all():
                    question_id = answer.question.id
                    if answer.question.type.startswith("rating"):
                        question_data[question_id]["ratings"].append(answer.rating)
                    elif answer.question.type == "descriptive_answer":
                        question_data[question_id]["descriptive_answers"].append(
                            answer.text_answer
                        )

            for question_id, data in question_data.items():
                ratings = data["ratings"]
                if ratings:
                    if data["type"] == "rating_0_to_10":

                        nps += calculate_nps(ratings)
                        total_questions += 1

        if total_questions > 0:
            average_nps = nps / total_questions
        else:
            average_nps = 0

        res[project.id] = average_nps
    return Response(res)


class GetAllNudgesOfSchedularProjects(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, project_id):
        try:
            hr_id = request.query_params.get("hr", None)
            data = []
            courses = None
            if project_id == "all":
                courses = Course.objects.all()
            else:
                courses = Course.objects.filter(batch__project__id=int(project_id))
            if hr_id:
                courses = courses.filter(batch__project__hr__id=hr_id)
            for course in courses:
                nudges = get_nudges_of_course(course)
                data = list(data) + list(nudges)

            return Response(data)
        except Exception as e:
            print(str(e))
            return Response({"error": "Failed to get data"}, status=500)


class CreateAssignmentLesson(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            course_template_id = request.data.get("course_template", "")
            course_id = request.data.get("course", "")
            lesson = None

            if course_template_id != "null":
                course_template = CourseTemplate.objects.get(id=int(course_template_id))
                lesson = Lesson.objects.create(
                    course_template=course_template,
                    name=request.data["name"],
                    status=request.data["status"],
                    lesson_type="assignment",
                    order=int(request.data["order"]),
                )
            elif course_id != "null":
                course = Course.objects.get(id=int(course_id))
                live_session_id = request.data["live_session"]
                live_session = None
                if live_session_id != "null":
                    live_session = LiveSessionSchedular.objects.get(id=live_session_id)
                lesson = Lesson.objects.create(
                    course=course,
                    name=request.data["name"],
                    status=request.data["status"],
                    lesson_type="assignment",
                    drip_date=request.data["drip_date"],
                    order=int(request.data["order"]),
                    live_session=live_session,
                )
            assignment_lesson = AssignmentLesson.objects.create(
                lesson=lesson,
                name=request.data["name"],
                description=request.data["description"],
            )

            return Response(
                {"message": "Assignment Lesson Created."}, status=status.HTTP_200_OK
            )
        except Exception as e:
            print(str(e))
            return Response(
                {"message": "Failed to create assignment."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class UpdateAssignmentLesson(APIView):
    def put(self, request, assignment_id):
        permission_classes = [IsAuthenticated]
        try:

            assignment_lesson = AssignmentLesson.objects.get(
                id=assignment_id,
            )
            assignment_lesson.name = request.data["name"]
            assignment_lesson.description = request.data["description"]

            assignment_lesson.save()
            lesson = Lesson.objects.get(id=assignment_lesson.lesson.id)
            lesson.name = request.data["name"]
            lesson.drip_date = request.data["drip_date"]
            live_session_id = request.data["live_session"]
            live_session = None
            if live_session_id != "null":
                live_session = LiveSessionSchedular.objects.get(id=live_session_id)
            lesson.live_session = live_session
            lesson.save()
            return Response(
                {"message": f"Assignment Lesson Updated."}, status=status.HTTP_200_OK
            )
        except Exception as e:
            print(str(e))
            return Response(
                {"message": f"Failed to update assignment."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class GetAllAssignmentsResponses(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, user_type, user_id):

        assignments = None

        if user_type == "pmo":
            assignments = AssignmentLessonResponse.objects.all()
        elif user_type == "learner":
            learner = Learner.objects.get(id=user_id)
            assignments = AssignmentLessonResponse.objects.filter(learner=learner)
        elif user_type == "coach":
            coach = Coach.objects.get(id=user_id)
            assignments = AssignmentLessonResponse.objects.filter(
                assignment_lesson__lesson__course__batch__coaches=coach
            )

        data = []
        for assignment in assignments:
            serializer = AssignmentResponseSerializer(assignment)
            temp = {
                "id": assignment.id,
                "name": assignment.assignment_lesson.name,
                "learner_name": assignment.learner.name,
                "batch_name": assignment.assignment_lesson.lesson.course.batch.name,
                "project_name": assignment.assignment_lesson.lesson.course.batch.project.name,
                "file": serializer.data["file"],
                "org_name": assignment.assignment_lesson.lesson.course.batch.project.organisation.name,
                "created_at": assignment.created_at,
                "edited_at": assignment.edited_at,
            }
            data.append(temp)

        return Response(data)


class CreateAssignmentLessonResponse(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            file = request.data["file"]
            assignment_id = request.data["assignment_id"]
            learner_id = request.data["learner_id"]

            assignment = AssignmentLesson.objects.get(id=assignment_id)

            learner = Learner.objects.get(id=learner_id)

            assignment_response = AssignmentLessonResponse.objects.create(
                assignment_lesson=assignment, file=file, learner=learner
            )

            return Response(
                {"message": f"File Uploaded Sucessfully."}, status=status.HTTP_200_OK
            )

        except Exception as e:
            print(str(e))
            return Response(
                {"message": f"Failed to upload file."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class GetAssignmentsResponses(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, assignment_id, learner_id):
        try:
            assignment_response = AssignmentLessonResponse.objects.get(
                assignment_lesson__id=assignment_id, learner__id=learner_id
            )

            serializer = AssignmentResponseSerializerDepthSix(assignment_response)
            return Response(serializer.data)
        except Exception as e:
            print(str(e))
            return Response(
                {"message": f"Failed to get data."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class UpdateAssignmentLessonFile(APIView):
    permission_classes = [IsAuthenticated]

    def put(self, request):
        try:
            file = request.data.get("file", None)
            assignment_response_id = request.data.get("assignment_response_id", None)
            assignment_response = AssignmentLessonResponse.objects.get(
                id=int(assignment_response_id)
            )
            assignment_response.file = file
            assignment_response.save()

            return Response({"message": "File updated sucessfully."}, status=200)

        except Exception as e:
            print(str(e))
            return Response(
                {"message": f"Failed to update file."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def submit_feedback(request, feedback_id, learner_id):
    try:

        feedback = get_object_or_404(Feedback, id=feedback_id)
        learner = get_object_or_404(Learner, id=learner_id)
    except (
        Feedback.DoesNotExist,
        Learner.DoesNotExist,
    ) as e:
        return Response(
            {"error": "Failed to submit feedback."}, status=status.HTTP_404_NOT_FOUND
        )

    caas_session_id = request.data.get("caas_session_id", "")
    schedular_session_id = request.data.get("schedular_session_id", "")
    caas_session = None
    schedular_session = None
    if caas_session_id:
        caas_session = SessionRequestCaas.objects.get(id=caas_session_id)
    if schedular_session_id:
        schedular_session = SchedularSessions.objects.get(id=schedular_session_id)
    if caas_session or schedular_session:
        answers_data = request.data.get(
            "answers",
        )
        serializer = AnswerSerializer(data=answers_data, many=True)
        if serializer.is_valid():
            answers = serializer.save()
            coaching_session_response = CoachingSessionsFeedbackResponse.objects.create(
                feedback=feedback,
                learner=learner,
                caas_session=caas_session,
                schedular_session=schedular_session,
            )
            coaching_session_response.answers.set(answers)
            coaching_session_response.save()
            return Response(
                {"detail": "Feedback submitted successfully"}, status=status.HTTP_200_OK
            )
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    else:
        return Response(
            {"error": "Failed to submit feedback"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_feedback(request, feedback_id):
    feedback = Feedback.objects.get(id=feedback_id)
    serializer = FeedbackDepthOneSerializer(feedback)
    return Response(serializer.data)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_end_meeting_feedback_response_data(request):

    try:
        coach_session_feedback_responses = (
            CoachingSessionsFeedbackResponse.objects.all()
        )
        data = []
        for coach_session_feedback_response in coach_session_feedback_responses:
            temp = {}
            cass_session = coach_session_feedback_response.caas_session
            if cass_session:
                if cass_session.coach:
                    coach_name = (
                        cass_session.coach.first_name
                        + " "
                        + cass_session.coach.last_name
                    )

                else:
                    coach_name = None

                temp = {
                    "feedback_responses_id": coach_session_feedback_response.id,
                    "coach_name": coach_name,
                    "project_name": cass_session.project.name,
                    "org_name": cass_session.project.organisation.name,
                    "coachee_name": cass_session.learner.name,
                    "session_type": cass_session.session_type,
                    "session_number": cass_session.session_number,
                    "type": "CAAS",
                }
            else:
                seeq_session = coach_session_feedback_response.schedular_session

                temp = {
                    "feedback_responses_id": coach_session_feedback_response.id,
                    "coach_name": seeq_session.availibility.coach.first_name
                    + " "
                    + seeq_session.availibility.coach.last_name,
                    "project_name": seeq_session.coaching_session.batch.project.name,
                    "org_name": seeq_session.coaching_session.batch.project.organisation.name,
                    "coachee_name": seeq_session.learner.name,
                    "session_type": seeq_session.coaching_session.session_type,
                    "session_number": seeq_session.coaching_session.coaching_session_number,
                    "type": "SEEQ",
                }

            for answer in coach_session_feedback_response.answers.all():
                if answer.question.type == "rating_1_to_5":
                    temp["sesson_rating"] = answer.rating
                    break

            data.append(temp)

        return Response(data)
    except Exception as e:
        print(str(e))
        return Response(
            {"error": "Failed to get data"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_coach_session_feedback_response_data(request, feedback_response_id):

    try:
        coach_session_feedback_responses = CoachingSessionsFeedbackResponse.objects.get(
            id=feedback_response_id
        )

        data = []

        for answer in coach_session_feedback_responses.answers.all():

            temp = {
                "question": answer.question.text,
                "rating": answer.rating,
                "selected_answer": answer.selected_options,
                "type": answer.question.type,
            }

            data.append(temp)
        return Response(data)
    except Exception as e:
        print(str(e))
        return Response(
            {"error": "Failed to get data"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


class FacilitatorWiseFeedback(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, feedback_id):
        try:
            feedback = FeedbackLesson.objects.get(id=feedback_id)
        except Exception as e:
            return Response(
                {"message": f"Failed to update file."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_live_sessions_by_course(request, course_id):
    live_sessions = LiveSession.objects.filter(batch__course__id=course_id)
    live_sessions_serializer = LiveSessionSchedularSerializer(live_sessions, many=True)
    return Response(live_sessions_serializer.data)
