from datetime import date, datetime, time
import requests
from django.http import JsonResponse
import calendar
from os import name
from django.shortcuts import render, get_object_or_404
from django.template.loader import render_to_string
from django.utils.safestring import mark_safe
from django.utils.timezone import make_aware
from django_celery_beat.models import PeriodicTask, ClockedSchedule
from django.db import transaction, IntegrityError
from django.core.mail import EmailMessage
from rest_framework.exceptions import ParseError, ValidationError
from django.core.exceptions import ObjectDoesNotExist
from operationsBackend import settings
from .serializers import (
    CoachSerializer,
    UserSerializer,
    LeaderDepthOneSerializer,
    LeaderSerializer,
    LearnerSerializer,
    PmoDepthOneSerializer,
    SessionRequestCaasSerializer,
    CoachDepthOneSerializer,
    ProjectDepthTwoSerializer,
    HrSerializer,
    OrganisationSerializer,
    ProjectSerializer,
    LearnerDepthOneSerializer,
    HrDepthOneSerializer,
    SessionRequestCaasDepthOneSerializer,
    SessionRequestCaasDepthTwoSerializer,
    AvailibilitySerializer,
    NotificationSerializer,
    EngagementDepthOneSerializer,
    GoalSerializer,
    GetGoalSerializer,
    CompetencySerializer,
    CompetencyDepthOneSerializer,
    ActionItemSerializer,
    GetActionItemDepthOneSerializer,
    PendingActionItemSerializer,
    EngagementSerializer,
    SessionRequestWithEngagementCaasDepthOneSerializer,
    ProfileEditActivitySerializer,
    UserLoginActivitySerializer,
    AddGoalActivitySerializer,
    AddCoachActivitySerializer,
    SentEmailActivitySerializer,
    CoachProfileTemplateSerializer,
    StandardizedFieldSerializer,
    StandardizedFieldRequestSerializer,
    StandardizedFieldRequestDepthOneSerializer,
    SessionRequestedActivitySerializer,
    DeleteCoachProfileActivitySerializer,
    RemoveCoachActivitySerializer,
    PastSessionActivitySerializer,
    TemplateSerializer,
    ProjectContractSerializer,
    CoachContractSerializer,
    UpdateSerializer,
    UpdateDepthOneSerializer,
    UserTokenSerializer,
    CalendarEventSerializer,
    ShareCoachProfileActivitySerializer,
    CreateProjectActivitySerializer,
    FinalizeCoachActivitySerializer,
    SessionDataSerializer,
    SessionRequestWithEngagementCaasAndIsSeeqProjectDepthOneSerializer,
    SuperAdminDepthOneSerializer,
    FinanceDepthOneSerializer,
    PmoSerializer,
    FacilitatorDepthOneSerializer,
    FacilitatorSerializer,
    APILogSerializer,
    PmoSerializerAll,
    CTTPmoSerializer,
    CTTPmoDepthOneSerializer,
    ProjectDepthTwoSerializerArchiveCheck,
    CustomUserSerializer,
    SalesSerializer,
    SalesDepthOneSerializer,
    GoalDescriptionSerializer,
    CoachProfileShareSerializer,
)
from zohoapi.serializers import (
    VendorDepthOneSerializer,
    PurchaseOrderSerializer,
    PurchaseOrderGetSerializer,
)
from zohoapi.views import get_organization_data, get_vendor, fetch_purchase_orders
from zohoapi.tasks import (
    organization_id,
    get_access_token,
    base_url,
    filter_purchase_order_data,
    purchase_orders_allowed,
    purchase_orders_allowed,
)
from .permissions import IsInRoles
from rest_framework import generics
from django.utils.crypto import get_random_string
import jwt
import jwt
import uuid
import pytz
import math
from django.db.models import IntegerField
from django.db.models.functions import Cast
from rest_framework.exceptions import AuthenticationFailed
from datetime import datetime, timedelta
from rest_framework.response import Response
from django.core.mail import EmailMessage, BadHeaderError
from django.contrib.auth.models import User
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated, AllowAny
from django.forms.models import model_to_dict
from .models import (
    Profile,
    Leader,
    Pmo,
    CTTPmo,
    Coach,
    OTP,
    Project,
    HR,
    Organisation,
    SessionRequestCaas,
    Availibility,
    Learner,
    CoachStatus,
    Notification,
    Engagement,
    Goal,
    Competency,
    ActionItem,
    ProfileEditActivity,
    UserLoginActivity,
    AddGoalActivity,
    AddCoachActivity,
    SentEmailActivity,
    CoachProfileTemplate,
    StandardizedField,
    StandardizedFieldRequest,
    Update,
    SessionRequestedActivity,
    DeleteCoachProfileActivity,
    RemoveCoachActivity,
    PastSessionActivity,
    Template,
    Role,
    UserToken,
    CalendarEvent,
    ShareCoachProfileActivity,
    CreateProjectActivity,
    FinalizeCoachActivity,
    APILog,
    Facilitator,
    SuperAdmin,
    Finance,
    Sales,
    TableHiddenColumn,
    CoachProfileShare,
)
from rest_framework.pagination import PageNumberPagination
from rest_framework.authtoken.models import Token
from django.contrib.auth import authenticate, login, logout
from django.core.mail import send_mail
from django.utils import timezone
from django.middleware.csrf import get_token
from django.views.decorators.csrf import ensure_csrf_cookie
import json
import string
import random
from django.db.models import Q, Min, F, Exists, OuterRef
from collections import defaultdict
from django.db.models import Avg
from rest_framework import status
from rest_framework.views import APIView

from django.db.models import Count, Sum, Case, When, IntegerField
import io
from openpyxl import Workbook
from openpyxl.styles import Font
from rest_framework import generics
from django.db.models import Subquery, OuterRef, Value, BooleanField
from schedularApi.models import (
    SchedularBatch,
    SchedularSessions,
    SchedularProject,
    SchedularBatch,
    SchedularSessions,
    CalendarInvites,
    CoachingSession,
    LiveSession,
    HandoverDetails,
    Task,
    Expense,
    CoachContract,
    ProjectContract,
)
from schedularApi.serializers import (
    SchedularProjectSerializer,
    TaskSerializer,
    ExpenseSerializerDepthOne,
)
from django_rest_passwordreset.models import ResetPasswordToken
from django_rest_passwordreset.serializers import EmailSerializer
from django_rest_passwordreset.tokens import get_token_generator
from zohoapi.models import Vendor, InvoiceData, OrdersAndProjectMapping, PurchaseOrder
from courses.models import CourseEnrollment, CoachingSessionsFeedbackResponse, Answer
from urllib.parse import urlencode
from django.http import HttpResponseRedirect
import pdfkit
import os

# Create your views here.
from collections import defaultdict
import pandas as pd
from django.http import HttpResponse
import environ
from time import sleep
from django.db.models import Max


env = environ.Env()

wkhtmltopdf_path = os.environ.get("WKHTMLTOPDF_PATH", r"/usr/local/bin/wkhtmltopdf")

pdfkit_config = pdfkit.configuration(wkhtmltopdf=f"{wkhtmltopdf_path}")


def get_current_date_timestamps():
    now = timezone.now()
    current_date = now.date()
    start_timestamp = str(
        int(datetime.combine(current_date, datetime.min.time()).timestamp() * 1000)
    )
    end_timestamp = str(
        int(datetime.combine(current_date, datetime.max.time()).timestamp() * 1000)
    )
    return start_timestamp, end_timestamp


def get_live_session_name(session_type):
    session_name = None
    if session_type == "live_session":
        session_name = "Live Session"
    elif session_type == "check_in_session":
        session_name = "Check In Session"
    elif session_type == "in_person_session":
        session_name = "In Person Session"
    elif session_type == "pre_study":
        session_name = "Pre Study"
    elif session_type == "kickoff_session":
        session_name = "Kickoff Session"
    elif session_type == "virtual_session":
        session_name = "Virtual Session"
    return session_name


def calculate_nps(ratings):
    promoters = sum(rating >= 9 for rating in ratings)
    detractors = sum(rating <= 6 for rating in ratings)
    nps = ((promoters - detractors) / len(ratings)) * 100
    return nps


def calculate_nps_from_answers(answers):
    ratings = [answer.rating for answer in answers]
    if ratings:
        return calculate_nps(ratings)
    return None


def add_contact_in_wati(user_type, name, phone):
    try:
        wati_api_endpoint = env("WATI_API_ENDPOINT")
        wati_authorization = env("WATI_AUTHORIZATION")
        wati_api_url = f"{wati_api_endpoint}/api/v1/addContact/{phone}"
        headers = {
            "content-type": "text/json",
            "Authorization": wati_authorization,
        }
        payload = {
            "customParams": [
                {
                    "name": "user_type",
                    "value": user_type,
                },
            ],
            "name": name,
        }
        response = requests.post(wati_api_url, headers=headers, json=payload)
        response.raise_for_status()  # Raise an HTTPError for bad responses
        print(response.json())
        return response.json()
    except Exception as e:
        pass


def create_send_email(user_email, file_name):
    try:
        user = User.objects.get(username=user_email)
        sent_email = SentEmailActivity.objects.create(
            user=user,
            email_subject=file_name,
            timestamp=timezone.now(),
        )
        sent_email.save()
    except Exception as e:
        pass


def send_mail_templates(file_name, user_email, email_subject, content, bcc_emails):
    try:
        email_message = render_to_string(file_name, content)

        email = EmailMessage(
            f"{env('EMAIL_SUBJECT_INITIAL',default='')} {email_subject}",
            email_message,
            settings.DEFAULT_FROM_EMAIL,
            user_email,
            bcc_emails,
        )
        email.content_subtype = "html"

        email.send(fail_silently=False)
        for email in user_email:
            create_send_email(email, file_name)
    except Exception as e:
        print(f"Error occurred while sending emails: {str(e)}")


def convert_to_24hr_format(time_str):
    time_obj = datetime.strptime(time_str, "%I:%M %p")
    time_24hr = time_obj.strftime("%H:%M")
    return time_24hr


def refresh_google_access_token(user_token):
    if not user_token:
        return None

    refresh_token = user_token.refresh_token
    access_token_expiry = user_token.access_token_expiry
    auth_code = user_token.authorization_code
    if not refresh_token:
        return None

    access_token_expiry = int(access_token_expiry)

    expiration_timestamp = user_token.updated_at + timezone.timedelta(
        seconds=access_token_expiry
    )

    if expiration_timestamp <= timezone.now():
        token_url = "https://oauth2.googleapis.com/token"
        token_data = {
            "refresh_token": refresh_token,
            "client_id": env("GOOGLE_OAUTH2_CLIENT_ID"),
            "client_secret": env("GOOGLE_OAUTH2_CLIENT_SECRET"),
            "grant_type": "refresh_token",
        }

        response = requests.post(token_url, data=token_data)
        token_json = response.json()

        if "access_token" in token_json:
            user_token.access_token = token_json["access_token"]
            user_token.access_token_expiry = token_json.get("expires_in")
            user_token.updated_at = timezone.now()
            user_token.save()

            return user_token.access_token

    return user_token.access_token


def refresh_microsoft_access_token(user_token):
    if not user_token:
        return None

    refresh_token = user_token.refresh_token
    access_token_expiry = user_token.access_token_expiry
    auth_code = user_token.authorization_code
    if not refresh_token:
        return None

    access_token_expiry = int(access_token_expiry)

    expiration_timestamp = user_token.updated_at + timezone.timedelta(
        seconds=access_token_expiry
    )

    if expiration_timestamp <= timezone.now():
        token_url = f"https://login.microsoftonline.com/common/oauth2/v2.0/token"

        token_data = {
            "client_id": env("MICROSOFT_CLIENT_ID"),
            "refresh_token": refresh_token,
            "grant_type": "refresh_token",
            "client_secret": env("MICROSOFT_CLIENT_SECRET"),
        }

        response = requests.post(token_url, data=token_data)
        token_json = response.json()

        if "access_token" in token_json:
            user_token.access_token = token_json["access_token"]
            user_token.access_token_expiry = token_json.get("expires_in")
            user_token.updated_at = timezone.now()
            user_token.save()

            return user_token.access_token

    return user_token.access_token


def create_google_calendar_event(access_token, event_details, attendee_email, session):
    try:
        formatted_date = datetime.strptime(
            event_details.get("startDate"), "%d-%m-%Y"
        ).strftime("%Y-%m-%d")

        start_time = convert_to_24hr_format(event_details.get("startTime"))
        end_time = convert_to_24hr_format(event_details.get("endTime"))

        event_details_title = event_details.get("title")
        if event_details.get("title") == "Coaching Session Session":
            event_details_title = "Coaching Session"

        event_data = {
            "summary": event_details_title,
            "description": event_details.get("description"),
            "start": {
                "dateTime": f"{formatted_date}T{start_time}:00",
                "timeZone": "IST",
            },
            "end": {
                "dateTime": f"{formatted_date}T{end_time}:00",
                "timeZone": "IST",
            },
            "attendees": [
                {"email": attendee_email},
            ],
        }

        user_token = UserToken.objects.get(access_token=access_token)
        new_access_token = refresh_google_access_token(user_token)
        if not new_access_token:
            new_access_token = access_token

        # Make a POST request to the Google Calendar API
        response = requests.post(
            "https://www.googleapis.com/calendar/v3/calendars/primary/events",
            json=event_data,
            headers={
                "Authorization": f"Bearer {new_access_token}",
                "Content-Type": "application/json",
            },
        )

        if response.status_code == 200:
            response_data = response.json()

            calendar_event = CalendarEvent(
                event_id=response_data.get("id"),
                title=event_details_title,
                description=event_details.get("description"),
                start_datetime=f"{formatted_date}T{start_time}:00",
                end_datetime=f"{formatted_date}T{end_time}:00",
                attendee=attendee_email,
                creator=response_data.get("creator", {}).get("email", ""),
                session=session,
                account_type="google",
            )
            calendar_event.save()

            return {
                "message": "Event created successfully",
                "event_data": response_data,
            }
        else:
            return {
                "error": "Failed to create event",
                "status_code": response.status_code,
            }

    except Exception as e:
        return {"error": "An error occurred", "details": str(e)}


def create_microsoft_calendar_event(
    access_token, event_details, attendee_email_name, session
):
    event_create_url = "https://graph.microsoft.com/v1.0/me/events"

    formatted_date = datetime.strptime(event_details["startDate"], "%d-%m-%Y").strftime(
        "%Y-%m-%d"
    )

    start_datetime = (
        f"{formatted_date}T{convert_to_24hr_format(event_details['startTime'])}:00"
    )
    end_datetime = (
        f"{formatted_date}T{convert_to_24hr_format(event_details['endTime'])}:00"
    )

    event_details_title = event_details["title"]
    if event_details["title"] == "Coaching Session Session":
        event_details_title = "Coaching Session"

    event_payload = {
        "subject": event_details_title,
        "body": {"contentType": "HTML", "content": event_details["description"]},
        "start": {"dateTime": start_datetime, "timeZone": "Asia/Kolkata"},
        "end": {"dateTime": end_datetime, "timeZone": "Asia/Kolkata"},
        "attendees": [{"emailAddress": attendee_email_name, "type": "required"}],
    }

    user_token = UserToken.objects.get(access_token=access_token)
    new_access_token = refresh_microsoft_access_token(user_token)
    if not new_access_token:
        new_access_token = access_token

    headers = {
        "Authorization": f"Bearer {new_access_token}",
        "Content-Type": "application/json",
    }

    response = requests.post(event_create_url, json=event_payload, headers=headers)

    if response.status_code == 201:
        microsoft_response_data = response.json()

        calendar_event = CalendarEvent(
            event_id=microsoft_response_data.get("id"),
            title=event_details_title,
            description=event_details.get("description"),
            start_datetime=start_datetime,
            end_datetime=end_datetime,
            attendee=attendee_email_name.get("address"),
            creator=microsoft_response_data.get("organizer", {})
            .get("emailAddress", {})
            .get("address", ""),
            session=session,
            account_type="microsoft",
        )
        calendar_event.save()

        print("Event created successfully.")
        return True
    else:
        print(f"Event creation failed. Status code: {response.status_code}")
        print(response.text)
        return False


def delete_google_calendar_event(access_token, event_id):
    try:
        response = requests.delete(
            f"https://www.googleapis.com/calendar/v3/calendars/primary/events/{event_id}",
            headers={
                "Authorization": f"Bearer {access_token}",
                "Content-Type": "application/json",
            },
        )

        if response.status_code == 204:
            return {"message": "Event deleted successfully"}
        elif response.status_code == 404:
            return {"error": "Event not found"}
        else:
            return {
                "error": "Failed to delete event",
                "status_code": response.status_code,
            }

    except Exception as e:
        return {"error": "An error occurred", "details": str(e)}


def delete_microsoft_calendar_event(access_token, event_id):
    try:
        event_delete_url = f"https://graph.microsoft.com/v1.0/me/events/{event_id}"

        headers = {
            "Authorization": f"Bearer {access_token}",
        }

        response = requests.delete(event_delete_url, headers=headers)

        if response.status_code == 204:
            return {"message": "Event deleted successfully"}
        elif response.status_code == 404:
            return {"error": "Event not found"}
        else:
            return {
                "error": "Failed to delete event",
                "status_code": response.status_code,
            }

    except Exception as e:
        return {"error": "An error occurred", "details": str(e)}


def create_teams_meeting(user_email, live_session_id, topic, start_time, end_time):
    try:
        event_create_url = "https://graph.microsoft.com/v1.0/me/onlineMeetings"
        user_token = UserToken.objects.get(user_profile__user__username=user_email)
        new_access_token = refresh_microsoft_access_token(user_token)
        if not new_access_token:
            new_access_token = user_token.access_token
        headers = {
            "Authorization": f"Bearer {new_access_token}",
            "Content-Type": "application/json",
        }
        event_payload = {
            "startDateTime": start_time,
            "endDateTime": end_time,
            "subject": topic,
        }
        response = requests.post(event_create_url, json=event_payload, headers=headers)
        print(response.json())
        if response.status_code == 201:
            meeting_info = response.json()
            meeting_link = meeting_info.get("joinWebUrl")
            live_session = LiveSession.objects.get(id=live_session_id)
            live_session.meeting_link = meeting_link
            live_session.teams_meeting_id = meeting_info.get("id")
            live_session.save()
            print("Meeting Link Generated")
            return True
        else:
            return False
    except Exception as e:
        print(str(e))
        return False


def delete_teams_meeting(user_email, live_session):
    user_token = UserToken.objects.get(user_profile__user__username=user_email)
    new_access_token = refresh_microsoft_access_token(user_token)
    if not new_access_token:
        new_access_token = user_token.access_token
    meeting_delete_url = f"https://graph.microsoft.com/v1.0/me/onlineMeetings/{live_session.teams_meeting_id}"
    headers = {
        "Authorization": f"Bearer {new_access_token}",
    }
    response = requests.delete(meeting_delete_url, headers=headers)
    if response.status_code == 204:
        # live_session.meeting_link = ""
        # live_session.save()
        return {"message": "Event deleted successfully"}
    elif response.status_code == 404:
        return {"error": "Event not found"}
    else:
        return {
            "error": "Failed to delete event",
            "status_code": response.status_code,
        }


def create_outlook_calendar_invite(
    subject,
    description,
    start_time_stamp,
    end_time_stamp,
    attendees,
    user_email,
    caas_session,
    schedular_session,
    live_session,
    meeting_location,
):
    event_create_url = "https://graph.microsoft.com/v1.0/me/events"
    try:

        user_token = UserToken.objects.get(user_profile__user__username=user_email)
        new_access_token = refresh_microsoft_access_token(user_token)
        if not new_access_token:
            new_access_token = user_token.access_token
        headers = {
            "Authorization": f"Bearer {new_access_token}",
            "Content-Type": "application/json",
        }
        start_datetime_obj = datetime.fromtimestamp(
            int(start_time_stamp) / 1000
        ) + timedelta(hours=5, minutes=30)
        end_datetime_obj = datetime.fromtimestamp(
            int(end_time_stamp) / 1000
        ) + timedelta(hours=5, minutes=30)
        start_datetime = start_datetime_obj.strftime("%Y-%m-%dT%H:%M:00")
        end_datetime = end_datetime_obj.strftime("%Y-%m-%dT%H:%M:00")
        event_payload = {
            "subject": subject,
            "body": {"contentType": "HTML", "content": description},
            "start": {"dateTime": start_datetime, "timeZone": "Asia/Kolkata"},
            "end": {"dateTime": end_datetime, "timeZone": "Asia/Kolkata"},
            "attendees": attendees,
            "location": {
                "displayName": meeting_location if meeting_location else "",
            },
        }
        response = requests.post(event_create_url, json=event_payload, headers=headers)
        if response.status_code == 201:
            microsoft_response_data = response.json()
            calendar_invite = CalendarInvites(
                event_id=microsoft_response_data.get("id"),
                title=subject,
                description=description,
                start_datetime=start_datetime,
                end_datetime=end_datetime,
                attendees=attendees,
                creator=user_email,
                caas_session=caas_session,
                schedular_session=schedular_session,
                live_session=live_session,
            )
            calendar_invite.save()
            print("Calendar invite sent successfully.")
            return True
        else:
            print(f"Calendar invitation failed. Status code: {response.status_code}")
            print(response.text)
            return False

    except UserToken.DoesNotExist:
        print(f"User token not found for email: {user_email}")
        return False

    except Exception as e:
        print(f"An error occurred: {str(e)}")
        return False


def delete_outlook_calendar_invite(calendar_invite):
    try:
        user_token = UserToken.objects.get(
            user_profile__user__username=calendar_invite.creator
        )
        new_access_token = refresh_microsoft_access_token(user_token)
        if not new_access_token:
            new_access_token = user_token.access_token
        event_delete_url = (
            f"https://graph.microsoft.com/v1.0/me/events/{calendar_invite.event_id}"
        )
        headers = {
            "Authorization": f"Bearer {new_access_token}",
        }
        response = requests.delete(event_delete_url, headers=headers)
        if response.status_code == 204:
            calendar_invite.delete()
            return {"message": "Event deleted successfully"}
        elif response.status_code == 404:
            return {"error": "Event not found"}
        else:
            return {
                "error": "Failed to delete event",
                "status_code": response.status_code,
            }

    except Exception as e:
        return {"error": "An error occurred", "details": str(e)}


def update_outlook_attendees(calendar_invite, new_attendees):
    try:
        user_token = UserToken.objects.get(
            user_profile__user__username=calendar_invite.creator
        )
        new_access_token = refresh_microsoft_access_token(user_token)
        if not new_access_token:
            new_access_token = user_token.access_token

        graph_api_url = (
            f"https://graph.microsoft.com/v1.0/me/events/{calendar_invite.event_id}"
        )

        headers = {
            "Authorization": f"Bearer {new_access_token}",
            "Content-Type": "application/json",
        }

        response = requests.patch(
            graph_api_url, json={"attendees": new_attendees}, headers=headers
        )

        if response.status_code == 200:
            print("Attendees updated successfully.")

        else:
            print(f"Failed to update attendees: {response.text}")

    except Exception as e:
        print(str(e))


def create_notification(user, path, message):
    notification = Notification.objects.create(user=user, path=path, message=message)
    return notification


#  this returns in dd-mm-yyyy hh:mm a
def format_timestamp(timestamp):
    dt = datetime.fromtimestamp(timestamp / 1000)  # Convert milliseconds to seconds
    return dt.strftime("%d-%m-%Y %I:%M %p")


def get_date(timestamp):
    dt = datetime.fromtimestamp(timestamp / 1000) + timedelta(
        hours=5, minutes=30
    )  # Convert milliseconds to seconds
    return dt.strftime("%d-%m-%Y")


def get_time(timestamp):
    dt = datetime.fromtimestamp(timestamp / 1000) + timedelta(
        hours=5, minutes=30
    )  # Convert milliseconds to seconds
    return dt.strftime("%I:%M %p")


def generateManagementToken():
    expires = 24 * 3600
    now = datetime.utcnow()
    exp = now + timedelta(seconds=expires)
    return jwt.encode(
        payload={
            "access_key": env("100MS_APP_ACCESS_KEY"),
            "type": "management",
            "version": 2,
            "jti": str(uuid.uuid4()),
            "iat": now,
            "exp": exp,
            "nbf": now,
        },
        key=env("100MS_APP_SECRET"),
    )


def generate_room_id(email):
    management_token = generateManagementToken()

    try:
        payload = {
            "name": email.replace(".", "-").replace("@", ""),
            "description": "This is a sample description for the room",
            "region": "in",
        }

        response_from_100ms = requests.post(
            "https://api.100ms.live/v2/rooms",
            headers={
                "Authorization": f"Bearer {management_token}",
                "Content-Type": "application/json",
            },
            json=payload,
        )
        if response_from_100ms.status_code == 200:
            room_id = response_from_100ms.json().get("id")
            return room_id
        else:
            return None
    except Exception as e:
        print(f"Error while generating meeting link: {str(e)}")
        return None


def get_trimmed_emails(emails):
    res = []
    for email in emails:
        res.append(email.strip().lower())
    return res


def add_so_to_project(project_type, project_id, sales_order_ids):
    if project_type == "CAAS":
        orders_and_project_mapping = OrdersAndProjectMapping.objects.filter(
            project=project_id
        )
        project = Project.objects.get(id=project_id)
        schedular_project = None
    elif project_type == "SEEQ":
        orders_and_project_mapping = OrdersAndProjectMapping.objects.filter(
            schedular_project=project_id
        )
        schedular_project = SchedularProject.objects.get(id=project_id)
        project = None

    if not orders_and_project_mapping.exists():
        for id in sales_order_ids:
            orders_and_project_mapping = OrdersAndProjectMapping.objects.filter(
                sales_order_ids__contains=id
            )
            if orders_and_project_mapping.exists():
                mapping = orders_and_project_mapping.first()
                if project_type == "CAAS":
                    if mapping.schedular_project:
                        raise Exception(
                            f"SO already exist in Schedular Project: {mapping.schedular_project.name}"
                        )
                    if mapping.project and mapping.project.id != project.id:
                        raise Exception(
                            f"SO already exist in project: {mapping.project.name}"
                        )

                if project_type == "SEEQ":
                    if mapping.project:
                        raise Exception(
                            f"SO already exist in Coaching Project: {mapping.project.name}"
                        )
                    if (
                        mapping.schedular_project
                        and mapping.schedular_project.id != schedular_project.id
                    ):
                        raise Exception(
                            f"SO already exist in project: {mapping.schedular_project.name}"
                        )
                break
    if orders_and_project_mapping.exists():
        mapping = orders_and_project_mapping.first()
        existing_sales_order_ids = mapping.sales_order_ids
        set_of_sales_order_ids = set(existing_sales_order_ids)
        for id in sales_order_ids:
            set_of_sales_order_ids.add(id)
        final_list_of_sales_order_ids = list(set_of_sales_order_ids)
        mapping.sales_order_ids = final_list_of_sales_order_ids
        mapping.project = project
        mapping.schedular_project = schedular_project
        mapping.save()
    else:
        OrdersAndProjectMapping.objects.create(
            project=project,
            sales_order_ids=sales_order_ids,
            schedular_project=schedular_project,
        )


def get_available_credit_for_project(project_id, status):
    try:
        project = Project.objects.get(id=int(project_id))
        filters = Q(project=project)

        if status == "both":
            filters &= Q(status__in=["booked", "completed"])
        else:
            filters &= Q(status=status)

        total_durations = (
            SessionRequestCaas.objects.filter(filters).aggregate(
                total_duration=Sum("session_duration")
            )["total_duration"]
            or 0
        )

        return project.total_credits - total_durations

    except Exception as e:
        print(str(e))
        return None


def credits_needed_for_an_engagement(engagement):
    try:
        total_durations = (
            SessionRequestCaas.objects.filter(engagement=engagement).aggregate(
                total_duration=Sum("session_duration")
            )["total_duration"]
            or 0
        )

        return total_durations

    except Exception as e:
        print(str(e))
        return None


def credits_needed_based_for_project_structure(project):
    try:
        total_duration = 0
        for session in project.project_structure:
            if session["billable"]:
                total_duration += (
                    session["session_duration"] * session["no_of_sessions"]
                )

        return total_duration

    except Exception as e:
        print(str(e))
        return None


SESSION_TYPE_VALUE = {
    "chemistry": "Chemistry",
    "tripartite": "Tripartite",
    "goal_setting": "Goal Setting",
    "coaching_session": "Coaching Session",
    "mid_review": "Mid Review",
    "end_review": "End Review",
    "closure_session": "Closure Session",
    "stakeholder_without_coach": "Tripartite Without Coach",
    "interview": "Interview",
    "stakeholder_interview": "Stakeholder Interview",
}

FIELD_NAME_VALUES = {
    "location": "Work Location",
    "other_certification": "Assessment Certification",
    "area_of_expertise": "Industry",
    "job_roles": "Job roles",
    "companies_worked_in": "Companies worked in",
    "language": "Language Proficiency",
    "education": "Education institutions",
    "domain": "Functional Domain",
    "client_companies": "Client companies",
    "educational_qualification": "Educational Qualification",
    "city": "City",
    "country": "Country",
    "topic": "Topic",
}

SESSIONS_WITH_STAKEHOLDERS = [
    "tripartite",
    "mid_review",
    "end_review",
    "stakeholder_without_coach",
    "stakeholder_interview",
]


def get_booked_session_of_user_confirmed_avalibility(user_type, user_id, date):
    date_obj = datetime.strptime(date, "%Y-%m-%d")
    start_time = date_obj.replace(hour=0, minute=0, second=0)
    # Set the end time of the given date at 23:59:59
    end_time = date_obj.replace(hour=23, minute=59, second=59)
    # Convert the datetime objects to timestamps
    start_timestamp = int(start_time.timestamp())
    end_timestamp = int(end_time.timestamp())

    if user_type == "hr":
        session_requests = SessionRequestCaas.objects.filter(
            hr__id=user_id,
            confirmed_availability__start_time__range=(start_timestamp, end_timestamp),
        )
    elif user_type == "learner":
        session_requests = SessionRequestCaas.objects.filter(
            learner__id=user_id,
            confirmed_availability__start_time__range=(start_timestamp, end_timestamp),
        )
    elif user_type == "coach":
        session_requests = SessionRequestCaas.objects.filter(
            coach__id=user_id,
            confirmed_availability__start_time__range=(start_timestamp, end_timestamp),
        )
    return session_requests


def check_if_the_avalibility_is_already_booked(user_id, availability):

    availability_start_time = datetime.fromtimestamp(
        int(availability.start_time) / 1000
    )

    availability_end_time = datetime.fromtimestamp(int(availability.end_time) / 1000)

    booked_sessions = get_booked_session_of_user_confirmed_avalibility(
        "coach", user_id, availability_start_time.strftime("%Y-%m-%d")
    )

    for session_request in booked_sessions:

        session_start_time = datetime.fromtimestamp(
            int(session_request.confirmed_availability.start_time) / 1000
        )
        session_end_time = datetime.fromtimestamp(
            int(session_request.confirmed_availability.end_time) / 1000
        )
        if (
            availability_start_time <= session_end_time
            and availability_end_time >= session_start_time
        ):
            return True
    return False


def create_task(task_details, number_of_days):
    triggered_date = datetime.now() + timedelta(days=number_of_days)
    triggered_date = triggered_date.replace(hour=8, minute=0, second=0, microsecond=0)
    task_details["trigger_date"] = triggered_date
    serializer = TaskSerializer(data=task_details)
    if serializer.is_valid():
        task_instance = serializer.save()
        return task_instance
    else:
        return None


def add_new_pmo(data):
    try:

        name = data.get("name")
        email = data.get("email", "").strip().lower()
        phone = data.get("phone")
        username = email  # username and email are the same
        password = data.get("password")
        sub_role = data.get("sub_role")
        room_id = generate_room_id(email)

        # Check if required data is provided
        if not all([name, email, phone, username, password,room_id]):
            return Response(
                {"error": "All required fields must be provided."}, status=400
            )

        with transaction.atomic():
            # Check if the user already exists
            user = User.objects.filter(email=email).first()

            if not user:
                # If the user does not exist, create a new user
                user = User.objects.create_user(
                    username=username, password=password, email=email
                )
                profile = Profile.objects.create(user=user)

            else:
                profile = Profile.objects.get(user=user)

            # Create or get the "pmo" role
            pmo_role, created = Role.objects.get_or_create(name="pmo")
            profile.roles.add(pmo_role)
            profile.save()

            # Create the PMO User using the Profile
            pmo_user = Pmo.objects.create(
                user=profile,
                name=name,
                email=email,
                phone=phone,
                sub_role=sub_role,
                room_id=room_id
            )

            name = pmo_user.name
            if pmo_user.phone:
                add_contact_in_wati("pmo", name, pmo_user.phone)

            # Return success response without room_id
            return True

    except Exception as e:
        print(str(e))
        return False


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("superadmin")])
def create_pmo(request):
    try:

        data = request.data
        added = add_new_pmo(data=data)

        if added:
            return Response({"message": "PMO added successfully."}, status=201)
        else:
            return Response({"error": "Failed to add pmo"}, status=500)
    except Exception as e:
        print(str(e))
        return Response({"error": "Failed to add pmo"}, status=500)


@api_view(["PUT"])
@permission_classes([AllowAny, IsInRoles("superadmin")])
def edit_pmo(request):
    # Get data from request
    name = request.data.get("name")
    email = request.data.get("email", "").strip().lower()
    phone = request.data.get("phone")
    sub_role = request.data.get("sub_role")
    pmo_email = request.data.get("pmo_email", "").strip().lower()
    try:
        with transaction.atomic():
            existing_user = (
                User.objects.filter(username=email).exclude(username=pmo_email).first()
            )
            if existing_user:
                return Response(
                    {"error": "User with this email already exists."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            pmo = Pmo.objects.get(email=pmo_email)
            pmo.user.user.username = email
            pmo.user.user.email = email
            pmo.user.user.save()
            pmo.email = email
            pmo.name = name
            pmo.phone = phone
            pmo.sub_role = sub_role
            pmo.save()
            if pmo.phone:
                add_contact_in_wati("pmo", pmo.name, pmo.phone)

            return Response({"message": "PMO updated successfully."}, status=201)
    except Exception as e:
        print(str(e))
        return Response({"error": "Failed to update pmo."}, status=500)


@api_view(["PUT"])
@permission_classes([AllowAny, IsInRoles("superadmin")])
def edit_ctt_pmo(request, ctt_pmo_id):
    name = request.data.get("name")
    email = request.data.get("email", "").strip().lower()
    phone = request.data.get("phone")
    ctt_pmo = CTTPmo.objects.get(id=ctt_pmo_id)

    try:
        with transaction.atomic():
            existing_user = (
                User.objects.filter(username=email)
                .exclude(username=ctt_pmo.user.user.username)
                .first()
            )
            if existing_user:
                return Response(
                    {"error": "User with this email already exists."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            ctt_pmo.user.user.username = email
            ctt_pmo.user.user.email = email
            ctt_pmo.user.user.save()
            ctt_pmo.email = email
            ctt_pmo.name = name
            ctt_pmo.phone = phone
            ctt_pmo.save()
            if ctt_pmo.phone:
                add_contact_in_wati("pmo", ctt_pmo.name, ctt_pmo.phone)

            return Response({"message": "CTT PMO updated successfully."}, status=201)
    except Exception as e:
        print(str(e))
        return Response({"error": "Failed to update CTT PMO."}, status=500)


@api_view(["PUT"])
@permission_classes([IsAuthenticated, IsInRoles("pmo")])
def approve_coach(request):
    try:

        unapproved_coach = request.data["coach"]
        room_id = request.data["room_id"]
        coach = Coach.objects.get(id=unapproved_coach["id"])

        coach.is_approved = True
        coach.is_rejected = False
        coach.active_inactive = True
        coach.room_id = room_id
        coach.save()

        path = f"/profile"

        message = f"Congratulations ! Your profile has been approved. You will be notified for projects that match your profile. Thank You !"

        create_notification(coach.user.user, path, message)

        send_mail_templates(
            "coach_templates/pmo_approves_profile.html",
            [coach.email],
            "Congratulations! Your Coach Registration is Approved",
            {
                "name": f"{coach.first_name.strip().title()} {coach.last_name.strip().title()}",
            },
            [],
        )
        return Response({"message": "Coach approved successfully."}, status=200)

    except Exception as e:
        print(str(e))

        return Response({"error": "Failed to approve coach."}, status=500)


@api_view(["PUT"])
@permission_classes([AllowAny])
def reject_coach(request, coach_id):
    try:
        with transaction.atomic():
            coach = Coach.objects.get(id=coach_id)

            update_data = {
                "pmo": request.data.get("pmo", ""),
                "coach": coach.id,
                "message": request.data.get("message", ""),
            }

            serializer = UpdateSerializer(data=update_data)
            if serializer.is_valid():
                serializer.save()
                coach.is_rejected = True
                coach.save()
                send_mail_templates(
                    "coach_templates/coach_is_rejected.html",
                    [coach.email],
                    "Meeraq | Profile Rejected",
                    {
                        "name": f"{coach.first_name.strip().title()}",
                    },
                    [],
                )
                return Response({"message": "Coach rejected successfully!"}, status=200)
            return Response(serializer.errors, status=500)
    except Exception as e:
        print(str(e))

        return Response({"error": "Failed to reject coach."}, status=500)


@api_view(["PUT"])
@permission_classes([AllowAny])
def reject_facilitator(request, facilitator_id):
    try:

        facilitator = Facilitator.objects.get(id=facilitator_id)

        facilitator.is_rejected = True
        facilitator.save()

        return Response({"message": "Facilitator rejected successfully!"}, status=200)

    except Exception as e:
        print(str(e))
        return Response({"error": "Failed to reject facilitator."}, status=500)


@api_view(["PUT"])
@permission_classes([IsAuthenticated])
def approve_facilitator(request):
    try:
        # Get the Coach object
        unapproved_coach = request.data["coach"]
        coach = Facilitator.objects.get(id=unapproved_coach["id"])

        # Change the is_approved field to True
        coach.is_approved = True
        coach.save()

        path = f"/profile"

        message = f"Congratulations ! Your profile has been approved. You will be notified for projects that match your profile. Thank You !"

        create_notification(coach.user.user, path, message)
        # Return success response
        # Send approval email to the coach
        send_mail_templates(
            "coach_templates/pmo_approves_profile.html",
            [coach.email],
            "Congratulations! Your Facilitator Registration is Approved",
            {
                "name": f"{coach.first_name.strip().title()} {coach.last_name.strip().title()}",
            },
            [],
        )
        return Response({"message": "Facilitator approved successfully."}, status=200)

    except Coach.DoesNotExist:
        # Return error response if Coach with the provided ID does not exist
        return Response({"error": "Facilitator does not exist."}, status=404)

    except Exception as e:
        # Return error response if any other exception occurs
        return Response({"error": str(e)}, status=500)


@api_view(["PUT"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "coach")])
def update_coach_profile(request, id):
    try:
        coach = Coach.objects.get(id=id)
        mutable_data = request.data.copy()

    except Coach.DoesNotExist:
        return Response(status=404)

    remove_education_upload_file = request.data.get(
        "remove_education_upload_file", False
    )
    internal_coach = json.loads(request.data["internal_coach"])
    organization_of_coach = request.data.get("organization_of_coach")
    user = coach.user.user
    new_email = request.data.get("email", "").strip().lower()
    #  other user exists with the new email
    if new_email and new_email != user.email:
        if (
            new_email
            and User.objects.filter(username=new_email).exclude(id=user.id).exists()
        ):
            existing_user_with_same_email = (
                User.objects.filter(username=new_email).exclude(id=user.id).first()
            )
            current_user_profile = user.profile
            existing_profile_with_same_email = existing_user_with_same_email.profile
            # coach exists with the new email
            if existing_profile_with_same_email.roles.filter(name="coach").exists():
                return Response(
                    {"error": "Coach with the same email already exists."},
                    status=400,
                )
            # another user exists with the new email
            else:
                coach_role, created = Role.objects.get_or_create(name="coach")
                existing_profile_with_same_email.roles.add(coach_role)
                coach.email = new_email
                current_user_profile.roles.remove(coach_role)
                coach.save()
                current_user_profile.save()
                existing_profile_with_same_email.save()

        # no other user exists with the new email
        else:
            user.email = new_email
            user.username = new_email
            user.save()
            # updating emails in all user's
            for role in user.profile.roles.all():
                if role.name == "pmo":
                    pmo = Pmo.objects.get(user=user.profile)
                    pmo.email = new_email
                    pmo.save()
                if role.name == "hr":
                    hr = HR.objects.get(user=user.profile)
                    hr.email = new_email
                    hr.save()
                if role.name == "learner":
                    learner = Learner.objects.get(user=user.profile)
                    learner.email = new_email
                    learner.save()
                if role.name == "vendor":
                    vendor = Vendor.objects.get(user=user.profile)
                    vendor.email = new_email
                    vendor.save()

    if internal_coach and not organization_of_coach:
        return Response(
            {
                "error": "Organization field must not be empty if internal coach is selected yes."
            },
            status=400,
        )
    profile_edit_start = ProfileEditActivity.objects.create(
        user=coach.user.user,
        timestamp=timezone.now(),
    )
    serializer = CoachSerializer(coach, data=request.data, partial=True)

    name = coach.first_name.strip().title() + " " + coach.last_name.strip().title()
    add_contact_in_wati("coach", name, coach.phone)

    if serializer.is_valid():
        coach_instance = serializer.save()
        if remove_education_upload_file:
            coach_instance.education_upload_file = None
            coach_instance.save()
        depth_serializer = CoachDepthOneSerializer(coach)
        is_caas_allowed = Project.objects.filter(
            coaches_status__coach=user.profile.coach
        ).exists()
        is_seeq_allowed = SchedularBatch.objects.filter(
            coaches=user.profile.coach
        ).exists()
        roles = []

        for role in user.profile.roles.all():
            roles.append(role.name)

        return Response(
            {
                **depth_serializer.data,
                "is_caas_allowed": is_caas_allowed,
                "is_seeq_allowed": is_seeq_allowed,
                "roles": roles,
                "last_login": coach.user.user.last_login,
                "user": {**depth_serializer.data["user"], "type": "coach"},
            }
        )
    return Response(serializer.errors, status=400)


def get_user_for_active_inactive(role, email):
    try:
        if role == "pmo":
            user = Pmo.objects.get(email=email)
        if role == "coach":
            user = Coach.objects.get(email=email)
        if role == "vendor":
            user = Vendor.objects.get(email=email)
        if role == "hr":
            user = HR.objects.get(email=email)
        if role == "learner":
            user = Learner.objects.get(email=email)
        if role == "superadmin":
            user = SuperAdmin.objects.get(email=email)
        if role == "facilitator":
            user = Facilitator.objects.get(email=email)
        if role == "finance":
            user = Finance.objects.get(email=email)
        if role == "sales":
            user = Sales.objects.get(email=email)
        if role == "ctt_pmo":
            user = CTTPmo.objects.get(email=email)
        if role == "leader":
            user = Leader.objects.get(email=email)
        return user
    except Exception as e:
        print(str(e))
        return None


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "coach")])
def get_coaches(request):
    try:
        # Get all the Coach objects
        coaches = Coach.objects.filter(is_approved=True)

        # Serialize the Coach objects
        serializer = CoachSerializer(coaches, many=True)

        # Return the serialized Coach objects as the response
        return Response(serializer.data, status=200)

    except Exception as e:
        # Return error response if any exception occurs
        return Response({"error": str(e)}, status=500)


def updateLastLogin(email):
    today = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    user = User.objects.get(username=email)
    user.last_login = today
    user.save()


@api_view(["POST"])
def pmo_login(request):
    email = request.data.get("email")
    password = request.data.get("password")

    if email is None or password is None:
        return Response({"error": "Please provide both email and password"}, status=400)

    user = authenticate(username=email, password=password, type="pmo")
    if not user:
        return Response({"error": "Invalid credentials"}, status=401)

    token, _ = Token.objects.get_or_create(user=user)
    pmo = Pmo.objects.get(user=user.profile)

    updateLastLogin(pmo.email)
    return Response(
        {
            "token": token.key,
            "pmo": {
                "name": pmo.name,
                "email": pmo.email,
                "phone": pmo.phone,
                "last_login": pmo.user.user.last_login,
            },
        }
    )


@api_view(["POST"])
def coach_login(request):
    email = request.data.get("email").strip().lower()
    password = request.data.get("password")

    if email is None or password is None:
        return Response({"error": "Please provide both email and password"}, status=400)

    user = authenticate(username=email, password=password)

    if not user:
        return Response({"error": "Invalid credentials"}, status=401)

    try:
        coach = Coach.objects.get(user=user.profile)
    except Coach.DoesNotExist:
        return Response({"error": "Coach not found"}, status=404)

    # Return the coach information in the response
    coach_serializer = CoachSerializer(coach)
    updateLastLogin(coach.email)
    return Response({"coach": coach_serializer.data}, status=200)


@api_view(["GET"])
@permission_classes([AllowAny])
def get_management_token(request):
    management_token = generateManagementToken()
    return Response(
        {"message": "Success", "management_token": management_token}, status=200
    )


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("pmo")])
@transaction.atomic
def create_project_cass(request):
    with transaction.atomic():
        organisation = Organisation.objects.filter(
            id=request.data["organisation_name"]
        ).first()
        junior_pmo = None
        if "junior_pmo" in request.data:
            junior_pmo = Pmo.objects.filter(id=request.data["junior_pmo"]).first()

        if not organisation:
            organisation = Organisation(
                name=request.data["organisation_name"],
                image_url=request.data["image_url"],
            )
        organisation.save()
        desc = request.data["project_description"]
        total_credits_in_minutes = 0
        request_expiry_time_in_minutes = 0
        duration_of_each_session = 0
        if "total_credits" in request.data and request.data["total_credits"]:
            total_credits_in_hours = float(request.data["total_credits"])
            total_credits_in_minutes = total_credits_in_hours * 60
        if not request.data["is_project_structure"]:
            duration_of_each_session = request.data["duration_of_each_session"]
            if request.data["is_session_expiry"]:
                if request.data["request_expiry_time"]:

                    request_expiry_time_in_hours = float(
                        request.data["request_expiry_time"]
                    )
                    request_expiry_time_in_minutes = request_expiry_time_in_hours * 60

        try:
            project = Project(
                # print(organisation.name, organisation.image_url, "details of org")
                name=request.data["project_name"],
                organisation=organisation,
                approx_coachee=request.data["approx_coachee"],
                frequency_of_session=request.data["frequency_of_session"],
                # currency=request.data["currency"],
                # price_per_hour=request.data["price_per_hour"],
                # coach_fees_per_hour=request.data["coach_fees_per_hour"],
                project_type=request.data["project_type"],
                interview_allowed=request.data["interview_allowed"],
                # chemistry_allowed= request.data['chemistry_allowed'],
                specific_coach=request.data["specific_coach"],
                empanelment=request.data["empanelment"],
                end_date=datetime.now() + timedelta(days=365),
                tentative_start_date=request.data["tentative_start_date"],
                mode=request.data["mode"],
                sold=request.data["sold"],
                project_description=desc,
                # updated_to_sold= request.data['updated_to_sold'],
                location=json.loads(request.data["location"]),
                enable_emails_to_hr_and_coachee=request.data.get(
                    "enable_emails_to_hr_and_coachee", True
                ),
                steps=dict(
                    project_structure={"status": "pending"},
                    coach_list={"status": "pending"},
                    coach_consent={"status": "pending"},
                    coach_list_to_hr={"status": "pending"},
                    interviews={"status": "pending"},
                    add_learners={"status": "pending"},
                    coach_approval={"status": "pending"},
                    chemistry_session={"status": "pending"},
                    coach_selected={"status": "pending"},
                    final_coaches={"status": "pending"},
                    project_live="pending",
                ),
                status="presales",
                masked_coach_profile=request.data["masked_coach_profile"],
                email_reminder=request.data["email_reminder"],
                whatsapp_reminder=request.data["whatsapp_reminder"],
                junior_pmo=junior_pmo,
                calendar_invites=request.data["calendar_invites"],
                finance=request.data["finance"],
                is_project_structure=request.data["is_project_structure"],
                total_credits=total_credits_in_minutes,
                duration_of_each_session=duration_of_each_session,
                request_expiry_time=request_expiry_time_in_minutes,
                is_session_expiry=(
                    False
                    if request.data["is_project_structure"]
                    else request.data["is_session_expiry"]
                ),
            )

            project.save()

            try:
                userId = request.data.get("user_id")
                user_who_created = User.objects.get(id=userId)
                timestamp = timezone.now()
                createProject = CreateProjectActivity.objects.create(
                    user_who_created=user_who_created,
                    project=project,
                    timestamp=timestamp,
                )
                createProject.save()
            except Exception as e:
                pass

            handover_id = request.data.get("handover")
            if handover_id:
                handover = HandoverDetails.objects.get(id=handover_id)
                handover.caas_project = project
                handover.save()
                project.project_structure = handover.project_structure
                project.save()
                add_so_to_project("CAAS", project.id, handover.sales_order_ids)
            else:
                sales_order_ids = request.data["sales_order_ids"]
                if sales_order_ids:
                    add_so_to_project("CAAS", project.id, sales_order_ids)

        except IntegrityError as e:
            print(str(e))
            return Response(
                {"error": "Project with this name already exists"}, status=400
            )
        except Exception as e:
            print(str(e))
            return Response({"error": "Failed to create project."}, status=400)

        if project.project_type == "COD" and not project.is_project_structure:
            if project.duration_of_each_session != 0:
                total_sessions = math.floor(
                    project.total_credits / project.duration_of_each_session
                )

            project.project_structure = [
                {
                    "f2f": False,
                    "index": 0,
                    "price": "100",
                    "billable": True,
                    "coach_price": "100",
                    "session_type": "coaching_session",
                    "no_of_sessions": total_sessions,
                    "session_duration": f"{project.duration_of_each_session}",
                }
            ]
            project.steps["project_structure"]["status"] = "complete"
            project.save()

        for hr in request.data["hr"]:
            single_hr = HR.objects.get(id=hr)
            project.hr.add(single_hr)

        # create tasks for the pmo
        try:
            print("creating tasks here")
            create_task(
                {
                    "task": "add_project_structure",
                    "caas_project": project.id,
                    "priority": "high",
                    "status": "pending",
                    "remarks": [],
                },
                30,
            )
            create_task(
                {
                    "task": "add_coach",
                    "caas_project": project.id,
                    "priority": "high",
                    "status": "pending",
                    "remarks": [],
                },
                7,
            )
            create_task(
                {
                    "task": "add_coach_contract",
                    "caas_project": project.id,
                    "priority": "high",
                    "status": "pending",
                    "remarks": [],
                },
                1,
            )
        except Exception as e:
            print("Error", str(e))

        try:
            path = f"/projects/caas/progress/{project.id}"
            message = f"A new project - {project.name} has been created for the organisation - {project.organisation.name}"
            for hr_member in project.hr.all():
                create_notification(hr_member.user.user, path, message)
        except Exception as e:
            print(f"Error occurred while creating notification: {str(e)}")
        return Response(
            {"message": "Project created successfully", "project_id": project.id},
            status=200,
        )


def create_learners(learners_data):
    try:
        with transaction.atomic():
            if not learners_data:
                raise ValueError("Learners data is required")
            learners = []
            for learner_data in learners_data:
                # Check if username field is provided
                email = learner_data.get("email", "").strip().lower()
                if "email" not in learner_data:
                    raise ValueError("Username field is required")
                # Check if user already exists
                user = User.objects.filter(username=email).first()
                if user:
                    # If user exists, check if learner already exists
                    learner = Learner.objects.filter(user__user=user).first()

                    if learner:
                        learner.name = learner_data.get("name").strip().title()
                        learner.phone = learner_data.get("phone")
                        try:
                            if learner_data.get("area_of_expertise", ""):
                                learner.area_of_expertise = learner_data.get(
                                    "area_of_expertise"
                                )

                            if learner_data.get("years_of_experience", ""):
                                learner.years_of_experience = learner_data.get(
                                    "years_of_experience"
                                )

                        except Exception as e:
                            print(str(e))
                        learner.save()
                        learners.append(learner)
                        continue
                    else:
                        profile = Profile.objects.get(user=user)

                else:
                    # If learner does not exist, create the user object with an unusable password
                    temp_password = "".join(
                        random.choices(
                            string.ascii_uppercase
                            + string.ascii_lowercase
                            + string.digits,
                            k=8,
                        )
                    )
                    user = User.objects.create_user(
                        username=email,
                        password=temp_password,
                        email=email,
                    )
                    # user.set_unusable_password()
                    user.save()
                    # Create the learner profile
                    profile = Profile.objects.create(user=user)
                learner_role, created = Role.objects.get_or_create(name="learner")
                profile.roles.add(learner_role)
                profile.save()

                learner = Learner.objects.create(
                    user=profile,
                    name=learner_data.get("name").strip().title(),
                    email=email,
                    phone=learner_data.get("phone"),
                )
                learners.append(learner)
                # Return response with learners created or already existing
                name = learner.name
                add_contact_in_wati("learner", name, learner.phone)
            serializer = LearnerSerializer(learners, many=True)
            return learners

    except ValueError as e:
        # Handle missing or invalid request data
        raise ValueError(str(e))

    except Exception as e:
        # Handle any other exceptions
        # transaction.set_rollback(True) # Rollback the transaction
        raise Exception(str(e))


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "sales")])
def get_ongoing_projects(request):
    try:
        pmo_id = request.query_params.get("pmo")
        projects = Project.objects.all()

        if pmo_id:
            pmo = Pmo.objects.get(id=int(pmo_id))
            if pmo.sub_role == "junior_pmo":
                projects = projects.filter(junior_pmo=int(pmo_id))

        projects = projects.filter(steps__project_live="pending").annotate(
            is_archive_enabled=Case(
                When(coaches_status__isnull=True, then=True),
                default=False,
                output_field=BooleanField(),
            )
        )
        projects = projects.distinct()

        serializer = ProjectDepthTwoSerializerArchiveCheck(projects, many=True)
        for project_data in serializer.data:
            latest_update = (
                Update.objects.filter(project__id=project_data["id"])
                .order_by("-created_at")
                .first()
            )
            project_data["latest_update"] = (
                latest_update.message if latest_update else None
            )
        return Response(serializer.data)
    except Exception as e:
        print(str(e))
        return Response({"error": "Failed to get data"}, status=500)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("pmo")])
def get_project_updates(request, project_id):
    updates = Update.objects.filter(project__id=project_id).order_by("-created_at")
    serializer = UpdateDepthOneSerializer(updates, many=True)
    return Response(serializer.data)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("pmo")])
def get_coach_updates(request, coach_id):
    updates = Update.objects.filter(coach__id=coach_id).order_by("-created_at")
    serializer = UpdateDepthOneSerializer(updates, many=True)
    return Response(serializer.data)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("pmo")])
def add_project_update(request, project_id):
    try:
        project = Project.objects.get(id=project_id)
    except Project.DoesNotExist:
        return Response({"error": "Project not found"}, status=404)
    # Assuming your request data has a "message" field for the update message
    update_data = {
        "pmo": request.data.get(
            "pmo", ""
        ),  # Assuming the PMO is associated with the user
        "project": project.id,
        "message": request.data.get("message", ""),
    }
    serializer = UpdateSerializer(data=update_data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Update added to project successfully!"}, status=201
        )
    return Response(serializer.errors, status=400)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("pmo")])
def add_coach_update(request, coach_id):
    try:
        coach = Coach.objects.get(id=coach_id)

        update_data = {
            "pmo": request.data.get("pmo", ""),
            "coach": coach.id,
            "message": request.data.get("message", ""),
        }
        serializer = UpdateSerializer(data=update_data)
        if serializer.is_valid():
            serializer.save()
            return Response(
                {"message": "Update added to coach successfully!"}, status=201
            )
        return Response(serializer.errors, status=500)
    except Exception as e:
        print(e)
        return Response({"error": "Failed to add update"}, status=500)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("learner", "pmo")])
def get_projects_of_learner(request, learner_id):
    projects = Project.objects.filter(engagement__learner__id=learner_id)
    serializer = ProjectDepthTwoSerializer(projects, many=True)
    return Response(serializer.data)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("hr")])
def get_ongoing_projects_of_hr(request, hr_id):
    projects = Project.objects.filter(hr__id=hr_id, steps__project_live="pending")
    schedular_projects = SchedularProject.objects.filter(hr__id=hr_id)
    schedular_project_serializer = SchedularProjectSerializer(
        schedular_projects, many=True
    )
    serializer = ProjectDepthTwoSerializer(projects, many=True)
    return Response(
        {
            "caas_projects": serializer.data,
            "schedular_projects": schedular_project_serializer.data,
        },
        status=200,
    )


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("hr")])
def get_schedular_projects_of_hr(request, hr_id):

    schedular_projects = SchedularProject.objects.filter(hr__id=hr_id)
    serializer = SchedularProjectSerializer(schedular_projects, many=True)
    return Response(serializer.data, status=200)


@api_view(["GET"])
@permission_classes(
    [IsAuthenticated, IsInRoles("coach", "pmo", "learner", "hr", "sales")]
)
def get_hr(request):
    try:
        # Get all the Coach objects
        hr = HR.objects.all()

        # Serialize the Coach objects
        serializer = HrSerializer(hr, many=True)

        # Return the serialized Coach objects as the response
        return Response(serializer.data, status=200)

    except Exception as e:
        # Return error response if any exception occurs
        return Response({"error": str(e)}, status=500)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("coach", "pmo", "learner", "hr")])
def get_projects_and_sessions_by_coach(request, coach_id):
    projects = Project.objects.filter(
        coaches_status__coach__id=coach_id, coaches_status__is_consent_asked=True
    )
    project_serializer = ProjectDepthTwoSerializer(projects, many=True)
    return Response({"projects": project_serializer.data})


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("coach", "pmo", "learner", "hr")])
def coach_session_list(request, coach_id):
    projects = Project.objects.filter(
        coaches_status__coach__id=coach_id, coaches_status__is_consent_asked=True
    )
    project_serializer = ProjectDepthTwoSerializer(projects, many=True)

    # Fetch sessions related to the coach
    sessions = SessionRequestCaas.objects.filter(coach__id=coach_id)
    session_serializer = SessionRequestCaasSerializer(sessions, many=True)

    # Group sessions by project ID
    sessions_dict = {}
    for session in session_serializer.data:
        project_id = session["project"]
        if project_id in sessions_dict:
            sessions_dict[project_id].append(session)
        else:
            sessions_dict[project_id] = [session]

    # Add the session data to the projects
    for project_data in project_serializer.data:
        project_id = project_data["id"]
        if project_id in sessions_dict:
            project_data["sessions"] = sessions_dict[project_id]
        else:
            project_data["sessions"] = []

    return Response({"projects": project_serializer.data})


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("pmo")])
def add_coach(request):
    # Get data from request
    first_name = request.data.get("first_name")
    last_name = request.data.get("last_name")
    email = request.data.get("email", "").strip().lower()
    age = request.data.get("age")
    gender = request.data.get("gender")
    domain = json.loads(request.data["domain"])
    room_id = request.data.get("room_id")
    phone = request.data.get("phone")
    phone_country_code = request.data.get("phone_country_code")
    level = request.data.get("level")
    currency = request.data.get("currency")
    education = json.loads(request.data["education"])
    rating = "5"
    min_fees = request.data["min_fees"]
    fee_remark = request.data.get("fee_remark", "")
    coaching_hours = request.data.get("coaching_hours")
    linkedin_profile_link = request.data["linkedin_profile_link"]
    companies_worked_in = json.loads(request.data["companies_worked_in"])
    other_certification = json.loads(request.data["other_certification"])
    active_inactive = json.loads(request.data["active_inactive"])
    area_of_expertise = json.loads(request.data["area_of_expertise"])
    location = json.loads(request.data["location"])
    language = json.loads(request.data["language"])
    job_roles = json.loads(request.data["job_roles"])
    ctt_nctt = json.loads(request.data["ctt_nctt"])
    years_of_coaching_experience = request.data.get("years_of_coaching_experience")
    years_of_corporate_experience = request.data.get("years_of_corporate_experience")
    username = (
        request.data.get("email", "").strip().lower()
    )  # keeping username and email same
    profile_pic = request.data.get("profile_pic", None)
    corporate_experience = request.data.get("corporate_experience", "")
    coaching_experience = request.data.get("coaching_experience", "")
    internal_coach = json.loads(request.data["internal_coach"])
    organization_of_coach = request.data.get("organization_of_coach")
    reason_for_inactive = json.loads(request.data["reason_for_inactive"])
    client_companies = json.loads(request.data["client_companies"])
    education_pic = request.data.get("education_pic", None)
    educational_qualification = json.loads(request.data["educational_qualification"])
    education_upload_file = request.data.get("education_upload_file", None)
    is_coach = json.loads(request.data.get("is_coach", False))
    is_mentor = json.loads(request.data.get("is_mentor", False))
    is_consultant = json.loads(request.data.get("is_consultant", False))

    # Check if required data is provided
    if not all(
        [
            first_name,
            last_name,
            email,
            gender,
            phone,
            phone_country_code,
            level,
            username,
            room_id,
            corporate_experience,
            coaching_experience,
        ]
    ):
        return Response({"error": "All required fields must be provided."}, status=400)
    if internal_coach and not organization_of_coach:
        return Response(
            {
                "error": "Organization field must not be empty if internal coach is selected yes."
            },
            status=400,
        )

    try:
        with transaction.atomic():
            # Check if the user already exists
            user = User.objects.filter(email=email).first()

            if not user:
                # If the user does not exist, create a new user
                temp_password = "".join(
                    random.choices(
                        string.ascii_uppercase + string.ascii_lowercase + string.digits,
                        k=8,
                    )
                )
                user = User.objects.create_user(
                    username=username, password=temp_password, email=email
                )
                profile = Profile.objects.create(user=user)
            else:
                profile = Profile.objects.get(user=user)

            coach_role, created = Role.objects.get_or_create(name="coach")
            # Create the Profile linked to the User
            profile.roles.add(coach_role)
            profile.save()

            # Create the Coach User using the Profile
            coach_user = Coach.objects.create(
                user=profile,
                room_id=room_id,
                first_name=first_name.strip().title(),
                last_name=last_name.strip().title(),
                email=email,
                phone=phone,
                phone_country_code=phone_country_code,
                level=level,
                currency=currency,
                education=education,
                rating=rating,
                area_of_expertise=area_of_expertise,
                age=age,
                gender=gender,
                domain=domain,
                years_of_corporate_experience=years_of_corporate_experience,
                ctt_nctt=ctt_nctt,
                years_of_coaching_experience=years_of_coaching_experience,
                profile_pic=profile_pic,
                language=language,
                min_fees=min_fees,
                fee_remark=fee_remark,
                job_roles=job_roles,
                location=location,
                coaching_hours=coaching_hours,
                linkedin_profile_link=linkedin_profile_link,
                companies_worked_in=companies_worked_in,
                other_certification=other_certification,
                active_inactive=active_inactive,
                internal_coach=internal_coach,
                corporate_experience=corporate_experience,
                coaching_experience=coaching_experience,
                organization_of_coach=organization_of_coach,
                reason_for_inactive=reason_for_inactive,
                client_companies=client_companies,
                education_pic=education_pic,
                educational_qualification=educational_qualification,
                education_upload_file=education_upload_file,
                is_coach=is_coach,
                is_mentor=is_mentor,
                is_consultant=is_consultant,
            )

            # Approve coach
            coach_user.is_approved = True
            coach_user.save()
            # Send email notification to the coach

            name = coach_user.first_name + " " + coach_user.last_name
            add_contact_in_wati("coach", name, coach_user.phone)

            full_name = coach_user.first_name + " " + coach_user.last_name
            microsoft_auth_url = (
                f'{env("BACKEND_URL")}/api/microsoft/oauth/{coach_user.email}/'
            )
            user_token_present = False
            try:
                user_token = UserToken.objects.get(
                    user_profile__user__username=coach_user.email
                )
                if user_token:
                    user_token_present = True
            except Exception as e:
                pass
            send_mail_templates(
                "coach_templates/pmo-adds-coach-as-user.html",
                [coach_user.email],
                "Meeraq Coaching | New Beginning !",
                {
                    "name": coach_user.first_name,
                    "email": coach_user.email,
                    "microsoft_auth_url": microsoft_auth_url,
                    "user_token_present": user_token_present,
                },
                [],  # no bcc emails
            )

        return Response({"message": "Coach added successfully."}, status=201)

    except IntegrityError as e:
        print(e)
        return Response(
            {"error": "A coach user with this email already exists."}, status=400
        )

    except Exception as e:
        print(e)
        return Response(
            {"error": "An error occurred while creating the coach user."}, status=500
        )


# @api_view(["POST"])
# @permission_classes([IsAuthenticated])
# def delete_coach(request):
#     coach_id = request.data.get("coach_id", None)
#     user_id = request.data.get("user_id")
#     if coach_id:
#         try:
#             coach = Coach.objects.get(id=coach_id)
#             coach_name = coach.first_name + " " + coach.last_name
#             coach_user_profile = coach.user
#             is_delete_user = True
#             for role in coach_user_profile.roles.all():
#                 if not role.name == "coach":
#                     # don't delete user if any other role exists
#                     is_delete_user = False
#                 else:
#                     coach_user_profile.roles.remove(role)
#                     coach_user_profile.save()
#             if is_delete_user:
#                 user = coach.user.user
#                 user.delete()
#             else:
#                 coach.delete()
#             timestamp = timezone.now()
#             current_user = User.objects.get(id=user_id)

#             deleteCoachProfile = DeleteCoachProfileActivity.objects.create(
#                 user_who_got_deleted=coach_name,
#                 user_who_deleted=current_user,
#                 timestamp=timestamp,
#             )
#             deleteCoachProfile.save()
#             return Response({"message": "Coach deleted."}, status=200)
#         except ObjectDoesNotExist:
#             return Response({"message": "Failed to delete coach profile"}, status=400)
#     else:
#         return Response({"message": "Failed to delete coach profile"}, status=400)


@api_view(["GET"])
@permission_classes([AllowAny])
@ensure_csrf_cookie
def get_csrf(request):
    response = Response({"detail": "CSRF cookie set"})
    response["X-CSRFToken"] = get_token(request)
    return response


@api_view(["POST"])
@permission_classes([AllowAny])
def login_view(request):
    data = request.data
    print(data)
    username = data.get("username")
    password = data.get("password")
    platform = data.get("platform", "unknown")
    if username is None or password is None:
        raise ValidationError({"detail": "Please provide username and password."})
    user = authenticate(request, username=username, password=password)

    if user is None:
        raise AuthenticationFailed({"detail": "Invalid credentials."})

    last_login = user.last_login
    login(request, user)
    user_data = get_user_data(user)
    if user_data:
        login_timestamp = timezone.now()
        UserLoginActivity.objects.create(
            user=user, timestamp=login_timestamp, platform=platform
        )
        user_token = None
        try:
            user_token = UserToken.objects.get(user_profile__user__username=username)
            if user_token.account_type == "google":
                refresh_google_access_token(user_token)
            else:
                refresh_microsoft_access_token(user_token)

        except ObjectDoesNotExist:
            print("Does not exist")

        response = Response(
            {
                "detail": "Successfully logged in.",
                "user": {**user_data, "last_login": last_login},
            }
        )
        response["X-CSRFToken"] = get_token(request)
        return response
    else:
        logout(request)
        return Response({"error": "Invalid user type"}, status=400)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def logout_view(request):
    if not request.user.is_authenticated:
        raise AuthenticationFailed({"detail": "You're not logged in."})

    logout(request)
    return Response({"detail": "Successfully logged out."})


@api_view(["GET"])
@permission_classes([AllowAny])
@ensure_csrf_cookie
def session_view(request):
    user = request.user
    last_login = user.last_login
    user_data = get_user_data(user)
    if user_data:
        response = Response(
            {
                "isAuthenticated": True,
                "user": {**user_data, "last_login": last_login},
            }
        )
        response["X-CSRFToken"] = get_token(request)
        return response
    else:
        return Response({"error": "Invalid user type"}, status=400)


def get_user_data(user):
    if not user.profile:
        return None
    elif user.profile.roles.exclude(name="vendor").count() == 0:
        return None
    user_profile_role = user.profile.roles.all().exclude(name="vendor").first().name
    roles = []
    for role in user.profile.roles.all():
        user_data = get_user_for_active_inactive(role.name, user.profile.user.username)
        if user_data and user_data.active_inactive:
            roles.append(role.name)

    if user_profile_role == "coach":
        if user.profile.coach.active_inactive or not user.profile.coach.is_approved:
            serializer = CoachDepthOneSerializer(user.profile.coach)
            is_caas_allowed = Project.objects.filter(
                coaches_status__coach=user.profile.coach
            ).exists()
            is_seeq_allowed = SchedularBatch.objects.filter(
                coaches=user.profile.coach
            ).exists()
            return {
                **serializer.data,
                "is_caas_allowed": is_caas_allowed,
                "is_seeq_allowed": is_seeq_allowed,
                "roles": roles,
                "user": {**serializer.data["user"], "type": user_profile_role},
            }
        else:

            return None
    elif user_profile_role == "facilitator":
        if not user.profile.facilitator.active_inactive:
            return None
        serializer = FacilitatorDepthOneSerializer(user.profile.facilitator)
        return {
            **serializer.data,
            "roles": roles,
            "user": {**serializer.data["user"], "type": user_profile_role},
        }

    elif user_profile_role == "pmo":
        if not user.profile.pmo.active_inactive:
            return None
        serializer = PmoDepthOneSerializer(user.profile.pmo)
    elif user_profile_role == "superadmin":
        if not user.profile.superadmin.active_inactive:
            return None

        serializer = SuperAdminDepthOneSerializer(user.profile.superadmin)
    elif user_profile_role == "finance":
        if not user.profile.finance.active_inactive:
            return None
        serializer = FinanceDepthOneSerializer(user.profile.finance)
    elif user_profile_role == "learner":
        if not user.profile.learner.active_inactive:
            return None
        serializer = LearnerDepthOneSerializer(user.profile.learner)
        is_caas_allowed = Engagement.objects.filter(
            learner=user.profile.learner
        ).exists()
        is_seeq_allowed = SchedularBatch.objects.filter(
            learners=user.profile.learner
        ).exists()
        return {
            **serializer.data,
            "is_caas_allowed": is_caas_allowed,
            "is_seeq_allowed": is_seeq_allowed,
            "roles": roles,
            "user": {**serializer.data["user"], "type": user_profile_role},
        }
    elif user_profile_role == "hr":
        if not user.profile.hr.active_inactive:
            return None
        serializer = HrDepthOneSerializer(user.profile.hr)
        is_caas_allowed = Project.objects.filter(hr=user.profile.hr).exists()
        is_seeq_allowed = SchedularProject.objects.filter(hr=user.profile.hr).exists()
        return {
            **serializer.data,
            "roles": roles,
            "is_caas_allowed": is_caas_allowed,
            "is_seeq_allowed": is_seeq_allowed,
            "user": {**serializer.data["user"], "type": user_profile_role},
        }
    elif user_profile_role == "sales":
        if not user.profile.sales.active_inactive:
            return None
        serializer = SalesDepthOneSerializer(user.profile.sales)
    elif user_profile_role == "ctt_pmo":
        if not user.profile.cttpmo.active_inactive:
            return None
        serializer = CTTPmoDepthOneSerializer(user.profile.cttpmo)
    elif user_profile_role == "leader":
        if not user.profile.leader.active_inactive:
            return None
        serializer = LeaderDepthOneSerializer(user.profile.leader)
    else:
        return None
    return {
        **serializer.data,
        "roles": roles,
        "user": {**serializer.data["user"], "type": user_profile_role},
    }


@api_view(["POST"])
@permission_classes([AllowAny])
def generate_otp(request):
    try:
        user = User.objects.get(username=request.data["email"])
        learner_roles = user.profile.roles.all().filter(name="learner")
        hr_roles = user.profile.roles.all().filter(name="hr")
        # for hr and coachee not allowing login when they are added in caas project where hr and coachee's platform is not provided/needed
        if learner_roles.exists():
            engagements = Engagement.objects.filter(
                learner=user.profile.learner,
                project__enable_emails_to_hr_and_coachee=False,
            )
            if engagements.exists():
                return Response(
                    {"error": "User with the given email does not exist."}, status=400
                )
        if hr_roles.exists():
            projects = Project.objects.filter(
                hr=user.profile.hr, enable_emails_to_hr_and_coachee=False
            )
            if projects.exists():
                return Response(
                    {"error": "User with the given email does not exist."}, status=400
                )
        print(user)
        try:
            # Check if OTP already exists for the user
            otp_obj = OTP.objects.get(user=user)
            otp_obj.delete()
        except OTP.DoesNotExist:
            pass
        # Generate OTP and save it to the database
        otp = get_random_string(length=6, allowed_chars="0123456789")
        created_otp = OTP.objects.create(user=user, otp=otp)
        user_data = get_user_data(user)
        name = user_data.get("name") or user_data.get("first_name") or "User"
        # Send OTP on email to learner
        subject = f"Meeraq Login OTP"
        message = (
            f"Dear {name} \n\n Your OTP for login on meeraq portal is {created_otp.otp}"
        )
        # send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, [user.username])
        microsoft_auth_url = (
            f'{env("BACKEND_URL")}/api/microsoft/oauth/{request.data["email"]}/'
        )
        user_token_present = False
        try:
            user_token = UserToken.objects.get(
                user_profile__user__username=request.data["email"]
            )
            if user_token:
                user_token_present = True
        except Exception as e:
            pass
        send_mail_templates(
            "hr_emails/login_with_otp.html",
            [user],
            subject,
            {
                "name": name,
                "otp": created_otp.otp,
                "email": request.data["email"],
                "microsoft_auth_url": microsoft_auth_url,
                "user_token_present": user_token_present,
            },
            [],  # no bcc
        )
        return Response({"message": f"OTP has been sent to {user.username}!"})

    except User.DoesNotExist:
        # Handle the case where the user with the given email does not exist
        print("hello 2 ", str(e))

        return Response(
            {"error": "User with the given email does not exist."}, status=400
        )

    except Exception as e:
        print("hello", str(e))
        # Handle any other exceptions
        return Response({"error": str(e)}, status=500)


@api_view(["POST"])
@permission_classes([AllowAny])
def validate_otp(request):
    otp_obj = (
        OTP.objects.filter(
            user__username=request.data["email"], otp=request.data["otp"]
        )
        .order_by("-created_at")
        .first()
    )
    data = request.data
    platform = data.get("platform", "unknown")

    if otp_obj is None:
        raise AuthenticationFailed("Invalid OTP")

    user = otp_obj.user
    # token, created = Token.objects.get_or_create(user=learner.user.user)
    # Delete the OTP object after it has been validated
    user_email = request.data["email"]
    otp_obj.delete()
    last_login = user.last_login
    login(request, user)
    user_data = get_user_data(user)
    if user_data:
        login_timestamp = timezone.now()
        UserLoginActivity.objects.create(
            user=user, timestamp=login_timestamp, platform=platform
        )
        user_token = None
        try:
            user_token = UserToken.objects.get(user_profile__user__username=user_email)
            if user_token.account_type == "google":
                refresh_google_access_token(user_token)
            else:
                refresh_microsoft_access_token(user_token)
        except ObjectDoesNotExist:
            print("Does not exist")

        response = Response(
            {
                "detail": "Successfully logged in.",
                "user": {**user_data, "last_login": last_login},
            }
        )
        response["X-CSRFToken"] = get_token(request)
        return response
    else:
        logout(request)
        return Response({"error": "Invalid user type"}, status=400)

    # learner_data = {'id':learner.id,'name':learner.name,'email': learner.email,'phone': learner.email,'last_login': learner.user.user.last_login ,'token': token.key}
    # updateLastLogin(learner.email)
    # return Response({ 'learner': learner_data},status=200)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_organisation(request):
    orgs = Organisation.objects.all()
    serializer = OrganisationSerializer(orgs, many=True)
    return Response(serializer.data, status=200)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("pmo")])
def add_organisation(request):
    print(request.data.get("image_url", ""))
    org = Organisation.objects.create(
        name=request.data.get("name", ""), image_url=request.data.get("image_url", "")
    )
    orgs = Organisation.objects.all()
    serializer = OrganisationSerializer(orgs, many=True)
    return Response(
        {"message": "Organisation added successfully.", "details": serializer.data},
        status=200,
    )


@api_view(["PUT"])
@permission_classes([IsAuthenticated, IsInRoles("pmo")])
def update_organisation(request, org_id):
    try:
        org = Organisation.objects.get(id=org_id)
    except Organisation.DoesNotExist:
        return Response({"error": "Organization not found"}, status=404)

    serializer = OrganisationSerializer(org, data=request.data)

    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Organization updated successfully", "data": serializer.data}
        )

    return Response(serializer.errors, status=400)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("pmo")])
def add_hr(request):
    try:
        email = request.data.get("email", "").strip().lower()
        with transaction.atomic():
            user = User.objects.filter(email=email).first()
            if not user:
                temp_password = "".join(
                    random.choices(
                        string.ascii_uppercase + string.ascii_lowercase + string.digits,
                        k=8,
                    )
                )
                user = User.objects.create_user(
                    username=email,
                    password=temp_password,
                    email=email,
                )
                profile = Profile.objects.create(user=user)
            else:
                profile = Profile.objects.get(user=user)

            hr_role, created = Role.objects.get_or_create(name="hr")
            profile.roles.add(hr_role)

            organisation = Organisation.objects.filter(
                id=request.data.get("organisation")
            ).first()

            hr = HR.objects.create(
                user=profile,
                first_name=request.data.get("first_name"),
                last_name=request.data.get("last_name"),
                email=email,
                phone=request.data.get("phone"),
                organisation=organisation,
            )
            name = hr.first_name + " " + hr.last_name
            add_contact_in_wati("hr", name, hr.phone)

            hrs = HR.objects.all()
            serializer = HrSerializer(hrs, many=True)
            return Response(
                {"message": "HR added successfully", "details": serializer.data},
                status=200,
            )

    except Exception as e:
        print(str(e))
        return Response({"error": "Failed to add HR"}, status=400)


@api_view(["PUT"])
@permission_classes([IsAuthenticated, IsInRoles("pmo")])
def update_hr(request, hr_id):
    try:
        hr = HR.objects.get(id=hr_id)
    except HR.DoesNotExist:
        return Response({"error": "HR not found"}, status=status.HTTP_404_NOT_FOUND)

    # Get the associated user profile
    with transaction.atomic():
        # Update HR instance
        serializer = HrSerializer(hr, data=request.data, partial=True)
        if serializer.is_valid():
            new_email = (
                request.data.get("email", "").strip().lower()
            )  # Get the new email from the request
            existing_user = (
                User.objects.filter(email=new_email).exclude(username=hr.email).first()
            )
            if existing_user:
                return Response(
                    {"error": "User with this email already exists."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # saving hr
            updated_hr = serializer.save()
            user = updated_hr.user.user

            name = hr.first_name + " " + hr.last_name
            add_contact_in_wati("hr", name, hr.phone)

            # if email if getting updated -> updating email in all other user present
            if not updated_hr.email.strip().lower() == user.email.strip().lower():
                user.email = new_email
                user.username = new_email
                user.save()
                for role in user.profile.roles.all():
                    if role.name == "pmo":
                        pmo = Pmo.objects.get(user=user.profile)
                        pmo.email = new_email
                        pmo.save()
                    if role.name == "coach":
                        coach = Coach.objects.get(user=user.profile)
                        coach.email = new_email
                        coach.save()
                    if role.name == "learner":
                        learner = Learner.objects.get(user=user.profile)
                        learner.email = new_email
                        learner.save()
                    if role.name == "vendor":
                        vendor = Vendor.objects.get(user=user.profile)
                        vendor.email = new_email
                        vendor.save()
            return Response(
                {"message": "HR updated successfully", "data": serializer.data}
            )

        # Handle serializer errors
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# @api_view(["DELETE"])
# @permission_classes([IsAuthenticated])
# def delete_hr(request, hr_id):
#     try:
#         hr = HR.objects.get(id=hr_id)
#         user_profile = hr.user
#         is_delete_user = True
#         for role in user_profile.roles.all():
#             if not role.name == "hr":
#                 is_delete_user = False
#             else:
#                 user_profile.roles.remove(role)
#                 user_profile.save()
#         if is_delete_user:
#             user = user_profile.user
#             user.delete()
#         else:
#             hr.delete()
#         return Response(
#             {"message": "HR deleted successfully"}, status=status.HTTP_200_OK
#         )
#     except ObjectDoesNotExist:
#         return Response(
#             {"message": "Failed to delete HR profile"},
#             status=status.HTTP_400_BAD_REQUEST,
#         )


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("pmo")])
def add_project_struture(request):
    try:
        project = Project.objects.get(id=request.data.get("project_id", ""))
    except Project.DoesNotExist:
        return Response({"message": "Project does not exist"}, status=400)
    project.project_structure = request.data.get("project_structure", [])
    project.currency = request.data.get("currency", "")
    project.save()
    tasks = Task.objects.filter(
        task="add_project_structure", status="pending", caas_project=project
    )
    tasks.update(status="completed")

    if project.steps["project_structure"]["status"] == "complete":
        #  update project structure for  all coaches
        for coach_status in project.coaches_status.all():
            coach_status.project_structure = project.project_structure
            coach_status.save()
        if not project.coach_consent_mandatory:
            for coach_status in project.coaches_status.filter(is_consent_asked=True):
                if not coach_status.status["consent"]["status"] == "reject":
                    coach_status.status["consent"]["status"] = "select"
                    coach_status.status["project_structure"]["status"] = "select"
                coach_status.save()
    return Response({"message": "Structure added", "details": ""}, status=200)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("pmo")])
def send_consent(request):
    # Get the project
    try:
        project = Project.objects.get(id=request.data.get("project_id", ""))
    except Project.DoesNotExist:
        return Response({"message": "Project does not exist"}, status=400)
    # Get the list of coach IDs
    coach_list = request.data.get("coach_list", [])
    for coach_status in project.coaches_status.filter(
        coach__id__in=coach_list, is_consent_asked=False
    ):
        coach_status.consent_expiry_date = request.data["consent_expiry_date"]
        coach_status.is_consent_asked = True
        coach_status.save()
        if project.coach_consent_mandatory:
            create_task(
                {
                    "task": "give_consent",
                    "caas_project": project.id,
                    "coach": coach_status.coach.id,
                    "priority": "medium",
                    "status": "pending",
                    "remarks": [],
                },
                7,
            )
            if project.steps["project_structure"]["status"] == "complete":
                create_task(
                    {
                        "task": "agree_project_structure",
                        "caas_project": project.id,
                        "coach": coach_status.coach.id,
                        "priority": "medium",
                        "status": "pending",
                        "remarks": [],
                    },
                    7,
                )
    # Update project's coach_status and steps
    project.steps["coach_list"]["status"] = "complete"
    project.save()

    # Send notifications and emails
    try:
        path = f"/projects"
        message = (
            f"Admin has requested your consent to share profile for a new project."
        )
        for status in coach_status:
            if project.coach_consent_mandatory:
                create_notification(status.coach.user.user, path, message)
            send_mail_templates(
                "coach_templates/pmo_ask_for_consent.html",
                [status.coach.email],
                "Meeraq Coaching | New Project!",
                {
                    "name": status.coach.first_name,
                    "email": status.coach.email,
                },
                [],  # no bcc
            )
    except Exception as e:
        print(f"Error occurred while creating notification: {str(e)}")

    if not project.coach_consent_mandatory:
        for coach_status in project.coaches_status.filter(
            is_consent_asked=True, coach__id__in=coach_list
        ):
            coach_status.status["consent"]["status"] = "select"
            if project.steps["project_structure"]["status"] == "complete":
                coach_status.status["project_structure"]["status"] = "select"
            coach_status.save()
    return Response({"message": "Consent sent successfully", "details": ""}, status=200)


@api_view(["GET"])
@permission_classes(
    [IsAuthenticated, IsInRoles("learner", "pmo", "hr", "coach", "sales", "finance")]
)
def get_project_details(request, project_type, project_id):
    try:
        if project_type == "caas":
            project = Project.objects.get(id=project_id)
            serializer = ProjectDepthTwoSerializer(project)
        else:
            project = SchedularProject.objects.get(id=project_id)
            serializer = SchedularProjectSerializer(project)
        return Response(serializer.data)
    except Exception as e:
        print(str(e))
        return Response({"error": "Failed to get data"}, status=400)


# Filter API for Coaches
# Expected input
# "project_id": 1
# "coach_id": 1
# "status": Consent Approved/Consent Rejected
@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("coach")])
def receive_coach_consent(request):
    try:
        project = Project.objects.get(id=request.data.get("project_id", ""))
    except Project.DoesNotExist:
        return Response({"message": "Project does not exist"}, status=400)
    for coach_status in project.coaches_status.all():
        try:
            if coach_status.coach.id == request.data.get("coach_id", ""):
                coach_status.status["consent"]["status"] = request.data["status"]
                tasks = Task.objects.filter(
                    task="give_consent",
                    status="pending",
                    coach=coach_status.coach,
                    caas_project=project,
                )
                tasks.update(status="completed")
                if project.steps["project_structure"]["status"] == "complete":
                    tasks = Task.objects.filter(
                        task="agree_project_structure",
                        status="pending",
                        coach=coach_status.coach,
                        caas_project=project,
                    )
                    tasks.update(status="completed")
                    coach_status.status["project_structure"]["status"] = request.data[
                        "status"
                    ]
                coach_status.save()

                try:
                    pmo_user = User.objects.filter(profile__roles__name="pmo").first()
                    pmo = Pmo.objects.get(email=pmo_user.email)
                    if request.data["status"] == "select":
                        if pmo_user:
                            path = f"/projects/caas/progress/{project.id}"
                            coach_name = (
                                coach_status.coach.first_name
                                + " "
                                + coach_status.coach.last_name
                            )
                            message = f"{coach_name.title() } has accepted your consent request for Project - {project.name}"
                            create_notification(pmo_user, path, message)

                            send_mail_templates(
                                "pmo_emails/coach_agrees-rejects_consent.html",
                                [pmo_user.email],
                                f"Meeraq Coaching | Coach {coach_status.coach.first_name} agreed to consent",
                                {
                                    "projectname": project.name,
                                    "name": pmo.name,
                                    "coachname": coach_status.coach.first_name,
                                    "agreeddisagreed": request.data["status"],
                                },
                                [],  # no bcc
                            )
                    if request.data["status"] == "reject":
                        send_mail_templates(
                            "pmo_emails/coach_agrees-rejects_consent.html",
                            [pmo_user.email],
                            f"Meeraq Coaching | Coach {coach_status.coach.first_name} reject to consent",
                            {
                                "projectname": project.name,
                                "name": pmo.name,
                                "coachname": coach_status.coach.first_name,
                                "agreeddisagreed": request.data["status"],
                            },
                            [],  # no bcc
                        )

                except Exception as e:
                    print(f"Error occurred while creating notification: {str(e)}")
                    continue

        except Exception as e:
            print(e)
            return Response({"message": "Coach not Found"}, status=400)
    return Response({"message": request.data.get("status", "")}, status=200)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("coach")])
def update_project_structure_consent_by_coach(request):
    try:
        project = Project.objects.get(id=request.data.get("project_id", ""))
    except Project.DoesNotExist:
        return Response({"message": "Project does not exist"}, status=400)
    for coach_status in project.coaches_status.all():
        try:
            if coach_status.coach.id == request.data.get("coach_id", ""):
                # coach_status.status[request.data.get('status','').split(" ")[0].lower()]=request.data.get('status','').split(" ")[1].lower()
                # if request.data.get('status','').split(" ")[0].lower()=='contract':
                #     coach_status.status['consent'] = "approved"
                # coach_status.save()
                coach_status.status["project_structure"]["status"] = request.data[
                    "status"
                ]
                coach_status.save()
                try:
                    if request.data["status"] == "select":
                        pmo_user = User.objects.filter(
                            profile__roles__name="pmo"
                        ).first()
                        if pmo_user:
                            path = f"/projects/caas/progress/{project.id}"
                            coach_name = (
                                coach_status.coach.first_name
                                + " "
                                + coach_status.coach.last_name
                            )
                            message = f"{coach_name.title() } has agreed to the project structure for Project - {project.name}"
                            create_notification(pmo_user, path, message)
                except Exception as e:
                    print(f"Error occurred while creating notification: {str(e)}")
                    continue

        except Exception as e:
            print(e)
            return Response({"message": "Coach not Found"}, status=400)
    return Response({"message": "Status updated."}, status=200)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("coach", "pmo", "learner", "hr")])
def complete_coach_consent(request):
    try:
        project = Project.objects.get(id=request.data.get("project_id", ""))
    except Project.DoesNotExist:
        return Response({"message": "Project does not exist"}, status=400)
    project.status["coach_consent"] = "complete"
    project.save()
    return Response({"message": "Coach list sent to HR", "details": ""}, status=200)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("coach", "pmo", "learner", "hr")])
def complete_coach_list_to_hr(request):
    try:
        project = Project.objects.get(id=request.data.get("project_id", ""))
    except Project.DoesNotExist:
        return Response({"message": "Project does not exist"}, status=400)
    project.steps["coach_list_to_hr"]["status"] = "complete"
    project.steps["coach_consent"]["status"] = "complete"
    if not project.empanelment:
        for coach_status in project.coaches_status.all():
            if coach_status.status["hr"]["status"] == "select":
                coach_status.status["learner"]["status"] = "sent"
                coach_status.save()
    project.save()
    return Response({"message": "Step marked as complete.", "details": {}}, status=200)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("coach", "pmo", "learner", "hr")])
def complete_interviews_step(request):
    try:
        project = Project.objects.get(id=request.data.get("project_id", ""))
    except Project.DoesNotExist:
        return Response({"message": "Project does not exist"}, status=400)
    project.steps["interviews"]["status"] = "complete"
    project.steps["coach_consent"]["status"] = "complete"
    if not project.empanelment:
        for coach_status in project.coaches_status.all():
            if coach_status.status["hr"]["status"] == "select":
                coach_status.status["learner"]["status"] = "sent"
                coach_status.save()
    project.save()
    return Response({"message": "Step marked as complete.", "details": {}}, status=200)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("coach", "pmo", "learner", "hr")])
def complete_empanelment(request):
    try:
        project = Project.objects.get(id=request.data.get("project_id", ""))
    except Project.DoesNotExist:
        return Response({"message": "Project does not exist"}, status=400)
    project.steps["add_learners"]["status"] = "complete"
    project.save()
    return Response({"message": "Empanelement completed.", "details": ""}, status=200)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("pmo")])
def complete_project_structure(request):
    try:
        project = Project.objects.get(id=request.data.get("project_id", ""))
    except Project.DoesNotExist:
        return Response({"message": "Project does not exist"}, status=400)
    project.status["project_structure"] = "complete"
    project.save()
    return Response({"message": "Project structure approved."}, status=200)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("coach", "pmo", "learner", "hr")])
def get_interview_data(request, project_id):
    sessions = SessionRequestCaas.objects.filter(
        project__id=project_id, session_type="interview"
    ).all()
    serializer = SessionRequestCaasDepthOneSerializer(sessions, many=True)
    return Response(serializer.data, status=200)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("coach", "pmo", "learner", "hr")])
def get_chemistry_session_data(request, project_id):
    sessions = SessionRequestCaas.objects.filter(
        project__id=project_id, session_type="chemistry"
    ).exclude(status="pending")
    serializer = SessionRequestCaasDepthTwoSerializer(sessions, many=True)
    return Response(serializer.data, status=200)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("coach", "pmo", "learner", "hr")])
def get_chemistry_session_data_for_engagement(request, engagement_id):
    print(engagement_id)
    sessions = SessionRequestCaas.objects.filter(
        engagement__id=engagement_id, session_type="chemistry"
    ).exclude(status="pending")
    print(sessions)
    serializer = SessionRequestCaasDepthTwoSerializer(sessions, many=True)
    return Response(serializer.data, status=200)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("coach", "pmo", "learner", "hr")])
def get_session_requests_of_hr(request, hr_id):
    sessions = SessionRequestCaas.objects.filter(hr__id=hr_id).all()
    serializer = SessionRequestCaasDepthOneSerializer(sessions, many=True)
    return Response(serializer.data, status=200)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("coach", "pmo", "learner", "hr")])
def get_session_requests_of_learner(request, learner_id):
    sessions = SessionRequestCaas.objects.filter(learner__id=learner_id).all()
    print(sessions, "session")
    serializer = SessionRequestCaasDepthOneSerializer(sessions, many=True)
    return Response(serializer.data, status=200)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("coach", "pmo", "learner", "hr")])
def get_upcoming_booked_session_of_coach(request, coach_id):
    current_time = int(timezone.now().timestamp() * 1000)
    # convert current time to milliseconds
    sessions = SessionRequestCaas.objects.filter(
        coach__id=coach_id,
        is_booked=True,
        confirmed_availability__start_time__gt=current_time,
    ).all()
    serializer = SessionRequestCaasDepthOneSerializer(sessions, many=True)
    return Response(serializer.data, status=200)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "coach", "learner", "hr")])
@transaction.atomic
def book_session_caas(request):
    session_request = SessionRequestCaas.objects.get(
        id=request.data.get("session_request")
    )
    coach = Coach.objects.get(id=request.data["coach"])
    session_request.coach = coach
    session_request.save()
    existing_calendar_invite = CalendarInvites.objects.filter(
        caas_session=session_request
    ).first()

    confirmed_availability = Availibility.objects.get(
        id=request.data.get("confirmed_availability")
    )
    check_booking = check_if_the_avalibility_is_already_booked(
        request.data["coach"], confirmed_availability
    )

    if check_booking:
        return Response(
            {
                "error": "Sorry, this time is already booked. Please choose a different time.."
            },
            status=500,
        )

    session_request.confirmed_availability = Availibility.objects.get(
        id=request.data.get("confirmed_availability")
    )
    if session_request.learner:
        coachee = session_request.learner
    #     sessionName = str(session_request.session_type).replace('_', ' ')
    #     if( sessionName == "stakeholder without coach"):
    #         sessionName = "tripartite without coach"
    session_request.is_booked = True
    session_request.status = "booked"
    session_request.invitees = get_trimmed_emails(request.data.get("invitees", []))
    session_request.save()

    # create tasks
    tasks = Task.objects.filter(
        task="update_session_status", session_caas=session_request
    )
    if not tasks.exists():
        create_task(
            {
                "task": "update_session_status",
                "caas_project": session_request.project.id,
                "coach": session_request.coach.id if session_request.coach else None,
                "session_caas": session_request.id,
                "priority": "high",
                "status": "pending",
                "remarks": [],
            },
            1,
        )
    tasks = Task.objects.filter(task="add_invitees", session_caas=session_request)
    if tasks.exists():
        if len(session_request.invitees) > 0:
            tasks.update(status="completed")
    else:
        if (
            session_request.session_type in SESSIONS_WITH_STAKEHOLDERS
            and len(session_request.invitees) == 0
        ):
            create_task(
                {
                    "task": "add_invitees",
                    "caas_project": session_request.project.id,
                    "coach": (
                        session_request.coach.id if session_request.coach else None
                    ),
                    "session_caas": session_request.id,
                    "priority": "high",
                    "status": "pending",
                    "remarks": [],
                },
                1,
            )

    # # Send email notification to the learner
    if session_request.session_type == "interview":
        subject = "Hello hr your session is booked."
        message = f"Dear {session_request.hr.first_name},\n\nThank you booking slots of hr.Please be ready on date and time to complete session. Best of luck!"
        # send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, [session_request.hr.email])
    if session_request.session_type == "chemistry":
        subject = "Hello learner your session is booked."
        message = f"Dear {session_request.learner.name},\n\nThank you booking slots of hr.Please be ready on date and time to complete session. Best of luck!"
        # send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, [session_request.learner.email])
    try:
        pmo_user = User.objects.filter(profile__roles__name="pmo").first()
        project = session_request.project
        coach = session_request.coach
        pmo = Pmo.objects.get(email=pmo_user)

        if pmo_user:
            path = f"/projects/caas/progress/{project.id}"
            coach_name = coach.first_name + " " + coach.last_name
            start_time = format_timestamp(
                int(session_request.confirmed_availability.start_time)
            )
            end_time = format_timestamp(
                int(session_request.confirmed_availability.end_time)
            )
            slot_message = f"{start_time} - {end_time}"

            session_date = get_date(
                int(session_request.confirmed_availability.start_time)
            )
            start_time = get_time(
                int(session_request.confirmed_availability.start_time)
            )
            end_time = get_time(int(session_request.confirmed_availability.end_time))

            session_time = f"{start_time} - {end_time} IST"

            if session_request.session_type == "interview":
                hr_user = session_request.hr.user.user
                # message = f"{coach_name.title()} has booked the interview session for Project - {project.name}.The booked slot is "
                message_for_hr = f"{coach_name.title()} has booked the slot {slot_message} from your interview request. You can join the meeting on scheduled time."
                create_notification(hr_user, path, message_for_hr)
            if session_request.session_type == "chemistry":
                learner_user = session_request.learner.user.user
                message_for_learner = f"{coach_name.title()} has booked the slot {slot_message} from your chemistry session request. You can join the meeting on scheduled time."
                message_for_hr = f"{coach_name.title()} has booked the slot {slot_message} as per the request from {session_request.learner.name.title()} for the project - {session_request.project.name}"
                create_notification(learner_user, path, message_for_learner)
            message = f"{coach_name.title()} has booked the {SESSION_TYPE_VALUE[session_request.session_type]} for Project - {project.name}.The booked slot is "
            message += slot_message
            create_notification(pmo_user, path, message)

            # WHATSAPP MESSAGE CHECK
            # before 5 mins whatsapp msg
            confirmed_availability_start_time = (
                session_request.confirmed_availability.start_time
            )
            start_datetime_obj = datetime.fromtimestamp(
                int(confirmed_availability_start_time) / 1000
            )
            # Decrease 5 minutes
            five_minutes_prior_start_datetime = start_datetime_obj - timedelta(
                minutes=5
            )
            clocked = ClockedSchedule.objects.create(
                clocked_time=five_minutes_prior_start_datetime
            )
            periodic_task = PeriodicTask.objects.create(
                name=uuid.uuid1(),
                task="schedularApi.tasks.send_whatsapp_reminder_to_users_before_5mins_in_caas",
                args=[session_request.id],
                clocked=clocked,
                one_off=True,
            )
            periodic_task.save()
            # after 3 minutes whatsapp message
            # increase 5 minutes
            three_minutes_ahead_start_datetime = start_datetime_obj + timedelta(
                minutes=3
            )
            clocked = ClockedSchedule.objects.create(
                clocked_time=three_minutes_ahead_start_datetime
            )
            periodic_task = PeriodicTask.objects.create(
                name=uuid.uuid1(),
                task="schedularApi.tasks.send_whatsapp_reminder_to_users_after_3mins_in_caas",
                args=[session_request.id],
                clocked=clocked,
                one_off=True,
            )
            periodic_task.save()

            # WHATSAPP MESSAGE CHECK
            booking_id = coach.room_id

            if coachee:
                if session_request.project.enable_emails_to_hr_and_coachee:
                    send_mail_templates(
                        "coachee_emails/session_booked.html",
                        [coachee.email],
                        "Meeraq Coaching | Session Booked",
                        {
                            "projectName": session_request.project.name,
                            "name": coachee.name,
                            "sessionName": SESSION_TYPE_VALUE[
                                session_request.session_type
                            ],
                            "slot_date": session_date,
                            "slot_time": session_time,
                            "booking_id": booking_id,
                            "email": coachee.email,
                        },
                        [],  # no bcc
                    )
                # add microsoft auth url before uncommenting
                try:
                    if existing_calendar_invite:
                        delete_outlook_calendar_invite(existing_calendar_invite)
                    attendees = []
                    # Adding learners to attendees list
                    attendees.append(
                        {
                            "emailAddress": {
                                "address": coach.email,
                            },
                            "type": "required",
                        }
                    )
                    if (
                        session_request.project.enable_emails_to_hr_and_coachee
                        and session_request.project.calendar_invites
                    ):
                        for invitee in session_request.invitees:
                            attendee = {
                                "emailAddress": {
                                    "address": invitee,
                                },
                                "type": "required",
                            }
                            attendees.append(attendee)

                        attendees.append(
                            {
                                "emailAddress": {
                                    "address": coachee.email,
                                },
                                "type": "required",
                            }
                        )
                    create_outlook_calendar_invite(
                        f"{SESSION_TYPE_VALUE[session_request.session_type]} Session",
                        "Session Scheduled",
                        int(session_request.confirmed_availability.start_time),
                        int(session_request.confirmed_availability.end_time),
                        attendees,
                        env("COACHING_CALENDAR_INVITATION_ORGANIZER"),
                        session_request,
                        None,
                        None,
                        f'{env("CAAS_APP_URL")}/call/{coach.room_id}',
                    )

                except Exception as e:
                    print(f"Calendar error {str(e)}")

    except Exception as e:
        print(f"Error occurred while creating notification: {str(e)}")

    return Response({"message": "Session booked successfully!"}, status=201)


def create_time_arr(availability):
    time_arr = []
    for time in availability:
        availibility_serilizer = AvailibilitySerializer(data=time)
        if availibility_serilizer.is_valid():
            avil_id = availibility_serilizer.save()
            time_arr.append(avil_id.id)
        else:
            return Response(
                {
                    "message": str(availibility_serilizer.errors),
                },
                status=401,
            )
    return time_arr


def get_slot_message(availability):
    slot_message = ""
    for i, slot in enumerate(availability):
        start_time = format_timestamp(slot["start_time"])
        end_time = format_timestamp(slot["end_time"])
        slot_message += f"Slot {i+1}: {start_time} - {end_time}"
        if i == (len(availability) - 1):
            slot_message += ". "
        elif i == (len(availability) - 2):
            slot_message += " and "
        else:
            slot_message += ", "
    return slot_message


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("coach", "pmo", "learner", "hr")])
def create_session_request_caas(request):
    time_arr = create_time_arr(request.data["availibility"])

    try:
        if request.data["session_type"] == "chemistry":
            session = SessionRequestCaas.objects.get(
                learner__id=request.data["learner_id"],
                project__id=request.data["project_id"],
                coach__id=request.data["coach_id"],
                session_type=request.data["session_type"],
                is_archive=False,
            )
        else:
            session = SessionRequestCaas.objects.get(
                project__id=request.data["project_id"],
                coach__id=request.data["coach_id"],
                session_type=request.data["session_type"],
                is_archive=False,
            )
        session.availibility.set(time_arr)
        session.save()
        return Response({"message": "Session updated successfully."}, status=201)
    except SessionRequestCaas.DoesNotExist:
        session = {
            "project": request.data["project_id"],
            "availibility": time_arr,
            "coach": request.data["coach_id"],
            "session_type": request.data["session_type"],
        }
        if session["session_type"] == "interview":
            session["hr"] = request.data["hr_id"]
        elif session["session_type"] == "chemistry":
            session["learner"] = request.data["learner_id"]
        session_serilizer = SessionRequestCaasSerializer(data=session)
        if session_serilizer.is_valid():
            session_created = session_serilizer.save()
            try:
                pmo_user = User.objects.filter(profile__roles__name="pmo").first()
                project = Project.objects.get(id=request.data["project_id"])
                coach = Coach.objects.get(id=request.data["coach_id"])
                if pmo_user:
                    path = f"/projects/caas/progress/{project.id}"
                    path_for_coach = f"/sessions"
                    coach_name = coach.first_name + " " + coach.last_name
                    slot_message = get_slot_message(request.data["availibility"])
                    if session["session_type"] == "interview":
                        message = f"HR has requested interview session to {coach_name.title()} for Project - {project.name}. The requested slots are "
                        message_for_coach = f"HR has requested for {slot_message} for Interview for Project - {project.name}. Please book one of the requested slots now"
                    elif session["session_type"] == "chemistry":
                        hr_user_in_project = (
                            session_created.project.hr.first().user.user
                        )
                        message_for_hr = f"{session_created.learner.name.title()} has requested for Chemistry session to the Coach - {coach_name.title()} for {slot_message} for the Project - {project.name}"
                        message_for_coach = f"Coachee has requested {slot_message} for Chemistry session for the Project - {project.name}. Please book one of the requested slots now"
                        for hr_user in hr_user_in_project:
                            create_notification(hr_user, path, message_for_hr)
                        message = f"Coachee has requested chemistry session to {coach_name.title()} for Project - {project.name}. The requested slots are "
                    message += " " + slot_message
                    create_notification(pmo_user, path, message)
                    create_notification(
                        coach.user.user, path_for_coach, message_for_coach
                    )
            except Exception as e:
                print(f"Error occurred while creating notification: {str(e)}")
            return Response({"message": "Session sequested successfully."}, status=201)
        else:
            return Response(
                {
                    "message": str(session_serilizer.errors),
                },
                status=401,
            )


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("coach", "pmo", "learner", "hr")])
def get_session_requests_of_coach(request, coach_id):
    sessions = SessionRequestCaas.objects.filter(coach__id=coach_id).all()
    serializer = SessionRequestCaasDepthTwoSerializer(sessions, many=True)
    return Response(serializer.data, status=200)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("coach", "pmo", "learner", "hr")])
def accept_coach_caas_hr(request):
    try:
        project = Project.objects.get(id=request.data.get("project_id", ""))
    except Project.DoesNotExist:
        return Response({"message": "Project does not exist"}, status=400)
    coaches_selected_count = 0
    for coach in project.coaches_status.all():
        if coach.coach_id == request.data.get("coach_id"):
            if (
                coach.status["consent"]["status"] == "select"
                and coach.status["hr"]["status"] == "sent"
            ):
                coach.status["hr"]["status"] = request.data["status"]
                coach.save()

            else:
                return Response({"error": "Failed to update status."}, status=400)
            # print(coach.id)
            # print(coach.status)
            # if coach.status not in ["HR Selected","HR Rejected"]:
            #     coach.status['hr']=request.data.get('status').split(" ")[1].lower()
            #     coach.save()
            #     print("->")
            #     print(coach.status)
            # else:
            #     return Response({"error": "Status Already Updated"}, status=400)

        if coach.status["hr"]["status"] == "select":
            coaches_selected_count += 1

    project.save()
    try:
        userId = request.data.get("user_id")
        coachId = request.data.get("coach_id")
        user_who_finalized = User.objects.get(id=userId)
        coach_who_got_finalized = Coach.objects.get(id=coachId)
        project = project
        timestamp = timezone.now()

        finalizeCoach = FinalizeCoachActivity.objects.create(
            user_who_finalized=user_who_finalized,
            coach_who_got_finalized=coach_who_got_finalized,
            project=project,
            timestamp=timestamp,
        )

        finalizeCoach.save()
    except Exception as e:
        pass
    # for i in range(0,len(project.coaches_status)):
    #     print(project.coaches_status[i])
    #     status=project.coaches_status[i].status.hr.status
    #     if status=="select":
    #         coach_count=coach_count+1

    message = ""
    if request.data.get("status") == "select":
        # Project
        try:
            contract = ProjectContract.objects.get(project=project.id)
            coach_for_contract = Coach.objects.get(id=request.data["coach_id"])
            contract_data = {}
            if not project.coach_consent_mandatory:
                contract_data = {
                    "project_contract": contract.id,
                    "project": project.id,
                    "status": "approved",
                    "coach": request.data["coach_id"],
                    "name_inputed": coach_for_contract.first_name
                    + " "
                    + coach_for_contract.last_name,
                    "response_date": timezone.now().date(),
                }
            else:
                contract_data = {
                    "project_contract": contract.id,
                    "project": project.id,
                    "status": "pending",
                    "coach": request.data["coach_id"],
                }

            contract_serializer = CoachContractSerializer(data=contract_data)

            if contract_serializer.is_valid():
                contract_serializer.save()
        except Exception as e:
            print(str(e))

        try:
            pmo_user = User.objects.filter(profile__roles__name="pmo").first()
            pmo = Pmo.objects.get(email=pmo_user.email)
            coach = Coach.objects.get(id=request.data["coach_id"])
            if pmo_user:
                path = f"/projects/caas/progress/{project.id}"
                coach_name = coach.first_name + " " + coach.last_name
                message = f"HR has selected {coach_name.title()} for the Project - {project.name}"
                message_for_coach = f"Congratulations! You have been selected by HR for the Project - {project.name}"
                create_notification(pmo_user, path, message)
                create_notification(coach.user.user, path, message)

                send_mail_templates(
                    "pmo_emails/hr_selects_a_coach.html",
                    [pmo_user.email],
                    f"Meeraq Coaching | HR selected {coach_name}",
                    {
                        "projectname": project.name,
                        "name": pmo.name,
                        "coaches_selected_count": coaches_selected_count,
                        "coachname": coach_name,
                    },
                    [],  # no bcc
                )
                microsoft_auth_url = (
                    f'{env("BACKEND_URL")}/api/microsoft/oauth/{coach.email}/'
                )
                user_token_present = False
                try:
                    user_token = UserToken.objects.get(
                        user_profile__user__username=coach.email
                    )
                    if user_token:
                        user_token_present = True
                except Exception as e:
                    pass
                send_mail_templates(
                    "coach_templates/intro_mail_to_coach.html",
                    [coach.email],
                    f"Meeraq Coaching | {project.organisation.name} has selected you",
                    {
                        "name": coach.first_name,
                        "orgName": project.organisation.name,
                        "email": coach.email,
                        "microsoft_auth_url": microsoft_auth_url,
                        "user_token_present": user_token_present,
                    },
                    [env("BCC_EMAIL")],
                )

        except Exception as e:
            print(f"Error occurred while creating notification: {str(e)}")
        message = "Coach selected."
    elif request.data.get("status") == "reject":
        try:
            coach = Coach.objects.get(id=request.data["coach_id"])
            path = f"/projects/caas/progress/{project.id}"
            message_for_coach = f"Unfortunately, your profile is not selected for the Project - {project.name}"
            create_notification(coach.user.user, path, message_for_coach)

        except Exception as e:
            print(f"Error occurred while creating notification: {str(e)}")
        message = "Coach rejected."
    return Response({"message": message}, status=200)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "hr")])
def add_learner_to_project(request):
    coacheeCounts = int(0)
    try:
        project = Project.objects.get(id=request.data["project_id"])
    except Project.DoesNotExist:
        return Response({"error": "Project does not exist."}, status=404)
    try:
        learners = create_learners(request.data["learners"])

        for learner in learners:

            create_engagement(learner, project)
            try:
                tasks = Task.objects.filter(task="add_coachee", caas_project=project)
                tasks.update(status="completed")
            except Exception as e:
                print(str(e))

            if project.enable_emails_to_hr_and_coachee:
                try:
                    path = f"/projects/caas/progress/{project.id}"
                    message = f"You have been added to Project - {project.name}"
                    create_notification(learner.user.user, path, message)
                    coacheeCounts = coacheeCounts + 1

                    send_mail_templates(
                        "coachee_emails/add_coachee.html",
                        [learner.email],
                        "Meeraq Coaching | Welcome to Meeraq",
                        {
                            "name": learner.name,
                            "orgname": project.organisation.name,
                            "email": learner.email,
                        },
                        [],
                    )
                except Exception as e:
                    print(f"Error occurred while creating notification: {str(e)}")
                continue
    except Exception as e:
        # Handle any exceptions from create_learners
        return Response({"error": str(e)}, status=500)
    try:
        pmo_user = User.objects.filter(profile__roles__name="pmo").first()
        pmo = Pmo.objects.get(email=pmo_user.email)

        if pmo_user:
            path = f"/projects/caas/progress/{project.id}"
            message = f"HR has added Coachees to the Project - {project.name}"
            create_notification(pmo_user, path, message)
            send_mail_templates(
                "pmo_emails/hr_adding_coachees.html",
                [pmo_user.email],
                "Meeraq Coaching | Coachees added",
                {
                    "projectName": project.name,
                    "name": pmo.name,
                    "coacheeCount": str(coacheeCounts),
                },
                [],  # no bcc
            )
    except Exception as e:
        print(f"Error occurred while creating notification: {str(e)}")
    return Response({"message": "Coachee added succesfully", "details": ""}, status=201)


def transform_project_structure(sessions):
    # convert project level
    #  to engagement level project structure
    # argument
    # sessions - array of objects where object has price, no. of sessions, session type, session durations
    # returns:
    # sessions - array of objects where object has - session type, session name (session type + n (numbered)) , session duration, status (pending)
    session_counts = {}
    transformed_sessions = []
    billable_session_number = 0

    for session in sessions:
        session_type = session["session_type"]
        session_duration = session["session_duration"]
        is_billable = session["billable"]
        if session_type not in session_counts:
            session_counts[session_type] = 1

        for i in range(session["no_of_sessions"]):
            session_name = f"{session_type}_{session_counts[session_type]}"
            if is_billable:
                billable_session_number = billable_session_number + 1

            transformed_session = {
                "session_name": session_name,
                "session_number": session_counts[session_type],
                "session_type": session_type,
                "session_duration": session_duration,
                "billable_session_number": (
                    billable_session_number if is_billable else None
                ),
                "status": "pending",
            }
            print(transformed_session)
            transformed_sessions.append(transformed_session)
            session_counts[session_type] += 1

    return transformed_sessions


def create_engagement(learner, project):
    existing_engagement = Engagement.objects.filter(
        learner__id=learner.id, project__id=project.id
    )
    if len(existing_engagement) == 0:
        engagemenet_project_structure = transform_project_structure(
            project.project_structure
        )
        engagement = Engagement(
            learner=learner,
            project=project,
            status="active",
            type="caas" if project.is_project_structure else "cod",
        )
        engagement.save()
        if project.is_project_structure:
            for index, session in enumerate(engagemenet_project_structure):
                session_data = SessionRequestCaas.objects.create(
                    learner=learner,
                    project=project,
                    session_duration=session["session_duration"],
                    session_number=session["session_number"],
                    session_type=session["session_type"],
                    billable_session_number=session["billable_session_number"],
                    status="pending",
                    order=index + 1,
                    engagement=engagement,
                )

            #  create task
            try:
                create_task(
                    {
                        "task": "select_a_coach",
                        "engagement": engagement.id,
                        "priority": "high",
                        "status": "pending",
                        "remarks": [],
                    },
                    30,
                )
            except Exception as e:
                print(str(e))
        return engagement
    return existing_engagement


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("coach", "pmo", "learner", "hr")])
def accept_coach_caas_learner(request):
    try:
        project = Project.objects.get(id=request.data.get("project_id", ""))
    except Project.DoesNotExist:
        return Response({"message": "Project does not exist"}, status=400)
    cnt = 0
    if not project.project_type == "COD":
        cnt = len(
            project.coaches_status.filter(
                learner_id__contains=request.data.get("learner_id")
            )
        )

    if cnt == 0:
        for coach in project.coaches_status.filter(
            coach__id=request.data.get("coach_id")
        ):
            coach.status["learner"]["status"] = request.data.get("status")
            if request.data.get("status") == "select":
                learner = Learner.objects.get(id=request.data.get("learner_id"))
                engagement = Engagement.objects.get(
                    learner__id=request.data.get("learner_id"),
                    project__id=project.id,
                    status="active",
                )
                engagement.coach = coach.coach
                engagement.save()
                # update select a coach task
                tasks = Task.objects.filter(
                    task="select_a_coach", engagement=engagement
                )
                tasks.update(status="completed")
                sessions = SessionRequestCaas.objects.filter(
                    learner__id=request.data.get("learner_id"),
                    session_type="chemistry_session",
                    project=project,
                ).exclude(coach=coach.coach)
                for session in sessions:
                    session.is_archive = True
                    session.save()
                coach.learner_id.append(request.data.get("learner_id"))
            coach.save()
    else:
        return Response({"error": "Coach Already Selected"}, status=400)
    message = ""
    if request.data.get("status") == "select":
        try:
            pmo_user = User.objects.filter(profile__roles__name="pmo").first()
            coach = Coach.objects.get(id=request.data["coach_id"])
            if pmo_user:
                path = f"/projects/caas/progress/{project.id}"
                coach_name = coach.first_name + " " + coach.last_name
                message = f"Coachee has selected {coach_name.title()} for the Project - {project.name}"
                create_notification(pmo_user, path, message)
                learner = Learner.objects.get(id=request.data.get("learner_id"))
                message_for_hr = f"{learner.name.title()} has selected {coach_name.title()} as their coach for the project - {project.name} "
                create_notification(learner.user.user, path, message_for_hr)
                message_for_coach = f"Congratulations! Coachee has selected you as their coach for the Project - {project.name}"
                create_notification(coach.user.user, path, message_for_coach)
        except Exception as e:
            print(f"Error occurred while creating notification: {str(e)}")
        message = "Coach selected succesfully."
    else:
        message = "Coach rejected."
    return Response({"message": message}, status=200)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("coach", "pmo", "learner", "hr")])
def complete_cass_step(request):
    try:
        step = request.data.get("step")
        project = Project.objects.get(id=request.data.get("project_id", ""))
    except Project.DoesNotExist:
        return Response({"message": "Project does not exist"}, status=400)
    project.steps[step]["status"] = "complete"
    project.save()
    return Response({"message": "Marked as completed."}, status=200)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("coach", "pmo", "learner", "hr")])
def mark_as_incomplete(request):
    stepList = [
        "coach_list",
        "coach_consent",
        "coach_list_to_hr",
        "interviews",
        "add_learners",
        "chemistry_session",
        "coach_selected",
        "final_coaches",
    ]
    try:
        step = request.data.get("step")
        project = Project.objects.get(id=request.data.get("project_id", ""))
    except Project.DoesNotExist:
        return Response({"message": "Project does not exist"}, status=400)
    flag = False
    steps = project.steps
    for item in stepList:
        print(item == step)
        print(flag)
        if step == item:
            flag = True
        if flag:
            if (steps[item]["status"]) == "complete":
                steps[item]["status"] = "incomplete"
    # print(statuses)
    project.steps = steps
    project.save()
    return Response({"message": "Marked as Incomplete."}, status=200)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("pmo")])
def send_project_strure_to_hr(request):
    try:
        project = Project.objects.get(id=request.data.get("project_id", ""))
    except Project.DoesNotExist:
        return Response({"message": "Project does not exist"}, status=400)
    project.steps["project_structure"]["status"] = "complete"
    project.save()

    #  adding project structure for coach if project structure not added yet.
    for coach_status in project.coaches_status.all():
        if len(coach_status.project_structure) == 0:
            coach_status.project_structure = project.project_structure
            coach_status.save()

    if not project.coach_consent_mandatory:
        for coach_status in project.coaches_status.filter(is_consent_asked=True):
            if not coach_status.status["consent"]["status"] == "reject":
                coach_status.status["consent"]["status"] = "select"
                if project.steps["project_structure"]["status"] == "complete":
                    coach_status.status["project_structure"]["status"] = "select"
                    create_task(
                        {
                            "task": "agree_project_structure",
                            "caas_project": project.id,
                            "coach": coach_status.coach.id,
                            "priority": "medium",
                            "status": "pending",
                            "remarks": [],
                        },
                        7,
                    )
            coach_status.save()
    try:
        path = f"/projects/caas/progress/{project.id}"
        message = f"Project structure has been added to the project - {project.name}."
        for hr_user in project.hr.all():
            create_notification(hr_user.user.user, path, message)
    except Exception as e:
        print(f"Error occurred while creating notification: {str(e)}")
    return Response(
        {
            "message": "Project Strcuture is Shared with HR and added to the Project successfully."
        },
        status=200,
    )


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("coach", "pmo", "learner", "hr")])
def send_reject_reason(request):
    try:
        project = Project.objects.get(id=request.data.get("project_id", ""))
    except Project.DoesNotExist:
        return Response({"message": "Project does not exist"}, status=400)
    project.steps["project_structure"]["status"] = "pending"
    rejection = dict(
        reason=request.data.get("reject_reason", ""),
        project_structure=request.data.get("project_structure", []),
    )
    if "details" not in project.steps["project_structure"]:
        project.steps["project_structure"]["details"] = []
    project.steps["project_structure"]["details"].append(rejection)
    project.save()
    return Response({"message": "Rejected."}, status=200)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("hr")])
def project_structure_agree_by_hr(request):
    try:
        project = Project.objects.get(id=request.data.get("project_id", ""))
    except Project.DoesNotExist:
        return Response({"message": "Project does not exist"}, status=400)
    project.steps["project_structure"]["status"] = "complete"
    project.save()
    return Response({"message": "Agreed."}, status=200)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("hr")])
def request_more_profiles_by_hr(request):
    try:
        project = Project.objects.get(id=request.data.get("project_id", ""))
    except Project.DoesNotExist:
        return Response({"message": "Project does not exist"}, status=400)
    if (
        request.data["step"] == "coach_list_to_hr"
        or request.data["step"] == "interviews"
    ):
        if "request_details" in project.steps["coach_consent"]:
            project.steps["coach_consent"]["request_details"].append(
                {"message": request.data["message"]}
            )
        else:
            project.steps["coach_consent"]["request_details"] = [
                {"message": request.data["message"]}
            ]
    project.steps["coach_consent"]["status"] = "incomplete"
    project.save()
    try:
        pmo_user = User.objects.filter(profile__roles__name="pmo").first()
        if pmo_user:
            path = f"/projects/caas/progress/{project.id}"
            message = (
                f"HR has requested for more coach profiles for Project - {project.name}"
            )
            create_notification(pmo_user, path, message)
    except Exception as e:
        print(f"Error occurred while creating notification: {str(e)}")
    return Response({"message": "Request sent successfully"})


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("coach", "pmo", "learner", "hr")])
def edit_learner(request):
    try:
        learner = Learner.objects.get(id=request.data.get("learner_id", ""))
    except Learner.DoesNotExist:
        return Response({"message": "Learner does not exist"}, status=400)
    email = request.data["email"].strip().lower()
    existing_user = (
        User.objects.filter(username=email).exclude(username=learner.email).first()
    )
    if existing_user:
        return Response(
            {"error": "User with this email already exists."},
            status=status.HTTP_400_BAD_REQUEST,
        )
    learner.user.user.username = email
    learner.user.user.email = email
    learner.user.user.save()
    learner.email = email
    learner.name = request.data["name"]
    learner.phone = request.data.get("phone", "")
    learner.save()
    if learner.phone:
        add_contact_in_wati("learner", learner.name, learner.phone)

    return Response({"message": "Learner details updated.", "details": ""}, status=200)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "superadmin", "learner")])
def edit_individual_learner(request, user_id):
    try:
        learner = Learner.objects.get(id=user_id)
        serializer = LearnerSerializer(learner, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            if learner.phone:
                add_contact_in_wati("learner", learner.name, learner.phone)

            return Response({"message": "Learner details updated."}, status=200)
        else:
            print(serializer.errors)
            return Response({"message": "Failed to edit learner"}, status=500)
    except Exception as e:
        return Response({"message": "Failed to edit learner"}, status=500)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("coach", "pmo", "learner", "hr")])
def mark_finalized_list_complete(request):
    try:
        project = Project.objects.get(id=request.data.get("project_id", ""))
    except Project.DoesNotExist:
        return Response({"message": "Project does not exist"}, status=400)
    project.steps["final_coaches"]["status"] = "complete"
    project.save()
    return Response({"message": "Step marked as Complete", "details": ""}, status=200)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("pmo")])
def send_list_to_hr(request):
    try:
        project = Project.objects.get(id=request.data.get("project_id", ""))
    except Project.DoesNotExist:
        return Response({"message": "Project does not exist"}, status=400)
    # project.status['coach_list_to_hr'] = 'pending'

    coaches = []

    for coach_id in request.data["coach_list"]:
        coach_status = project.coaches_status.get(coach__id=coach_id)
        coach_status.status["hr"]["status"] = "sent"
        coaches.append(Coach.objects.get(id=coach_id))
        coach_status.save()

    project.save()
    try:
        user_who_shared = User.objects.get(id=request.data.get("user_id", ""))
        project_name = project
        coaches = coaches
        timestamp = timezone.now()
        shareCoachProfile = ShareCoachProfileActivity.objects.create(
            user_who_shared=user_who_shared, project=project_name, timestamp=timestamp
        )

        shareCoachProfile.coaches.set(coaches)
        shareCoachProfile.save()
    except Exception as e:
        pass
    if project.enable_emails_to_hr_and_coachee:
        try:
            path = f"/projects/caas/progress/{project.id}"
            message = f"Admin has shared {len(request.data['coach_list'])} coach profile with you for the Project - {project.name}."
            hr_users = project.hr.all()
            for hr_user in hr_users:
                hr_email = hr_user.email
                hr_name = hr_user.first_name
                microsoft_auth_url = (
                    f'{env("BACKEND_URL")}/api/microsoft/oauth/{hr_email}/'
                )
                user_token_present = False
                try:
                    user_token = UserToken.objects.get(
                        user_profile__user__username=hr_email
                    )
                    if user_token:
                        user_token_present = True
                except Exception as e:
                    pass
                send_mail_templates(
                    "hr_emails/pmo_share_coach_list.html",
                    [hr_email],
                    "Welcome to the Meeraq Platform",
                    {
                        "name": hr_name,
                        "email": hr_email,
                        "microsoft_auth_url": microsoft_auth_url,
                        "user_token_present": user_token_present,
                    },
                    json.loads(env("BCC_EMAIL_SALES_TEAM")),  # bcc
                )
            for hr_user in project.hr.all():
                create_notification(hr_user.user.user, path, message)
        except Exception as e:
            print(f"Error occurred while creating notification: {str(e)}")
    return Response({"message": "Sent Successfully", "details": {}}, status=200)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("coach", "pmo", "learner", "hr")])
def finalized_coach_from_coach_consent(request):
    try:
        project = Project.objects.get(id=request.data.get("project_id", ""))
    except Project.DoesNotExist:
        return Response({"message": "Project does not exist"}, status=400)

    for coach_id in request.data["coach_list"]:
        coach_status = project.coaches_status.get(coach__id=coach_id)
        coach_status.status["hr"]["status"] = "select"
        coach_status.save()

    project.steps["coach_consent"]["status"] = "complete"
    project.save()

    return Response({"message": "Sent Successfully", "details": {}}, status=200)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("coach", "pmo", "learner", "hr")])
def get_coach_field_values(request):
    job_roles = set()
    languages = set()
    locations = set()
    companies_worked_in = set()
    other_certifications = set()
    industries = set()
    functional_domain = set()
    institute = set()
    for coach in Coach.objects.all():
        # 1st coach
        for role in coach.job_roles:
            job_roles.add(role)
        for language in coach.language:
            languages.add(language)
        for location in coach.location:
            locations.add(location)
        for company in coach.companies_worked_in:
            companies_worked_in.add(company)
        for certificate in coach.other_certification:
            other_certifications.add(certificate)
        for industry in coach.area_of_expertise:
            industries.add(industry)
        for functional_dom in coach.domain:
            functional_domain.add(functional_dom)
        for edu in coach.education:
            institute.add(edu)

        # domains.add(coach.domain)
        # educations.add(coach.education)
    return Response(
        {
            "job_roles": list(job_roles),
            "languages": list(languages),
            "educations": list(institute),
            "locations": list(locations),
            "companies_worked_in": list(companies_worked_in),
            "other_certifications": list(other_certifications),
            "domains": list(functional_domain),
            "industries": list(industries),
        },
        status=200,
    )


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("coach", "pmo", "learner", "hr")])
def add_mulitple_coaches(request):
    # Get data from request
    coaches = request.data.get("coaches")
    coaches_count_bulk_upload = 0
    # Check if coaches data is provided
    if not coaches or not isinstance(coaches, list):
        return Response(
            {"error": "Coaches data must be provided as a list."}, status=400
        )
    total_coaches_in_excel = len(coaches)
    for coach_data in coaches:
        email = coach_data.get("email", "").strip()
        if Coach.objects.filter(email=email).exists():
            coaches_count_bulk_upload += 1
    try:
        for coach_data in coaches:
            with transaction.atomic():
                # Extract coach details from the coach_data dictionary
                first_name = coach_data.get("first_name")
                last_name = coach_data.get("last_name")
                age = coach_data.get("age", "")
                gender = coach_data.get("gender")
                level = coach_data.get("level")
                min_fees = coach_data.get("min_fees", "")
                active_inactive = coach_data.get("active_inactive")
                corporate_yoe = coach_data.get("corporate_yoe", "")
                coaching_yoe = coach_data.get("coaching_yoe", "")
                domain = coach_data.get("functional_domain", [])
                email = coach_data.get("email", "").strip()
                phone = coach_data.get("mobile")
                phone_country_code = coach_data.get("phone_country_code")
                job_roles = coach_data.get("job_roles", [])
                companies_worked_in = coach_data.get("companies_worked_in", [])
                language = coach_data.get("language", [])
                area_of_expertise = coach_data.get("industries", [])
                location = coach_data.get("location", [])
                linkedin_profile_link = coach_data.get("linkedin_profile", "")
                coaching_hours = coach_data.get("coaching_hours", "")
                fee_remark = coach_data.get("fee_remark", "")
                client_companies = coach_data.get("client_companies", [])
                educational_qualification = coach_data.get(
                    "educational_qualification", []
                )
                corporate_experience = coach_data.get("corporate_experience", "")
                coaching_experience = coach_data.get("coaching_experience", "")
                education = coach_data.get("education", [])
                if coach_data.get("ctt_nctt") == "Yes":
                    ctt_nctt = True
                else:
                    ctt_nctt = False
                if coach_data.get("active_inactive") == "Yes":
                    active_inactive = True
                else:
                    active_inactive = False

                # Perform validation on required fields
                if not all(
                    [
                        first_name,
                        last_name,
                        gender,
                        level,
                        email,
                        phone,
                        phone_country_code,
                    ]
                ):
                    return Response(
                        {
                            "error": "All required fields must be provided for each coach."
                        },
                        status=400,
                    )

                user = User.objects.filter(email=email).first()
                if not user:
                    temp_password = "".join(
                        random.choices(
                            string.ascii_uppercase
                            + string.ascii_lowercase
                            + string.digits,
                            k=8,
                        )
                    )
                    user = User.objects.create_user(
                        username=email, password=temp_password, email=email
                    )
                    profile = Profile.objects.create(user=user)
                else:
                    profile = Profile.objects.get(user=user)

                coach_role, created = Role.objects.get_or_create(name="coach")
                profile.roles.add(coach_role)
                profile.save()

                # api call
                room_id = ""
                management_token = generateManagementToken()
                try:
                    response_from_100ms = requests.post(
                        "https://api.100ms.live/v2/rooms",
                        headers={
                            "Authorization": f"Bearer {management_token}",
                            "Content-Type": "application/json",
                        },
                        json={
                            "name": email.replace(".", "-").replace("@", ""),
                            "description": "This is a sample description for the room",
                            "region": "in",
                        },
                    )
                    if response_from_100ms.status_code == 200:
                        room_id = response_from_100ms.json().get("id")
                except Exception as e:
                    print(f"Error while generating meeting link, {str(e)}")

                # Create the Coach User using the Profile
                coach_user = Coach.objects.create(
                    user=profile,
                    first_name=first_name,
                    last_name=last_name,
                    age=age,
                    gender=gender,
                    level=level,
                    room_id=room_id,
                    min_fees=min_fees,
                    fee_remark=fee_remark,
                    ctt_nctt=ctt_nctt,
                    active_inactive=active_inactive,
                    years_of_corporate_experience=corporate_yoe,
                    years_of_coaching_experience=coaching_yoe,
                    domain=domain,
                    email=email,
                    phone=phone,
                    phone_country_code=phone_country_code,
                    job_roles=job_roles,
                    companies_worked_in=companies_worked_in,
                    language=language,
                    area_of_expertise=area_of_expertise,
                    location=location,
                    linkedin_profile_link=linkedin_profile_link,
                    coaching_hours=coaching_hours,
                    client_companies=client_companies,
                    educational_qualification=educational_qualification,
                    corporate_experience=corporate_experience,
                    coaching_experience=coaching_experience,
                    education=education,
                )

                # Approve coach
                coach = Coach.objects.get(id=coach_user.id)
                coach.is_approved = True
                coach.save()
                name = coach_user.first_name + " " + coach_user.last_name
                add_contact_in_wati("coach", name, coach_user.phone)

        return Response(
            {
                "message": f"{total_coaches_in_excel-coaches_count_bulk_upload} Coaches added successfully."
            },
            status=201,
        )
    except IntegrityError as e:
        print(str(e))
        if "Duplicate entry" in str(e):
            coaches_message = (
                f"{coaches_count_bulk_upload} coach"
                if coaches_count_bulk_upload == 1
                else f"{coaches_count_bulk_upload} coaches"
            )
            added_coaches_message = (
                f"{total_coaches_in_excel-coaches_count_bulk_upload} coach"
                if total_coaches_in_excel - coaches_count_bulk_upload == 1
                else f"{total_coaches_in_excel-coaches_count_bulk_upload} coaches"
            )
            return Response(
                {
                    "message": f"{added_coaches_message} added successfully and {coaches_message} already exist."
                },
                status=201,
            )

        return Response(
            {"error": "A coach user with this email already exists."}, status=400
        )

    except Exception as e:
        # Return error response if any other exception occurs
        print(e)
        return Response(
            {"error": "An error occurred while creating the coach user."}, status=500
        )


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("coach", "pmo", "learner", "hr")])
def get_notifications(request, user_id):
    notifications = Notification.objects.filter(user__id=user_id).order_by(
        "-created_at"
    )
    serializer = NotificationSerializer(notifications, many=True)
    return Response(serializer.data)


@api_view(["PUT"])
@permission_classes([IsAuthenticated, IsInRoles("coach", "pmo", "learner", "hr")])
def mark_all_notifications_as_read(request):
    notifications = Notification.objects.filter(
        read_status=False, user__id=request.data["user_id"]
    )
    notifications.update(read_status=True)
    return Response("Notifications marked as read.")


@api_view(["PUT"])
@permission_classes([IsAuthenticated, IsInRoles("coach", "pmo", "learner", "hr")])
def mark_notifications_as_read(request):
    user_id = request.data.get("user_id")
    notification_ids = request.data.get("notification_ids")

    if user_id is None or notification_ids is None:
        return Response("Both user_id and notification_ids are required.", status=400)

    print("abcd")

    notifications = Notification.objects.filter(
        id=notification_ids, user__id=user_id, read_status=False
    )

    notifications.update(read_status=True)
    return Response("Notifications marked as read.")


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("coach", "pmo", "learner", "hr")])
def unread_notification_count(request, user_id):
    count = Notification.objects.filter(user__id=user_id, read_status=False).count()
    return Response({"count": count})


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("coach", "pmo", "learner", "hr")])
def mark_project_as_sold(request):
    try:
        project = Project.objects.get(id=request.data.get("project_id", ""))
    except Project.DoesNotExist:
        return Response({"message": "Project does not exist"}, status=400)
    project.updated_to_sold = True
    project.steps["project_structure"]["status"] = "complete"
    project.save()
    try:
        path = f"/projects/caas/progress/{project.id}"
        message = f"Project structure has been added to the project - {project.name}."
        message_for_coach = f"Admin has added project structure. Please agree to the project structure to go forward with the Project - {project.name}"
        for hr_user in project.hr.all():
            create_notification(hr_user.user.user, path, message)
        for coach_status in project.coaches_status.all():
            create_notification(coach_status.coach.user.user, path, message_for_coach)
    except Exception as e:
        print(f"Error occurred while creating notification: {str(e)}")
    return Response({"message": "Project marked as sold"}, status=200)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("coach", "pmo", "learner", "hr")])
def get_session_requests_of_user_on_date(request, user_type, user_id, date):
    date_obj = datetime.strptime(date, "%Y-%m-%d")
    start_time = date_obj.replace(hour=0, minute=0, second=0)
    # Set the end time of the given date at 23:59:59
    end_time = date_obj.replace(hour=23, minute=59, second=59)
    # Convert the datetime objects to timestamps
    start_timestamp = int(start_time.timestamp())
    end_timestamp = int(end_time.timestamp())

    if user_type == "hr":
        session_requests = SessionRequestCaas.objects.filter(
            hr__id=user_id,
            availibility__start_time__range=(start_timestamp, end_timestamp),
        )
    elif user_type == "learner":
        session_requests = SessionRequestCaas.objects.filter(
            learner__id=user_id,
            availibility__start_time__range=(start_timestamp, end_timestamp),
        )
    elif user_type == "coach":
        session_requests = SessionRequestCaas.objects.filter(
            coach__id=user_id,
            availibility__start_time__range=(start_timestamp, end_timestamp),
        )
    serializer = SessionRequestCaasDepthOneSerializer(session_requests, many=True)
    return Response(serializer.data, status=200)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("coach", "pmo", "learner", "hr")])
def request_reschedule(request, session_id):
    session = SessionRequestCaas.objects.get(id=session_id)
    session.reschedule_request.append(
        {
            "requested_by": "coach",
            "requested_on_timestamp": request.data["requested_on"],
            "message": request.data["message"],
        }
    )
    session.save()
    return Response({"message": "Requested for reschedule"})


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("coach", "pmo", "learner", "hr")])
def reschedule_session(request):
    existing_session = SessionRequestCaas.objects.get(
        id=request.data["existing_session_id"]
    )
    existing_session.is_archive = True
    existing_session.save()
    time_arr = create_time_arr(request.data["availibility"])

    try:
        if request.data["session_type"] == "chemistry":
            session = SessionRequestCaas.objects.get(
                learner__id=request.data["learner_id"],
                project__id=request.data["project_id"],
                coach__id=request.data["coach_id"],
                session_type=request.data["session_type"],
                is_archive=False,
            )
        else:
            session = SessionRequestCaas.objects.get(
                project__id=request.data["project_id"],
                coach__id=request.data["coach_id"],
                session_type=request.data["session_type"],
                is_archive=False,
            )
        session.availibility.set(time_arr)
        session.save()
        return Response({"message": "Session updated successfully."}, status=201)
    except SessionRequestCaas.DoesNotExist:
        session = {
            "project": request.data["project_id"],
            "availibility": time_arr,
            "coach": request.data["coach_id"],
            "session_type": request.data["session_type"],
        }
        if session["session_type"] == "interview":
            session["hr"] = request.data["hr_id"]
        elif session["session_type"] == "chemistry":
            session["learner"] = request.data["learner_id"]
        session_serilizer = SessionRequestCaasSerializer(data=session)
        if session_serilizer.is_valid():
            session_created = session_serilizer.save()
            try:
                pmo_user = User.objects.filter(profile__roles__name="pmo").first()
                project = Project.objects.get(id=request.data["project_id"])
                coach = Coach.objects.get(id=request.data["coach_id"])
                if pmo_user:
                    path = f"/projects/caas/progress/{project.id}"
                    coach_name = coach.first_name + " " + coach.last_name
                    slot_message = get_slot_message(request.data["availibility"])
                    if session["session_type"] == "interview":
                        message = f"HR has requested interview session to {coach_name.title()} for Project - {project.name}. The requested slots are "
                        message_for_coach = f"HR has requested for {slot_message} for Interview for Project - {project.name}. Please book one of the requested slots now"
                    elif session["session_type"] == "chemistry":
                        hr_user_in_project = (
                            session_created.project.hr.first().user.user
                        )
                        message_for_hr = f"{session_created.learner.name.title()} has requested for Chemistry session to the Coach - {coach_name.title()} for {slot_message} for the Project - {project.name}"
                        message_for_coach = f"Coachee has requested {slot_message} for Chemistry session for the Project - {project.name}. Please book one of the requested slots now"
                        for hr_user in hr_user_in_project:
                            create_notification(hr_user, path, message_for_hr)
                        message = f"Coachee has requested chemistry session to {coach_name.title()} for Project - {project.name}. The requested slots are "
                    message += " " + slot_message
                    create_notification(pmo_user, path, message)
                    create_notification(coach.user.user, path, message_for_coach)
            except Exception as e:
                print(f"Error occurred while creating notification: {str(e)}")
            return Response({"message": "Session sequested successfully."}, status=201)
        else:
            return Response(
                {
                    "message": str(session_serilizer.errors),
                },
                status=401,
            )


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "learner", "coach", "hr")])
def get_engagement_in_projects(request, project_id):
    engagements = Engagement.objects.filter(project__id=project_id)
    engagements_data = []

    for engagement in engagements:
        completed_sessions_count = SessionRequestCaas.objects.filter(
            status="completed",
            engagement=engagement,
        ).count()

        total_sessions_count = SessionRequestCaas.objects.filter(
            engagement=engagement,
            is_archive=False,
        ).count()

        serializer = EngagementDepthOneSerializer(engagement)
        data = serializer.data
        data["completed_sessions_count"] = completed_sessions_count
        data["total_sessions_count"] = total_sessions_count
        engagements_data.append(data)

    return Response(engagements_data, status=200)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("hr")])
def get_engagements_of_hr(request, user_id):
    engagements = Engagement.objects.filter(project__hr__id=user_id)
    engagements_data = []
    current_time = int(timezone.now().timestamp() * 1000)
    for engagement in engagements:
        completed_sessions_count = SessionRequestCaas.objects.filter(
            status="completed",
            project__id=engagement.project.id,
            learner__id=engagement.learner.id,
        ).count()

        total_sessions_count = SessionRequestCaas.objects.filter(
            project__id=engagement.project.id,
            learner__id=engagement.learner.id,
            is_archive=False,
        ).count()
        upcoming_session = SessionRequestCaas.objects.filter(
            Q(is_booked=True),
            Q(confirmed_availability__end_time__gt=current_time),
            Q(project__hr__id=user_id),
            Q(project__id=engagement.project.id),
            Q(learner__id=engagement.learner.id),
            Q(is_archive=False),
            ~Q(status="completed"),
        ).aggregate(Min("confirmed_availability__end_time"))
        serializer = EngagementDepthOneSerializer(engagement)
        data = serializer.data
        data["completed_sessions_count"] = completed_sessions_count
        data["total_sessions_count"] = total_sessions_count
        data["next_session_time"] = upcoming_session.get(
            "confirmed_availability__end_time__min"
        )
        engagements_data.append(data)

    return Response(engagements_data, status=200)


class SessionCountsForAllLearners(APIView):
    permission_classes = [IsAuthenticated, IsInRoles("pmo", "coach", "hr")]

    def get(self, request, user_type, user_id, format=None):
        try:
            if user_type == "pmo":
                learners = Learner.objects.all()
            elif user_type == "coach":
                learners = Learner.objects.filter(
                    engagement__coach__id=user_id
                ).distinct()
            elif user_type == "hr":
                learners = Learner.objects.filter(
                    engagement__project__hr__id=user_id
                ).distinct()

            engagements = Engagement.objects.all()
            learner_session_counts = {}

            for engagement in engagements:
                learner_id = engagement.learner.id

                completed_sessions_count = SessionRequestCaas.objects.filter(
                    status="completed",
                    billable_session_number__isnull=False,
                    learner__id=learner_id,
                    is_archive=False,
                ).count()

                total_sessions_count = SessionRequestCaas.objects.filter(
                    learner__id=learner_id,
                    # project_id = engagement.project_id
                    billable_session_number__isnull=False,
                    is_archive=False,
                ).count()

                if learner_id in learners.values_list("id", flat=True):
                    if learner_id not in learner_session_counts:
                        learner_data = {
                            "completed_sessions_count": completed_sessions_count,
                            "total_sessions_count": total_sessions_count,
                        }
                        learner_session_counts[learner_id] = learner_data

            return Response(learner_session_counts, status=status.HTTP_200_OK)

        except ObjectDoesNotExist as e:
            return Response({"error": str(e)}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response(
                {"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_learner_engagement_of_project(request, project_id, learner_id):
    engagement = Engagement.objects.get(learner__id=learner_id, project__id=project_id)
    serializer = EngagementDepthOneSerializer(engagement)
    return Response(serializer.data, status=200)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_engagement_detail(request, engagement_id):
    engagement = Engagement.objects.get(id=engagement_id)
    serializer = EngagementDepthOneSerializer(engagement)
    return Response(serializer.data, status=200)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "learner", "coach", "hr")])
def get_learners_engagement(request, learner_id):
    engagements = Engagement.objects.filter(learner__id=learner_id)
    serializer = EngagementDepthOneSerializer(engagements, many=True)
    return Response(serializer.data, status=200)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_coaches_of_learner(request, learner_id):
    engagements = Engagement.objects.filter(learner__id=learner_id)
    data = []

    for engagement in engagements:
        if engagement.coach:
            coach_name = f"{engagement.coach.first_name} {engagement.coach.last_name}"
            project_name = engagement.project.name
            profile_pic_url = None  # Default to None
            if engagement.coach.profile_pic:
                coach_serializer = CoachSerializer(engagement.coach)
            data.append(
                {
                    "coach": engagement.coach.id,
                    "project": engagement.project.id,
                    "coach_name": coach_name,
                    "project_name": project_name,
                    "profile_pic": coach_serializer.data["profile_pic"],
                }
            )

    return Response(data, status=200)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("learner")])
def create_session_request_by_learner(request, session_id):
    # print(request.user)
    # if request.user.profile.type != "learner":
    #     return Response({"message": "Unauthorized"}, status=401)
    time_arr = create_time_arr(request.data["availibility"])
    session = SessionRequestCaas.objects.get(id=session_id)
    session.availibility.set(time_arr)
    session.status = "requested"
    session.save()
    return Response({"message": "Session requested successfully"})


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "learner", "coach", "hr")])
def get_session_requests_of_user(request, user_type, user_id):
    session_requests = []
    if user_type == "pmo":
        session_requests = SessionRequestCaas.objects.filter(
            Q(confirmed_availability=None) & ~Q(status="pending") & ~Q(status="expired")
        )
    if user_type == "learner":
        session_requests = SessionRequestCaas.objects.filter(
            Q(confirmed_availability=None)
            & Q(learner__id=user_id)
            & ~Q(status="expired")
            & ~Q(session_type="chemistry")
            & ~Q(status="pending")
        )
    if user_type == "coach":
        # session_requests = SessionRequestCaas.objects.filter(
        #     Q(confirmed_availability=None)
        #     & (Q(coach__id=user_id) | Q(coach__isnull=True))
        #     & ~Q(status="pending")
        #     & Q(project__coaches_status__coach__id=user_id)
        # # )
        # session_requests = []
        # for session in SessionRequestCaas.objects.all():
        #     if not session.confirmed_availability and not session.status == "pending":
        #         if session.coach and session.coach.id == user_id:
        #             session_requests.append(session)
        #         # is contract approved
        #         # is selected and confirmed, and coach is in the project
        #         elif session.engagement and session.engagement.type=="cod" and CoachContract.objects.filter(project=session.project, coach__id=user_id, status="approved").exists() and  session.project.coaches_status.filter(coach__id = user_id , status__hr__status="select").exists():
        #             session_requests.append(session)
        project_id = request.query_params.get("project")
        if project_id:
            session_requests = SessionRequestCaas.objects.filter(
                Q(confirmed_availability=None)
                & ~Q(status="expired")
                & Q(project__id=int(project_id))
                & ~Q(status="pending")
                & (
                    Q(coach__id=user_id)
                    | (
                        Q(project__engagement__type="cod")
                        & Q(project__coaches_status__coach__id=user_id)
                        & Q(project__coaches_status__status__hr__status="select")
                        & Exists(
                            CoachContract.objects.filter(
                                project_id=OuterRef("project_id"),
                                coach_id=user_id,
                                status="approved",
                            )
                        )
                    )
                )
            ).distinct()
        else:

            session_requests = SessionRequestCaas.objects.filter(
                Q(confirmed_availability=None)
                & ~Q(status="expired")
                & ~Q(status="pending")
                & (
                    Q(coach__id=user_id)
                    | (
                        Q(project__engagement__type="cod")
                        & Q(project__coaches_status__coach__id=user_id)
                        & Q(project__coaches_status__status__hr__status="select")
                        & Exists(
                            CoachContract.objects.filter(
                                project_id=OuterRef("project_id"),
                                coach_id=user_id,
                                status="approved",
                            )
                        )
                    )
                )
            ).distinct()

    if user_type == "hr":
        session_requests = SessionRequestCaas.objects.filter(
            Q(confirmed_availability=None)
            & ~Q(status="expired")
            & Q(project__hr__id=user_id)
            & ~Q(status="pending")
        )
    session_requests = session_requests.annotate(
        engagement_status=Subquery(
            Engagement.objects.filter(
                project=OuterRef("project"),
                learner=OuterRef("learner"),
                status="active",
            ).values("status")[:1]
        )
    )
    serializer = SessionRequestWithEngagementCaasDepthOneSerializer(
        session_requests, many=True
    )
    return Response(serializer.data, status=200)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "hr")])
def get_session_pending_of_user(request, user_type, user_id):
    session_requests = []
    if user_type == "pmo":
        session_requests = SessionRequestCaas.objects.filter(
            Q(confirmed_availability=None)
            & Q(status="pending")
            & ~Q(session_type="interview")
        )
    if user_type == "hr":
        session_requests = SessionRequestCaas.objects.filter(
            Q(confirmed_availability=None),
            Q(project__hr__id=user_id),
            Q(status="pending"),
            Q(is_archive=False),
            ~Q(session_type="interview"),
        )

    serializer = SessionRequestCaasDepthOneSerializer(session_requests, many=True)
    res = []
    for session in serializer.data:
        engagement = Engagement.objects.filter(
            learner__id=session["learner"]["id"],
            project__id=session["project"]["id"],
            status="active",
        )
        if len(engagement) > 0 and engagement[0].coach:
            coach_serializer = CoachSerializer(engagement[0].coach)
            res.append({**session, "coach": coach_serializer.data})
        else:
            res.append({**session})
    return Response(res, status=200)


# used  for hr report section
@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "learner", "coach", "hr")])
def get_all_sessions_of_user(request, user_type, user_id):
    session_requests = []
    if user_type == "pmo":
        session_requests = SessionRequestCaas.objects.filter(
            ~Q(session_type="interview"),
            ~Q(session_type="chemistry", billable_session_number=None),
            # ~Q(billable_session_number=None),
            is_archive=False,
        )
    elif user_type == "hr":
        session_requests = SessionRequestCaas.objects.filter(
            ~Q(session_type="interview"),
            ~Q(session_type="chemistry", billable_session_number=None),
            # ~Q(billable_session_number=None),
            is_archive=False,
            project__hr__id=user_id,
        )
    sessions_serializer = SessionRequestCaasDepthOneSerializer(
        session_requests, many=True
    )
    res = []
    for session in sessions_serializer.data:
        engagement = Engagement.objects.filter(
            learner__id=session["learner"]["id"],
            project__id=session["project"]["id"],
            status="active",
        )
        if len(engagement) > 0 and engagement[0].coach:
            coach_serializer = CoachSerializer(engagement[0].coach)
            res.append({**session, "coach": coach_serializer.data})
        else:
            res.append({**session})
    return Response(res, status=200)


# used for pmo report section
@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "learner", "coach", "hr")])
def get_all_sessions_of_user_for_pmo(request, user_type, user_id):
    session_requests = []
    schedular_sessions = []

    if user_type == "pmo":
        pmo = Pmo.objects.get(id=user_id)
        if pmo.sub_role == "manager":
            session_requests = SessionRequestCaas.objects.filter(
                ~Q(session_type="interview"),
                ~Q(session_type="chemistry", billable_session_number=None),
                is_archive=False,
            )
            schedular_sessions = SchedularSessions.objects.all()
        else:
            session_requests = SessionRequestCaas.objects.filter(
                ~Q(session_type="interview"),
                Q(project__junior_pmo=pmo),
                ~Q(session_type="chemistry", billable_session_number=None),
                is_archive=False,
            )
            schedular_sessions = SchedularSessions.objects.filter(
                coaching_session__batch__project__junior_pmo=pmo
            )
    elif user_type == "hr":
        session_requests = SessionRequestCaas.objects.filter(
            ~Q(session_type="interview"),
            ~Q(session_type="chemistry", billable_session_number=None),
            is_archive=False,
            project__hr__id=user_id,
        )
        schedular_sessions = SchedularSessions.objects.filter(
            coaching_session__batch__project__hr__id=user_id
        )
    res = []
    for session_request in session_requests:
        project_name = session_request.project.name
        project = ProjectSerializer(session_request.project).data
        project_type = (
            "caas" if session_request.project.project_type == "CAAS" else "COD"
        )
        organisation = session_request.project.organisation.name
        engagement = Engagement.objects.filter(
            learner_id=session_request.learner.id,
            project_id=session_request.project.id,
            status="active",
        )
        if len(engagement) > 0 and engagement[0].coach:
            coach_name = (
                engagement[0].coach.first_name + " " + engagement[0].coach.last_name
            )
            coach_email = engagement[0].coach.email
            coach = CoachSerializer(engagement[0].coach).data
        else:
            coach = None
            coach_name = None
            coach_email = None
        learner = LearnerSerializer(session_request.learner).data
        learner_name = session_request.learner.name
        learner_email = session_request.learner.email
        session_type = session_request.session_type
        session_number = session_request.session_number
        billable_session_number = session_request.billable_session_number
        status = session_request.status
        status_updated_at = session_request.status_updated_at
        if session_request.confirmed_availability:
            slot_start_time = session_request.confirmed_availability.start_time
        else:
            slot_start_time = None
        if session_request.confirmed_availability:
            slot_end_time = session_request.confirmed_availability.end_time
        else:
            slot_end_time = None
        session_duration = session_request.session_duration
        is_archive = session_request.is_archive
        invitees = session_request.invitees

        res.append(
            {
                "project_name": project_name,
                "project_type": project_type,
                "organisation": organisation,
                "coach_name": coach_name,
                "coach_email": coach_email,
                "learner": learner,
                "learner_name": learner_name,
                "learner_email": learner_email,
                "session_type": session_type,
                "session_number": session_number,
                "billable_session_number": billable_session_number,
                "status": status,
                "status_updated_at": status_updated_at,
                "slot_start_time": slot_start_time,
                "slot_end_time": slot_end_time,
                "session_duration": session_duration,
                "is_archive": is_archive,
                "invitees": invitees,
                "coach": coach,
                "project": project,
            }
        )
    for schedular_session in schedular_sessions:
        project_name = schedular_session.coaching_session.batch.project.name
        project_type = "seeq"
        organisation = (
            schedular_session.coaching_session.batch.project.organisation.name
        )
        coach_name = (
            schedular_session.availibility.coach.first_name
            + " "
            + schedular_session.availibility.coach.last_name
        )
        coach_email = schedular_session.availibility.coach.email
        learner = LearnerSerializer(schedular_session.learner).data
        learner_name = schedular_session.learner.name
        learner_email = schedular_session.learner.email
        session_type = schedular_session.coaching_session.session_type
        session_number = schedular_session.coaching_session.coaching_session_number
        billable_session_number = None
        status = schedular_session.status
        status_updated_at = schedular_session.updated_at
        slot_start_time = schedular_session.availibility.start_time
        slot_end_time = schedular_session.availibility.end_time
        session_duration = schedular_session.coaching_session.duration
        coach = CoachSerializer(schedular_session.availibility.coach).data
        is_archive = None
        invitees = []

        res.append(
            {
                "project_name": project_name,
                "project_type": project_type,
                "organisation": organisation,
                "coach_name": coach_name,
                "coach_email": coach_email,
                "learner": learner,
                "learner_name": learner_name,
                "learner_email": learner_email,
                "session_type": session_type,
                "session_number": session_number,
                "billable_session_number": billable_session_number,
                "status": status,
                "status_updated_at": status_updated_at,
                "slot_start_time": slot_start_time,
                "slot_end_time": slot_end_time,
                "session_duration": session_duration,
                "is_archive": is_archive,
                "invitees": invitees,
                "coach": coach,
            }
        )

    return Response({"sessions": res})


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "learner", "coach", "hr")])
def get_upcoming_sessions_of_user(request, user_type, user_id):
    current_time = int(timezone.now().timestamp() * 1000)
    session_requests = []
    if user_type == "pmo":
        pmo = Pmo.objects.get(id=user_id)
        if pmo.sub_role == "manager":
            session_requests = SessionRequestCaas.objects.filter(
                Q(is_booked=True),
                Q(confirmed_availability__end_time__gt=current_time),
                ~Q(status="completed"),
            )
        else:
            session_requests = SessionRequestCaas.objects.filter(
                Q(is_booked=True),
                Q(confirmed_availability__end_time__gt=current_time),
                ~Q(status="completed"),
                Q(project__junior_pmo=pmo),
            )
    if user_type == "learner":
        session_requests = SessionRequestCaas.objects.filter(
            Q(is_booked=True),
            Q(confirmed_availability__end_time__gt=current_time),
            Q(learner__id=user_id),
            ~Q(session_type="chemistry"),
            Q(is_archive=False),
            ~Q(status="completed"),
        )
    if user_type == "coach":
        session_requests = SessionRequestCaas.objects.filter(
            Q(is_booked=True),
            Q(confirmed_availability__end_time__gt=current_time),
            Q(coach__id=user_id),
            Q(is_archive=False),
            ~Q(status="completed"),
        )
    if user_type == "hr":
        session_requests = SessionRequestCaas.objects.filter(
            Q(is_booked=True),
            Q(confirmed_availability__end_time__gt=current_time),
            Q(project__hr__id=user_id),
            Q(is_archive=False),
            ~Q(status="completed"),
        )

    session_requests = session_requests.annotate(
        engagement_status=Subquery(
            Engagement.objects.filter(
                project=OuterRef("project"),
                learner=OuterRef("learner"),
                status="active",
            ).values("status")[:1]
        )
    )
    serializer = SessionRequestWithEngagementCaasDepthOneSerializer(
        session_requests, many=True
    )
    return Response(serializer.data, status=200)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "learner", "coach", "hr")])
def new_get_upcoming_sessions_of_user(request, user_type, user_id):
    current_time = int(timezone.now().timestamp() * 1000)
    session_requests = []
    current_time_seeq = timezone.now()
    timestamp_milliseconds = str(int(current_time_seeq.timestamp() * 1000))
    avaliable_sessions = []

    if user_type == "pmo":
        pmo = Pmo.objects.get(id=user_id)
        if pmo.sub_role == "manager":

            session_requests = SessionRequestCaas.objects.filter(
                Q(is_booked=True),
                Q(confirmed_availability__end_time__gt=current_time),
                ~Q(status="completed"),
            )
            schedular_sessions = SchedularSessions.objects.all()
            avaliable_sessions = schedular_sessions.filter(
                availibility__end_time__gt=timestamp_milliseconds
            )
        else:
            session_requests = SessionRequestCaas.objects.filter(
                Q(is_booked=True),
                Q(confirmed_availability__end_time__gt=current_time),
                ~Q(status="completed"),
                Q(project__junior_pmo=pmo),
            )
            schedular_sessions = SchedularSessions.objects.filter(
                coaching_session__batch__project__junior_pmo=pmo
            )
            avaliable_sessions = schedular_sessions.filter(
                availibility__end_time__gt=timestamp_milliseconds
            )

    if user_type == "learner":
        session_requests = SessionRequestCaas.objects.filter(
            Q(is_booked=True),
            Q(confirmed_availability__end_time__gt=current_time),
            Q(learner__id=user_id),
            ~Q(session_type="chemistry"),
            Q(is_archive=False),
            ~Q(status="completed"),
        )
        learner = Learner.objects.get(id=user_id)
        schedular_sessions = SchedularSessions.objects.filter(learner=learner)
        avaliable_sessions = schedular_sessions.filter(
            availibility__end_time__gt=timestamp_milliseconds
        )
    if user_type == "coach":
        project_id = request.query_params.get("project")
        if project_id:
            session_requests = SessionRequestCaas.objects.filter(
                Q(is_booked=True),
                Q(confirmed_availability__end_time__gt=current_time),
                Q(coach__id=user_id),
                Q(is_archive=False),
                ~Q(status="completed"),
                Q(project__id=int(project_id)),
            )

            schedular_sessions = SchedularSessions.objects.filter(
                availibility__coach__id=user_id,
                coaching_session__batch__project__id=int(project_id),
            )
        else:

            session_requests = SessionRequestCaas.objects.filter(
                Q(is_booked=True),
                Q(confirmed_availability__end_time__gt=current_time),
                Q(coach__id=user_id),
                Q(is_archive=False),
                ~Q(status="completed"),
            )
            schedular_sessions = SchedularSessions.objects.filter(
                availibility__coach__id=user_id
            )
        avaliable_sessions = schedular_sessions.filter(
            availibility__end_time__gt=timestamp_milliseconds
        )
    if user_type == "hr":
        session_requests = SessionRequestCaas.objects.filter(
            Q(is_booked=True),
            Q(confirmed_availability__end_time__gt=current_time),
            Q(project__hr__id=user_id),
            Q(is_archive=False),
            ~Q(status="completed"),
        )
        schedular_sessions = SchedularSessions.objects.filter(
            coaching_session__batch__project__hr__id=user_id
        )
        avaliable_sessions = schedular_sessions.filter(
            availibility__end_time__gt=timestamp_milliseconds
        )

    session_requests = session_requests.annotate(
        engagement_status=Subquery(
            Engagement.objects.filter(
                project=OuterRef("project"),
                learner=OuterRef("learner"),
                status="active",
            ).values("status")[:1]
        ),
        is_seeq_project=Value(False, output_field=BooleanField()),
    )
    session_details = []
    coach_id = None
    if user_type == "coach":
        coach_id = user_id
    for session in avaliable_sessions:
        session_detail = {
            "id": session.id,
            "batch_name": (
                session.coaching_session.batch.name if coach_id is None else None
            ),
            "project_name": session.coaching_session.batch.project.name,
            "organisation_name": session.coaching_session.batch.project.organisation.name,
            "project_id": (
                session.coaching_session.batch.project.id if coach_id is None else None
            ),
            "coach_name": session.availibility.coach.first_name
            + " "
            + session.availibility.coach.last_name,
            "coach_email": session.availibility.coach.email,
            "coach_phone": "+"
            + session.availibility.coach.phone_country_code
            + session.availibility.coach.phone,
            "participant_name": session.learner.name,
            "participant_email": session.learner.email,
            "participant_phone": session.learner.phone,
            "coaching_session_number": (
                session.coaching_session.coaching_session_number
                if coach_id is None
                else None
            ),
            "meeting_link": f"{env('CAAS_APP_URL')}/call/{session.availibility.coach.room_id}",
            "start_time": session.availibility.start_time,
            "room_id": f"{session.availibility.coach.room_id}",
            "status": session.status,
            "session_type": session.coaching_session.session_type,
            "end_time": session.availibility.end_time,
            "is_seeq_project": True,
            "coaching_session_id": session.coaching_session.id,
        }
        session_details.append(session_detail)

    serializer = SessionRequestWithEngagementCaasAndIsSeeqProjectDepthOneSerializer(
        session_requests, many=True
    )
    return Response(
        {
            "caas_session_details": serializer.data,
            "seeq_session_details": session_details,
        },
        status=200,
    )


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "learner", "coach", "hr")])
def get_past_sessions_of_user(request, user_type, user_id):
    current_time = int(timezone.now().timestamp() * 1000)
    session_requests = []
    if user_type == "pmo":
        pmo = Pmo.objects.get(id=user_id)
        if pmo.sub_role == "manager":
            session_requests = SessionRequestCaas.objects.filter(
                Q(is_booked=True),
                Q(confirmed_availability__end_time__lt=current_time)
                | Q(status="completed"),
            )
        else:
            session_requests = SessionRequestCaas.objects.filter(
                Q(is_booked=True),
                Q(confirmed_availability__end_time__lt=current_time)
                | Q(status="completed"),
                Q(project__junior_pmo=pmo),
            )

    if user_type == "learner":
        session_requests = SessionRequestCaas.objects.filter(
            Q(is_booked=True),
            Q(confirmed_availability__end_time__lt=current_time)
            | Q(status="completed"),
            Q(learner__id=user_id),
            ~Q(session_type="chemistry"),
            Q(is_archive=False),
        )
    if user_type == "coach":
        session_requests = SessionRequestCaas.objects.filter(
            Q(is_booked=True),
            Q(confirmed_availability__end_time__lt=current_time)
            | Q(status="completed"),
            Q(coach__id=user_id),
            Q(is_archive=False),
        )
    if user_type == "hr":
        session_requests = SessionRequestCaas.objects.filter(
            Q(is_booked=True),
            Q(confirmed_availability__end_time__lt=current_time)
            | Q(status="completed"),
            Q(project__hr__id=user_id),
            Q(is_archive=False),
        )

    session_requests = session_requests.annotate(
        engagement_status=Subquery(
            Engagement.objects.filter(
                project=OuterRef("project"),
                learner=OuterRef("learner"),
                status="active",
            ).values("status")[:1]
        )
    )
    for session_request in session_requests:
        print(session_request.engagement_status)
    serializer = SessionRequestWithEngagementCaasDepthOneSerializer(
        session_requests, many=True
    )
    return Response(serializer.data, status=200)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "learner", "coach", "hr")])
def new_get_past_sessions_of_user(request, user_type, user_id):
    current_time = int(timezone.now().timestamp() * 1000)
    session_requests = []
    current_time_seeq = timezone.now()
    timestamp_milliseconds = int(current_time_seeq.timestamp() * 1000)
    avaliable_sessions = []
    if user_type == "pmo":
        pmo = Pmo.objects.get(id=user_id)
        if pmo.sub_role == "manager":
            session_requests = SessionRequestCaas.objects.filter(
                Q(is_booked=True),
                Q(confirmed_availability__end_time__lt=current_time)
                | Q(status="completed"),
            )
            schedular_sessions = SchedularSessions.objects.all()
            avaliable_sessions = schedular_sessions.filter(
                availibility__end_time__lt=timestamp_milliseconds
            )
        else:
            session_requests = SessionRequestCaas.objects.filter(
                Q(is_booked=True),
                Q(confirmed_availability__end_time__lt=current_time)
                | Q(status="completed"),
                Q(project__junior_pmo=pmo),
            )
            schedular_sessions = SchedularSessions.objects.filter(
                coaching_session__batch__project__junior_pmo=pmo
            )
            avaliable_sessions = schedular_sessions.filter(
                availibility__end_time__lt=timestamp_milliseconds
            )
    if user_type == "learner":
        session_requests = SessionRequestCaas.objects.filter(
            Q(is_booked=True),
            Q(confirmed_availability__end_time__lt=current_time)
            | Q(status="completed"),
            Q(learner__id=user_id),
            ~Q(session_type="chemistry"),
            Q(is_archive=False),
        )
        learner = Learner.objects.get(id=user_id)
        schedular_sessions = SchedularSessions.objects.filter(learner=learner)
        avaliable_sessions = schedular_sessions.filter(
            availibility__end_time__lt=timestamp_milliseconds
        )
    if user_type == "coach":
        project_id = request.query_params.get("project")

        if project_id:
            session_requests = SessionRequestCaas.objects.filter(
                Q(is_booked=True),
                Q(confirmed_availability__end_time__lt=current_time)
                | Q(status="completed"),
                Q(coach__id=user_id),
                Q(is_archive=False),
                Q(project__id=int(project_id)),
            )
            schedular_sessions = SchedularSessions.objects.filter(
                availibility__coach__id=user_id,
                coaching_session__batch__project__id=int(project_id),
            )
        else:
            session_requests = SessionRequestCaas.objects.filter(
                Q(is_booked=True),
                Q(confirmed_availability__end_time__lt=current_time)
                | Q(status="completed"),
                Q(coach__id=user_id),
                Q(is_archive=False),
            )
            schedular_sessions = SchedularSessions.objects.filter(
                availibility__coach__id=user_id
            )
        avaliable_sessions = schedular_sessions.filter(
            availibility__end_time__lt=timestamp_milliseconds
        )
    if user_type == "hr":
        session_requests = SessionRequestCaas.objects.filter(
            Q(is_booked=True),
            Q(confirmed_availability__end_time__lt=current_time)
            | Q(status="completed"),
            Q(project__hr__id=user_id),
            Q(is_archive=False),
        )
        schedular_sessions = SchedularSessions.objects.filter(
            coaching_session__batch__project__hr__id=user_id
        )
        avaliable_sessions = schedular_sessions.filter(
            availibility__end_time__lt=timestamp_milliseconds
        )

    session_requests = session_requests.annotate(
        engagement_status=Subquery(
            Engagement.objects.filter(
                project=OuterRef("project"),
                learner=OuterRef("learner"),
                status="active",
            ).values("status")[:1]
        ),
        is_seeq_project=Value(False, output_field=BooleanField()),
    )
    session_details = []
    coach_id = None
    if user_type == "coach":
        coach_id = user_id
    for session in avaliable_sessions:
        session_detail = {
            "id": session.id,
            "batch_name": (
                session.coaching_session.batch.name if coach_id is None else None
            ),
            "project_name": session.coaching_session.batch.project.name,
            "organisation_name": session.coaching_session.batch.project.organisation.name,
            "project_id": (
                session.coaching_session.batch.project.id if coach_id is None else None
            ),
            "coach_name": session.availibility.coach.first_name
            + " "
            + session.availibility.coach.last_name,
            "coach_email": session.availibility.coach.email,
            "coach_phone": "+"
            + session.availibility.coach.phone_country_code
            + session.availibility.coach.phone,
            "participant_name": session.learner.name,
            "participant_email": session.learner.email,
            "participant_phone": session.learner.phone,
            "coaching_session_number": (
                session.coaching_session.coaching_session_number
                if coach_id is None
                else None
            ),
            "meeting_link": f"{env('CAAS_APP_URL')}/call/{session.availibility.coach.room_id}",
            "start_time": session.availibility.start_time,
            "room_id": f"{session.availibility.coach.room_id}",
            "status": session.status,
            "session_type": session.coaching_session.session_type,
            "end_time": session.availibility.end_time,
            "session_duration": session.coaching_session.duration,
            "is_seeq_project": True,
            "auto_generated_status": session.auto_generated_status,
            "coaching_session_id": session.coaching_session.id,
            "learner_id": session.learner.id,
            "learner_profile_pic": (
                session.learner.profile_pic.url if session.learner.profile_pic else None
            ),
        }
        session_details.append(session_detail)

    serializer = SessionRequestWithEngagementCaasAndIsSeeqProjectDepthOneSerializer(
        session_requests, many=True
    )
    return Response(
        {
            "caas_session_details": serializer.data,
            "seeq_session_details": session_details,
        },
        status=200,
    )


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "coach", "hr", "learner")])
def edit_session_status(request, session_id):
    try:
        session_request = SessionRequestCaas.objects.get(id=session_id)
    except SessionRequestCaas.DoesNotExist:
        return Response({"error": "Session request not found."}, status=404)
    new_status = request.data.get("status")
    if not new_status:
        return Response({"error": "Status field is required."}, status=400)
    session_request.status = new_status
    current_time = datetime.now()
    session_request.status_updated_at = current_time
    session_request.save()
    try:
        engagement = Engagement.objects.get(
            learner=session_request.learner,
            project=session_request.project,
            status="active",
        )
        if session_request.session_type == "goal_setting":
            tasks = Task.objects.filter(task="add_goal", engagement=engagement)
            if not tasks.exists():
                create_task(
                    {
                        "task": "add_goal",
                        "engagement": engagement.id,
                        "priority": "high",
                        "status": "pending",
                        "remarks": [],
                    },
                    7,
                )
    except Exception as e:
        print(str(e))
    return Response({"message": "Session status updated successfully."}, status=200)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "coach", "hr", "learner")])
def edit_session_availability(request, session_id):
    time_arr = create_time_arr(request.data["availibility"])
    try:
        # changing availability - edit request.
        session = SessionRequestCaas.objects.get(id=session_id)

        existing_calendar_invite = CalendarInvites.objects.filter(
            caas_session=session
        ).first()

        if session.is_booked:
            return Response({"message": "Session edit failed."}, status=401)
        session.availibility.set(time_arr)
        session.status = "requested"
        session.save()
        if existing_calendar_invite:
            existing_calendar_invite.caas_session = session

        return Response({"message": "Session updated successfully."}, status=201)
    except:
        return Response({"message": "Session edit failed."}, status=401)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "coach", "hr")])
def get_coachee_of_user(request, user_type, user_id):
    learners_data = []
    if user_type == "pmo":
        learners = Learner.objects.all()
    elif user_type == "coach":
        learners = Learner.objects.filter(engagement__coach__id=user_id).distinct()
    elif user_type == "hr":
        learners = Learner.objects.filter(
            Q(engagement__project__hr__id=user_id)
            | Q(schedularbatch__project__hr__id=user_id)
        ).distinct()
    for learner in learners:
        learner_dict = {
            "id": learner.id,
            "name": learner.name,
            "email": learner.email,
            "phone": learner.phone,
            "area_of_expertise": learner.area_of_expertise,
            "years_of_experience": learner.years_of_experience,
            "projects": [],
            "organisation": set(),
        }
        if user_type == "pmo":
            projects = Project.objects.filter(engagement__learner=learner)
            schedular_batches = SchedularBatch.objects.filter(
                learners__email=learner.email
            )
            course_enrollments = CourseEnrollment.objects.filter(learner__id=learner.id)
            courses_names = []
            for course_enrollment in course_enrollments:
                courses_names.append(course_enrollment.course.name)
            learner_dict["coursesEnrolled"] = courses_names
        elif user_type == "coach":
            projects = Project.objects.filter(
                Q(engagement__learner=learner) & Q(engagement__coach__id=user_id)
            )
        elif user_type == "hr":
            projects = Project.objects.filter(
                Q(engagement__learner=learner) & Q(hr__id=user_id)
            )
            schedular_batches = SchedularBatch.objects.filter(
                Q(learners__email=learner.email) & Q(project__hr__id=user_id)
            )
            course_enrollments = CourseEnrollment.objects.filter(learner__id=learner.id)
            courses_names = []
            for course_enrollment in course_enrollments:
                courses_names.append(course_enrollment.course.name)
            learner_dict["coursesEnrolled"] = courses_names
        for project in projects:
            project_dict = {
                "project_id": project.id,
                "name": project.name,
                "type": "CAAS",
            }
            learner_dict["organisation"].add(project.organisation.name)
            learner_dict["projects"].append(project_dict)
        if user_type == "pmo" or user_type == "hr":
            for batch in schedular_batches:
                project_dict = {
                    "project_id": batch.project.id,
                    "batch_id": batch.id,
                    "name": batch.project.name,
                    "type": "SEEQ",
                }
                learner_dict["organisation"].add(batch.project.organisation.name)
                if project_dict["name"] not in [
                    proj["name"] for proj in learner_dict["projects"]
                ]:
                    learner_dict["projects"].append(project_dict)
        learners_data.append(learner_dict)
    return Response(learners_data)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "coach", "hr")])
def get_learner_of_user_optimized(request, user_type, user_id):
    try:
        learners = None
        if user_type == "pmo":
            learners = Learner.objects.all()
        elif user_type == "coach":
            learners = Learner.objects.filter(engagement__coach__id=user_id).distinct()
        elif user_type == "hr":
            learners = Learner.objects.filter(
                Q(engagement__project__hr__id=user_id)
                | Q(schedularbatch__project__hr__id=user_id)
            ).distinct()

        serializer = LearnerSerializer(learners, many=True)
        return Response(serializer.data)
    except Exception as e:
        print(str(e))
        return Response({"error": str(e)}, status=500)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "coach", "hr")])
def get_coachee_of_coach(request):
    try:
        engagements = Engagement.objects.filter(
            coach__user__user__id=request.user.id
        ).distinct()
        serializer = EngagementDepthOneSerializer(engagements, many=True)
        return Response(serializer.data)
    except Exception as e:
        print(str(e))
        return Response({"error": str(e)}, status=500)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "coach", "learner", "hr")])
def get_learner_course_enrolled_of_user_optimized(request, user_type, user_id):
    try:
        learners = None
        learner_dict = {}

        if user_type == "pmo":
            learners = Learner.objects.all()
        elif user_type == "coach":
            learners = Learner.objects.filter(engagement__coach__id=user_id).distinct()
        elif user_type == "hr":
            learners = Learner.objects.filter(
                Q(engagement__project__hr__id=user_id)
                | Q(schedularbatch__project__hr__id=user_id)
            ).distinct()

        for learner in learners:
            if user_type == "pmo" or user_type == "hr":
                course_enrollments = CourseEnrollment.objects.filter(learner=learner)
                courses_names = [
                    course_enrollment.course.name
                    for course_enrollment in course_enrollments
                ]
                learner_dict[learner.id] = courses_names

        return Response(learner_dict)

    except Exception as e:
        print(str(e))
        return Response({"error": str(e)}, status=500)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "coach", "learner", "hr")])
def get_project_organisation_learner_of_user_optimized(request, user_type, user_id):
    try:
        learners = None
        learner_dict_project = {}
        learner_dict_organisation = {}

        if user_type == "pmo":
            learners = Learner.objects.all()
        elif user_type == "coach":
            learners = Learner.objects.filter(engagement__coach__id=user_id).distinct()
        elif user_type == "hr":
            learners = Learner.objects.filter(
                Q(engagement__project__hr__id=user_id)
                | Q(schedularbatch__project__hr__id=user_id)
            ).distinct()

        for learner in learners:
            projects = None
            schedular_batches = None
            if user_type == "pmo":
                projects = Project.objects.filter(engagement__learner=learner)
                schedular_batches = SchedularBatch.objects.filter(
                    learners__email=learner.email
                )
            elif user_type == "coach":
                projects = Project.objects.filter(
                    Q(engagement__learner=learner) & Q(engagement__coach__id=user_id)
                )
            elif user_type == "hr":
                projects = Project.objects.filter(
                    Q(engagement__learner=learner) & Q(hr__id=user_id)
                )
                schedular_batches = SchedularBatch.objects.filter(
                    Q(learners__email=learner.email) & Q(project__hr__id=user_id)
                )

            learner_dict_organisation[learner.id] = set()
            learner_dict_project[learner.id] = []

            for project in projects:
                project_dict = {
                    "project_id": project.id,
                    "name": project.name,
                    "type": "CAAS" if project.project_type == "CAAS" else "COD",
                }

                learner_dict_organisation[learner.id].add(project.organisation.name)
                learner_dict_project[learner.id].append(project_dict)

            if user_type == "pmo" or user_type == "hr":
                for batch in schedular_batches:
                    project_dict = {
                        "project_id": batch.project.id,
                        "batch_id": batch.id,
                        "name": batch.project.name,
                        "type": "SEEQ",
                    }
                    learner_dict_organisation[learner.id].add(
                        batch.project.organisation.name
                    )
                    if project_dict["name"] not in [
                        proj["name"] for proj in learner_dict_project[learner.id]
                    ]:
                        learner_dict_project[learner.id].append(project_dict)

        return Response(
            {
                "learner_dict_project": learner_dict_project,
                "learner_dict_organisation": learner_dict_organisation,
            }
        )

    except Exception as e:
        print(str(e))
        return Response({"error": str(e)}, status=500)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("coach", "pmo", "hr", "learner")])
def get_learner_data(request, learner_id):
    learner = Learner.objects.get(id=learner_id)
    serializer = LearnerSerializer(learner)
    return Response(serializer.data)


# updating the availability and coach in the first pending chemistry available for learner
@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("learner", "hr", "pmo")])
def request_chemistry_session(request, project_id, learner_id):
    session = SessionRequestCaas.objects.filter(
        learner__id=learner_id,
        project__id=project_id,
        session_type="chemistry",
        status="pending",
    )

    if len(session) == 0:
        return Response({"error": "Max sessions exceeded."}, status=400)
    else:
        coach = Coach.objects.get(id=request.data["coach_id"])
        time_arr = create_time_arr(request.data["availibility"])
        session_to_update = session[0]
        print("****************************", session_to_update.id, session_to_update)
        session_to_update.availibility.set(time_arr)
        session_to_update.coach = coach
        session_to_update.status = "requested"
        session_to_update.save()
        path_for_coach = f"/sessions"
        slot_message = get_slot_message(request.data["availibility"])
        message_for_coach = f"Coachee has requested {slot_message} for Chemistry session for the Project - {session_to_update.project.name}. Please book one of the requested slots now"
        create_notification(coach.user.user, path_for_coach, message_for_coach)

        project = Project.objects.get(id=project_id)
        time_of_request = timezone.now()
        coachee = Learner.objects.get(id=learner_id)
        session_name = "chemistry"

        session_request = SessionRequestedActivity.objects.create(
            project=project,
            time_of_request=time_of_request,
            coach=coach,
            coachee=coachee,
            session_name=session_name,
        )

    session_request.save()

    return Response({"message": "Session requested successfully"}, status=200)


# updating the availability and coach in the first pending chemistry available for learner
@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("learner", "hr", "pmo")])
def request_chemistry_session_for_engagement(request, engagement_id):
    with transaction.atomic():
        engagement = Engagement.objects.get(id=engagement_id)
        session = SessionRequestCaas.objects.filter(
            engagement=engagement,
            session_type="chemistry",
            status="pending",
        )

        if len(session) == 0:
            return Response({"error": "Max sessions exceeded."}, status=400)
        else:
            coach = Coach.objects.get(id=request.data["coach_id"])
            time_arr = create_time_arr(request.data["availibility"])
            session_to_update = session[0]
            print(
                "****************************", session_to_update.id, session_to_update
            )
            session_to_update.availibility.set(time_arr)
            session_to_update.coach = coach
            session_to_update.status = "requested"
            session_to_update.save()
            path_for_coach = f"/sessions"
            slot_message = get_slot_message(request.data["availibility"])
            message_for_coach = f"Coachee has requested {slot_message} for Chemistry session for the Project - {session_to_update.project.name}. Please book one of the requested slots now"
            create_notification(coach.user.user, path_for_coach, message_for_coach)

            project = Project.objects.get(id=engagement.project.id)
            time_of_request = timezone.now()
            coachee = Learner.objects.get(id=engagement.project.id)
            session_name = "chemistry"

            session_request = SessionRequestedActivity.objects.create(
                project=project,
                time_of_request=time_of_request,
                coach=coach,
                coachee=coachee,
                session_name=session_name,
            )

        session_request.save()

        return Response({"message": "Session requested successfully"}, status=200)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "learner", "hr", "coach")])
def get_learner_sessions_in_project(request, project_id, learner_id):
    sessions = SessionRequestCaas.objects.filter(
        project__id=project_id, learner__id=learner_id
    ).order_by("order")
    serializer = SessionRequestCaasDepthOneSerializer(sessions, many=True)
    return Response(serializer.data, status=200)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "learner", "hr", "coach")])
def get_learner_sessions_in_project_from_engagement(request, engagement_id):
    sessions = SessionRequestCaas.objects.filter(engagement__id=engagement_id).order_by(
        "order"
    )
    serializer = SessionRequestCaasDepthOneSerializer(sessions, many=True)
    return Response(serializer.data, status=200)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("learner")])
def request_session(request, session_id, coach_id):
    session = SessionRequestCaas.objects.get(id=session_id)
    coach = Coach.objects.get(id=coach_id)
    time_arr = create_time_arr(request.data["availibility"])
    session.availibility.set(time_arr)
    session.coach = coach
    session.status = "requested"
    session.save()

    project = session.project
    time_of_request = timezone.now()
    coachee = session.learner
    session_name = session.session_type

    session_request = SessionRequestedActivity.objects.create(
        project=project,
        time_of_request=time_of_request,
        coach=coach,
        coachee=coachee,
        session_name=session_name,
    )

    session_request.save()

    return Response({"message": "Session requested successfully"}, status=200)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("learner")])
def request_session_without_project_structure(request, engagement_id):
    try:
        with transaction.atomic():
            engagement = Engagement.objects.get(id=engagement_id)

            sessions = SessionRequestCaas.objects.filter(
                engagement=engagement
            ).order_by("order")

            max_order = sessions.aggregate(Max("order"))["order__max"]

            if max_order is None:
                max_order = 0

            session_data = SessionRequestCaas.objects.create(
                learner=engagement.learner,
                project=engagement.project,
                session_duration=engagement.project.duration_of_each_session,
                session_number=max_order + 1,
                session_type="coaching_session",
                billable_session_number=max_order + 1,
                status="requested",
                order=max_order + 1,
                engagement=engagement,
                requested_at=timezone.now(),
            )

            time_arr = create_time_arr(request.data["availibility"])
            session_data.availibility.set(time_arr)

            session_data.save()

            return Response({"message": "Session requested successfully"}, status=200)

    except Exception as e:
        print(str(e))
        return Response({"error": "Failed to create request."}, status=500)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("learner", "coach", "hr", "pmo")])
def reschedule_session_of_coachee(request, session_id):
    session = SessionRequestCaas.objects.get(id=session_id)
    session.is_archive = True

    time_arr = create_time_arr(request.data["availibility"])
    new_session = SessionRequestCaas.objects.create(
        learner=session.learner,
        project=session.project,
        coach=session.coach,
        session_type=session.session_type,
        status="requested",
        session_number=session.session_number,
        session_duration=session.session_duration,
        order=session.order,
    )
    new_session.availibility.set(time_arr)
    new_session.save()
    session.save()

    return Response({"message": "Session reschedule successfully"}, status=200)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("learner", "coach")])
def create_goal(request):
    user_email = request.data.get("email")
    serializer = GoalSerializer(data=request.data)
    goal_name = request.data["name"]
    engagement_id = request.data.get("engagement")
    if not Goal.objects.filter(name=goal_name, engagement__id=engagement_id).exists():
        if Goal.objects.filter(engagement__id=engagement_id).count() < 10:
            if serializer.is_valid():
                goal_instance = serializer.save()
                try:
                    tasks = Task.objects.filter(
                        task="add_goal", engagement=engagement_id
                    )
                    tasks.update(status="completed")
                except Exception as e:
                    print(str(e))
                try:
                    create_task(
                        {
                            "task": "add_competency",
                            "engagement": int(engagement_id),
                            "goal": goal_instance.id,
                            "priority": "high",
                            "status": "pending",
                            "remarks": [],
                        },
                        1,
                    )
                    user_instance = User.objects.get(username=user_email)
                    AddGoalActivity.objects.create(
                        user=user_instance,
                        timestamp=timezone.now(),
                    )
                except Exception as e:
                    print("AddGoalActivity error", str(e))

                return Response({"message": "Goal created successfully."}, status=201)
            return Response(serializer.errors, status=400)
        else:
            return Response(
                {"error": "The project already has 10 goals. Cannot add more."},
                status=400,
            )
    else:
        return Response({"error": "Goal already exists for the project."}, status=400)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("learner", "coach", "hr", "pmo")])
def get_engagement_goals(request, engagement_id):
    goals = Goal.objects.filter(engagement__id=engagement_id)
    serializer = GetGoalSerializer(goals, many=True)
    return Response(serializer.data, status=200)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("learner", "coach")])
def edit_goal(request, goal_id):
    try:
        goal = Goal.objects.get(id=goal_id)
    except Goal.DoesNotExist:
        return Response({"error": "Goal not found."}, status=404)

    serializer = GoalSerializer(instance=goal, data=request.data)
    goal_name = request.data.get("name", goal.name)
    engagement_id = request.data.get("engagement", goal.engagement.id)

    if (
        not Goal.objects.filter(name=goal_name, engagement_id=engagement_id)
        .exclude(id=goal_id)
        .exists()
    ):
        if serializer.is_valid():
            serializer.save()
            return Response({"message": "Goal updated successfully."}, status=200)
        return Response(serializer.errors, status=400)
    else:
        return Response(
            {
                "error": "Another goal with the same name already exists in the engagement."
            },
            status=400,
        )


@api_view(["DELETE"])
@permission_classes([IsAuthenticated, IsInRoles("learner", "coach")])
def delete_goal(request, goal_id):
    try:
        goal = Goal.objects.get(id=goal_id)
    except Goal.DoesNotExist:
        return Response({"error": "Goal not found."}, status=404)
    goal.delete()
    return Response({"message": "Goal deleted successfully."}, status=204)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("learner", "coach")])
def create_competency(request):
    serializer = CompetencySerializer(data=request.data)
    goal_id = request.data.get("goal")
    competency_count = Competency.objects.filter(goal__id=goal_id).count()

    if competency_count < 10:
        if serializer.is_valid():
            competency = serializer.save()
            # update tasks
            try:
                tasks = Task.objects.filter(task="add_competency", goal__id=goal_id)
                tasks.update(status="completed")
            except Exception as e:
                print(str(e))
            return Response(
                {
                    "message": "Competency added successfully",
                    "competency": serializer.data,
                },
                status=201,
            )
        return Response(serializer.errors, status=400)
    else:
        return Response(
            {"error": "The goal already has 10 competencies. Cannot add more."},
            status=400,
        )


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("learner", "coach")])
def edit_competency(request, competency_id):
    try:
        competency = Competency.objects.get(id=competency_id)
    except Competency.DoesNotExist:
        return Response({"error": "Competency not found."}, status=404)

    serializer = CompetencySerializer(instance=competency, data=request.data)
    competency_name = request.data["name"]
    if (
        not Competency.objects.filter(name=competency_name, goal__id=competency.goal.id)
        .exclude(id=competency.goal.id)
        .exists()
    ):
        if serializer.is_valid():
            serializer.save()
            return Response({"message": "Competency updated successfully."}, status=200)
        return Response(serializer.errors, status=400)
    else:
        return Response(
            {
                "error": "Another competency with the same name already exists in the goal."
            },
            status=400,
        )


@api_view(["DELETE"])
@permission_classes([IsAuthenticated, IsInRoles("learner", "coach", "pmo")])
def delete_competency(request, competency_id):
    try:
        competency = Competency.objects.get(id=competency_id)
    except Competency.DoesNotExist:
        return Response({"error": "Competency not found."}, status=404)
    competency.delete()
    return Response({"message": "Goal deleted successfully."}, status=204)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("learner", "coach", "hr", "pmo")])
def get_engagement_competency(request, engagement_id):
    competentcy = Competency.objects.filter(goal__engagement__id=engagement_id)
    serializer = CompetencyDepthOneSerializer(competentcy, many=True)
    return Response(serializer.data, status=200)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("learner", "coach")])
def add_score_to_competency(request, competency_id):
    try:
        competency = Competency.objects.get(id=competency_id)
        scoring_data = {
            "date": request.data.get("date"),
            "score": request.data.get("score"),
        }
        if len(competency.scoring) >= 20:
            return Response(
                {
                    "error": "The maximum number of scores (20) for this competency has been reached."
                },
                status=400,
            )

        competency.scoring.append(scoring_data)
        competency.save()
        return Response(
            {"message": "Competency scoring updated successfully."}, status=201
        )
    except Competency.DoesNotExist:
        return Response({"error": "Competency not found."}, status=404)
    except Exception as e:
        return Response({"error": str(e)}, status=400)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("learner", "coach", "hr", "pmo")])
def get_competency_by_goal(request, goal_id):
    competentcy = Competency.objects.filter(goal__id=goal_id)
    serializer = CompetencyDepthOneSerializer(competentcy, many=True)
    return Response(serializer.data, status=200)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("learner", "coach")])
def create_action_item(request):
    serializer = ActionItemSerializer(data=request.data)
    competency_id = request.data.get("competency")

    if ActionItem.objects.filter(competency__id=competency_id).count() < 20:
        if serializer.is_valid():
            serializer.save()
            return Response({"message": "Action created successfully."}, status=201)
        return Response(serializer.errors, status=400)
    else:
        return Response(
            {"error": "The competency already has 20 action items. Cannot add more."},
            status=400,
        )


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("learner", "coach", "hr", "pmo")])
def get_engagement_action_items(request, engagement_id):
    action_items = ActionItem.objects.filter(
        competency__goal__engagement__id=engagement_id
    )
    serializer = GetActionItemDepthOneSerializer(action_items, many=True)
    return Response(serializer.data, status=200)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("learner", "coach", "hr", "pmo")])
def get_action_items_by_competency(request, competency_id):
    action_items = ActionItem.objects.filter(competency__id=competency_id)
    serializer = GetActionItemDepthOneSerializer(action_items, many=True)
    return Response(serializer.data, status=200)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("learner", "coach")])
def edit_action_item(request, action_item_id):
    try:
        action_item = ActionItem.objects.get(id=action_item_id)
    except ActionItem.DoesNotExist:
        return Response({"error": "Action item not found."}, status=404)
    serializer = ActionItemSerializer(instance=action_item, data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response({"message": "Action item updated successfully."}, status=200)
    return Response(serializer.errors, status=400)


@api_view(["DELETE"])
@permission_classes([IsAuthenticated, IsInRoles("learner", "coach")])
def delete_action_item(request, action_item_id):
    try:
        action_item = ActionItem.objects.get(id=action_item_id)
    except ActionItem.DoesNotExist:
        return Response({"error": "Action item not found."}, status=404)
    action_item.delete()
    return Response({"message": "Action item deleted successfully"}, status=204)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "learner", "coach", "hr")])
def mark_session_as_complete(request, session_id):
    try:
        session = SessionRequestCaas.objects.get(id=session_id)
    except SessionRequestCaas.DoesNotExist:
        return Response({"error": "Session not found."}, status=404)
    session.status = "completed"
    session.save()
    return Response({"message": "Session marked as complete."}, status=201)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "learner", "coach", "hr")])
def update_engagement_status(request, status, engagement_id):
    try:
        engagement = Engagement.objects.get(id=engagement_id)
    except Engagement.DoesNotExist:
        return Response({"error": "Engagement not found."}, status=404)
    engagement.status = status
    engagement.save()
    return Response({"message": f"Engagement marked as {status}."}, status=201)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "learner", "coach", "hr")])
def get_all_competencies(request):
    goals_with_competencies = Goal.objects.prefetch_related("competency_set").all()

    competency_list = []

    for goal in goals_with_competencies:
        goal_name = goal.name
        project_name = (
            goal.engagement.project.name
            if goal.engagement and goal.engagement.project
            else "N/A"
        )
        coachee_name = (
            goal.engagement.learner.name
            if goal.engagement and goal.engagement.learner
            else "N/A"
        )
        coachee_email = (
            goal.engagement.learner.email
            if goal.engagement and goal.engagement.learner
            else "N/A"
        )
        coach_name = (
            goal.engagement.coach.first_name + " " + goal.engagement.coach.last_name
            if goal.engagement and goal.engagement.coach
            else "N/A"
        )
        coach_email = (
            goal.engagement.coach.email
            if goal.engagement and goal.engagement.coach
            else "N/A"
        )

        # Include goals without competencies
        if goal.competency_set.exists():
            for competency in goal.competency_set.all():
                competency_data = {
                    "id": competency.id,
                    "goal_id": goal_name,
                    "name": competency.name,
                    "scoring": competency.scoring,
                    "created_at": competency.created_at.isoformat(),
                    "project_name": project_name,
                    "learner_name": coachee_name,
                    "learner_email": coachee_email,
                    "coach_name": coach_name,
                    "coach_email": coach_email,
                }
                competency_list.append(competency_data)
        else:
            # Include goals with no competencies
            competency_data = {
                "id": None,
                "goal_id": goal_name,
                "name": "-",
                "scoring": [],
                "created_at": None,
                "project_name": project_name,
                "learner_name": coachee_name,
                "learner_email": coachee_email,
                "coach_name": coach_name,
                "coach_email": coach_email,
            }
            competency_list.append(competency_data)

    return Response({"competencies": competency_list})


@api_view(["GET"])
@permission_classes(
    [IsAuthenticated, IsInRoles("coach", "learner", "pmo", "hr", "facilitator")]
)
def get_current_session(request, user_type, room_id, user_id):
    five_minutes_in_milliseconds = 300000
    current_time = int(timezone.now().timestamp() * 1000)
    five_minutes_plus_current_time = current_time + five_minutes_in_milliseconds
    caas_sessions = None
    seeq_sessions = None
    if user_type == "coach":
        caas_sessions = SessionRequestCaas.objects.filter(
            Q(is_booked=True),
            Q(confirmed_availability__end_time__gt=current_time),
            Q(coach__id=user_id),
            Q(coach__room_id=room_id),
            Q(is_archive=False),
            ~Q(status="completed"),
        ).order_by("confirmed_availability__start_time")

        coach = Coach.objects.get(id=user_id)
        seeq_sessions = SchedularSessions.objects.filter(
            availibility__end_time__gt=current_time,
            availibility__coach__email=coach.email,
            availibility__coach__room_id=room_id,
        ).order_by("availibility__start_time")

    elif user_type == "learner":
        caas_sessions = SessionRequestCaas.objects.filter(
            Q(is_booked=True),
            Q(confirmed_availability__end_time__gt=current_time),
            Q(learner__id=user_id),
            (Q(coach__room_id=room_id) | Q(pmo__room_id=room_id)),
            Q(is_archive=False),
            ~Q(status="completed"),
        ).order_by("confirmed_availability__start_time")

        learner = Learner.objects.get(id=user_id)
        seeq_sessions = SchedularSessions.objects.filter(
            availibility__end_time__gt=current_time,
            learner__email=learner.email,
            availibility__coach__room_id=room_id,
        ).order_by("availibility__start_time")

    elif user_type in ["hr", "pmo"]:

        caas_sessions = SessionRequestCaas.objects.filter(
            Q(is_booked=True),
            Q(confirmed_availability__end_time__gt=current_time),
            Q(hr__id=user_id) if user_type == "hr" else Q(pmo__id=user_id),
            (Q(coach__room_id=room_id) | Q(pmo__room_id=room_id)),
            Q(is_archive=False),
            ~Q(status="completed"),
        ).order_by("confirmed_availability__start_time")

    if not seeq_sessions and not caas_sessions:
        return Response({"error": "You don't have any upcoming sessions."}, status=404)

    nearest_session = None
    if caas_sessions:
        nearest_session = caas_sessions.first()
    if seeq_sessions:
        if nearest_session:
            if (
                seeq_sessions.first().availibility.start_time
                < nearest_session.confirmed_availability.start_time
            ):
                nearest_session = seeq_sessions.first()
        else:
            nearest_session = seeq_sessions.first()

    if nearest_session:
        session_details = {
            "session_id": nearest_session.id,
            "type": (
                "CAAS" if isinstance(nearest_session, SessionRequestCaas) else "SEEQ"
            ),
            "start_time": (
                nearest_session.confirmed_availability.start_time
                if isinstance(nearest_session, SessionRequestCaas)
                else nearest_session.availibility.start_time
            ),
            "end_time": (
                nearest_session.confirmed_availability.end_time
                if isinstance(nearest_session, SessionRequestCaas)
                else nearest_session.availibility.end_time
            ),
        }
        response_data = {
            "message": "success",
            "upcoming_session": session_details,
        }
        return Response(response_data, status=200)
    else:
        return Response({"error": "You don't have any upcoming sessions."}, status=404)


@api_view(["GET"])
@permission_classes(
    [IsAuthenticated, IsInRoles("coach", "learner", "pmo", "hr", "facilitator")]
)
def get_current_session_for_coach(request, user_type, user_id):
    current_time = int(timezone.now().timestamp() * 1000)
    caas_sessions = None
    seeq_sessions = None
    if user_type == "coach":

        coach = Coach.objects.get(id=user_id)
        seeq_sessions = SchedularSessions.objects.filter(
            availibility__end_time__gt=current_time,
            availibility__coach__email=coach.email,
        ).order_by("availibility__start_time")
    current_time = int(timezone.now().timestamp() * 1000)
    caas_sessions = None
    seeq_sessions = None
    if user_type == "coach":
        caas_sessions = SessionRequestCaas.objects.filter(
            Q(is_booked=True),
            Q(confirmed_availability__end_time__gt=current_time),
            Q(coach__id=user_id),
            Q(is_archive=False),
            ~Q(status="completed"),
        ).order_by("confirmed_availability__start_time")

        coach = Coach.objects.get(id=user_id)
        seeq_sessions = SchedularSessions.objects.filter(
            availibility__end_time__gt=current_time,
            availibility__coach__email=coach.email,
        ).order_by("availibility__start_time")
    if not seeq_sessions and not caas_sessions:
        return Response({"error": "You don't have any upcoming sessions."}, status=404)

    nearest_session = None
    if caas_sessions:
        nearest_session = caas_sessions.first()
    if seeq_sessions:
        print("seeq", seeq_sessions)
        if nearest_session:
            if (
                seeq_sessions.first().availibility.start_time
                < nearest_session.confirmed_availability.start_time
            ):
                nearest_session = seeq_sessions.first()
        else:
            nearest_session = seeq_sessions.first()

    if nearest_session:
        session_details = {
            "session_id": nearest_session.id,
            "type": (
                "CAAS" if isinstance(nearest_session, SessionRequestCaas) else "SEEQ"
            ),
            "start_time": (
                nearest_session.confirmed_availability.start_time
                if isinstance(nearest_session, SessionRequestCaas)
                else nearest_session.availibility.start_time
            ),
            "end_time": (
                nearest_session.confirmed_availability.end_time
                if isinstance(nearest_session, SessionRequestCaas)
                else nearest_session.availibility.end_time
            ),
            "project_name": (
                nearest_session.project.name
                if isinstance(nearest_session, SessionRequestCaas)
                else nearest_session.coaching_session.batch.project.name
            ),
            "session_name": (
                f"{nearest_session.session_type}"
                if isinstance(nearest_session, SessionRequestCaas)
                else nearest_session.coaching_session.session_type
            ),
            "room_id": (
                nearest_session.coach.room_id
                if isinstance(nearest_session, SessionRequestCaas)
                else nearest_session.availibility.coach.room_id
            ),
        }
        response_data = {
            "message": "success",
            "upcoming_session": session_details,
        }
        return Response(response_data, status=200)
    else:
        return Response({"error": "You don't have any upcoming sessions."}, status=404)


@api_view(["POST"])
@permission_classes([AllowAny])
def get_current_session_of_stakeholder(request, room_id):
    five_minutes_in_milliseconds = 300000
    current_time = int(timezone.now().timestamp() * 1000)
    five_minutes_plus_current_time = current_time + five_minutes_in_milliseconds
    sessions = SessionRequestCaas.objects.filter(
        Q(is_booked=True),
        Q(confirmed_availability__start_time__lt=five_minutes_plus_current_time)
        & Q(confirmed_availability__end_time__gt=current_time),
        (Q(coach__room_id=room_id) | Q(pmo__room_id=room_id)),
        Q(session_type="tripartite")
        | Q(session_type="mid_review")
        | Q(session_type="end_review")
        | Q(session_type="stakeholder_without_coach")
        | Q(session_type="stakeholder_interview"),
        Q(invitees__contains=request.data["email"]),
        Q(is_archive=False),
        ~Q(status="completed"),
    )
    if len(sessions) == 0:
        return Response({"error": "You don't have any sessions right now."}, status=404)
    return Response({"message": "success"}, status=200)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "coach", "learner", "hr")])
def schedule_session_directly(request, session_id):
    with transaction.atomic():
        try:
            session = SessionRequestCaas.objects.get(id=session_id)
        except SessionRequestCaas.DoesNotExist:
            return Response({"error": "Session not found."}, status=404)

        existing_calendar_invite = CalendarInvites.objects.filter(
            caas_session=session
        ).first()

        if session.learner:
            coachee = session.learner

            sessionName = str(session.session_type).replace("_", " ")
            if sessionName == "stakeholder without coach":
                sessionName = "tripartite without coach"
        time_arr = create_time_arr(request.data.get("availability", []))
        if len(time_arr) == 0:
            return Response({"error": "Please provide the availability."}, status=404)

        availability = Availibility.objects.get(id=time_arr[0])
        if request.data["user_type"] == "coach":
            check_booking = check_if_the_avalibility_is_already_booked(
                request.data["user_id"], availability
            )
        if check_booking:
            return Response(
                {
                    "error": "Sorry, this time is already booked. Please choose a different time.."
                },
                status=500,
            )
        if request.data["user_type"] == "pmo":
            pmo = Pmo.objects.get(id=request.data["user_id"])
            session.pmo = pmo

        if request.data["user_type"] == "coach":
            coach = Coach.objects.get(id=request.data["user_id"])
            session.coach = coach

        if session.session_type == "stakeholder_interview":
            engagement = Engagement.objects.get(
                learner=session.learner, project=session.project, status="active"
            )
            session.coach = engagement.coach
            session.hr = session.project.hr.first()

        session.availibility.add(availability)
        session.confirmed_availability = availability
        start_time = format_timestamp(int(session.confirmed_availability.start_time))
        end_time = format_timestamp(int(session.confirmed_availability.end_time))
        slot_message = f"{start_time} - {end_time}"

        session_date = get_date(int(session.confirmed_availability.start_time))
        start_time = get_time(int(session.confirmed_availability.start_time))
        end_time = get_time(int(session.confirmed_availability.end_time))

        session_time = f"{start_time} - {end_time} IST"

        session.is_booked = True

        session.status = "booked"
        session.invitees = get_trimmed_emails(request.data.get("invitees", []))
        session.save()
        tasks = Task.objects.filter(task="update_session_status", session_caas=session)
        if not tasks.exists():
            create_task(
                {
                    "task": "update_session_status",
                    "caas_project": session.project.id,
                    "coach": session.coach.id if session.coach else None,
                    "session_caas": session.id,
                    "priority": "high",
                    "status": "pending",
                    "remarks": [],
                },
                1,
            )
        if (
            session.session_type in SESSIONS_WITH_STAKEHOLDERS
            and len(session.invitees) == 0
        ):
            create_task(
                {
                    "task": "add_invitees",
                    "caas_project": session.project.id,
                    "coach": session.coach.id if session.coach else None,
                    "session_caas": session.id,
                    "priority": "high",
                    "status": "pending",
                    "remarks": [],
                },
                1,
            )
        if session.coach and session.learner and session.project:
            tasks = Task.objects.filter(
                task="book_session",
                engagement__project=session.project,
                engagement__learner=session.learner,
            )
            tasks.update(status="completed")

        coach = None
        if request.data["user_type"] == "coach":
            coach = Coach.objects.get(id=request.data["user_id"])
        if coachee:

            # print(,start_time,end_time)
            # WHATSAPP MESSAGE CHECK
            start_datetime_obj = datetime.fromtimestamp(
                int(session.confirmed_availability.start_time) / 1000
            )
            # Decrease 5 minutes
            five_minutes_prior_start_datetime = start_datetime_obj - timedelta(
                minutes=5
            )
            clocked = ClockedSchedule.objects.create(
                clocked_time=five_minutes_prior_start_datetime
            )
            periodic_task = PeriodicTask.objects.create(
                name=uuid.uuid1(),
                task="schedularApi.tasks.send_whatsapp_reminder_to_users_before_5mins_in_caas",
                args=[session_id],
                clocked=clocked,
                one_off=True,
            )
            periodic_task.save()

            # after 3 minutes
            three_minutes_ahead_start_datetime = start_datetime_obj + timedelta(
                minutes=3
            )
            clocked = ClockedSchedule.objects.create(
                clocked_time=three_minutes_ahead_start_datetime
            )
            periodic_task = PeriodicTask.objects.create(
                name=uuid.uuid1(),
                task="schedularApi.tasks.send_whatsapp_reminder_to_users_after_3mins_in_caas",
                args=[session_id],
                clocked=clocked,
                one_off=True,
            )
            periodic_task.save()
            # WHATSAPP MESSAGE CHECK

            if session.project.enable_emails_to_hr_and_coachee:
                send_mail_templates(
                    "coachee_emails/session_booked.html",
                    [coachee.email],
                    "Meeraq Coaching | Session Booked",
                    {
                        "projectName": session.project.name,
                        "name": coachee.name,
                        "sessionName": SESSION_TYPE_VALUE[session.session_type],
                        "slot_date": session_date,
                        "slot_time": session_time,
                        "booking_id": (
                            coach.room_id
                            if request.data["user_type"] == "coach"
                            else pmo.room_id
                        ),
                        "email": coachee.email,
                    },
                    [],  # no bcc
                )

                try:
                    if existing_calendar_invite:
                        delete_outlook_calendar_invite(existing_calendar_invite)

                    attendees = []
                    if request.data["user_type"] == "coach":
                        attendees.append(
                            {
                                "emailAddress": {
                                    "address": coach.email,
                                },
                                "type": "required",
                            }
                        )
                    else:
                        attendees.append(
                            {
                                "emailAddress": {
                                    "address": pmo.email,
                                },
                                "type": "required",
                            }
                        )
                    if session.project.calendar_invites:
                        for invitee in session.invitees:
                            attendee = {
                                "emailAddress": {
                                    "address": invitee,
                                },
                                "type": "required",
                            }
                            attendees.append(attendee)
                        attendees.append(
                            {
                                "emailAddress": {
                                    "address": coachee.email,
                                },
                                "type": "required",
                            }
                        )
                    create_outlook_calendar_invite(
                        f"{SESSION_TYPE_VALUE[session.session_type]} Session",
                        "Session Scheduled",
                        int(session.confirmed_availability.start_time),
                        int(session.confirmed_availability.end_time),
                        attendees,
                        env("COACHING_CALENDAR_INVITATION_ORGANIZER"),
                        session,
                        None,
                        None,
                        f'{env("CAAS_APP_URL")}/call/{coach.room_id if request.data["user_type"] == "coach" else pmo.room_id}',
                    )

                except Exception as e:
                    print(f"Calendar error {str(e)}")

        return Response({"message": "Session booked successfully."})


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("pmo")])
def delete_learner_from_project(request, engagement_id):
    try:
        engagement = Engagement.objects.get(id=engagement_id)
    except Engagement.DoesNotExist:
        return Response({"error": "Learner not found in project."}, status=404)
    try:
        print("outside the for loop ")
        # removing the learner id from the project coaches statuses
        for coach_statuses in engagement.project.coaches_status.all():
            print("inside for loop")
            if engagement.learner.id in coach_statuses.learner_id:
                coach_statuses.learner_id.remove(engagement.learner.id)
            coach_statuses.save()
        sessions_to_delete = SessionRequestCaas.objects.filter(
            learner__id=engagement.learner.id, project__id=engagement.project.id
        )
        sessions_to_delete.delete()
        engagement.delete()
        return Response({"message": "Learner deleted successfully"})
    except Exception as e:
        print(str(e))
        return Response(
            {"error": "Failed to remove learner from the project."}, status=400
        )


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("pmo")])
def reset_consent(request):
    try:
        coach_status = get_object_or_404(
            CoachStatus, id=request.data.get("coach_status_id")
        )
    except CoachStatus.DoesNotExist:
        return Response({"error": "Could not find coach consent"}, status=404)

    if request.data.get("type") == "consent":
        coach_status.status["consent"]["status"] = "sent"
        coach_status.consent_expiry_date = request.data["consent_expiry_date"]

    coach_status.status["project_structure"]["status"] = "sent"
    coach_status.save()

    return Response({"message": "Consent reset successfully"})


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "coach", "learner", "hr")])
def get_competency_averages(request, hr_id):
    # Step 1: Retrieve the data from the Competency model
    competencies = Competency.objects.filter(goal__engagement__project__hr__id=hr_id)
    # Step 2 and 3: Calculate the average score for each competency and store in a dictionary
    competency_averages = defaultdict(lambda: {"total_score": 0, "count": 0})

    for competency in competencies:
        competency_name = competency.name
        scoring_data = competency.scoring
        if scoring_data and len(scoring_data) > 0:
            # taking only the latest score
            total_score = scoring_data[-1]["score"]
            # sum(entry["score"] for entry in scoring_data)
            # count = len(scoring_data)
            competency_averages[competency_name]["total_score"] += total_score
            competency_averages[competency_name]["count"] += 1
    # Step 4: Calculate the final average for each competency, on their latest score
    final_averages = {}
    for competency_name, data in competency_averages.items():
        total_score = data["total_score"]
        count = data["count"]
        average_score = total_score / count if count > 0 else 0
        final_averages[competency_name] = average_score
    top_5_competencies = dict(
        sorted(final_averages.items(), key=lambda item: item[1], reverse=True)[:5]
    )
    return Response(top_5_competencies, status=200)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("learner", "coach", "pmo", "hr")])
def get_learner_competency_averages(request, learner_id):
    competencies = Competency.objects.filter(goal__engagement__learner__id=learner_id).order_by('-created_at')
    serializer = CompetencyDepthOneSerializer(competencies, many=True)
    return Response(serializer.data, status=200)



@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("hr")])
def get_upcoming_session_count(request, hr_id):
    current_time = int(timezone.now().timestamp() * 1000)
    session_requests = []
    # Get the start and end of the current month
    # current_month = timezone.now().replace(
    #     day=1, hour=0, minute=0, second=0, microsecond=0
    # )
    # current_month_timestamp = int(current_month.timestamp() * 1000)
    # next_month = current_month.replace(month=current_month.month + 1, day=1)
    # if current_month.month == 12:  # Handle December case
    #     next_month = next_month.replace(year=current_month.year + 1)
    # next_month_timestamp = int(next_month.timestamp() * 1000)
    session_requests = SessionRequestCaas.objects.filter(
        Q(is_booked=True),
        Q(confirmed_availability__end_time__gt=current_time),
        Q(project__hr__id=hr_id),
        Q(is_archive=False),
        ~Q(status="completed"),
    )
    upcoming_session_count = session_requests.count()
    serializer = SessionRequestCaasDepthOneSerializer(session_requests, many=True)
    return Response(
        {
            "upcoming_sessions": serializer.data,
            "upcoming_session_count": upcoming_session_count,
        },
        status=200,
    )


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("hr")])
def get_requests_count(request, hr_id):
    session_requests = SessionRequestCaas.objects.filter(
        Q(confirmed_availability=None) & Q(project__hr__id=hr_id) & ~Q(status="pending")
    )
    requests_count = session_requests.count()
    serializer = SessionRequestCaasDepthOneSerializer(session_requests, many=True)
    return Response(
        {
            "requestes": serializer.data,
            "requests_count": requests_count,
        },
        status=200,
    )


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("hr")])
def get_completed_sessions_count(request, hr_id):
    session_requests = SessionRequestCaas.objects.filter(
        Q(project__hr__id=hr_id) & Q(status="completed")
    )
    sessions_count = session_requests.count()
    serializer = SessionRequestCaasDepthOneSerializer(session_requests, many=True)
    return Response(
        {
            "sessions": serializer.data,
            "sessions_count": sessions_count,
        },
        status=200,
    )


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("learner")])
def get_completed_learner_sessions_count(request, learner_id):
    learner_session_requests = SessionRequestCaas.objects.filter(
        Q(learner__id=learner_id) & Q(status="completed")
    )
    sessions_count = learner_session_requests.count()
    serializer = SessionRequestCaasDepthOneSerializer(
        learner_session_requests, many=True
    )
    return Response(
        {
            "sessions": serializer.data,
            "sessions_count": sessions_count,
        },
        status=200,
    )


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("learner")])
def get_total_goals_for_learner(request, learner_id):
    try:
        total_goals = Goal.objects.filter(engagement__learner_id=learner_id).count()
        return Response(
            {"total_goals": total_goals},
            status=200,
        )
    except Exception as e:
        return Response(
            {"error": str(e)},
            status=400,
        )


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("learner")])
def get_total_competencies_for_learner(request, learner_id):
    try:
        total_competencies = Competency.objects.filter(
            goal__engagement__learner_id=learner_id
        ).count()
        return Response(
            {"total_competency": total_competencies},
            status=200,
        )
    except Exception as e:
        return Response(
            {"error": str(e)},
            status=400,
        )


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("hr")])
def get_learners_without_sessions(request, hr_id):
    # Get the learners associated with the given hr_id who don't have any sessions with status = "requested" or "booked".
    learners = (
        Learner.objects.filter(engagement__project__hr__id=hr_id)
        .exclude(sessionrequestcaas__status__in=["requested", "booked"])
        .distinct()
    )
    # Serialize the filtered learner data.
    learners_count = learners.count()
    serializer = LearnerSerializer(learners, many=True)
    return Response(
        {"learners": serializer.data, "learners_count": learners_count}, status=200
    )


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "hr")])
def select_coach_for_coachee(request):
    try:
        coach_status = CoachStatus.objects.get(id=request.data["coach_status_id"])
    except CoachStatus.DoesNotExist:
        return Response({"error": "Coach not found"}, status=404)
    try:
        engagement = Engagement.objects.get(id=request.data["engagement_id"])
    except Engagement.DoesNotExist:
        return Response({"error": "Unable to find learner"}, status=404)
    coach_status.learner_id.append(engagement.learner.id)
    coach_status.save()
    engagement.coach = coach_status.coach
    engagement.save()
    tasks = Task.objects.filter(task="select_a_coach", engagement=engagement)
    tasks.update(status="completed")
    return Response({"message": "Coach finalized for coachee"}, status=201)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "hr", "coach")])
def add_past_session(request, session_id):
    # print("request data",request.data)
    try:
        session = SessionRequestCaas.objects.get(id=session_id)
    except SessionRequestCaas.DoesNotExist:
        return Response({"error": "Session not found."}, status=404)

    time_arr = create_time_arr(request.data.get("availability", []))

    if len(time_arr) == 0:
        return Response({"error": "Please provide the availability."}, status=404)
    availability = Availibility.objects.get(id=time_arr[0])
    session.availibility.add(availability)
    session.confirmed_availability = availability
    session.is_booked = True
    if session.session_type != "stakeholder_without_coach":
        coach_id = request.data.get("coach_id")
        if coach_id:
            try:
                coach = Coach.objects.get(id=coach_id)
                session.coach = coach

            except Coach.DoesNotExist:
                return Response({"error": "Coach not found."}, status=404)

    session.status = "completed"
    session.invitees = get_trimmed_emails(request.data.get("invitees", []))
    session.status_updated_at = datetime.now()
    session.save()

    coach = session.coach
    coachee = session.learner
    project = session.project
    timestamp = timezone.now()
    user_id = request.data.get("userId")
    user_who_added = User.objects.get(id=user_id)

    addPastSession = PastSessionActivity.objects.create(
        project=project,
        user_who_added=user_who_added,
        coach=coach,
        coachee=coachee,
        timestamp=timestamp,
        session_name=session.session_type,
    )
    addPastSession.save()

    return Response({"message": "Session booked successfully."})


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("learner")])
def get_pending_action_items_by_competency(request, learner_id):
    action_items = ActionItem.objects.filter(
        competency__goal__engagement__learner_id=learner_id, status="not_done"
    )
    serializer = PendingActionItemSerializer(action_items, many=True)
    return Response(serializer.data, status=200)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("hr")])
def get_all_competencies_of_hr(request, hr_id):
    goals_with_competencies = Goal.objects.prefetch_related("competency_set").filter(
        engagement__project__hr=hr_id
    )

    competency_list = []

    for goal in goals_with_competencies:
        goal_name = goal.name
        project_name = (
            goal.engagement.project.name
            if goal.engagement and goal.engagement.project
            else "N/A"
        )
        coachee_name = (
            goal.engagement.learner.name
            if goal.engagement and goal.engagement.learner
            else "N/A"
        )

        coach_name = (
            goal.engagement.coach.first_name + " " + goal.engagement.coach.last_name
            if goal.engagement and goal.engagement.coach
            else "N/A"
        )

        # Include goals without competencies
        if goal.competency_set.exists():
            for competency in goal.competency_set.all():
                competency_data = {
                    "id": competency.id,
                    "goal_id": goal_name,
                    "name": competency.name,
                    "scoring": competency.scoring,
                    "created_at": competency.created_at.isoformat(),
                    "project_name": project_name,
                    "learner_name": coachee_name,
                    "coach_name": coach_name,
                }
                competency_list.append(competency_data)
        else:
            # Include goals with no competencies
            competency_data = {
                "id": None,
                "goal_id": goal_name,
                "name": "-",
                "scoring": [],
                "created_at": None,
                "project_name": project_name,
                "learner_name": coachee_name,
                "coach_name": coach_name,
            }
            competency_list.append(competency_data)

    return Response({"competencies": competency_list})


class UpdateInviteesView(APIView):
    permission_classes = [IsAuthenticated, IsInRoles("pmo", "coach", "hr")]

    def put(self, request, session_request_id):
        try:
            session_request = SessionRequestCaas.objects.get(pk=session_request_id)
        except SessionRequestCaas.DoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND)

        try:
            invitee_emails = request.data.get("inviteeEmails", [])

            existing_calendar_invite = CalendarInvites.objects.filter(
                caas_session=session_request
            ).first()

            session_request.invitees = get_trimmed_emails(invitee_emails)
            session_request.save()
            try:
                attendees = []

                for invitee in session_request.invitees:
                    attendee = {
                        "emailAddress": {
                            "address": invitee,
                        },
                        "type": "required",
                    }
                    attendees.append(attendee)
                if session_request.coach:
                    attendees.append(
                        {
                            "emailAddress": {
                                "address": session_request.coach.email,
                            },
                            "type": "required",
                        }
                    )
                if session_request.learner:
                    attendees.append(
                        {
                            "emailAddress": {
                                "address": session_request.learner.email,
                            },
                            "type": "required",
                        }
                    )
                if session_request.pmo:
                    attendees.append(
                        {
                            "emailAddress": {
                                "address": session_request.pmo.email,
                            },
                            "type": "required",
                        }
                    )

                update_outlook_attendees(existing_calendar_invite, attendees)

            except Exception as e:
                print(str(e))

            return Response(
                {"message": "Invitees Updated Sucessfully"}, status=status.HTTP_200_OK
            )
        except Exception as e:
            return Response(
                data={"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


# @api_view(["GET"])
# def coach_session_list(request, coach_id):
#     try:
#         coach = Coach.objects.get(id=coach_id)
#     except Coach.DoesNotExist:
#         return Response({"error": "Coach not found"}, status=status.HTTP_404_NOT_FOUND)

#     sessions = SessionRequestCaas.objects.filter(coach=coach)
#     serializer = SessionRequestCaasSerializer(sessions, many=True)

#     return Response(serializer.data, status=status.HTTP_200_OK)


# @api_view(["GET"])
# def coach_session_list(request, coach_id):
#     try:
#         coach = Coach.objects.get(id=coach_id)
#     except Coach.DoesNotExist:
#         return Response({"error": "Coach not found"}, status=status.HTTP_404_NOT_FOUND)

#     sessions = SessionRequestCaas.objects.filter(coach=coach)
#     serializer_data = []

#     for session in sessions:
#         print("4575", session)
#         session_data = {
#             "session_id": session.id,
#             "session_status": session.status,
#             "project_id": session.project.id,
#             "project_name": session.project.name,
#             "organisation_name": session.project.organisation.name,
#             # Include other relevant project details here
#         }
#         serializer_data.append(session_data)

#     return Response(serializer_data, status=status.HTTP_200_OK)


@api_view(["DELETE"])
@permission_classes([IsAuthenticated, IsInRoles("pmo")])
def remove_coach_from_project(request, project_id):
    try:
        project = Project.objects.get(id=project_id)
        coach_id = request.data.get("coachIdToDelete")
        coach = Coach.objects.get(id=coach_id)
        currentuser_id = request.data.get("currentUserId")
        currentuser = User.objects.get(id=currentuser_id)

    except Project.DoesNotExist:
        return Response(
            {"message": "Project not found"}, status=status.HTTP_404_NOT_FOUND
        )
    except Coach.DoesNotExist:
        return Response(
            {"message": "Coach not found"}, status=status.HTTP_404_NOT_FOUND
        )

    engagements_with_coach = Engagement.objects.filter(project=project, coach=coach)

    for engagement in engagements_with_coach:
        sessions = SessionRequestCaas.objects.filter(
            project=project,
            learner=engagement.learner,
        )
        for session in sessions:
            session.status = "pending"
            session.hr = None
            session.pmo = None
            session.coach = None
            session.invitees = []
            session.availibility.clear()
            session.confirmed_availability = None
            session.is_booked = False
            session.reschedule_request = []
            session.is_archive = False
            session.save()

        engagement.coach = None
        engagement.save()

    for coach_status in project.coaches_status.all():
        if coach_status.coach.id == coach_id:
            project.coaches_status.remove(coach_status)
            project.coaches.remove(coach)
            project.save()
            coach_status.delete()

            user = "default user"
            time_of_removal = timezone.now()
            removed_coach = coach
            removed_from_project = project

            removeCoachProfile = RemoveCoachActivity.objects.create(
                user=currentuser,
                time_of_removal=time_of_removal,
                removed_coach=removed_coach,
                removed_from_project=removed_from_project,
            )

            removeCoachProfile.save()

            return Response(
                {"message": "Coach has been removed from the project."},
                status=status.HTTP_201_CREATED,
            )

    return Response(
        {"message": "Coach is not associated with the project"},
        status=status.HTTP_400_BAD_REQUEST,
    )

    # if request.method == "POST":


@api_view(["POST"])
@permission_classes(
    [IsAuthenticated, IsInRoles("pmo", "coach", "facilitator", "learner")]
)
def standard_field_request(request, user_id):
    try:
        with transaction.atomic():
            value = request.data.get("value").strip()
            userType = request.data.get("userType")

            field_name = request.data.get(
                "field_name"
            )  # Adjust this based on your field name
            standardized_field, created = StandardizedField.objects.get_or_create(
                field=field_name
            )

            if value not in standardized_field.values:
                standardized_field.values.append(value)
                standardized_field.save()
            else:
                return Response({"error": "Value already present."}, status=404)

            if userType == "coach":
                user_instance = Coach.objects.get(id=user_id)
                standardized_field_request = StandardizedFieldRequest(
                    standardized_field_name=standardized_field,
                    coach=user_instance,
                    value=value,
                    status="pending",
                )
                standardized_field_request.save()

            elif userType == "facilitator":
                user_instance = Facilitator.objects.get(id=user_id)
                standardized_field_request = StandardizedFieldRequest(
                    standardized_field_name=standardized_field,
                    facilitator=user_instance,
                    value=value,
                    status="pending",
                )
                standardized_field_request.save()
            elif userType == "learner":
                user_instance = Learner.objects.get(id=user_id)
                standardized_field_request = StandardizedFieldRequest(
                    standardized_field_name=standardized_field,
                    learner=user_instance,
                    value=value,
                    status="pending",
                )
                standardized_field_request.save()
            return Response({"message": "Request sent."}, status=200)
    except Exception as e:
        print(str(e))
        return Response({"error": "Failed to create request."}, status=500)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("pmo")])
def coaches_which_are_included_in_projects(request):
    coachesId = []
    projects = Project.objects.all()
    schedular_projects = SchedularProject.objects.all()
    for project in projects:
        for coach_status in project.coaches_status.all():
            if (
                coach_status.status["hr"]["status"] == "select"
                and coach_status.status["project_structure"]["status"] == "select"
            ):
                coachesId.append(coach_status.coach.id)
    for project in schedular_projects:
        batches = SchedularBatch.objects.filter(project=project)
        for batch in batches:
            coaches = batch.coaches.all()
            for coach in coaches:
                coachesId.append(coach.id)
    coachesId = set(coachesId)
    return Response(coachesId)


def calculate_session_progress_data_for_hr(user_id):
    session_requests = (
        SessionRequestCaas.objects.filter(project__hr__id=user_id)
        .exclude(
            Q(session_type="chemistry", billable_session_number__isnull=True)
            | Q(session_type="interview")
        )
        .annotate(
            completed_sessions_count=Count(
                Case(When(status="completed", then=1), output_field=IntegerField())
            ),
            total_sessions_count=Count("pk"),
            billable_count=Count(
                Case(
                    When(billable_session_number__isnull=False, then=1),
                    output_field=IntegerField(),
                )
            ),
        )
        .prefetch_related("project", "learner")
    )

    session_data = []
    for session_request in session_requests:
        session_data.append(
            {
                "session_type": session_request.session_type,
                "project_data": ProjectDepthTwoSerializer(session_request.project).data,
                "learner": LearnerDepthOneSerializer(session_request.learner).data,
                "billable": session_request.billable_count > 0,
                "duration": session_request.session_duration,
                "completed_sessions": session_request.completed_sessions_count > 0,
            }
        )
    return session_data


def calculate_and_send_session_data(user_id):
    try:
        filtered_sessions = calculate_session_progress_data_for_hr(user_id)

        session_type_data = defaultdict(
            lambda: defaultdict(
                lambda: {
                    "session_type": None,
                    "billable": "No",
                    "duration": None,
                    "total_sessions": 0,
                    "completed_sessions": 0,
                }
            )
        )

        for session in filtered_sessions:
            project_data = session.get("project_data")
            project_name = project_data.get("name")
            session_type = session.get("session_type")

            session_type_data[project_name][session_type]["session_type"] = session_type
            session_type_data[project_name][session_type]["billable"] = (
                "Yes" if session.get("billable") else "No"
            )
            session_type_data[project_name][session_type]["duration"] = session.get(
                "duration"
            )

            session_type_data[project_name][session_type]["total_sessions"] += 1
            session_type_data[project_name][session_type][
                "completed_sessions"
            ] += session.get("completed_sessions", 0)

        calculated_data = {
            project_name: {
                session_type: {
                    **session,
                    "completion_percentage": (
                        format(
                            (session["completed_sessions"] / session["total_sessions"])
                            * 100,
                            ".1f",
                        )
                        if session["total_sessions"] > 0
                        else 0
                    ),
                }
                for session_type, session in project_data.items()
            }
            for project_name, project_data in session_type_data.items()
        }

        hr = HR.objects.get(id=user_id)

        email_message = render_to_string(
            "hr_emails/progress_data_excel_send.html",
            {"name": hr.first_name},
        )
        email = EmailMessage(
            f"{env('EMAIL_SUBJECT_INITIAL',default='')}{'Progress Data'}",
            email_message,
            settings.DEFAULT_FROM_EMAIL,
            [hr.email],
            bcc=[],
        )

        excel_buffer = io.BytesIO()

        wb = Workbook()

        for project_name, project_data in calculated_data.items():
            ws = wb.create_sheet(title=project_name)

            df = pd.DataFrame.from_dict(project_data, orient="index")
            df.columns = [col.replace("_", " ").capitalize() for col in df.columns]
            # Add column names as the first row in the worksheet
            for c_idx, col_name in enumerate(df.columns, start=1):
                if col_name == "Duration":
                    col_name = "Duration (Min)"
                cell = ws.cell(row=1, column=c_idx, value=col_name)
                cell.font = Font(bold=True)

            # Iterate over DataFrame rows and add data to the worksheet
            for r_idx, row in enumerate(df.iterrows(), start=2):
                for c_idx, value in enumerate(row[1], start=1):
                    if isinstance(value, str):
                        value = str(value).replace("_", " ").capitalize()
                    if value == "Stakeholder without coach":
                        value = "Tripartate without coach"
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)

        default_sheet = wb["Sheet"]
        wb.remove(default_sheet)

        wb.save(excel_buffer)

        email.attach(
            "progress_data.xlsx",
            excel_buffer.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        email.send(fail_silently=False)

    except Exception as e:
        print(f"Error occurred while sending email with attachment: {str(e)}")


class SessionsProgressOfAllCoacheeForAnHr(APIView):
    permission_classes = [IsAuthenticated, IsInRoles("hr")]

    def get(self, request, user_id, format=None):
        session_data = calculate_session_progress_data_for_hr(user_id)

        return Response({"session_data": session_data})


class AddRegisteredCoach(APIView):
    permission_classes = [AllowAny]

    def post(self, request, *args, **kwargs):
        # Get data from request

        first_name = request.data.get("first_name")
        last_name = request.data.get("last_name")
        email = request.data.get("email")
        phone = request.data.get("phone")
        is_approved = request.data.get("is_approved")
        phone_country_code = request.data.get("phone_country_code")

        if not all([first_name, last_name, email, phone, phone_country_code]):
            return Response(
                {"error": "All required fields must be provided."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            # Create the Django User
            if Coach.objects.filter(email=email).exists():
                return Response(
                    {"error": "User with this email already exists."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            with transaction.atomic():
                # Check if the user already exists
                user = User.objects.filter(email=email).first()
                if not user:
                    temp_password = "".join(
                        random.choices(
                            string.ascii_uppercase
                            + string.ascii_lowercase
                            + string.digits,
                            k=8,
                        )
                    )
                    user = User.objects.create_user(
                        username=email, password=temp_password, email=email
                    )
                    profile = Profile.objects.create(user=user)
                    print("createing new user and profile.")
                else:
                    print("hello.")
                    profile = Profile.objects.get(user=user)

                coach_role, created = Role.objects.get_or_create(name="coach")
                profile.roles.add(coach_role)
                profile.save()

                # Create the Coach User using the Profile
                coach_user = Coach.objects.create(
                    user=profile,
                    first_name=first_name.strip().title(),
                    last_name=last_name.strip().title(),
                    email=email,
                    phone=phone,
                    phone_country_code=phone_country_code,
                    is_approved=is_approved,
                    active_inactive=True,
                )

                # Approve coach
                coach = Coach.objects.get(id=coach_user.id)
                # Change the is_approved field to True
                coach.is_approved = False
                coach.save()
                name = coach_user.first_name + " " + coach_user.last_name
                add_contact_in_wati("coach", name, coach_user.phone)
                coach_serializer = CoachSerializer(coach)

                path = f"/profile"
                message = f"Welcome to Meeraq. As next step, you need to fill out your details. Admin will look into your profile and contact you for profile approval. Thank You!"

                create_notification(coach.user.user, path, message)
                pmo_user = User.objects.filter(profile__roles__name="pmo").first()
                pmo = Pmo.objects.get(email=pmo_user.username)
                create_notification(
                    pmo_user,
                    f"/registeredcoach",
                    f"{coach.first_name} {coach.last_name} has registered as a coach. Please go through his Profile.",
                )
                send_mail_templates(
                    "pmo_emails/coach_register.html",
                    [pmo_user.username],
                    f"{coach.first_name} {coach.last_name} has Registered as a Coach",
                    {
                        "name": pmo.name,
                        "coachName": f"{coach.first_name} {coach.last_name} ",
                    },
                    json.loads(env("BCC_EMAIL_RAJAT_SUJATA")),
                )
                # Send profile completion tips to the coach
                send_mail_templates(
                    "coach_templates/profile_creation_tips.html",
                    [coach.email],
                    "Profile Completion Tips for Success on Meeraq Platform",
                    {
                        "name": f"{coach.first_name} {coach.last_name}",
                    },
                    [],
                )

            return Response({"coach": coach_serializer.data})

        except IntegrityError as e:
            return Response(
                {"error": "A user with this email already exists."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        except Exception as e:
            # Return error response if any other exception occurs
            print(e)
            return Response(
                {"error": "An error occurred while registering."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class AddRegisteredFacilitator(APIView):
    permission_classes = [AllowAny]

    def post(self, request, *args, **kwargs):
        # Get data from request

        first_name = request.data.get("first_name")
        last_name = request.data.get("last_name")
        email = request.data.get("email")
        phone = request.data.get("phone")
        is_approved = request.data.get("is_approved")
        phone_country_code = request.data.get("phone_country_code")

        if not all([first_name, last_name, email, phone, phone_country_code]):
            return Response(
                {"error": "All required fields must be provided."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            # Create the Django User
            if Facilitator.objects.filter(email=email).exists():
                return Response(
                    {"error": "User with this email already exists."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            with transaction.atomic():
                # Check if the user already exists
                user = User.objects.filter(email=email).first()
                if not user:
                    temp_password = "".join(
                        random.choices(
                            string.ascii_uppercase
                            + string.ascii_lowercase
                            + string.digits,
                            k=8,
                        )
                    )
                    user = User.objects.create_user(
                        username=email, password=temp_password, email=email
                    )
                    profile = Profile.objects.create(user=user)
                else:
                    profile = Profile.objects.get(user=user)

                facilitator_role, created = Role.objects.get_or_create(
                    name="facilitator"
                )
                profile.roles.add(facilitator_role)
                profile.save()
                # Create the Facilitator User using the Profile
                facilitator_user = Facilitator.objects.create(
                    user=profile,
                    first_name=first_name.strip().title(),
                    last_name=last_name.strip().title(),
                    email=email,
                    phone=phone,
                    phone_country_code=phone_country_code,
                    is_approved=is_approved,
                )
                # Approve facilitator
                facilitator = Facilitator.objects.get(id=facilitator_user.id)
                # Change the is_approved field to True
                facilitator.is_approved = False
                facilitator.save()
                name = facilitator_user.first_name + " " + facilitator_user.last_name
                add_contact_in_wati("facilitator", name, facilitator_user.phone)
                facilitator_serializer = FacilitatorSerializer(facilitator)

                path = f"/profile"
                message = f"Welcome to Meeraq. As next step, you need to fill out your details. Admin will look into your profile and contact you for profile approval. Thank You!"

                create_notification(facilitator.user.user, path, message)
                pmo_user = User.objects.filter(profile__roles__name="pmo").first()
                pmo = Pmo.objects.get(email=pmo_user.username)
                create_notification(
                    pmo_user,
                    f"/registeredcoach",
                    f"{facilitator.first_name} {facilitator.last_name} has registered as a Facilitator. Please go through his Profile.",
                )
                send_mail_templates(
                    "pmo_emails/facilitator_register.html",
                    [pmo_user.username],
                    f"{facilitator.first_name} {facilitator.last_name} has Registered as a Facilitator",
                    {
                        "name": pmo.name,
                        "facilitatorName": f"{facilitator.first_name} {facilitator.last_name} ",
                    },
                    json.loads(env("BCC_EMAIL_RAJAT_SUJATA")),
                )
                # Send profile completion tips to the coach
                send_mail_templates(
                    "facilitator_templates/profile_creation_tips_facilitator.html",
                    [facilitator.email],
                    "Profile Completion Tips for Success on Meeraq Platform",
                    {
                        "name": f"{facilitator.first_name} {facilitator.last_name}",
                    },
                    [],
                )

            return Response({"coach": facilitator_serializer.data})

        except Exception as e:
            # Return error response if any other exception occurs
            print(e)
            return Response(
                {"error": "An error occurred while registering."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("pmo")])
def get_registered_coaches(request):
    try:
        # Filter coaches where isapproved is False
        unapproved_coaches = Coach.objects.filter(is_approved=False)

        # Serialize the unapproved coaches
        serializer = CoachSerializer(unapproved_coaches, many=True)

        # Return the serialized unapproved coaches as the response
        return Response(serializer.data, status=200)

    except Exception as e:
        # Return error response if any exception occurs
        return Response({"error": str(e)}, status=500)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("pmo")])
def get_registered_facilitators(request):
    try:
        # Filter facilitators where isapproved is False
        unapproved_facilitators = Facilitator.objects.filter(is_approved=False)

        # Serialize the unapproved facilitators
        serializer = FacilitatorSerializer(unapproved_facilitators, many=True)

        # Return the serialized unapproved facilitators as the response
        return Response(serializer.data, status=200)

    except Exception as e:
        # Return error response if any exception occurs
        return Response({"error": str(e)}, status=500)


# @api_view(["GET"])
# def pmo_dashboard(request):
#     try:
#         organisations = Organisation.objects.all()
#         serialized_data = []

#         for organisation in organisations:
#             projects = Project.objects.filter(organisation=organisation)
#             organisation_data = OrganisationSerializer(organisation).data
#             organisation_data["projects"] = []

#             for project in projects:
#                 project_data = ProjectSerializer(project).data

#                 # Get all engagements in the project
#                 engagements = Engagement.objects.filter(project=project)
#                 engagement_data = EngagementSerializer(engagements, many=True).data

#                 project_data["engagements"] = engagement_data
#                 organisation_data["projects"].append(project_data)

#             serialized_data.append(organisation_data)

#         return Response(serialized_data, status=200)
#     except Exception as e:
#         return Response({"error": str(e)}, status=500)


# @api_view(["GET"])
# def get_all_engagements(request):
#     engagements = Engagement.objects.all()
#     serializer = EngagementSerializer(engagements, many=True)
#     return Response(serializer.data)


# @api_view(["GET"])
# def get_all_completed_sessions(request):
#     completed_sessions = SessionRequestCaas.objects.filter(status="completed")
#     serializer = SessionRequestCaasSerializer(completed_sessions, many=True)
#     return Response(serializer.data)


# @api_view(["GET"])
# def get_all_engagements(request):
#     # Get all engagements
#     engagements = Engagement.objects.all()

#     # Create a list to store engagement data with session counts
#     engagement_data_list = []

#     for engagement in engagements:
#         # Get the engagement ID
#         engagement_id = engagement.id

#         # Count completed sessions for the learner in this engagement
#         completed_sessions_count = SessionRequestCaas.objects.filter(
#             status="completed",
#             billable_session_number__isnull=False,
#             learner__id=engagement.learner.id,
#             project__id=engagement.project.id,
#             is_archive=False,
#         ).count()

#         # Count total sessions for the learner in this engagement
#         total_sessions_count = SessionRequestCaas.objects.filter(
#             learner__id=engagement.learner.id,
#             billable_session_number__isnull=False,
#             project__id=engagement.project.id,
#             is_archive=False,
#         ).count()

#         # Create a dictionary with session counts for this engagement
#         engagement_data = {
#             "completed_sessions_count": completed_sessions_count,
#             "total_sessions_count": total_sessions_count,
#         }

#         # Add engagement data to the list
#         engagement_data_list.append(engagement_data)

#     # Serialize the engagements along with session counts
#     serialized_engagements = [
#         {"engagement": engagement, "session_counts": data}
#         for engagement, data in zip(engagements, engagement_data_list)
#     ]

#     # Serialize and return the data
#     serializer = EngagementSerializer(serialized_engagements, many=True)
#     return Response(serializer.data)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("pmo")])
def get_all_engagements(request):
    # Get all engagements
    engagements = Engagement.objects.all()

    # Create a list to store serialized engagement data with session counts
    engagement_data_list = []

    for engagement in engagements:
        # Get the engagement ID
        engagement_id = engagement.id

        completed_sessions_count = (
            SessionRequestCaas.objects.filter(
                status="completed",
                learner__id=engagement.learner.id,
                project__id=engagement.project.id,
                is_archive=False,
            )
            .exclude(Q(billable_session_number__isnull=True, session_type="chemistry"))
            .count()
        )
        # Count total sessions for the learner in this engagement
        total_sessions_count = SessionRequestCaas.objects.filter(
            learner__id=engagement.learner.id,
            billable_session_number__isnull=False,
            project__id=engagement.project.id,
            is_archive=False,
        ).count()

        completed_sessions = SessionRequestCaas.objects.filter(
            project__id=engagement.project.id,
            learner=engagement.learner,
            status="completed",
            is_archive=False,
        ).exclude(billable_session_number__isnull=True, session_type="chemistry")

        sessions_data = [
            {
                "start_time": datetime.utcfromtimestamp(
                    int(session.confirmed_availability.start_time) / 1000
                ),
                "end_time": datetime.utcfromtimestamp(
                    int(session.confirmed_availability.end_time) / 1000
                ),
            }
            for session in completed_sessions
        ]
        print(sessions_data)

        # Sort the availabilities_data based on start time in descending order
        sorted_availabilities = sorted(
            sessions_data,
            key=lambda availability: availability["start_time"],
            reverse=True,
        )

        # Extract the date of the most recent completed session
        last_session_date = (
            sorted_availabilities[0]["start_time"].date()
            if sorted_availabilities
            else None
        )

        # Serialize the engagement along with session counts
        serialized_engagement = EngagementSerializer(engagement).data
        serialized_engagement["completed_sessions_count"] = completed_sessions_count
        serialized_engagement["total_sessions_count"] = total_sessions_count
        serialized_engagement["last_completed_session_date"] = last_session_date

        # Add serialized engagement data to the list
        engagement_data_list.append(serialized_engagement)

    # Return the list of serialized engagement data with session counts
    # print("engagement_data_list--------------------------------",engagement_data_list)
    return Response(engagement_data_list)


@api_view(["PUT"])
@permission_classes([IsAuthenticated, IsInRoles("pmo")])
def edit_project_caas(request, project_id):
    organisation = Organisation.objects.filter(
        id=request.data["organisation_id"]
    ).first()
    junior_pmo = None
    if "junior_pmo" in request.data:
        junior_pmo = Pmo.objects.filter(id=request.data["junior_pmo"]).first()

    try:
        # Retrieve the existing project from the database
        project = get_object_or_404(Project, pk=project_id)
        # Update project attributes based on the data in the PUT request
        project.name = request.data.get("project_name", project.name)
        project.approx_coachee = request.data.get(
            "approx_coachee", project.approx_coachee
        )
        project.organisation = organisation
        project.frequency_of_session = request.data.get(
            "frequency_of_session", project.frequency_of_session
        )
        project.interview_allowed = request.data.get(
            "interview_allowed", project.interview_allowed
        )
        project.specific_coach = request.data.get(
            "specific_coach", project.specific_coach
        )
        project.empanelment = request.data.get("empanelment", project.empanelment)
        project.tentative_start_date = request.data.get(
            "tentative_start_date", project.tentative_start_date
        )
        project.mode = request.data.get("mode", project.mode)
        project.sold = request.data.get("sold", project.sold)
        project.location = json.loads(request.data.get("location", "[]"))
        project.project_description = request.data.get(
            "project_description", project.project_description
        )
        project.coach_consent_mandatory = request.data.get(
            "coach_consent_mandatory", project.coach_consent_mandatory
        )
        project.enable_emails_to_hr_and_coachee = request.data.get(
            "enable_emails_to_hr_and_coachee", project.enable_emails_to_hr_and_coachee
        )

        project.masked_coach_profile = request.data.get(
            "masked_coach_profile", project.masked_coach_profile
        )

        project.email_reminder = request.data.get(
            "email_reminder", project.email_reminder
        )
        project.whatsapp_reminder = request.data.get(
            "whatsapp_reminder", project.whatsapp_reminder
        )
        project.calendar_invites = request.data.get(
            "calendar_invites", project.calendar_invites
        )

        if project.project_type == "COD":
            project.is_project_structure = request.data.get(
                "is_project_structure", project.is_project_structure
            )
            total_credits_in_hours = int(request.data.get("total_credits"))
            total_credits_in_minutes = total_credits_in_hours * 60
            project.total_credits = total_credits_in_minutes
            if not project.is_project_structure:
                project.duration_of_each_session = request.data.get(
                    "duration_of_each_session", project.duration_of_each_session
                )
            if not project.is_project_structure and request.data["is_session_expiry"]:
                project.is_session_expiry = request.data.get(
                    "is_session_expiry", project.is_session_expiry
                )
                request_expiry_time_in_hours = int(request.data["request_expiry_time"])
                request_expiry_time_in_minutes = request_expiry_time_in_hours * 60
                project.request_expiry_time = request_expiry_time_in_minutes
            if project.total_credits != request.data.get("total_credits"):
                project.credit_history.append(request.data.get("total_credits"))

        project.finance = request.data.get("finance", project.finance)
        project.junior_pmo = junior_pmo

        project.hr.clear()
        for hr in request.data["hr"]:
            single_hr = HR.objects.get(id=hr)
            project.hr.add(single_hr)

        # Save the updated project

        project.save()

        # You can return a success response with the updated project details
        return Response(
            {"message": "Project updated successfully", "project_id": project.id}
        )

    except Project.DoesNotExist:
        return Response({"error": "Project not found"}, status=404)

    except Exception as e:
        return Response({"error": str(e)}, status=500)


@api_view(["PUT"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "coach")])
def project_status(request, project_id):
    try:
        # Use get_object_or_404 to retrieve the project or return a 404 response if it doesn't exist
        project = get_object_or_404(Project, id=project_id)
        # Retrieve engagements related to the project
        engagements = Engagement.objects.filter(project=project)
        # Extract the desired status from the request data
        new_status = request.data.get("status")

        if not new_status:
            return Response({"error": "Status cannot be empty."}, status=400)

        # Check if the new status is "completed" and if any engagement is not completed
        if new_status == "completed" and any(
            engagement.status != "completed" for engagement in engagements
        ):
            return Response(
                {
                    "error": "Cannot set project status to 'completed' if there are active engagements."
                },
                status=400,
            )

        # Update the project status
        project.status = new_status
        # Save the updated project status
        project.save()
        # Return a success response
        return Response(
            {"message": f"Project status updated successfully to {project.status}"},
            status=200,
        )

    except Exception as e:
        # Return a 500 response for other exceptions
        return Response({"error": str(e)}, status=500)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("coach", "pmo")])
def completed_projects(request, user_id):
    projects = Project.objects.filter(
        coaches_status__coach__id=user_id, coaches_status__is_consent_asked=True
    )
    completed_project = []
    for project in projects:
        if project.status == "completed":
            completed_project.append(project)
    project_serializer = ProjectDepthTwoSerializer(completed_project, many=True)
    return Response({"completed_project": project_serializer.data})


class ActivitySummary(APIView):
    permission_classes = [IsAuthenticated, IsInRoles("pmo")]

    def get(self, request):
        try:
            user_login_activities = UserLoginActivity.objects.all()
            total_login_count = user_login_activities.count()
            user_login_serializer = UserLoginActivitySerializer(
                user_login_activities, many=True
            )
        except Exception as e:
            print("user_login_activities", str(e))
            user_login_serializer = []

        try:
            profile_edit_activities = ProfileEditActivity.objects.all()
            total_profile_edit_count = profile_edit_activities.count()
            profile_edit_serializer = ProfileEditActivitySerializer(
                profile_edit_activities, many=True
            )
        except Exception as e:
            print("profile_edit_activities", str(e))
            profile_edit_serializer = []

        try:
            goal_add_activities = AddGoalActivity.objects.all()
            total_goal_add_count = goal_add_activities.count()
            goal_add_serializer = AddGoalActivitySerializer(
                goal_add_activities, many=True
            )
        except Exception as e:
            print("goal_add_activities", str(e))
            goal_add_serializer = []

        try:
            coach_add_activities = AddCoachActivity.objects.all()
            total_coach_add_count = coach_add_activities.count()
            coach_add_serializer = AddCoachActivitySerializer(
                coach_add_activities, many=True
            )
        except Exception as e:
            print("coach_add_activities", str(e))
            coach_add_serializer = []

        try:
            sent_email_activities = SentEmailActivity.objects.all()
            total_sent_email_count = sent_email_activities.count()
            sent_email_serializer = SentEmailActivitySerializer(
                sent_email_activities, many=True
            )
        except Exception as e:
            print("sent_email_activities", str(e))
            sent_email_serializer = []

        try:
            session_requested_activities = SessionRequestedActivity.objects.all()
            total_session_requested_count = session_requested_activities.count()
            session_requested_serializer = SessionRequestedActivitySerializer(
                session_requested_activities, many=True
            )
        except Exception as e:
            print("session_requested_activities", str(e))
            session_requested_serializer = []

        try:
            delete_coach_profile_activities = DeleteCoachProfileActivity.objects.all()
            total_coach_profile_delete_count = delete_coach_profile_activities.count()
            delete_coach_profile_serializer = DeleteCoachProfileActivitySerializer(
                delete_coach_profile_activities, many=True
            )
        except Exception as e:
            print("delete_coach_activities", str(e))
            delete_coach_profile_serializer = []

        try:
            remove_coach_profile_activities = RemoveCoachActivity.objects.all()
            total_remove_coach_count = remove_coach_profile_activities.count()
            remove_coach_profile_serializer = RemoveCoachActivitySerializer(
                remove_coach_profile_activities, many=True
            )
        except Exception as e:
            print("remove_coach_profile_activities", str(e))
            remove_coach_profile_serializer = []

        try:
            add_past_session_activities = PastSessionActivity.objects.all()
            total_past_session_count = add_past_session_activities.count()
            past_session_activity_serializer = PastSessionActivitySerializer(
                add_past_session_activities, many=True
            )
        except Exception as e:
            print("add_past_session_activities", str(e))
            past_session_activity_serializer = []

        try:
            share_coach_profile_activities = ShareCoachProfileActivity.objects.all()
            total_share_coach_profile_count = share_coach_profile_activities.count()
            share_coach_profile_serializer = ShareCoachProfileActivitySerializer(
                share_coach_profile_activities, many=True
            )
        except Exception as e:
            print("share_coach_profile_activities", str(e))
            share_coach_profile_serializer = []

        try:
            create_project_activities = CreateProjectActivity.objects.all()
            total_create_project_count = create_project_activities.count()
            create_project_serializer = CreateProjectActivitySerializer(
                create_project_activities, many=True
            )
        except Exception as e:
            print("share_coach_profile_activities", str(e))
            create_project_serializer = []

        try:
            finalize_coach_activities = FinalizeCoachActivity.objects.all()
            total_finalized_coach_activity_count = finalize_coach_activities.count()
            finalize_coach_serializer = FinalizeCoachActivitySerializer(
                finalize_coach_activities, many=True
            )
        except Exception as e:
            print("share_coach_profile_activities", str(e))
            finalize_coach_serializer = []

        response_data = {
            "user_login": {
                "total_count": total_login_count,
                "activity": user_login_serializer.data,
            },
            "profile_edit": {
                "total_count": total_profile_edit_count,
                "activity": profile_edit_serializer.data,
            },
            "goal_add": {
                "total_count": total_goal_add_count,
                "activity": goal_add_serializer.data,
            },
            "coach_add": {
                "total_count": total_coach_add_count,
                "activity": coach_add_serializer.data,
            },
            "sent_email": {
                "total_count": total_sent_email_count,
                "activity": sent_email_serializer.data,
            },
            "session_requested": {
                "total_count": total_session_requested_count,
                "activity": session_requested_serializer.data,
            },
            "delete_coach_profile": {
                "total_count": total_coach_profile_delete_count,
                "activity": delete_coach_profile_serializer.data,
            },
            "remove_coach_profile": {
                "total_count": total_remove_coach_count,
                "activity": remove_coach_profile_serializer.data,
            },
            "add_past_session": {
                "total_count": total_past_session_count,
                "activity": past_session_activity_serializer.data,
            },
            "share_coach_profile": {
                "total_count": total_share_coach_profile_count,
                "activity": share_coach_profile_serializer.data,
            },
            "create_project": {
                "total_count": total_create_project_count,
                "activity": create_project_serializer.data,
            },
            "finalize_coach": {
                "total_count": total_finalized_coach_activity_count,
                "activity": finalize_coach_serializer.data,
            },
        }

        return Response(response_data)


@api_view(["POST"])
@permission_classes([AllowAny])
def send_reset_password_link(request):
    users = request.data["users"]
    for user_data in users:
        try:
            user = User.objects.get(email=user_data["email"])
            token = get_token_generator().generate_token()
            ResetPasswordToken.objects.create(user=user, key=token)
            reset_password_link = f"{env('ZOHO_APP_URL')}/reset-password/{token}"
            send_mail_templates(
                "vendors/vendor_welcome.html",
                [user_data["email"]],
                "Meeraq - Exciting News! New Vendor Platform Launch.",
                {"name": user_data["name"], "link": reset_password_link},
                [],
            )
        except Exception as e:
            print(f"Error sending link to {user_data['email']}: {str(e)}")
        sleep(5)
    return Response({"message": "Reset password links sent successfully"})


@api_view(["POST"])
@permission_classes(
    [IsAuthenticated, IsInRoles("pmo", "learner", "coach", "hr", "facilitator")]
)
def create_coach_profile_template(request):
    coach_id = request.data.get("coach")
    project_id = request.data.get("project")
    templates = request.data.get("templates")

    coaching_experience = templates.get("coaching_experience")
    corporate_experience = templates.get("corporate_experience")

    try:
        # Fetch the existing CoachProfileTemplate object based on coach and project IDs
        template = CoachProfileTemplate.objects.get(
            coach__id=coach_id, project__id=project_id
        )
    except CoachProfileTemplate.DoesNotExist:
        # If the object doesn't exist, create a new one
        template = CoachProfileTemplate(coach__id=coach_id, project__id=project_id)

    # Get the existing templates data
    existing_templates = template.templates

    if coaching_experience is not None:
        # Update the "coaching_experience" key with the new value
        existing_templates["coaching_experience"] = coaching_experience

    if corporate_experience is not None:
        # Update the "corporate_experience" key with the new value
        existing_templates["corporate_experience"] = corporate_experience

    # Update the "templates" field with the updated data
    template.templates = existing_templates
    template.save()

    return Response("Coach profile template updated successfully")


@api_view(["GET"])
@permission_classes(
    [IsAuthenticated, IsInRoles("pmo", "learner", "coach", "hr", "facilitator")]
)
def get_coach_profile_template(request, project_id):
    try:
        coach_profile_templates = CoachProfileTemplate.objects.filter(
            project__id=project_id
        )
        serializer = CoachProfileTemplateSerializer(coach_profile_templates, many=True)
        return Response(serializer.data)
    except CoachProfileTemplate.DoesNotExist:
        return Response(status=404)


class StandardizedFieldAPI(APIView):
    permission_classes = [
        IsAuthenticated,
        IsInRoles("coach", "facilitator", "pmo", "hr", "learner"),
    ]

    def get(self, request):
        standardized_fields = StandardizedField.objects.all()

        standardized_fields_serializer = StandardizedFieldSerializer(
            standardized_fields, many=True
        )

        field_data = {
            field_data["field"]: field_data["values"]
            for field_data in standardized_fields_serializer.data
        }

        return Response(field_data)


class StandardizedFieldRequestAPI(APIView):
    permission_classes = [IsAuthenticated, IsInRoles("coach", "facilitator", "pmo")]

    def get(self, request):
        today_start = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

        today_requests = StandardizedFieldRequest.objects.filter(
            requested_at__gte=today_start, status="pending"
        ).order_by("-requested_at")

        other_requests = StandardizedFieldRequest.objects.filter(
            Q(status="pending") & Q(requested_at__lt=today_start)
        ).order_by("-requested_at")

        today_requests_serializer = StandardizedFieldRequestDepthOneSerializer(
            today_requests, many=True
        )
        other_requests_serializer = StandardizedFieldRequestDepthOneSerializer(
            other_requests, many=True
        )

        return Response(
            {
                "today_requests": today_requests_serializer.data,
                "other_requests": other_requests_serializer.data,
            }
        )


class StandardFieldAddValue(APIView):
    permission_classes = [IsAuthenticated, IsInRoles("pmo")]

    def post(self, request):
        try:
            with transaction.atomic():
                # Extracting data from request body
                field_name = request.data.get("field_name")
                option_value = request.data.get("optionValue").strip()

                # Get or create the StandardizedField instance for the given field_name
                standardized_field, created = StandardizedField.objects.get_or_create(
                    field=field_name
                )

                # Check if the option_value already exists in the values list of the standardized_field
                if option_value not in standardized_field.values:
                    # Add the option_value to the values list and save the instance
                    standardized_field.values.append(option_value)
                    standardized_field.save()
                else:
                    # Return error response if the option_value already exists
                    return Response({"error": "Value already present."}, status=400)

                # Return success response
                return Response(
                    {
                        "message": f"Value Added to {FIELD_NAME_VALUES[field_name]} field."
                    },
                    status=200,
                )

        except Exception as e:
            print(str(e))
            # Return error response if any exception occurs
            return Response(
                {"error": "Failed to add value."},
                status=500,
            )


class StandardFieldEditValue(APIView):
    permission_classes = [IsAuthenticated, IsInRoles("pmo")]

    def put(self, request):
        try:
            with transaction.atomic():
                # Extracting data from request body
                field_name = request.data.get("field_name")
                previous_value = request.data.get("previous_value")
                new_value = request.data.get("new_value").strip()

                # Retrieve the StandardizedField instance corresponding to the provided field_name
                standardized_field = StandardizedField.objects.filter(
                    field=field_name
                ).first()

                # Check if the field exists
                if standardized_field:
                    # Check if the previous_value exists in the values list of the standardized_field
                    if previous_value in standardized_field.values:
                        # Update the value if it exists

                        index = standardized_field.values.index(previous_value)
                        standardized_field.values[index] = new_value
                        standardized_field.save()

                        # Return success response
                        return Response(
                            {
                                "message": f"Value Updated in {FIELD_NAME_VALUES[field_name]} field."
                            },
                            status=200,
                        )
                    else:
                        # Return error response if the previous_value does not exist
                        return Response(
                            {
                                "message": f"{previous_value} not found in {FIELD_NAME_VALUES[field_name]} field."
                            },
                            status=404,
                        )
                else:
                    # Return error response if the field does not exist
                    return Response(
                        {"message": f"{FIELD_NAME_VALUES[field_name]} not found."},
                        status=404,
                    )

        except Exception as e:
            print(str(e))
            # Return error response if any exception occurs
            return Response(
                {"error": "Failed to update."},
                status=500,
            )


models_to_update = {
    "Coach": [
        "location",
        "other_certification",
        "area_of_expertise",
        "job_roles",
        "companies_worked_in",
        "language",
        "education",
        "domain",
        "client_companies",
        "educational_qualification",
    ],
    "Facilitator": [
        "location",
        "other_certification",
        "area_of_expertise",
        "job_roles",
        "companies_worked_in",
        "language",
        "education",
        "domain",
        "client_companies",
        "educational_qualification",
        "city",
        "country",
        "topic",
    ],
    "Learner": [
        "job_roles",
    ],
}


class StandardizedFieldRequestAcceptReject(APIView):
    permission_classes = [IsAuthenticated, IsInRoles("pmo")]

    def put(self, request):
        status = request.data.get("status")
        request_id = request.data.get("request_id")

        try:
            with transaction.atomic():
                request_instance = StandardizedFieldRequest.objects.get(id=request_id)
                field_name = request_instance.standardized_field_name.field
                value = request_instance.value

                standardized_field, created = StandardizedField.objects.get_or_create(
                    field=field_name
                )
                if status == "accepted":
                    request_instance.status = status
                    request_instance.save()

                    # if value not in standardized_field.values:
                    #     standardized_field.values.append(value)
                    #     standardized_field.save()
                    # else:
                    #     return Response({"error": "Value already present."}, status=404)
                    return Response({"message": f"Request {status}"}, status=200)
                else:
                    request_instance.status = status
                    request_instance.save()

                    if value in standardized_field.values:
                        standardized_field.values.remove(value)
                        standardized_field.save()

                    for model_name, fields in models_to_update.items():
                        model_class = globals()[model_name]
                        instances = model_class.objects.all()

                        for instance in instances:
                            for field in fields:
                                field_value = getattr(instance, field, None)
                                if field_value is not None:
                                    if (
                                        isinstance(field_value, list)
                                        and value in field_value
                                    ):
                                        field_value.remove(value)
                                        instance.save()
                    if request_instance.coach:
                        send_mail_templates(
                            "coach_templates/reject_feild_item_request.html",
                            [request_instance.coach.email],
                            "Meeraq | Field Rejected",
                            {
                                "name": f"{request_instance.coach.first_name} {request_instance.coach.last_name}",
                                "value": value,
                                "feild": field_name.replace(" ", "_").title(),
                            },
                            [],
                        )
                    return Response({"message": f"Request {status}"}, status=200)
        except Exception as e:
            print(str(e))
            return Response({"error": f"Failed to perform operation."}, status=500)


class StandardFieldDeleteValue(APIView):
    permission_classes = [IsAuthenticated, IsInRoles("pmo")]

    def delete(self, request):
        try:
            with transaction.atomic():
                # Extracting data from request body
                field_name = request.data.get("field_name")
                option_value = request.data.get("optionValue")

                # Retrieve the StandardizedField instance for the given field_name
                standardized_field = StandardizedField.objects.get(field=field_name)

                # Check if the option_value exists in the values list of the standardized_field
                if option_value in standardized_field.values:
                    # Remove the option_value from the values list and save the instance
                    standardized_field.values.remove(option_value)
                    standardized_field.save()
                else:
                    # Return error response if the option_value does not exist
                    return Response({"error": "Value not present."}, status=404)

                # Return success response
                return Response(
                    {
                        "message": f"Value deleted from {FIELD_NAME_VALUES[field_name]} field."
                    },
                    status=200,
                )

        except StandardizedField.DoesNotExist:
            # Return error response if the StandardizedField instance does not exist
            return Response({"error": "Field not found."}, status=404)

        except Exception as e:
            print(str(e))
            # Return error response if any other exception occurs
            return Response(
                {"error": "Failed to delete value."},
                status=500,
            )


@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated, IsInRoles("pmo")])
def template_list_create_view(request):
    if request.method == "GET":
        templates = Template.objects.all()
        serializer = TemplateSerializer(templates, many=True)
        return Response(serializer.data)

    elif request.method == "POST":
        serializer = TemplateSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(["GET", "PUT", "DELETE"])
@permission_classes([IsAuthenticated, IsInRoles("pmo")])
def template_retrieve_update_destroy_view(request, pk):
    try:
        template = Template.objects.get(pk=pk)
    except Template.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)

    if request.method == "GET":
        serializer = TemplateSerializer(template)
        return Response(serializer.data)

    elif request.method == "PUT":
        serializer = TemplateSerializer(template, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    elif request.method == "DELETE":
        template.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "coach")])
def create_project_contract(request):
    serializer = ProjectContractSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response({"message": "Contract Assigned Successfully."})
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class ProjectContractAPIView(APIView):
    permission_classes = [IsAuthenticated, IsInRoles("coach", "pmo", "learner", "hr")]

    def get(self, request, format=None):
        contracts = ProjectContract.objects.all()
        serializer = ProjectContractSerializer(contracts, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


class ProjectContractDetailView(APIView):
    permission_classes = [IsAuthenticated, IsInRoles("coach", "pmo", "hr")]

    def get(self, request, project_id, format=None):
        project_type = request.query_params.get("project_type", None)
        try:
            if project_type == "skill":
                project_contract = ProjectContract.objects.get(
                    schedular_project=project_id
                )
            else:
                project_contract = ProjectContract.objects.get(project=project_id)

        except ProjectContract.DoesNotExist:
            return Response(
                {"error": "Project contract not found."},
                status=status.HTTP_404_NOT_FOUND,
            )

        serializer = ProjectContractSerializer(project_contract)
        return Response(serializer.data, status=status.HTTP_200_OK)


class CoachContractList(APIView):
    permission_classes = [IsAuthenticated, IsInRoles("coach", "pmo")]

    def get(self, request, format=None):
        contracts = CoachContract.objects.all()
        serializer = CoachContractSerializer(contracts, many=True)
        return Response(serializer.data)

    def post(self, request, format=None):
        serializer = CoachContractSerializer(data=request.data)
        if serializer.is_valid():
            coach = serializer.validated_data["coach"]
            project = serializer.validated_data["project"]

            if CoachContract.objects.filter(coach=coach, project=project).exists():
                return Response(
                    {
                        "message": f"Coach {coach.username} already has a contract for {project.name}"
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            serializer.save()
            return Response(
                {"message": "Contract created successfully", "data": serializer.data},
                status=status.HTTP_201_CREATED,
            )
        return Response(
            {"message": "Invalid data", "errors": serializer.errors},
            status=status.HTTP_400_BAD_REQUEST,
        )


class ProjectContractListWithDepth(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, format=None):
        try:
            contracts = ProjectContract.objects.all()
            data = []
            for contract in contracts:
                if contract.project:
                    project = contract.project
                else:
                    project = contract.schedular_project

                coach_contracts = CoachContract.objects.filter(
                    project_contract=contract
                )
                coaches_selected_count = coach_contracts.count()
                pending_contracts = []
                approved_contracts = []
                rejected_contracts = []
                for coach_contract in coach_contracts:
                    contract_object = {
                        "name": coach_contract.coach.first_name
                        + " "
                        + coach_contract.coach.last_name,
                        "email": coach_contract.coach.email,
                        "profile_pic": (
                            coach_contract.coach.profile_pic
                            if coach_contract.coach.profile_pic
                            else None
                        ),
                    }
                    if coach_contract.status == "pending":
                        pending_contracts.append(contract_object)
                    elif coach_contract.status == "approved":
                        approved_contracts.append(contract_object)
                    elif coach_contract.status == "rejected":
                        rejected_contracts.append(contract_object)

                temp = {
                    "id": contract.id,
                    "template_id": contract.template_id,
                    "title": contract.title,
                    "content": contract.content,
                    "project_id": project.id,
                    "project_name": project.name,
                    "organisation_name": project.organisation.name,
                    "organisation_image": (
                        project.organisation.image_url
                        if project.organisation.image_url
                        else None
                    ),
                    "created_at": contract.created_at,
                    "updated_at": contract.updated_at,
                    "reminder_timestamp": contract.reminder_timestamp,
                    "project_type": "caas" if contract.project else "skill",
                    "selected_coaches": coaches_selected_count,
                    "pending_contracts_status": (
                        "Pending"
                        if not len(pending_contracts) == 0
                        or coaches_selected_count == 0
                        else "Done"
                    ),
                    "pending_contracts": pending_contracts,
                    "approved_contracts": approved_contracts,
                    "rejected_contracts": rejected_contracts,
                    "pending_contracts_count": len(pending_contracts),
                    "approved_contracts_count": len(approved_contracts),
                    "rejected_contracts_count": len(rejected_contracts),
                }
                data.append(temp)
            return Response(data)
        except Exception as e:
            print(str(e))
            return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class CoachContractDetail(APIView):
    permission_classes = [IsAuthenticated]

    def get_object(self, pk):
        try:
            return CoachContract.objects.get(pk=pk)
        except CoachContract.DoesNotExist:
            return None

    def get(self, request, pk, format=None):
        contract = self.get_object(pk)
        if contract is not None:
            serializer = CoachContractSerializer(contract)
            return Response(serializer.data)
        return Response(status=status.HTTP_404_NOT_FOUND)

    def put(self, request, pk, format=None):
        contract = self.get_object(pk)
        if contract is not None:
            serializer = CoachContractSerializer(contract, data=request.data)
            if serializer.is_valid():
                serializer.save()
                return Response(
                    {
                        "message": "Contract updated successfully",
                        "data": serializer.data,
                    }
                )
            return Response(
                {"message": "Invalid data", "errors": serializer.errors},
                status=status.HTTP_400_BAD_REQUEST,
            )
        return Response(
            {"message": "Contract not found"}, status=status.HTTP_404_NOT_FOUND
        )

    def delete(self, request, pk, format=None):
        contract = self.get_object(pk)
        if contract is not None:
            contract.delete()
            return Response(
                {"message": "Contract deleted successfully"},
                status=status.HTTP_204_NO_CONTENT,
            )
        return Response(
            {"message": "Contract not found"}, status=status.HTTP_404_NOT_FOUND
        )


class GetCoachContractFromProject(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, project_id):
        try:

            contract = ProjectContract.objects.filter(
                schedular_project__id=project_id
            ).first()
            coach_contracts = CoachContract.objects.filter(project_contract=contract)
            data = []
            index = 1
            for coach_contract in coach_contracts:
                contract_object = {
                    "id": coach_contract.id,
                    "sno": index,
                    "name": coach_contract.coach.first_name
                    + " "
                    + coach_contract.coach.last_name,
                    "email": coach_contract.coach.email,
                    "profile_pic": (
                        coach_contract.coach.profile_pic
                        if coach_contract.coach.profile_pic
                        else None
                    ),
                    "status": coach_contract.status,
                    "send_date": coach_contract.send_date,
                    "response_date": coach_contract.response_date,
                    "contract_name": contract.title,
                }
                index += 1

                data.append(contract_object)
            return Response(data)
        except Exception as e:
            print(str(e))
            return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class UpdateCoachContract(APIView):
    permission_classes = [IsAuthenticated, IsInRoles("pmo", "coach")]

    def put(self, request, format=None):
        coach_id = request.data.get("coach")
        project_id = request.data.get("project")
        schedular_project_id = request.data.get("schedular_project")
        try:
            if schedular_project_id:
                contract = CoachContract.objects.get(
                    coach__id=coach_id, schedular_project__id=schedular_project_id
                )
            else:
                contract = CoachContract.objects.get(
                    coach__id=coach_id, project__id=project_id
                )
        except CoachContract.DoesNotExist:
            return Response(
                {"error": "Coach Contract not found."}, status=status.HTTP_404_NOT_FOUND
            )

        coach_name = (
            f"{contract.coach.first_name.strip()} {contract.coach.last_name.strip()}"
        )
        provided_name_inputed = request.data.get("name_inputed").strip()

        if provided_name_inputed.lower() != coach_name.lower():
            return Response(
                {"error": "Provided name input does not match coach's name."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        serializer = CoachContractSerializer(contract, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(
                {"message": "Contract Accepted Successfully", "data": serializer.data},
                status=status.HTTP_200_OK,
            )
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class AssignCoachContractAndProjectContract(APIView):
    permission_classes = [IsAuthenticated, IsInRoles("pmo")]

    def post(self, request, format=None):
        try:
            try:
                tasks = Task.objects.filter(
                    task="add_coach_contract", status="pending", caas_project=project_id
                )
                tasks.update(status="completed")
                # create task
                create_task(
                    {
                        "task": "add_coachee",
                        "project": project_id,
                        "priority": "medium",
                        "status": "pending",
                        "remarks": [],
                    },
                    7,
                )
            except Exception as e:
                print(str(e))
                pass
            with transaction.atomic():
                print(request.data)
                project_type = request.data.get("project_type")
                project_id = request.data.get("project")
                template_id = request.data.get("template_id")
                title = request.data.get("title")
                content = request.data.get("content")
                schedular_project_id = request.data.get("schedular_project")
                project = None
                schedular_project = None
                existing_contract = None

                if schedular_project_id:
                    schedular_project = SchedularProject.objects.get(
                        id=schedular_project_id
                    )
                    existing_contract = ProjectContract.objects.filter(
                        schedular_project=schedular_project
                    ).first()
                if project_id:
                    project = Project.objects.get(id=project_id)

                    existing_contract = ProjectContract.objects.filter(
                        project=project
                    ).first()

                if not existing_contract:
                    contract = ProjectContract.objects.create(
                        template_id=template_id,
                        title=title,
                        content=content,
                        project=project,
                        schedular_project=schedular_project,
                    )
                else:
                    contract = existing_contract

                if project_type == "caas":

                    current_date = timezone.now().date()
                    project_id = request.data.get("project")

                    try:
                        project = Project.objects.get(id=project_id)
                    except Project.DoesNotExist:
                        return Response(
                            {"message": "Project not found."},
                            status=status.HTTP_404_NOT_FOUND,
                        )

                    coaches = project.coaches_status.all()

                    for coach_status in coaches:
                        hr_status = coach_status.status.get("hr", {}).get("status")

                        if hr_status == "select":
                            coach = coach_status.coach

                            existing_coach_contract = CoachContract.objects.filter(
                                project=project, coach=coach.id
                            ).exists()

                            if not existing_coach_contract:
                                contract_data = {}
                                if not project.coach_consent_mandatory:
                                    contract_data = {
                                        "project_contract": contract.id,
                                        "project": project_id,
                                        "status": "approved",
                                        "coach": coach.id,
                                        "name_inputed": coach.first_name
                                        + " "
                                        + coach.last_name,
                                        "response_date": timezone.now().date(),
                                    }
                                else:
                                    contract_data = {
                                        "project_contract": contract.id,
                                        "project": project_id,
                                        "status": "pending",
                                        "coach": coach.id,
                                    }
                                contract_serializer = CoachContractSerializer(
                                    data=contract_data
                                )

                                if contract_serializer.is_valid():
                                    contract_serializer.save()
                                else:
                                    return Response(
                                        contract_serializer.errors,
                                        status=status.HTTP_400_BAD_REQUEST,
                                    )

                # update the tasks

                return Response(
                    {
                        "message": "Project Contract Saved and Coach contracts assigned successfully."
                    },
                    status=status.HTTP_201_CREATED,
                )
        except Exception as e:
            print(str(e))
            return Response(
                {"error": "Failed to assign contract"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class ApprovedCoachContract(APIView):
    permission_classes = [IsAuthenticated, IsInRoles("pmo", "coach")]

    def get(self, request, project_id, coach_id, format=None):
        try:
            project_type = request.query_params.get("project_type", None)
            if project_type == "skill":
                coach_contract = CoachContract.objects.get(
                    schedular_project__id=project_id,
                    coach__id=coach_id,
                    status="approved",
                )
            else:
                coach_contract = CoachContract.objects.get(
                    project__id=project_id, coach__id=coach_id, status="approved"
                )
        except CoachContract.DoesNotExist:
            return Response(
                {"error": "Coach contract not found1."},
                status=status.HTTP_404_NOT_FOUND,
            )

        serializer = CoachContractSerializer(coach_contract)
        return Response(serializer.data, status=status.HTTP_200_OK)


class CoachWithApprovedContractsInProject(APIView):
    permission_classes = [IsAuthenticated, IsInRoles("coach", "pmo", "learner")]

    def get(self, request, project_id, format=None):
        try:
            coach_contracts = CoachContract.objects.filter(
                project=project_id, status="approved"
            )

            coach_ids = [contract.coach.id for contract in coach_contracts]

            coaches = Coach.objects.filter(id__in=coach_ids)

            serializer = CoachSerializer(coaches, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)

        except CoachContract.DoesNotExist:
            return Response(
                {"error": "Coach contracts not found."},
                status=status.HTTP_404_NOT_FOUND,
            )


class SendContractReminder(APIView):
    permission_classes = [IsAuthenticated, IsInRoles("pmo")]

    def post(self, request, format=None):
        try:
            project_contract = ProjectContract.objects.get(
                id=request.data.get("contract_id")
            )
            timestamp = request.data.get("timestamp")
            coach_contracts = CoachContract.objects.filter(
                project_contract=project_contract
            )
            for coach_contract in coach_contracts:

                send_mail_templates(
                    "coach_templates/contract_reminder.html",
                    [coach_contract.coach.email],
                    "Meeraq Coaching | Coach Contract Reminder",
                    {"name": coach_contract.coach.first_name},
                    [],  # no bcc emails
                )
                sleep(5)
            notification_message = "This is a reminder to accept the Coach contract."
            create_notification(
                coach_contract.coach.user.user, "/projects", notification_message
            )

            project_contract.reminder_timestamp = timestamp
            project_contract.save()

            return Response(
                {"message": "Emails and notifications sent successfully"},
                status=status.HTTP_201_CREATED,
            )
        except Exception as e:
            print(str(e))
            return Response(
                {"error": "Failed to send reminders"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def change_user_role(request, user_id):
    user_role = request.data.get("user_role", "")
    user = User.objects.get(id=user_id)
    if not user.profile:
        print("hello")
        return Response({"error": "No user profile."}, status=400)
    elif user.profile.roles.count() == 0:
        print("second")
        return Response({"error": "No user role."}, status=400)
    try:
        user_profile_role = user.profile.roles.all().get(name=user_role).name
    except Exception as e:
        print(e)
        return Response({"error": "User role not found."}, status=400)
    roles = []
    for role in user.profile.roles.all():
        user_data = get_user_for_active_inactive(role.name, user.profile.user.username)
        if user_data and user_data.active_inactive:
            roles.append(role.name)
    if user_profile_role == "coach":
        if user.profile.coach.active_inactive or not user.profile.coach.is_approved:
            serializer = CoachDepthOneSerializer(user.profile.coach)
            is_caas_allowed = Project.objects.filter(
                coaches_status__coach=user.profile.coach
            ).exists()
            is_seeq_allowed = SchedularBatch.objects.filter(
                coaches=user.profile.coach
            ).exists()
            return Response(
                {
                    **serializer.data,
                    "is_caas_allowed": is_caas_allowed,
                    "is_seeq_allowed": is_seeq_allowed,
                    "roles": roles,
                    "user": {**serializer.data["user"], "type": user_profile_role},
                    "last_login": user.last_login,
                    "message": "Role changed to Coach",
                }
            )
        else:
            return None
    elif user_profile_role == "pmo":
        if not user.profile.pmo.active_inactive:
            return None
        serializer = PmoDepthOneSerializer(user.profile.pmo)
    elif user_profile_role == "superadmin":
        if not user.profile.superadmin.active_inactive:
            return None
        serializer = SuperAdminDepthOneSerializer(user.profile.superadmin)
    elif user_profile_role == "finance":
        if not user.profile.finance.active_inactive:
            return None
        serializer = FinanceDepthOneSerializer(user.profile.finance)
    elif user_profile_role == "facilitator":
        if not user.profile.facilitator.active_inactive:
            return None
        serializer = FacilitatorDepthOneSerializer(user.profile.facilitator)
    elif user_profile_role == "vendor":
        if not user.profile.vendor.active_inactive:
            return None
        serializer = VendorDepthOneSerializer(user.profile.vendor)
        organization = get_organization_data()
        zoho_vendor = get_vendor(serializer.data["vendor_id"])
        login_timestamp = timezone.now()
        UserLoginActivity.objects.create(
            user=user, timestamp=login_timestamp, platform="caas"
        )

        return Response(
            {
                **serializer.data,
                "roles": roles,
                "user": {
                    **serializer.data["user"],
                    "type": user_profile_role,
                    "last_login": user.last_login,
                },
                "last_login": user.last_login,
                "organization": organization,
                "zoho_vendor": zoho_vendor,
                "message": f"Role changed to billing",
            }
        )
    elif user_profile_role == "learner":
        if not user.profile.learner.active_inactive:
            return None
        serializer = LearnerDepthOneSerializer(user.profile.learner)
        is_caas_allowed = Engagement.objects.filter(
            learner=user.profile.learner
        ).exists()
        is_seeq_allowed = SchedularBatch.objects.filter(
            learners=user.profile.learner
        ).exists()

        return Response(
            {
                **serializer.data,
                "is_caas_allowed": is_caas_allowed,
                "is_seeq_allowed": is_seeq_allowed,
                "roles": roles,
                "user": {**serializer.data["user"], "type": user_profile_role},
                "last_login": user.last_login,
                "message": "Role changed to Learner",
            }
        )
    elif user_profile_role == "hr":
        if not user.profile.hr.active_inactive:
            return None
        serializer = HrDepthOneSerializer(user.profile.hr)
        is_caas_allowed = Project.objects.filter(hr=user.profile.hr).exists()
        is_seeq_allowed = SchedularProject.objects.filter(hr=user.profile.hr).exists()
        return Response(
            {
                **serializer.data,
                "roles": roles,
                "is_caas_allowed": is_caas_allowed,
                "is_seeq_allowed": is_seeq_allowed,
                "user": {**serializer.data["user"], "type": user_profile_role},
            }
        )
    elif user_profile_role == "sales":
        if not user.profile.sales.active_inactive:
            return None
        serializer = SalesDepthOneSerializer(user.profile.sales)
    elif user_profile_role == "ctt_pmo":
        if not user.profile.cttpmo.active_inactive:
            return None
        serializer = CTTPmoDepthOneSerializer(user.profile.cttpmo)
    elif user_profile_role == "leader":
        if not user.profile.leader.active_inactive:
            return None
        serializer = LeaderDepthOneSerializer(user.profile.leader)
    else:
        return Response({"error": "Unknown user role."}, status=400)
    return Response(
        {
            **serializer.data,
            "roles": roles,
            "user": {**serializer.data["user"], "type": user_profile_role},
            "last_login": user.last_login,
            "message": f"Role changed to {user_profile_role}",
        }
    )


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "superadmin")])
def get_users(request):
    user_profiles = Profile.objects.all()
    res = []
    for profile in user_profiles:
        active_roles = []
        inactive_roles = []
        # existing_roles = [item.name for item in profile.roles.all()]
        for role in profile.roles.all():
            user = get_user_for_active_inactive(role.name, profile.user.username)
            if user:
                if user.active_inactive:
                    active_roles.append(role.name)
                else:
                    inactive_roles.append(role.name)
        email = profile.user.email
        res.append(
            {
                "id": profile.id,
                "email": email,
                "roles": active_roles,
                "inactive_roles": inactive_roles,
            }
        )
    return Response(res)


def get_weeks_for_current_month():
    current_year = datetime.now().year
    current_month = datetime.now().month
    current_date = datetime.now()
    first_day_of_current_month = current_date.replace(day=1)
    cal = calendar.monthcalendar(current_year, current_month)
    weeks = []

    for week in cal:
        days_in_week = [day for day in week if day != 0]
        if days_in_week:
            start_day = min(days_in_week)
            end_day = max(days_in_week)

            # Check if Saturday is the last day of the week
            if (
                calendar.weekday(current_year, current_month, end_day)
                == calendar.SATURDAY
            ):
                end_day += 1

            start_date = datetime(current_year, current_month, start_day)
            end_date = datetime(current_year, current_month, end_day)

            weeks.append(
                {
                    "start_day": start_day,
                    "end_day": end_day,
                    "start_date": start_date,
                    "end_date": end_date,
                }
            )

    return weeks


class SessionData(APIView):
    permission_classes = [IsAuthenticated, IsInRoles("pmo")]

    def get(self, request, format=None):
        now = date.today()
        current_datetime = datetime.now()
        first_day_of_current_month = datetime(
            current_datetime.year, current_datetime.month, 1
        )
        last_day_of_previous_month = first_day_of_current_month - timedelta(days=1)
        first_day_of_previous_month = datetime(
            last_day_of_previous_month.year, last_day_of_previous_month.month, 1
        )
        start_timestamp_prev_month = int(first_day_of_previous_month.timestamp() * 1000)
        end_timestamp_prev_month = int(
            last_day_of_previous_month.timestamp() * 1000 + 86400000
        )
        if now.month == 12:
            last_day_of_current_month = datetime(now.year + 1, 1, 1) - timedelta(days=1)
        else:
            last_day_of_current_month = datetime(
                now.year, now.month + 1, 1
            ) - timedelta(days=1)

        start_timestamp_current_month = int(
            first_day_of_current_month.timestamp() * 1000
        )
        end_timestamp_current_month = int(last_day_of_current_month.timestamp() * 1000)

        start_str_prev_mon = str(start_timestamp_prev_month)
        end_str_prev_mon = str(end_timestamp_prev_month)

        start_str_current_mon = str(start_timestamp_current_month)
        end_str_current_mon = str(end_timestamp_current_month)

        projects = Project.objects.all()
        sessiondata = []
        weeks = get_weeks_for_current_month()
        for project in projects:
            engagements = Engagement.objects.filter(project=project)
            planned_learner_count = 0
            for engagement in engagements:
                completed_sessions_count = SessionRequestCaas.objects.filter(
                    status="completed",
                    project__id=engagement.project.id,
                    learner__id=engagement.learner.id,
                ).count()

                total_sessions_count = SessionRequestCaas.objects.filter(
                    project__id=engagement.project.id,
                    learner__id=engagement.learner.id,
                    is_archive=False,
                ).count()
                if (
                    engagement.status == "active"
                    and completed_sessions_count != total_sessions_count
                ):
                    planned_learner_count += 1
            res_obj = {}
            is_involved_in_sessions = SessionRequestCaas.objects.filter(
                project=project, is_booked=True
            ).exists()
            if is_involved_in_sessions:
                res_obj["project_name"] = project.name
            else:
                continue
            previous_month_sessions = SessionRequestCaas.objects.filter(
                confirmed_availability__start_time__gte=start_str_prev_mon,
                confirmed_availability__end_time__lte=end_str_prev_mon,
                project__id=project.id,
            )
            res_obj["last_month"] = previous_month_sessions.count()

            current_month_sessions = SessionRequestCaas.objects.filter(
                confirmed_availability__start_time__gte=start_str_current_mon,
                confirmed_availability__end_time__lte=end_str_current_mon,
                project__id=project.id,
            )
            res_obj["current_month"] = planned_learner_count

            for i, week in enumerate(weeks, start=1):
                start_timestamp = int(week["start_date"].timestamp())
                end_timestamp = int(week["end_date"].timestamp())
                start_day_of_ith_week_of_curr_month = str(start_timestamp)
                last_day_of_ith_week_of_curr_month = str(end_timestamp)
                weekly_sessions = SessionRequestCaas.objects.filter(
                    confirmed_availability__start_time__gte=start_day_of_ith_week_of_curr_month,
                    confirmed_availability__end_time__lte=last_day_of_ith_week_of_curr_month,
                    project__id=project.id,
                )
                key = f"current_month_week_{i}"
                res_obj[key] = weekly_sessions.count()
            actual_sessions_in_current_month = SessionRequestCaas.objects.filter(
                confirmed_availability__start_time__gte=start_str_current_mon,
                confirmed_availability__end_time__lte=end_str_current_mon,
                project__id=project.id,
                status="completed",
            )
            res_obj["total_actuals"] = actual_sessions_in_current_month.count()
            balance = (
                current_month_sessions.count()
                - actual_sessions_in_current_month.count()
            )
            res_obj["balance"] = balance
            sessiondata.append(res_obj)

        return Response(
            {"sessiondata": sessiondata, "weeks": weeks},
            status=status.HTTP_200_OK,
        )


@api_view(["GET"])
@permission_classes([AllowAny])
def google_oauth(request, user_email):
    oauth2_endpoint = "https://accounts.google.com/o/oauth2/v2/auth"

    auth_params = {
        "client_id": env("GOOGLE_OAUTH2_CLIENT_ID"),
        "redirect_uri": env("GOOGLE_OAUTH2_REDIRECT_URI"),
        "response_type": "code",
        "scope": "https://www.googleapis.com/auth/calendar https://www.googleapis.com/auth/userinfo.email",
        "access_type": "offline",
        "login_hint": user_email,
    }

    auth_url = f"{oauth2_endpoint}?{urlencode(auth_params)}"

    return HttpResponseRedirect(auth_url)


@api_view(["GET"])
@permission_classes([AllowAny])
def google_auth_callback(request):
    code = request.GET.get("code")

    if code:
        token_url = "https://oauth2.googleapis.com/token"
        token_data = {
            "code": code,
            "client_id": env("GOOGLE_OAUTH2_CLIENT_ID"),
            "client_secret": env("GOOGLE_OAUTH2_CLIENT_SECRET"),
            "redirect_uri": env("GOOGLE_OAUTH2_REDIRECT_URI"),
            "grant_type": "authorization_code",
        }

        response = requests.post(token_url, data=token_data)
        token_json = response.json()

        if "access_token" in token_json and "refresh_token" in token_json:
            access_token = token_json["access_token"]
            refresh_token = token_json["refresh_token"]
            expires_in = token_json["expires_in"]
            auth_code = code
            user_info_url = "https://www.googleapis.com/oauth2/v3/userinfo"
            headers = {"Authorization": f"Bearer {access_token}"}
            user_info_response = requests.get(user_info_url, headers=headers)

            if user_info_response.status_code == 200:
                user_info_data = user_info_response.json()

                user_email = user_info_data.get("email", "")

                user = User.objects.get(username=user_email)
                user_profile = Profile.objects.get(user=user)

                user_token, created = UserToken.objects.get_or_create(
                    user_profile=user_profile
                )
                user_token.access_token = access_token
                user_token.refresh_token = refresh_token
                user_token.access_token_expiry = expires_in
                user_token.authorization_code = auth_code
                user_token.account_type = "google"
                user_token.save()

            return HttpResponseRedirect(env("APP_URL"))
        else:
            return JsonResponse({"error": "Token exchange failed."}, status=400)
    else:
        return JsonResponse(
            {"error": "Authentication failed. Code not found."}, status=400
        )


@api_view(["GET"])
@permission_classes([AllowAny])
def microsoft_auth(request, user_mail_address):
    oauth2_endpoint = f"https://login.microsoftonline.com/common/oauth2/v2.0/authorize"

    auth_params = {
        "client_id": env("MICROSOFT_CLIENT_ID"),
        "response_type": "code",
        "redirect_uri": env("MICROSOFT_REDIRECT_URI"),
        "response_mode": "query",
        "scope": "openid offline_access User.Read Calendars.ReadWrite profile email  OnlineMeetings.Read OnlineMeetings.ReadWrite",
        "state": "shashankmeeraq",
        "login_hint": user_mail_address,
    }

    auth_url = f"{oauth2_endpoint}?{urlencode(auth_params)}"

    return HttpResponseRedirect(auth_url)


@api_view(["POST", "GET"])
@permission_classes([AllowAny])
def microsoft_callback(request):
    try:
        authorization_code = request.GET.get("code")

        token_url = f"https://login.microsoftonline.com/common/oauth2/v2.0/token"
        token_data = {
            "client_id": env("MICROSOFT_CLIENT_ID"),
            "scope": "User.Read",
            "code": authorization_code,
            "redirect_uri": env("MICROSOFT_REDIRECT_URI"),
            "grant_type": "authorization_code",
            "client_secret": env("MICROSOFT_CLIENT_SECRET"),
        }

        response = requests.post(token_url, data=token_data)

        token_json = response.json()

        if "access_token" in token_json and "refresh_token" in token_json:
            access_token = token_json["access_token"]
            refresh_token = token_json["refresh_token"]
            expires_in = token_json["expires_in"]
            auth_code = authorization_code
            user_email_url = "https://graph.microsoft.com/v1.0/me"
            headers = {"Authorization": f"Bearer {access_token}"}

            user_email_response = requests.get(user_email_url, headers=headers)

            if user_email_response.status_code == 200:
                user_info_data = user_email_response.json()
                user_email = user_info_data.get("mail", "")
                user = User.objects.get(username=user_email)
                user_profile = Profile.objects.get(user=user)
                user_token, created = UserToken.objects.get_or_create(
                    user_profile=user_profile
                )
                user_token.access_token = access_token
                user_token.refresh_token = refresh_token
                user_token.access_token_expiry = expires_in
                user_token.authorization_code = auth_code
                user_token.account_type = "microsoft"
                user_token.save()
            return HttpResponseRedirect(env("APP_URL"))
        else:
            error_json = response.json()
            return JsonResponse(error_json, status=response.status_code)

    except Exception as e:
        # Handle exceptions here, you can log the exception for debugging
        print(f"An exception occurred: {str(e)}")
        # You might want to return an error response or redirect to an error page.
        return JsonResponse({"error": "An error occurred"}, status=500)


class UserTokenAvaliableCheck(APIView):
    permission_classes = [
        IsAuthenticated,
        IsInRoles("pmo", "coach", "facilitator", "hr", "learner"),
    ]

    def get(self, request, user_mail, format=None):
        user_token_present = False
        try:
            user_token = UserToken.objects.get(user_profile__user__username=user_mail)
            if user_token:
                user_token_present = True
        except Exception as e:
            pass
        return Response({"user_token_present": user_token_present})


class DownloadCoachContract(APIView):
    permission_classes = [IsAuthenticated, IsInRoles("pmo")]

    def get(self, request, coach_contract_id, format=None):
        try:
            coach_contract = CoachContract.objects.get(id=coach_contract_id)
            coach_contract.project_contract.project.project_structure
            coach_status = CoachStatus.objects.get(
                coach=coach_contract.coach,
                project=coach_contract.project_contract.project,
            )
            data = coach_status.project_structure
            for item in data:
                if (
                    "session_type" in item
                    and item["session_type"] in SESSION_TYPE_VALUE
                ):

                    result = (
                        float(item["coach_price"])
                        * float(item["session_duration"])
                        * float(item["no_of_sessions"])
                    ) / 60

                    item["coach_price"] = float(result)

                    item["session_type"] = SESSION_TYPE_VALUE[item["session_type"]]

            total_sessions = sum(session["no_of_sessions"] for session in data)
            total_duration = sum(int(session["session_duration"]) for session in data)
            total_coach_fees = sum(float(session["coach_price"]) for session in data)

            html_content = render_to_string(
                "contract/contract_template.html",
                {
                    "name": coach_contract.coach.first_name
                    + " "
                    + coach_contract.coach.last_name,
                    "data": data,
                    "content": coach_contract.project_contract.content,
                    "name_inputed": coach_contract.name_inputed.capitalize(),
                    "signed_date": coach_contract.response_date.strftime("%d-%m-%Y"),
                    "total_sessions": total_sessions,
                    "total_duration": total_duration,
                    "total_coach_fees": total_coach_fees,
                },
            )
            pdf = pdfkit.from_string(
                html_content,
                False,
                configuration=pdfkit_config,
            )
            response = HttpResponse(pdf, content_type="application/pdf")
            response["Content-Disposition"] = f'attachment; filename={f"Contract.pdf"}'
            return response

        except Exception as e:
            print(str(e))
            return Response(
                {"error": "Failed to download Contract."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("superadmin", "pmo")])
def add_pmo(request):
    try:
        with transaction.atomic():
            data = request.data.copy()
            data["user"] = {"user_types": "pmo"}  # Assigning the role 'pmo' to the user

            pmo_serializer = PmoSerializer(data=data)

            if pmo_serializer.is_valid():
                name = data.get("name")
                email = data.get("email")
                phone = data.get("phone")
                sub_role = data.get("subRole")

                if not (name and phone):
                    return Response(
                        {"error": "Name and phone are mandatory fields."},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                user = User.objects.filter(email=email).first()
                if not user:
                    user = User.objects.create_user(
                        username=email,
                        email=email,
                        password=User.objects.make_random_password(),
                    )

                    profile = Profile.objects.create(user=user)
                else:
                    profile = Profile.objects.get(user=user)
                pmo_role, created = Role.objects.get_or_create(name="pmo")
                profile.roles.add(pmo_role)
                profile.save()
                pmo_serializer.save(user=profile)
                return Response(pmo_serializer.data, status=status.HTTP_201_CREATED)
            else:
                return Response(
                    pmo_serializer.errors, status=status.HTTP_400_BAD_REQUEST
                )

    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("superadmin", "pmo")])
def get_pmo(request):
    try:
        pmos = Pmo.objects.all()
        serializer = PmoSerializer(pmos, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("superadmin", "pmo")])
def get_junior_pmo(request, user_id):
    try:
        pmos = Pmo.objects.filter(sub_role="junior_pmo").exclude(id=user_id)
        serializer = PmoSerializerAll(pmos, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("superadmin", "pmo")])
def add_ctt_pmo(request):
    try:
        with transaction.atomic():
            data = request.data
            ctt_pmo_serializer = CTTPmoSerializer(data=data)
            if ctt_pmo_serializer.is_valid():
                name = data.get("name")
                email = data.get("email", "").strip().lower()
                phone = data.get("phone")

                if not (name and phone and email):
                    return Response(
                        {"error": "Name and phone are mandatory fields."},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                user = User.objects.filter(email=email).first()
                if not user:
                    user = User.objects.create_user(
                        username=email,
                        email=email,
                        password=User.objects.make_random_password(),
                    )

                    profile = Profile.objects.create(user=user)
                else:
                    profile = Profile.objects.get(user=user)
                ctt_pmo_role, created = Role.objects.get_or_create(name="ctt_pmo")
                profile.roles.add(ctt_pmo_role)
                profile.save()
                ctt_pmo_serializer.save(user=profile)
                return Response(ctt_pmo_serializer.data, status=status.HTTP_201_CREATED)
            else:
                return Response(
                    ctt_pmo_serializer.errors, status=status.HTTP_400_BAD_REQUEST
                )

    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("superadmin", "pmo")])
def get_ctt_pmos(request):
    try:
        ctt_pmos = CTTPmo.objects.all()
        serializer = CTTPmoSerializer(ctt_pmos, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


PATH_ACTIVITY_MAPPING = {
    "/api/login/": "Login",  # for all
    "/api/send_list_to_hr/": "Send profile to HR",  # Pmo
    "/schedular/give_availibilty/": "Give availability",  # coach
    "/api/accept-coach-caas/hr": "Finalize Coach",  # hr
    "/schedular/send_coaching_session_mail": "Send booking link email manually",  # pmo
    "/api/competency/": "Add competency",  # coach
    "/api/competency/score/": "Add score",  # coach
    "/schedular/schedule-session/": "Book slot",  # coachee
    "/api/otp/validate/": "Login with OTP",  # for all
}


ACTIVITIES_PER_USER_TYPE = {
    "hr": ["Login", "Finalize Coach", "Login with OTP"],
    "pmo": [
        "Login",
        "Send profile to HR",
        "Send booking link email manually",
        "Login with OTP",
    ],
    "coach": [
        "Login",
        "Give availability",
        "Add competency",
        "Add score",
        "Login with OTP",
    ],
    "learner": ["Login", "Book slot", "Login with OTP"],
}


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("superadmin", "pmo")])
def get_api_logs(request):
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    if not start_date or not end_date:
        logs = APILog.objects.all()
    else:
        try:
            start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
            end_date = datetime.strptime(end_date, "%Y-%m-%d").date()
            logs = APILog.objects.filter(created_at__date__range=(start_date, end_date))
        except ValueError:
            return JsonResponse(
                {"error": "Invalid date format. Please use YYYY-MM-DD."}, status=400
            )

    result_dict = {}

    for log in logs:
        matching_key = next(
            (key for key in PATH_ACTIVITY_MAPPING if log.path.startswith(key)), None
        )
        if matching_key:
            activity = PATH_ACTIVITY_MAPPING[matching_key]
            user_type = (
                log.user.profile.roles.all().exclude(name="vendor").first().name.lower()
                if log.user
                and log.user.profile
                and log.user.profile.roles.all().exclude(name="vendor").first()
                else None
            )
            if (
                user_type in ACTIVITIES_PER_USER_TYPE
                and activity in ACTIVITIES_PER_USER_TYPE[user_type]
            ):
                key = (user_type, activity)
                result_dict[key] = result_dict.get(key, 0) + 1

    # Create a nested dictionary with user types and activities as keys and counts as values
    user_activity_count_dict = {
        user_type: {
            activity: sum(
                value
                for key, value in result_dict.items()
                if key[0] == user_type and key[1] == activity
            )
            for activity in ACTIVITIES_PER_USER_TYPE[user_type]
        }
        for user_type in set(key[0] for key in result_dict)
    }
    output_list = []
    for user_type, activities in user_activity_count_dict.items():
        for activity, count in activities.items():
            output_list.append(
                {"user_type": user_type, "activity": activity, "count": count}
            )
    return Response(output_list)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "hr")])
def get_skill_training_projects(request):
    hr_id = request.query_params.get("hr", None)
    projects = SchedularProject.objects.all()
    if hr_id:
        projects = projects.filter(hr__id=hr_id)
    project_serializer = SchedularProjectSerializer(projects, many=True)

    # Modify the reminder status directly in the serialized data
    for project in project_serializer.data:
        project["project_name"] = project["name"]
        project["email_reminder"] = bool(project["email_reminder"])
        project["whatsapp_reminder"] = bool(project["whatsapp_reminder"])
        project["calendar_invites"] = bool(project["calendar_invites"])
    return Response(project_serializer.data)


@api_view(["PUT"])
@permission_classes([IsAuthenticated, IsInRoles("pmo")])
def update_reminders_of_project(request):

    try:
        project_id = request.data.get("id")
        reminder_type = request.data.get("reminder_type")
        changes_status = request.data.get("changed_status")
        schedular_project = SchedularProject.objects.get(id=project_id)

        if reminder_type == "email_reminder":
            schedular_project.email_reminder = changes_status
        elif reminder_type == "whatsapp_reminder":
            schedular_project.whatsapp_reminder = changes_status
        elif reminder_type == "calendar_invites":
            schedular_project.calendar_invites = changes_status

        schedular_project.save()
        return Response({"message": "Reminders updated successfully"})
    except SchedularProject.DoesNotExist:
        return Response(
            {"error": "Project not found"}, status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["PUT"])
@permission_classes([IsAuthenticated, IsInRoles("pmo")])
def update_reminders_of_caas_project(request):

    try:
        project_id = request.data.get("id")
        reminder_type = request.data.get("reminder_type")
        changes_status = request.data.get("changed_status")
        project = Project.objects.get(id=project_id)

        if reminder_type == "email_reminder":
            project.email_reminder = changes_status
        elif reminder_type == "whatsapp_reminder":
            project.whatsapp_reminder = changes_status
        elif reminder_type == "calendar_invites":
            project.calendar_invites = changes_status

        project.save()
        return Response({"message": "Reminders updated successfully"})
    except SchedularProject.DoesNotExist:
        return Response(
            {"error": "Project not found"}, status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("pmo")])
def add_extra_session_in_caas(request, learner_id, project_id):
    try:

        session_data = request.data
        learner = Learner.objects.get(id=learner_id)
        project = Project.objects.get(id=project_id)
        sessions = SessionRequestCaas.objects.filter(
            project__id=project_id, learner__id=learner_id
        ).order_by("order")
        engagement = Engagement.objects.filter(project=project, learner=learner).first()
        filtered_sessions = sessions.filter(session_type=session_data["session_type"])
        max_session_number = (
            filtered_sessions.aggregate(Max("session_number"))["session_number__max"]
            if filtered_sessions.count() > 0
            else 0
        )
        max_billable_session_number = sessions.aggregate(
            Max("billable_session_number")
        )["billable_session_number__max"]
        max_order = sessions.aggregate(Max("order"))["order__max"]
        session_data = SessionRequestCaas.objects.create(
            engagement=engagement,
            learner=learner,
            project=project,
            session_duration=session_data["session_duration"],
            session_number=(max_session_number + 1),
            session_type=session_data["session_type"],
            billable_session_number=(max_billable_session_number + 1),
            status="pending",
            order=(max_order + 1),
            is_extra=True,
        )
        return Response(
            {"detail": "Extra session added successfully"},
            status=status.HTTP_201_CREATED,
        )

    except Exception as e:
        # Handle the exception as per your application's requirements
        return Response(
            {"detail": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("pmo")])
def add_coaches_to_project(request):
    try:
        project = Project.objects.get(id=request.data.get("project_id", ""))
    except Project.DoesNotExist:
        return Response({"message": "Project does not exist"}, status=400)
    # Get the list of coach IDs
    coach_list = request.data.get("coach_list", [])
    # Initialize a list to store coach status
    coach_status = []
    # Iterate through each coach
    for coach_id in coach_list:
        try:
            coach = Coach.objects.get(id=coach_id)
            # Check if CoachStatus already exists for this coach and project
            existing_coach_statuses = CoachStatus.objects.filter(
                coach=coach, project=project
            )
            if not existing_coach_statuses.exists():
                coach_status_instance = CoachStatus.objects.create(
                    coach=coach,
                    project=project,
                    status={
                        "consent": {
                            "status": "sent",
                            "response_date": None,
                        },
                        "project_structure": {
                            "status": "sent",
                            "response_date": None,
                        },
                        "hr": {
                            "status": None,
                            "session_id": None,
                            "response_date": None,
                        },
                        "learner": {
                            "status": None,
                            "session_id": None,
                            "response_date": None,
                        },
                    },
                    # consent_expiry_date=request.data["consent_expiry_date"],
                    is_consent_asked=False,
                    project_structure=(
                        project.project_structure
                        if project.steps["project_structure"]["status"] == "complete"
                        else []
                    ),  # adding project structure only if structure is finalized.
                )
                coach_status.append(coach_status_instance)
            # Create or update CoachProfileTemplate
            profile_template, created = CoachProfileTemplate.objects.get_or_create(
                coach=coach,
                project=project,
                defaults={
                    "templates": {
                        "coaching_experience": coach.coaching_experience,
                        "corporate_experience": coach.corporate_experience,
                        # # Add other fields here to include all coach details
                    }
                },
            )

            # If the template already exists, update the coach details
            if not created:
                profile_template.templates["templates"] = {
                    "coaching_experience": coach.coaching_experience,
                    "corporate_experience": coach.corporate_experience,
                }

            profile_template.save()

        except Coach.DoesNotExist:
            pass
    # Update project's coach_status and steps
    project.coaches_status.add(*coach_status)
    project.save()

    tasks = Task.objects.filter(
        task="add_coach", status="pending", caas_project=project
    )
    tasks.update(status="completed")

    return Response({"message": "Coaches added to project successfully!"})


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("pmo")])
def update_coach_project_structure(request, coach_id):
    try:
        project = Project.objects.get(id=request.data.get("project_id", ""))
    except Project.DoesNotExist:
        return Response({"message": "Project does not exist"}, status=400)
    try:
        coach_status = project.coaches_status.get(coach__id=coach_id)
        coach_status.project_structure = request.data.get("project_structure", [])
        coach_status.save()
        return Response(
            {"message": "Project structure updated for coach successfully."}
        )
    except Exception as e:
        print(str(e))
        return Response({"error": "Failed to update project structure."}, status=401)


@api_view(["PUT"])
@permission_classes([IsAuthenticated, IsInRoles("pmo")])
def blacklist_coach(request):
    try:
        coach_id = request.data.get("coach_id")
        coach = Coach.objects.get(id=int(coach_id))
        if coach.active_inactive:
            coach.active_inactive = False
        else:
            coach.active_inactive = True
        coach.save()
        return Response(
            {
                "message": f"Coach {'whitelist' if coach.active_inactive else 'blacklisted'} successfully"
            },
            status=200,
        )
    except Exception as e:
        print(str(e))
        return Response(
            {"error": "Failed to blacklist coach"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "superadmin")])
def get_all_api_logs(request):
    paginator = PageNumberPagination()
    paginator.page_size = 200
    logs = APILog.objects.exclude(path__startswith="/admin").order_by("-created_at")

    start_date = request.query_params.get("start_date")
    end_date = request.query_params.get("end_date")
    username = request.query_params.get("username")

    if start_date and end_date:
        start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
        end_date = datetime.strptime(end_date, "%Y-%m-%d").date()
        logs = logs.filter(created_at__date__range=(start_date, end_date))

    if username:
        logs = logs.filter(user__username__contains=username)

    result_page = paginator.paginate_queryset(logs, request)
    serializer = APILogSerializer(result_page, many=True)
    return paginator.get_paginated_response(serializer.data)


# @api_view(["DELETE"])
# @permission_classes([IsAuthenticated])
# def delete_pmo(request):
#     try:
#         with transaction.atomic():
#             pmo_email = request.data.get("email").strip().lower()
#             if pmo_email:

#                 pmo = Pmo.objects.get(email=pmo_email)

#                 pmo_user_profile = pmo.user
#                 is_delete_user = True
#                 for role in pmo_user_profile.roles.all():
#                     if not role.name == "pmo":
#                         # don't delete user if any other role exists
#                         is_delete_user = False
#                     else:
#                         pmo_user_profile.roles.remove(role)
#                         pmo_user_profile.save()
#                 if is_delete_user:
#                     user = pmo.user.user
#                     user.delete()
#                 else:
#                     pmo.delete()

#                 return Response({"message": "Pmo deleted successfully"})

#     except Exception as e:
#         print(str(e))
#         return Response(
#             {"error": "Failed to delete pmo."},
#             status=status.HTTP_500_INTERNAL_SERVER_ERROR,
#         )


class UpdateUserRoles(APIView):
    permission_classes = [IsAuthenticated]

    def put(self, request):
        try:
            with transaction.atomic():
                user_id = request.data.get("user_id")
                removed_roles = request.data.get("removed_roles")
                added_roles = request.data.get("added_roles")

                profile = Profile.objects.get(id=user_id)

                for role in removed_roles:
                    user = get_user_for_active_inactive(role, profile.user.username)

                    user.active_inactive = False
                    user.save()

                for role in added_roles:
                    user = get_user_for_active_inactive(role, profile.user.username)

                    user.active_inactive = True
                    user.save()

                return Response({"message": "Roles updated successfully!"})
        except Exception as e:
            print(str(e))
            return Response(
                {"error": "Failed to update."},
                status=500,
            )


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("pmo")])
def add_remark_to_task(request):
    task_id = request.data.get("task_id", "")
    try:
        task = Task.objects.get(id=task_id)
    except Task.DoesNotExist:
        return Response({"error": "Task not found"}, status=status.HTTP_404_NOT_FOUND)
    message = request.data.get("message", "")
    if not message:
        return Response(
            {"error": "message is required"}, status=status.HTTP_400_BAD_REQUEST
        )
    task.add_remark(message)
    return Response(
        {"message": "Remark added successfully"}, status=status.HTTP_201_CREATED
    )


def get_formatted_tasks(tasks):
    try:
        task_details = []
        for task in tasks:
            # Extract the latest remark message or set it to None if remarks are empty
            latest_remark = task.remarks[-1]["message"] if task.remarks else None

            # Determine the coach name associated with the task
            coach_name = (
                task.coach.__str__()
                if task.coach
                else (
                    task.engagement.coach.__str__()
                    if task.engagement and task.engagement.coach
                    else (
                        task.goal.engagement.coach.__str__()
                        if task.goal and task.goal.engagement.coach
                        else (
                            task.session_caas.coach.__str__()
                            if task.session_caas and task.session_caas.coach
                            else ""
                        )
                    )
                )
            )
            learner_name = (
                task.engagement.learner.name
                if task.engagement
                else (
                    task.goal.engagement.learner.name
                    if task.goal
                    else (
                        task.session_caas.learner.name
                        if task.session_caas and task.session_caas.learner
                        else ""
                    )
                )
            )
            session_name = (
                SESSION_TYPE_VALUE[task.session_caas.session_type]
                if task.session_caas
                else ""
            )
            # Retrieve vendor name associated with the task, handling exceptions gracefully
            vendor_name = None
            if task.vendor_user:
                try:
                    vendor = Vendor.objects.get(user__user=task.vendor_user)
                    vendor_name = vendor.name
                except Vendor.DoesNotExist:
                    pass

            # Determine the project name associated with the task
            project_name = (
                task.caas_project.name
                if task.caas_project
                else (
                    task.session_caas.project.name
                    if task.session_caas
                    else (
                        task.engagement.project.name
                        if task.engagement
                        else task.goal.engagement.project.name if task.goal else ""
                    )
                )
            )

            # Serialize task data
            serialized_data = TaskSerializer(task).data

            # Append task details to the list
            task_details.append(
                {
                    **serialized_data,
                    "learner_name": learner_name,
                    "project_name": project_name,
                    "latest_remark": latest_remark,
                    "coach_name": coach_name,
                    "vendor_name": vendor_name,
                    "session_name": session_name,
                }
            )
        return task_details
    except Exception as e:
        print(str(e))
        return None


def get_formatted_skill_training_tasks(tasks):
    try:
        task_details = []
        for task in tasks:
            # Extract the latest remark message or set it to None if remarks are empty
            latest_remark = task.remarks[-1]["message"] if task.remarks else None

            # Determine the coach name associated with the task
            coach_name = (
                task.coach.__str__()
                if task.coach
                else (
                    task.schedular_session.coach.__str__()
                    if task.schedular_session and task.schedular_session.coach
                    else ""
                )
            )
            facilitator_name = task.facilitator.__str__() if task.facilitator else ""
            learner_name = (
                task.schedular_session.learner.name
                if task.schedular_session and task.schedular_session.learner
                else ""
            )
            coaching_session = (
                task.coaching_session
                or (
                    task.schedular_session.coaching_session
                    if task.schedular_session
                    else None
                )
                or None
            )
            session_name = (
                (
                    coaching_session.session_type.replace("_", " ").capitalize()
                    if not coaching_session.session_type == "laser_coaching_session"
                    else "Coaching Session"
                    + " "
                    + str(coaching_session.coaching_session_number)
                )
                if coaching_session
                else ""
            )
            # Retrieve vendor name associated with the task, handling exceptions gracefully
            vendor_name = None
            if task.vendor_user:
                try:
                    vendor = Vendor.objects.get(user__user=task.vendor_user)
                    vendor_name = vendor.name
                except Vendor.DoesNotExist:
                    pass

            # Determine the project name associated with the task
            project_name = (
                task.schedular_project.name
                if task.schedular_project
                else (task.schedular_batch.project.name if task.schedular_batch else "")
            )
            batch_name = (
                task.schedular_batch.name
                if task.schedular_batch
                else (
                    task.live_session.batch.name
                    if task.live_session
                    else (
                        task.coaching_session.batch.name
                        if task.coaching_session
                        else ""
                    )
                )
            )
            live_session_name = (
                get_live_session_name(task.live_session.session_type)
                if task.live_session
                else ""
            )

            # Serialize task data
            serialized_data = TaskSerializer(task).data

            # Append task details to the list
            task_details.append(
                {
                    **serialized_data,
                    "learner_name": learner_name,
                    "project_name": project_name,
                    "latest_remark": latest_remark,
                    "coach_name": coach_name,
                    "vendor_name": vendor_name,
                    "session_name": session_name,
                    "batch_name": batch_name,
                    "facilitator_name": facilitator_name,
                    "live_session_name": live_session_name,
                }
            )
        return task_details
    except Exception as e:
        print(str(e))
        return None


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("pmo")])
def get_tasks(request):
    # Retrieve tasks that are pending and have a trigger date before or equal to current time
    tasks = Task.objects.filter(
        Q(trigger_date__lte=timezone.now()),
        ~Q(project_type="skill_training"),
        ~Q(status="completed"),
    )
    # Initialize a list to store task details
    task_details = []
    junior_pmo_id = request.query_params.get("junior_pmo")
    if junior_pmo_id:
        tasks = tasks.filter(
            Q(caas_project__junior_pmo=junior_pmo_id)
            | Q(session_caas__project__junior_pmo=junior_pmo_id)
            | Q(engagement__project__junior_pmo=junior_pmo_id)
            | Q(goal__engagement__project__junior_pmo=junior_pmo_id)
        )

    task_details = get_formatted_tasks(tasks)

    return Response(task_details)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("pmo")])
def get_skill_training_tasks(request):
    # Retrieve tasks that are pending and have a trigger date before or equal to current time
    tasks = Task.objects.filter(
        Q(trigger_date__lte=timezone.now()),
        Q(project_type="skill_training"),
        ~Q(status="completed"),
    )
    # Initialize a list to store task details
    task_details = []
    junior_pmo_id = request.query_params.get("junior_pmo")
    if junior_pmo_id:
        tasks = tasks.filter(
            Q(schedular_project__junior_pmo=junior_pmo_id)
            | Q(schedular_batch__project__junior_pmo=junior_pmo_id)
            | Q(live_session__batch__project__junior_pmo=junior_pmo_id)
            | Q(coaching_session__batch__project__junior_pmo=junior_pmo_id)
            | Q(
                schedular_session__coaching_session__batch__project__junior_pmo=junior_pmo_id
            )
        )

    task_details = get_formatted_skill_training_tasks(tasks)

    return Response(task_details)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("coach")])
def get_coach_tasks(request):

    coach_id = request.query_params.get("coach_id")

    coach = Coach.objects.get(id=int(coach_id))

    tasks = Task.objects.filter(
        Q(coach=coach)
        | Q(engagement__coach=coach)
        | Q(goal__engagement__coach=coach)
        | Q(session_caas__coach=coach),
        Q(trigger_date__lte=timezone.now()),
    )

    task_details = get_formatted_tasks(tasks)

    return Response(task_details)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("learner")])
def get_learner_tasks(request):

    learenr_id = request.query_params.get("learenr_id")

    learner = Learner.objects.get(id=int(learenr_id))

    tasks = Task.objects.filter(
        Q(engagement__learner=learner)
        | Q(goal__engagement__learner=learner)
        | Q(session_caas__learner=learner),
        Q(trigger_date__lte=timezone.now()),
    )

    task_details = get_formatted_tasks(tasks)

    return Response(task_details)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("pmo")])
def complete_task(request):
    task_id = request.data.get("task_id", "")
    try:
        task = Task.objects.get(id=task_id)
    except Task.DoesNotExist:
        return Response({"error": "Task not found"}, status=status.HTTP_404_NOT_FOUND)
    task.status = "completed"
    task.save()
    message = f"{request.user.username} completed this task."
    task.add_remark(message)
    return Response(
        {"message": "Task marked as completed."}, status=status.HTTP_201_CREATED
    )


def get_purchase_order(purchase_orders, purchase_order_id):
    for po in purchase_orders:
        if po.get("purchaseorder_id") == purchase_order_id:
            return po
    return None


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_coaches_in_project_is_vendor(request, project_id):
    try:
        project = Project.objects.get(id=project_id)
        data = {}
        purchase_orders = PurchaseOrderGetSerializer(
            PurchaseOrder.objects.filter(
                Q(created_time__year__gte=2024)
                | Q(purchaseorder_number__in=purchase_orders_allowed)
            ),
            many=True,
        ).data
        # filter_purchase_order_data(PurchaseOrderGetSerializer(PurchaseOrder.objects.all(), many=True).data)
        # fetch_purchase_orders(organization_id)
        for coach_status in project.coaches_status.all():
            is_vendor = coach_status.coach.user.roles.filter(name="vendor").exists()
            vendor_id = None
            is_delete_purchase_order_allowed = True
            purchase_order_id = (
                coach_status.purchase_order_id
                if coach_status.purchase_order_id
                else None
            )
            purchase_order_no = (
                coach_status.purchase_order_id
                if coach_status.purchase_order_no
                else None
            )
            if is_vendor:
                vendor_id = Vendor.objects.get(user=coach_status.coach.user).vendor_id

            purchase_order = None
            if purchase_order_id:
                purchase_order = get_purchase_order(purchase_orders, purchase_order_id)
                if not purchase_order:
                    coach_status.purchase_order_id = ""
                    coach_status.purchase_order_no = ""
                    coach_status.save()
                    purchase_order_id = None
                    purchase_order_no = None
                else:
                    invoices = InvoiceData.objects.filter(
                        purchase_order_id=purchase_order_id
                    )
                    if invoices.exists():
                        is_delete_purchase_order_allowed = False

            data[coach_status.coach.id] = {
                "vendor_id": vendor_id,
                "purchase_order_id": purchase_order_id,
                "purchase_order_no": purchase_order_no,
                "purchase_order": purchase_order,
                "is_delete_purchase_order_allowed": is_delete_purchase_order_allowed,
            }
        return Response(data)
    except Exception as e:
        print(str(e))
        return Response(
            {"error": "Failed to get data"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_all_users(request):
    users = User.objects.filter(profile__isnull=False)
    serializer = CustomUserSerializer(users, many=True)
    return Response(serializer.data)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def archive_project(request):
    try:
        project_id = request.data.get("project_id")
        project_type = request.data.get("project_type")
        if project_type == "SEEQ":
            project = SchedularProject.objects.get(id=project_id)
        elif project_type == "CAAS":
            project = Project.objects.get(id=project_id)

        if project.is_archive:
            project.is_archive = False
        else:
            project.is_archive = True

        project.save()

        return Response(
            {"message": "Project Archived Sucessfully!"}, status=status.HTTP_201_CREATED
        )
    except Exception as e:
        print(str(e))
        return Response({"error": "Failed to archive project!"}, status=500)


@api_view(["PUT"])
@permission_classes([IsAuthenticated])
def change_user_password(request):
    try:

        with transaction.atomic():
            email = request.data.get("email", None)
            password = request.data.get("password", None)

            if email and password:
                user = User.objects.get(username=email)
                user.set_password(password)
                user.save()
                return Response({"message": "Password reset successful!"}, status=200)
            else:
                return Response({"error": "Failed to reset password!"}, status=500)
    except Exception as e:
        print(str(e))


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def add_sales_user(request):
    name = request.data.get("name")
    email = request.data.get("email", "").strip().lower()
    phone = request.data.get("phone")
    business = request.data.get("business", "")
    sales_person_id = request.data.get("sales_person_id", "").strip()
    username = email  # username and email are the same
    # Check if required data is provided
    if not all([name, email, phone, username, sales_person_id]):
        return Response(
            {"error": "All required fields must be provided."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    existing_sales_user = Sales.objects.filter(sales_person_id=sales_person_id)
    if existing_sales_user.exists():
        return Response(
            {"error": "Selected zoho sales person already exists."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        with transaction.atomic():
            # Check if the user already exists
            user = User.objects.filter(email=email).first()

            if not user:
                # If the user does not exist, create a new user
                user = User.objects.create_user(username=username, email=email)
                temp_password = "".join(
                    random.choices(
                        string.ascii_uppercase + string.ascii_lowercase + string.digits,
                        k=8,
                    )
                )
                user.set_password(temp_password)
                user.save()
                profile = Profile.objects.create(user=user)

            else:
                profile = Profile.objects.get(user=user)

            # Create or get the "pmo" role
            sales_role, created = Role.objects.get_or_create(name="sales")
            profile.roles.add(sales_role)
            profile.save()

            # Create the PMO User using the Profile
            sales_user = Sales.objects.create(
                user=profile,
                name=name,
                email=email,
                phone=phone,
                sales_person_id=sales_person_id,
                business=business,
            )

            name = sales_user.name
            add_contact_in_wati("sales", name, sales_user.phone)

            # Return success response without room_id
            return Response({"message": "Sales User added successfully."}, status=201)

    except IntegrityError as e:
        return Response({"error": "User with this email already exists."}, status=400)

    except Exception as e:
        # Return error response if any exception occurs
        return Response({"error": str(e)}, status=500)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_sales_user(request):
    try:
        sales_user = Sales.objects.all()
        serializer = SalesSerializer(sales_user, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    except Exception as e:
        print(str(e))
        return Response(
            {"detail": f"Error fetching Sales user: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


def get_coach_overall_rating(coach_id):
    try:
        coach_feedback_responses = CoachingSessionsFeedbackResponse.objects.filter(
            Q(caas_session__coach_id=coach_id)
            | Q(schedular_session__availibility__coach_id=coach_id)
        )

        overall_rating = coach_feedback_responses.aggregate(
            avg_rating=Avg(
                "answers__rating", filter=Q(answers__question__type="rating_1_to_5")
            )
        )["avg_rating"]

        return overall_rating if overall_rating is not None else 0
    except Exception as e:
        print(str(e))
        return 0


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_coach_summary_data(request, coach_id):
    try:
        coach = Coach.objects.get(id=coach_id)

        coach_status_count = CoachStatus.objects.filter(coach=coach).count()

        coach_batches = SchedularBatch.objects.filter(coaches=coach)
        distinct_project_count = coach_batches.values("project").distinct().count()

        total_coaching_session = SchedularSessions.objects.filter(
            availibility__coach=coach,
        ).count()

        laser_coaching_sessions = SchedularSessions.objects.filter(
            availibility__coach=coach,
            coaching_session__session_type="laser_coaching_session",
        ).count()

        total_mentoring_sessions = SchedularSessions.objects.filter(
            availibility__coach=coach,
            coaching_session__session_type="mentoring_session",
        ).count()

        pending_schedular_session = SchedularSessions.objects.filter(
            availibility__coach=coach, status="pending"
        ).count()

        return Response(
            {
                "last_login": coach.user.user.last_login,
                "coaching_projects": coach_status_count,
                "training_projects": distinct_project_count,
                "coach_rating": get_coach_overall_rating(coach.id),
                "total_coaching_session": total_coaching_session,
                "laser_coaching_sessions": laser_coaching_sessions,
                "pending_session": pending_schedular_session,
                "total_mentoring_sessions": total_mentoring_sessions,
            }
        )

    except Exception as e:
        print(str(e))
        return Response(
            {"detail": f"Failed to get data"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_facilitator_summary_data(request, facilitator_id):
    try:
        facilitator = Facilitator.objects.get(id=facilitator_id)

        live_sessions = LiveSession.objects.filter(facilitator=facilitator)
        distinct_project_count = (
            live_sessions.values("batch__project").distinct().count()
        )

        overall_answer = Answer.objects.filter(
            question__type="rating_0_to_10",
            question__feedbacklesson__live_session__facilitator__id=facilitator_id,
        )

        overall_nps = calculate_nps_from_answers(overall_answer)

        return Response(
            {
                "last_login": facilitator.user.user.last_login,
                "training_projects": distinct_project_count,
                "overall_nps": overall_nps,
                "live_sessions": live_sessions.count(),
            }
        )

    except Exception as e:
        print(str(e))
        return Response(
            {"error": f"Failed to get data"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def hide_columns(request):
    try:
        user_id = request.data.get("user_id")
        hidden_columns = request.data.get("hidden_columns")
        table_name = request.data.get("table_name")

        user = User.objects.get(id=user_id)

        table_hidden_column, created = TableHiddenColumn.objects.get_or_create(
            table_name=table_name, user=user
        )
        table_hidden_column.hidden_columns = hidden_columns
        table_hidden_column.save()
        return Response(table_hidden_column.hidden_columns)

    except Exception as e:
        print(str(e))
        return Response(
            {"error": f"Failed to hide column"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_table_hide_columns(request, table_name, user_id):
    try:

        table_hidden_column = TableHiddenColumn.objects.get(
            table_name=table_name, user__id=user_id
        )

        return Response(table_hidden_column.hidden_columns)

    except Exception as e:
        print(str(e))
        return Response(
            {"detail": f"Failed to get data"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_expenses_for_coaching_project(request, project_id, coach_id):
    try:
        expense = Expense.objects.filter(
            session__project__id=project_id, coach__id=coach_id
        )
        serializer = ExpenseSerializerDepthOne(expense, many=True)
        serialized_data = serializer.data

        for data in serialized_data:
            session = data.get("session")
            if session:
                session_new = SessionRequestCaas.objects.filter(
                    id=session.get("id")
                ).first()
                if session_new:
                    learner = session_new.learner
                    data["learner"] = {
                        "id": learner.id,
                        "name": learner.name,
                    }
                else:
                    data["learner"] = None
            else:
                data["learner"] = None
        return Response(serializer.data)
    except Exception as e:
        print(str(e))
        return Response(
            {"error": "Failed to get expense"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_coach_with_vendor_id_in_project(request, project_id):
    try:
        project = Project.objects.get(id=project_id)
        all_coach = {}
        purchase_orders = fetch_purchase_orders(organization_id)
        for coach_status in project.coaches_status.all():
            coach = coach_status.coach
            vendor = Vendor.objects.filter(user=coach.user).first()
            if vendor:
                expense = Expense.objects.filter(
                    session__project__id=project_id, coach=coach
                ).first()
                purchase_order = None
                if expense and expense.purchase_order_id:
                    purchase_order = get_purchase_order(
                        purchase_orders, expense.purchase_order_id
                    )

                is_delete_purchase_order_allowed = True
                invoices = InvoiceData.objects.filter(
                    purchase_order_id=expense.purchase_order_id
                )
                if invoices.exists():
                    is_delete_purchase_order_allowed = False

                all_coach[coach.id] = {
                    "is_vendor": True,
                    "vendor_id": vendor.vendor_id,
                    "purchase_order_id": expense.purchase_order_id if expense else None,
                    "purchase_order_no": expense.purchase_order_no if expense else None,
                    "purchase_order": purchase_order,
                    "is_delete_purchase_order_allowed": is_delete_purchase_order_allowed,
                }
        return Response(all_coach)
    except Exception as e:
        print(str(e))
        return Response(
            {"error": "Failed to get data"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def edit_remark(request):
    try:
        coach_id = request.data.get("coach_id")
        remark = request.data.get("remark")
        coach = Coach.objects.get(id=int(coach_id))
        coach.remark = remark
        coach.save()
        return Response({"message": f"Remark updated successfully!"}, status=200)

    except Exception as e:
        print(str(e))
        return Response(
            {"error": f"Failed to update remark"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_all_goals(request):
    try:
        goals = Goal.objects.all()
        serializer = GoalDescriptionSerializer(goals, many=True)
        return Response(serializer.data)
    except Exception as e:
        print(str(e))
        return Response(
            {"error": f"Failed to get data"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def create_goal_without_enagagement(request):
    try:
        goal_name = request.data.get("name")
        description = request.data.get("description")
        goal = Goal.objects.create(name=goal_name, description=description)

        return Response({"message": "Goal created sucessfully!"})
    except Exception as e:
        print(str(e))
        return Response(
            {"error": f"Failed to create goal"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def add_new_user(request):
    try:
        with transaction.atomic():
            data = request.data
            print(request.data)
            name = data.get("name", None)
            email = data.get("email", "").strip().lower()
            roles = data.get("role", None)
            sub_role = data.get("sub_role", "")
            phone = data.get("phone", "")
            password = str(uuid.uuid4())

            if not email or not name or not roles:
                return Response(
                    {"error": "Please fill all the required feilds."}, status=500
                )

            if "pmo" in roles:

                add_new_pmo(
                    {
                        "name": name,
                        "email": email,
                        "phone": phone,
                        "password": password,
                        "sub_role": sub_role,
                    }
                )
            # if "sales" in roles:
            #     add_new_sales_user({})

            if "finance" in roles:

                # Check if the user already exists
                user = User.objects.filter(email=email).first()

                if not user:
                    # If the user does not exist, create a new user
                    user = User.objects.create_user(
                        username=email, password=password, email=email
                    )
                    profile = Profile.objects.create(user=user)

                else:
                    profile = Profile.objects.get(user=user)

                # Create or get the "finance" role
                finance_role, created = Role.objects.get_or_create(name="finance")
                profile.roles.add(finance_role)
                profile.save()

                # Create the finance User using the Profile
                finance_user = Finance.objects.create(
                    user=profile,
                    name=name,
                    email=email,
                )

            if "sales" in roles:

                # Check if the user already exists
                user = User.objects.filter(email=email).first()

                if not user:
                    # If the user does not exist, create a new user
                    user = User.objects.create_user(
                        username=email, password=password, email=email
                    )
                    profile = Profile.objects.create(user=user)

                else:
                    profile = Profile.objects.get(user=user)

                # Create or get the "sales" role
                sales_role, created = Role.objects.get_or_create(name="sales")
                profile.roles.add(sales_role)
                profile.save()

                # Create the sales User using the Profile
                sales_user = Sales.objects.create(
                    user=profile,
                    name=name,
                    email=email,
                )

        return Response({"message": "User added sucessfully!"}, status=200)
    except Exception as e:
        print(str(e))
        return Response({"error": "Failed to add user!"}, status=500)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_competency_of_goal(request, goal_id):
    try:
        competency = Competency.objects.filter(goal__id=goal_id)
        serializer = CompetencySerializer(competency, many=True)
        return Response(serializer.data)
    except Exception as e:
        print(str(e))
        return Response(
            {"error": f"Failed to get data"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def create_competency_for_goal(request, goal_id):
    try:
        name = request.data.get("name")

        goal = Goal.objects.get(id=goal_id)
        competency = Competency.objects.create(goal=goal, name=name)
        return Response({"message": "Competency created sucessfully!"})
    except Exception as e:
        print(str(e))
        return Response(
            {"error": f"Failed to create competency"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_goal_detail(request, goal_id):
    try:
        goal = Goal.objects.get(id=goal_id)
        serializer = GoalDescriptionSerializer(goal)
        return Response(serializer.data)

    except Exception as e:
        print(str(e))
        return Response(
            {"error": f"Failed to create competency"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_competency_action_items(request, comp_id):
    try:
        action_items = ActionItem.objects.filter(competency__id=comp_id)
        serializer = ActionItemSerializer(action_items, many=True)
        return Response(serializer.data)

    except Exception as e:
        print(str(e))
        return Response(
            {"error": f"Failed to get data"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["PUT"])
@permission_classes([IsAuthenticated])
def edit_pmo_goal(request):
    try:
        goal_name = request.data.get("name")
        description = request.data.get("description")
        goal_id = request.data.get("goal_id")
        goal = Goal.objects.get(id=goal_id)
        goal.name = goal_name
        goal.description = description
        goal.save()

        return Response({"message": "Goal updated sucessfully!"})
    except Exception as e:
        print(str(e))
        return Response(
            {"error": f"Failed to update goal"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_all_competency(request):
    try:
        competency = Competency.objects.all()
        serializer = CompetencySerializer(competency, many=True)
        return Response(serializer.data)
    except Exception as e:
        print(str(e))
        return Response(
            {"error": f"Failed to get data"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_all_po_of_project(request, project_id):
    try:
        po_ids = set()
        project = Project.objects.get(id=project_id)
        for coach_status in project.coaches_status:
            if coach_status.purchase_order_id:
                po_ids.add(coach_status.purchase_order_id)
        purchase_orders = []
        for po_id in po_ids:
            access_token = get_access_token(env("ZOHO_REFRESH_TOKEN"))
            if not access_token:
                raise Exception(
                    "Access token not found. Please generate an access token first."
                )
            api_url = (
                f"{base_url}/purchaseorders/{po_id}?organization_id={organization_id}"
            )
            auth_header = {"Authorization": f"Bearer {access_token}"}
            response = requests.put(api_url, headers=auth_header, data=request.data)
            if response.status_code == 200:
                purchase_orders.append(response.json())

        return Response(purchase_orders)
    except Exception as e:
        print(str(e))
        return Response(
            {"error": f"Failed to get data"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


# def create_teams_meeting():
#     event_create_url = "https://graph.microsoft.com/v1.0/me/onlineMeetings"
#     try:
#         user_token = UserToken.objects.get(user_profile__user__username="pankaj@meeraq.com")
#         new_access_token = refresh_microsoft_access_token(user_token)
#         if not new_access_token:
#             new_access_token = user_token.access_token
#         headers = {
#             "Authorization": f"Bearer {new_access_token}",
#             "Content-Type": "application/json",
#         }
#         start_datetime_obj = datetime(2024, 4, 5, 8, 0, 0) + timedelta(hours=5, minutes=30)
#         end_datetime_obj = datetime(2024, 4, 5, 11, 0, 0) + timedelta(hours=5, minutes=30)
#         start_datetime = start_datetime_obj.isoformat()
#         end_datetime = end_datetime_obj.isoformat()
#         event_payload = {
#             "startDateTime":"2024-04-05T14:30:34.2444915-07:00",
#             "endDateTime":"2024-04-05T15:00:34.2464912-07:00",
#             "subject": "User Token Meeting"
#         }
#         response = requests.post(event_create_url, json=event_payload, headers=headers)
#         print(response.json())
#         if response.status_code == 201:
#             microsoft_response_data = response.json()
#             print("Meeting created successfully.")
#         else:
#             print(f"Meeting creation failed. Status code: {response.status_code}")
#             print(response.text)
#             return False

#     except UserToken.DoesNotExist:
#         print("User token not found for email.")
#         return False

#     except Exception as e:
#         print(f"An error occurred: {str(e)}")
#         return False

#     return True


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "learner")])
def get_all_to_be_booked_sessions_for_coachee(request, learner_id):
    sessions = SessionRequestCaas.objects.filter(
        learner__id=learner_id,
        status="pending",
        engagement__coach__isnull=False,
    ).order_by("order")
    serializer = SessionRequestCaasDepthOneSerializer(sessions, many=True)
    return Response(serializer.data, status=200)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_engagement_of_a_coachee(request, learner_id):
    engagement = Engagement.objects.get(learner__id=learner_id)
    serializer = EngagementDepthOneSerializer(engagement)
    return Response(serializer.data, status=200)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("superadmin")])
def add_leader(request):
    try:
        with transaction.atomic():
            data = request.data
            leader_serializer = LeaderSerializer(data=data)
            if leader_serializer.is_valid():
                name = data.get("name")
                email = data.get("email", "").strip().lower()
                phone = data.get("phone")

                if not (name and phone and email):
                    return Response(
                        {"error": "Name and phone are mandatory fields."},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                user = User.objects.filter(email=email).first()
                if not user:
                    user = User.objects.create_user(
                        username=email,
                        email=email,
                        password=User.objects.make_random_password(),
                    )

                    profile = Profile.objects.create(user=user)
                else:
                    profile = Profile.objects.get(user=user)
                leader_role, created = Role.objects.get_or_create(name="leader")
                profile.roles.add(leader_role)
                profile.save()
                leader_serializer.save(user=profile)
                return Response(leader_serializer.data, status=status.HTTP_201_CREATED)
            else:
                return Response(
                    leader_serializer.errors, status=status.HTTP_400_BAD_REQUEST
                )

    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("superadmin")])
def get_leaders(request):
    try:
        leaders = Leader.objects.all()
        serializer = LeaderSerializer(leaders, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["PUT"])
@permission_classes([AllowAny, IsInRoles("superadmin")])
def edit_leader(request, leader_id):
    name = request.data.get("name")
    email = request.data.get("email", "").strip().lower()
    phone = request.data.get("phone")
    leader = Leader.objects.get(id=leader_id)
    try:
        with transaction.atomic():
            existing_user = (
                User.objects.filter(username=email)
                .exclude(username=leader.user.user.username)
                .first()
            )
            if existing_user:
                return Response(
                    {"error": "User with this email already exists."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            leader.user.user.username = email
            leader.user.user.email = email
            leader.user.user.save()
            leader.email = email
            leader.name = name
            leader.phone = phone
            leader.save()
            return Response({"message": "Leader updated successfully."}, status=201)
    except Exception as e:
        print(str(e))
        return Response({"error": "Failed to update Leader."}, status=500)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def create_engagement_of_learner(request):
    try:
        with transaction.atomic():

            engagement_id = request.data.get("engagement_id")

            engagement = Engagement.objects.get(id=engagement_id)

            engagemenet_project_structure = transform_project_structure(
                engagement.project.project_structure
            )
            new_engagement = Engagement(
                learner=engagement.learner, project=engagement.project, status="active"
            )
            new_engagement.save()
            for index, session in enumerate(engagemenet_project_structure):
                session_data = SessionRequestCaas.objects.create(
                    learner=engagement.learner,
                    project=engagement.project,
                    session_duration=session["session_duration"],
                    session_number=session["session_number"],
                    session_type=session["session_type"],
                    billable_session_number=session["billable_session_number"],
                    status="pending",
                    order=index + 1,
                    engagement=new_engagement,
                )

        try:
            available_credits = get_available_credit_for_project(
                engagement.project.id, "both"
            )

            total_durations = credits_needed_for_an_engagement(engagement)
            needed_credits_present = True
            if total_durations is not None and available_credits is not None:
                needed_credits_present = (
                    available_credits >= total_durations
                    if total_durations != 0
                    else False
                )

            if not needed_credits_present:
                for hr in engagement.project.hr.all():
                    send_mail_templates(
                        "hr_emails/engagement_created_with_less_credit.html",
                        [hr.email],
                        "Meeraq | Insufficient Credits Remaining",
                        {
                            "name": hr.first_name.strip().title(),
                            "learner_name": engagement.learner.name.strip().title(),
                            "project_name": engagement.project.name.title(),
                        },
                        [],
                    )
            create_task(
                {
                    "task": "select_a_coach",
                    "engagement": engagement.id,
                    "priority": "high",
                    "status": "pending",
                    "remarks": [],
                },
                30,
            )
        except Exception as e:
            print(str(e))

        return Response(
            {"message": f"Engagement created successfully."},
            status=status.HTTP_200_OK,
        )

    except Exception as e:
        print(str(e))
        return Response(
            {"error": f"Failed to get data"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_available_credits(request, engagement_id):
    try:
        engagement = Engagement.objects.get(id=engagement_id)
        available_credits = get_available_credit_for_project(
            engagement.project.id, "both"
        )

        total_durations = credits_needed_for_an_engagement(engagement)

        if total_durations is not None and available_credits is not None:
            needed_credits_present = (
                available_credits >= total_durations if total_durations != 0 else False
            )
        else:
            needed_credits_present = False

        return Response(
            {
                "available_credits": available_credits,
                "total_durations": total_durations,
                "needed_credits_present": needed_credits_present,
            }
        )

    except Exception as e:
        print(str(e))
        return Response(
            {"error": "Failed to retrieve data"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_available_credit_of_project(request, project_id):
    try:

        available_credits = get_available_credit_for_project(project_id, "both")

        return Response(
            {
                "available_credits": available_credits,
            }
        )

    except Exception as e:
        print(str(e))
        return Response(
            {"error": "Failed to retrieve data"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_available_credits_without_project_structure(request, engagement_id):
    try:
        engagement = Engagement.objects.get(id=engagement_id)
        available_credits = get_available_credit_for_project(
            engagement.project.id, "both"
        )

        total_durations = engagement.project.duration_of_each_session

        if total_durations is not None and available_credits is not None:
            needed_credits_present = (
                available_credits >= total_durations if total_durations != 0 else False
            )
        else:
            needed_credits_present = False

        return Response(
            {
                "available_credits": available_credits,
                "total_durations": total_durations,
                "needed_credits_present": needed_credits_present,
            }
        )

    except Exception as e:
        print(str(e))
        return Response(
            {"error": "Failed to retrieve data"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_available_credits_of_all_cod_projects(request):
    try:
        projects = Project.objects.filter(project_type="COD")
        all_project_credit = {}
        for project in projects:

            engagement = Engagement.objects.filter(project=project).first()
            available_credits = get_available_credit_for_project(
                engagement.project.id, "both"
            )

            all_project_credit[project.id] = available_credits

        return Response(all_project_credit)

    except Exception as e:
        print(str(e))
        return Response(
            {"error": "Failed to retrieve data"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_sharable_emails(request):
    try:
        all_emails = set()
        all_coach_profiles = CoachProfileShare.objects.all()
        for profile in all_coach_profiles:
            all_emails.update(profile.emails)

        return Response({"emails": list(all_emails)})
    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("pmo")])
def coach_profile_sharable_email(request):
    try:
        emails = request.data.get("emails", [])
        if not emails:
            return Response({"message": "No emails provided."}, status=400)

        masked_coach_profile = request.data.get("masked_coach_profile", False)
        unique_id = request.data.get("unique_id", "")
        coaches = request.data.get("coaches", [])
        name = request.data.get("name", "")

        coach_profile_share = CoachProfileShare.objects.create(
            masked_coach_profile=masked_coach_profile,
            unique_id=unique_id,
            emails=emails,
            name=name,
        )

        for coach_id in coaches:
            coach_profile_share.coaches.add(coach_id)

        coach_profile_share.save()

        for email in emails:
            print("emails", email)
            print("id", unique_id)
            send_mail_templates(
                "coach_profile_share.html",
                [email],
                "Meeraq Coaching | Shared Coach Profiles!",
                {
                    "profiles_id": unique_id,
                },
                [],
            )
        return Response({"message": "Coach Profile Shared successfully."})

    except Exception as e:
        print(str(e))
        return Response({"message": str(e)}, status=500)


@api_view(["POST"])
@permission_classes([AllowAny])
def coach_profile_share_email_validation(request):
    try:
        unique_id = request.data.get("unique_id")
        email = request.data.get("email").strip().lower()
        coach_profile_share = CoachProfileShare.objects.get(unique_id=unique_id)
        coach_serializer = CoachProfileShareSerializer(coach_profile_share)
        if email in coach_serializer.data["emails"]:
            return Response(
                {"message": "Email is verified."},
                status=status.HTTP_200_OK,
            )
        else:
            return Response(
                {"error": "Email is not verified."},
                status=status.HTTP_400_BAD_REQUEST,
            )
    except CoachProfileShare.DoesNotExist:
        return Response(
            {"error": "Coach Profile Share not found."},
            status=status.HTTP_404_NOT_FOUND,
        )
    except Exception as e:
        print(str(e))
        return Response(
            {"error": "Failed to validate email."},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["GET"])
@permission_classes([AllowAny])
def get_coach_profile_sharing_form(request, unique_id):
    try:
        coach_profile_share = CoachProfileShare.objects.get(unique_id=unique_id)
        coach_serializer = CoachSerializer(coach_profile_share.coaches.all(), many=True)
        return Response(
            {
                "coaches": coach_serializer.data,
                "masked_coach_profile": coach_profile_share.masked_coach_profile,
            },
            status=status.HTTP_200_OK,
        )
    except Exception as e:
        print(str(e))
        return Response(
            {"error": "Failed to get Coach Profiles."},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["GET"])
@permission_classes([AllowAny])
def get_coach_profile_shared_with(request):
    try:
        # Retrieve all CoachProfileShare instances
        coach_profile_shares = CoachProfileShare.objects.all()

        # Dictionary to store coach emails mapping
        coach_email_mapping = {}

        # Iterate through each CoachProfileShare instance
        for profile_share in coach_profile_shares:
            emails = profile_share.emails
            coaches = profile_share.coaches.all()
            for coach in coaches:
                # If the coach is not already in the mapping, add them
                if coach.id not in coach_email_mapping:
                    coach_email_mapping[coach.id] = {"emails": set()}
                # Add the emails shared with this coach
                coach_email_mapping[coach.id]["emails"].update(emails)

        return Response(coach_email_mapping)

    except Exception as e:
        print(str(e))
        return Response(
            {"error": "Failed to get Coach Profiles."},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_coach_shared_links(request):
    try:
        # Retrieve all CoachProfileShare instances
        coach_profile_shares = CoachProfileShare.objects.all().order_by("-created_at")
        coach_profile_shares_serializer = CoachProfileShareSerializer(
            coach_profile_shares, many=True
        )
        return Response(coach_profile_shares_serializer.data)
    except Exception as e:
        print(str(e))
        return Response(
            {"error": "Failed to get coach profile shares."},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )
